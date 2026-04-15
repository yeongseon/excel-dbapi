from pathlib import Path

import pytest
from excel_dbapi.exceptions import DatabaseError
from openpyxl import Workbook

from excel_dbapi.engines.openpyxl.backend import OpenpyxlBackend
from excel_dbapi.executor import SharedExecutor
from excel_dbapi.parser import parse_sql


def _create_window_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "scores"
    sheet.append(["id", "team", "player", "points"])
    sheet.append([1, "A", "Ann", 10])
    sheet.append([2, "A", "Bob", 20])
    sheet.append([3, "A", "Cam", 20])
    sheet.append([4, "B", "Dan", 5])
    sheet.append([5, "B", "Eve", 15])
    workbook.save(path)


def _create_window_join_workbook(path: Path) -> None:
    workbook = Workbook()

    scores = workbook.active
    assert scores is not None
    scores.title = "scores"
    scores.append(["id", "team_id", "points"])
    scores.append([1, 1, 10])
    scores.append([2, 1, 20])
    scores.append([3, 2, 15])

    teams = workbook.create_sheet("teams")
    teams.append(["team_id", "team_name"])
    teams.append([1, "Alpha"])
    teams.append([2, "Beta"])

    workbook.save(path)


def test_row_number_over_order_by(tmp_path: Path) -> None:
    file_path = tmp_path / "window_row_number.xlsx"
    _create_window_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT id, ROW_NUMBER() OVER (ORDER BY points ASC) AS rn FROM scores ORDER BY id ASC"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(1, 2), (2, 4), (3, 5), (4, 1), (5, 3)]


def test_row_number_over_partition_by(tmp_path: Path) -> None:
    file_path = tmp_path / "window_row_number_partition.xlsx"
    _create_window_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT id, ROW_NUMBER() OVER (PARTITION BY team ORDER BY points DESC) AS rn "
        "FROM scores ORDER BY id ASC"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(1, 3), (2, 1), (3, 2), (4, 2), (5, 1)]


def test_rank_with_ties(tmp_path: Path) -> None:
    file_path = tmp_path / "window_rank.xlsx"
    _create_window_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT id, RANK() OVER (ORDER BY points DESC) AS rnk FROM scores ORDER BY id ASC"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(1, 4), (2, 1), (3, 1), (4, 5), (5, 3)]


def test_dense_rank_with_ties(tmp_path: Path) -> None:
    file_path = tmp_path / "window_dense_rank.xlsx"
    _create_window_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT id, DENSE_RANK() OVER (ORDER BY points DESC) AS drnk "
        "FROM scores ORDER BY id ASC"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(1, 3), (2, 1), (3, 1), (4, 4), (5, 2)]


def test_sum_over_partition(tmp_path: Path) -> None:
    file_path = tmp_path / "window_sum_partition.xlsx"
    _create_window_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT id, SUM(points) OVER (PARTITION BY team) AS team_sum "
        "FROM scores ORDER BY id ASC"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(1, 50.0), (2, 50.0), (3, 50.0), (4, 20.0), (5, 20.0)]


def test_avg_over_partition(tmp_path: Path) -> None:
    file_path = tmp_path / "window_avg_partition.xlsx"
    _create_window_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT id, AVG(points) OVER (PARTITION BY team) AS team_avg "
        "FROM scores ORDER BY id ASC"
    )
    results = SharedExecutor(engine).execute(parsed)

    expected = [
        (1, pytest.approx(50.0 / 3.0)),
        (2, pytest.approx(50.0 / 3.0)),
        (3, pytest.approx(50.0 / 3.0)),
        (4, pytest.approx(10.0)),
        (5, pytest.approx(10.0)),
    ]
    assert len(results.rows) == len(expected)
    for actual_row, expected_row in zip(results.rows, expected):
        assert actual_row[0] == expected_row[0]
        assert actual_row[1] == expected_row[1]


def test_count_filter_clause(tmp_path: Path) -> None:
    file_path = tmp_path / "aggregate_filter_count.xlsx"
    _create_window_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql("SELECT COUNT(*) FILTER (WHERE points >= 15) FROM scores")
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(3,)]


def test_sum_filter_clause(tmp_path: Path) -> None:
    file_path = tmp_path / "aggregate_filter_sum.xlsx"
    _create_window_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql("SELECT SUM(points) FILTER (WHERE team = 'A') FROM scores")
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(50.0,)]


def test_window_function_with_join(tmp_path: Path) -> None:
    file_path = tmp_path / "window_join.xlsx"
    _create_window_join_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT s.id, t.team_name, "
        "ROW_NUMBER() OVER (PARTITION BY t.team_name ORDER BY s.points DESC) AS rn "
        "FROM scores s JOIN teams t ON s.team_id = t.team_id "
        "ORDER BY s.id ASC"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(1, "Alpha", 2), (2, "Alpha", 1), (3, "Beta", 1)]


def test_multiple_window_functions_in_select(tmp_path: Path) -> None:
    file_path = tmp_path / "window_multiple.xlsx"
    _create_window_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT id, "
        "ROW_NUMBER() OVER (ORDER BY points DESC) AS rn, "
        "DENSE_RANK() OVER (ORDER BY points DESC) AS drnk, "
        "SUM(points) OVER (PARTITION BY team) AS team_sum "
        "FROM scores ORDER BY id ASC"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [
        (1, 4, 3, 50.0),
        (2, 1, 1, 50.0),
        (3, 2, 1, 50.0),
        (4, 5, 4, 20.0),
        (5, 3, 2, 20.0),
    ]


def test_window_function_in_order_by(tmp_path: Path) -> None:
    file_path = tmp_path / "window_order_by.xlsx"
    _create_window_workbook(file_path)

    with pytest.raises(DatabaseError, match="Unsupported SQL syntax"):
        parse_sql(
            "SELECT id FROM scores "
            "ORDER BY ROW_NUMBER() OVER (ORDER BY points DESC) ASC"
        )
