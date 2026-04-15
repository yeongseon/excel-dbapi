from pathlib import Path
from typing import Any, cast

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from excel_dbapi.connection import ExcelConnection
from excel_dbapi.cursor import ExcelCursor
from excel_dbapi.engines.result import ExecutionResult
from excel_dbapi.exceptions import (
    DatabaseError,
    InterfaceError,
    NotSupportedError,
    OperationalError,
    ProgrammingError,
)

def test_cursor_execute_and_fetchall():
    with ExcelConnection("tests/data/sample.xlsx") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM Sheet1")
        results = cursor.fetchall()
        assert isinstance(results, list)
        assert isinstance(results[0], tuple)
        assert cursor.description is not None
        assert cursor.rowcount == len(results)


def test_cursor_fetchone():
    with ExcelConnection("tests/data/sample.xlsx") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM Sheet1")
        row = cursor.fetchone()
        assert isinstance(row, tuple)


def test_cursor_closed():
    with ExcelConnection("tests/data/sample.xlsx") as conn:
        cursor = conn.cursor()
        cursor.close()
        with pytest.raises(Exception):
            cursor.execute("SELECT * FROM Sheet1")


def test_cursor_error_translation():
    with ExcelConnection("tests/data/sample.xlsx") as conn:
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError):
            cursor.execute("INVALID SQL")

        with pytest.raises(NotSupportedError):
            cursor.execute("SELECT * FROM Sheet1 WHERE id LIKE 1")


def _create_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id", "name"])
    ws.append([1, "Alice"])
    ws.append([2, "Bob"])
    wb.save(path)


def test_fetchmany_and_arraysize(tmp_path: Path) -> None:
    file_path = tmp_path / "sample.xlsx"
    _create_workbook(file_path)

    with ExcelConnection(str(file_path)) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id, name FROM Sheet1 ORDER BY id ASC")
        cursor.arraysize = 1
        first = cursor.fetchmany()
        assert first == [(1, "Alice")]
        second = cursor.fetchmany(2)
        assert second == [(2, "Bob")]


def test_executemany_autocommit(tmp_path: Path) -> None:
    file_path = tmp_path / "sample.xlsx"
    _create_workbook(file_path)

    with ExcelConnection(str(file_path), autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.executemany(
            "INSERT INTO Sheet1 (id, name) VALUES (?, ?)",
            [(3, "Cara"), (4, "Dane")],
        )
        assert cursor.rowcount == 2

    wb = load_workbook(file_path, data_only=True)
    rows = list(wb["Sheet1"].iter_rows(values_only=True))
    assert rows[-2:] == [(3, "Cara"), (4, "Dane")]


def test_select_with_params_and_limit(tmp_path: Path) -> None:
    file_path = tmp_path / "sample.xlsx"
    _create_workbook(file_path)

    with ExcelConnection(str(file_path)) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM Sheet1 WHERE id >= ? LIMIT ?", (1, 1))
        assert cursor.fetchall() == [(1,)]


def test_pandas_insert_column_mismatch(tmp_path: Path) -> None:
    file_path = tmp_path / "sample.xlsx"
    df = pd.DataFrame([{"id": 1, "name": "Alice"}])
    df.to_excel(file_path, index=False, sheet_name="Sheet1")

    with ExcelConnection(str(file_path), engine="pandas") as conn:
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError):
            cursor.execute("INSERT INTO Sheet1 (id) VALUES (1, 'A')")



def _create_r7_workbook(
    path: Path, headers: list[object], rows: list[list[object]]
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Sheet1"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()

def test_pandas_lastrowid_matches_openpyxl(tmp_path: Path) -> None:
    openpyxl_file = tmp_path / "openpyxl-lastrowid.xlsx"
    pandas_file = tmp_path / "pandas-lastrowid.xlsx"
    _create_r7_workbook(openpyxl_file, ["id", "name"], [[1, "Alice"]])
    _create_r7_workbook(pandas_file, ["id", "name"], [[1, "Alice"]])

    with ExcelConnection(str(openpyxl_file), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute("INSERT INTO Sheet1 VALUES (2, 'Bob')")
        openpyxl_lastrowid = cursor.lastrowid

    with ExcelConnection(str(pandas_file), engine="pandas") as conn:
        cursor = conn.cursor()
        cursor.execute("INSERT INTO Sheet1 VALUES (2, 'Bob')")
        pandas_lastrowid = cursor.lastrowid

    assert openpyxl_lastrowid == 3
    assert pandas_lastrowid == openpyxl_lastrowid



def _create_r10_workbook(
    path: Path,
    *,
    headers: list[object],
    rows: list[list[object]],
    sheet_name: str,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = sheet_name
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()

def test_fetchone_raises_after_connection_close(tmp_path: Path) -> None:
    file_path = tmp_path / "closed-connection-fetch.xlsx"
    _create_r10_workbook(
        file_path,
        headers=["id", "name"],
        rows=[[1, "Alice"]],
        sheet_name="people",
    )

    conn = ExcelConnection(str(file_path), engine="openpyxl")
    cursor = conn.cursor()
    cursor.execute("SELECT id, name FROM people")
    conn.close()

    with pytest.raises(InterfaceError, match="Cannot operate on a closed connection"):
        cursor.fetchone()

def test_failed_execute_clears_prior_result_set(tmp_path: Path) -> None:
    file_path = tmp_path / "stale-results.xlsx"
    _create_r10_workbook(
        file_path,
        headers=["id", "name"],
        rows=[[1, "Alice"], [2, "Bob"]],
        sheet_name="people",
    )

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM people ORDER BY id")
        assert cursor.fetchone() == (1,)

        with pytest.raises(ProgrammingError):
            cursor.execute("INVALID SQL")

        with pytest.raises(ProgrammingError, match="No result set"):
            cursor.fetchone()



def _create_r9_workbook(
    path: Path,
    *,
    headers: list[object] | None = None,
    rows: list[list[object]] | None = None,
    sheet_name: str = "Sheet1",
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = sheet_name
    if headers is not None:
        sheet.append(headers)
    for row in rows or []:
        sheet.append(row)
    workbook.save(path)
    workbook.close()

def test_fetch_methods_raise_before_execute(tmp_path: Path) -> None:
    file_path = tmp_path / "fetch-before-execute.xlsx"
    _create_r9_workbook(file_path, headers=["id"], rows=[[1]])

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="No result set"):
            cursor.fetchone()
        with pytest.raises(ProgrammingError, match="No result set"):
            cursor.fetchall()
        with pytest.raises(ProgrammingError, match="No result set"):
            cursor.fetchmany()

def test_fetch_methods_raise_after_update_and_ddl(tmp_path: Path) -> None:
    file_path = tmp_path / "fetch-after-write.xlsx"
    _create_r9_workbook(file_path, headers=["id", "name"], rows=[[1, "Alice"]])

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()

        cursor.execute("UPDATE Sheet1 SET name = 'Bob' WHERE id = 1")
        with pytest.raises(ProgrammingError, match="No result set"):
            cursor.fetchall()

        cursor.execute("CREATE TABLE t (id INTEGER)")
        with pytest.raises(ProgrammingError, match="No result set"):
            cursor.fetchmany(1)



def _create_round11_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Sheet1"
    sheet.append(["id", "name"])
    sheet.append([1, "Alice"])
    sheet.append([2, "Bob"])

    table = workbook.create_sheet("t")
    table.append(["id", "a", "b", "c"])
    table.append([1, 10, 0, None])
    table.append([2, "alice", 0, None])

    workbook.save(path)

def test_executemany_failure_clears_prior_select_results(tmp_path: Path) -> None:
    file_path = tmp_path / "round11_cursor_state.xlsx"
    _create_round11_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=False) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id, name FROM Sheet1 ORDER BY id")
        assert cursor.fetchone() == (1, "Alice")

        with pytest.raises(ProgrammingError):
            cursor.executemany(
                "INSERT INTO Sheet1 (id, name) VALUES (?, ?)",
                [(3, "Cara"), (4,)],
            )

        with pytest.raises(ProgrammingError, match="No result set"):
            cursor.fetchone()



def _create_round14_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "t"
    sheet.append(["id", "name"])
    sheet.append([1, "a"])
    sheet.append([2, "b"])
    sheet.append([3, "c"])
    workbook.save(path)

def test_delete_without_placeholders_rejects_extra_parameters(tmp_path: Path) -> None:
    file_path = tmp_path / "round14_delete_extra_params.xlsx"
    _create_round14_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        with pytest.raises(
            ProgrammingError, match="Too many parameters for placeholders"
        ):
            cursor.execute("DELETE FROM t", (123,))

def test_drop_without_placeholders_rejects_extra_parameters(tmp_path: Path) -> None:
    file_path = tmp_path / "round14_drop_extra_params.xlsx"
    _create_round14_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        with pytest.raises(
            ProgrammingError, match="Too many parameters for placeholders"
        ):
            cursor.execute("DROP TABLE t", (123,))

def test_create_without_placeholders_rejects_extra_parameters(tmp_path: Path) -> None:
    file_path = tmp_path / "round14_create_extra_params.xlsx"
    _create_round14_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        with pytest.raises(
            ProgrammingError, match="Too many parameters for placeholders"
        ):
            cursor.execute("CREATE TABLE u (id INTEGER)", (1,))



def _create_r16_workbook(
    path: Path, sheet: str, headers: list[object], rows: list[list[object]]
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = sheet
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()

def test_lastrowid_is_cleared_after_failed_execute(tmp_path: Path) -> None:
    file_path = tmp_path / "lastrowid-reset.xlsx"
    _create_r16_workbook(file_path, "Sheet1", ["id", "name"], [[1, "Alice"]])

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute("INSERT INTO Sheet1 VALUES (2, 'Bob')")
        assert cursor.lastrowid is not None

        with pytest.raises(ProgrammingError):
            cursor.execute("SELECT * FROM MissingSheet")

        assert cursor.lastrowid is None



def _make_xlsx(
    path: Path,
    sheet: str = "users",
    headers: list[str] | None = None,
    rows: list[list[Any]] | None = None,
) -> str:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet
    for h in headers or ["id", "name"]:
        pass
    ws.append(headers or ["id", "name"])
    for row in rows or []:
        ws.append(row)
    fpath = str(path)
    wb.save(fpath)
    wb.close()
    return fpath

class TestExceptionMapping:
    """Fix 5: Cursor wraps raw exceptions into PEP 249 hierarchy."""

    def test_value_error_becomes_programming_error(self, tmp_path: Path) -> None:
        """ValueError from executor → ProgrammingError."""
        fpath = _make_xlsx(tmp_path / "test.xlsx")
        with ExcelConnection(fpath, engine="openpyxl") as conn:
            cursor = conn.cursor()
            with pytest.raises(ProgrammingError):
                cursor.execute("INVALID SQL GIBBERISH")

class TestExceptionMappingDirect:
    """Gap 2: Direct tests for each exception type mapping in execute() and executemany()."""

    def _make_conn_with_raising_executor(
        self, tmp_path: Path, exc: Exception
    ) -> ExcelConnection:
        """Create a connection whose executor raises the given exception."""
        fpath = _make_xlsx(tmp_path / "test.xlsx")
        conn = ExcelConnection(fpath, engine="openpyxl", autocommit=True)

        def raising_execute(query: str, params: Any = None) -> Any:
            raise exc

        conn._executor.execute_with_params = raising_execute  # type: ignore[assignment]
        return conn

    def test_key_error_maps_to_programming_error_execute(self, tmp_path: Path) -> None:
        conn = self._make_conn_with_raising_executor(tmp_path, KeyError("missing"))
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="missing"):
            cursor.execute("SELECT * FROM users")
        conn.close()

    def test_type_error_maps_to_programming_error_execute(self, tmp_path: Path) -> None:
        conn = self._make_conn_with_raising_executor(tmp_path, TypeError("bad type"))
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="bad type"):
            cursor.execute("SELECT * FROM users")
        conn.close()

    def test_index_error_maps_to_programming_error_execute(
        self, tmp_path: Path
    ) -> None:
        conn = self._make_conn_with_raising_executor(
            tmp_path, IndexError("out of range")
        )
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="out of range"):
            cursor.execute("SELECT * FROM users")
        conn.close()

    def test_os_error_maps_to_operational_error_execute(self, tmp_path: Path) -> None:
        conn = self._make_conn_with_raising_executor(tmp_path, OSError("disk full"))
        cursor = conn.cursor()
        with pytest.raises(OperationalError, match="disk full"):
            cursor.execute("SELECT * FROM users")
        conn.close()

    def test_generic_exception_maps_to_database_error_execute(
        self, tmp_path: Path
    ) -> None:
        conn = self._make_conn_with_raising_executor(
            tmp_path, RuntimeError("unexpected")
        )
        cursor = conn.cursor()
        with pytest.raises(DatabaseError, match="unexpected"):
            cursor.execute("SELECT * FROM users")
        conn.close()

    def test_database_error_passes_through_execute(self, tmp_path: Path) -> None:
        conn = self._make_conn_with_raising_executor(
            tmp_path, ProgrammingError("bad sql")
        )
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="bad sql"):
            cursor.execute("SELECT * FROM users")
        conn.close()

    def test_key_error_maps_to_programming_error_executemany(
        self, tmp_path: Path
    ) -> None:
        fpath = _make_xlsx(tmp_path / "test.xlsx")
        conn = ExcelConnection(fpath, engine="openpyxl", autocommit=True)
        original = conn._executor.execute_with_params
        call_count = 0

        def raising_on_second(query: str, params: Any = None) -> Any:
            nonlocal call_count
            call_count += 1
            if call_count >= 2:
                raise KeyError("missing key")
            return original(query, params)

        conn._executor.execute_with_params = raising_on_second  # type: ignore[assignment]
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="missing key"):
            cursor.executemany("INSERT INTO users VALUES (?, ?)", [(1, "a"), (2, "b")])
        conn.close()

    def test_os_error_maps_to_operational_error_executemany(
        self, tmp_path: Path
    ) -> None:
        fpath = _make_xlsx(tmp_path / "test.xlsx")
        conn = ExcelConnection(fpath, engine="openpyxl", autocommit=True)

        def raising_execute(query: str, params: Any = None) -> Any:
            raise OSError("permission denied")

        conn._executor.execute_with_params = raising_execute  # type: ignore[assignment]
        cursor = conn.cursor()
        with pytest.raises(OperationalError, match="permission denied"):
            cursor.executemany("INSERT INTO users VALUES (?, ?)", [(1, "a")])
        conn.close()

    def test_generic_exception_maps_to_database_error_executemany(
        self, tmp_path: Path
    ) -> None:
        fpath = _make_xlsx(tmp_path / "test.xlsx")
        conn = ExcelConnection(fpath, engine="openpyxl", autocommit=True)

        def raising_execute(query: str, params: Any = None) -> Any:
            raise RuntimeError("boom")

        conn._executor.execute_with_params = raising_execute  # type: ignore[assignment]
        cursor = conn.cursor()
        with pytest.raises(DatabaseError, match="boom"):
            cursor.executemany("INSERT INTO users VALUES (?, ?)", [(1, "a")])
        conn.close()

class TestExecutemanyMidBatchRestore:
    """Gap 3: Mid-batch executemany failure restores snapshot after partial mutation."""

    def test_mid_batch_failure_restores_snapshot(self, tmp_path: Path) -> None:
        """After 1 successful INSERT, a failure in batch 2 restores original state."""
        fpath = _make_xlsx(tmp_path / "test.xlsx", rows=[[1, "Original"]])
        conn = ExcelConnection(fpath, engine="openpyxl", autocommit=False)
        original = conn._executor.execute_with_params
        call_count = 0

        def fail_on_second(query: str, params: Any = None) -> Any:
            nonlocal call_count
            call_count += 1
            if call_count == 2:
                raise ValueError("bad value")
            return original(query, params)

        conn._executor.execute_with_params = fail_on_second  # type: ignore[assignment]
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="bad value"):
            cursor.executemany(
                "INSERT INTO users VALUES (?, ?)",
                [(2, "Second"), (3, "Third")],
            )
        # After error + snapshot restore, only original row should exist
        conn._executor.execute_with_params = original  # type: ignore[assignment]
        result = conn.execute("SELECT * FROM users")
        assert len(result.rows) == 1
        assert result.rows[0] == (1, "Original")
        conn.close()

class TestExecutemanyAutocommitRestore:
    """Regression: executemany with autocommit=True must also restore on failure."""

    def test_autocommit_true_mid_batch_failure_restores(self, tmp_path: Path) -> None:
        """Under autocommit=True, a mid-batch failure restores the snapshot.

        Previously snapshot was only taken when autocommit=False, leaving
        partial mutations in memory under autocommit=True.
        """
        fpath = _make_xlsx(tmp_path / "test.xlsx", rows=[[1, "Original"]])
        conn = ExcelConnection(fpath, engine="openpyxl", autocommit=True)
        original = conn._executor.execute_with_params
        call_count = 0

        def fail_on_second(query: str, params: Any = None) -> Any:
            nonlocal call_count
            call_count += 1
            if call_count == 2:
                raise ValueError("bad value")
            return original(query, params)

        conn._executor.execute_with_params = fail_on_second  # type: ignore[assignment]
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="bad value"):
            cursor.executemany(
                "INSERT INTO users VALUES (?, ?)",
                [(2, "Second"), (3, "Third")],
            )
        # After error + snapshot restore, only original row should exist
        conn._executor.execute_with_params = original  # type: ignore[assignment]
        result = conn.execute("SELECT * FROM users")
        assert len(result.rows) == 1
        assert result.rows[0] == (1, "Original")
        conn.close()



def test_cursor_paths_for_executemany_and_fetch() -> None:
    class FakeEngine:
        supports_transactions = True
        readonly = False

        def __init__(self) -> None:
            self.restored: Any = None
            self.saved = False

        def snapshot(self) -> str:
            return "snap"

        def restore(self, snapshot: Any) -> None:
            self.restored = snapshot

        def save(self) -> None:
            self.saved = True

    class FakeExecutor:
        def execute_with_params(
            self, query: str, params: tuple[Any, ...] | None = None
        ) -> ExecutionResult:
            raise NotImplementedError("not supported")

    class FakeConnection:
        def __init__(self) -> None:
            self.closed = False
            self.autocommit = False
            self.engine = FakeEngine()
            self._snapshot = None
            self._executor = FakeExecutor()

        def execute(
            self, query: str, params: tuple[Any, ...] | None = None
        ) -> ExecutionResult:
            raise NotImplementedError("not supported")

        def executemany(
            self, query: str, seq_of_params: Any
        ) -> ExecutionResult:
            raise NotImplementedError("not supported")

        def _finalize_autocommit(self, action: str) -> None:
            pass

    cursor = ExcelCursor(cast(Any, FakeConnection()))
    with pytest.raises(NotSupportedError):
        cursor.executemany("SELECT 1", [(1,)])
    with pytest.raises(ProgrammingError, match="No result set"):
        cursor.fetchone()
    with pytest.raises(ProgrammingError, match="No result set"):
        cursor.fetchmany(0)
