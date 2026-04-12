"""Tests for PEP 249 compliance details."""

import pytest


class TestModuleLevelExports:
    """PEP 249 §1 requires module-level constants and exception exports."""

    def test_apilevel(self) -> None:
        import excel_dbapi

        assert excel_dbapi.apilevel == "2.0"

    def test_threadsafety(self) -> None:
        import excel_dbapi

        assert excel_dbapi.threadsafety == 1

    def test_paramstyle(self) -> None:
        import excel_dbapi

        assert excel_dbapi.paramstyle == "qmark"

    @pytest.mark.parametrize(
        "name",
        [
            "Error",
            "Warning",
            "InterfaceError",
            "DatabaseError",
            "DataError",
            "OperationalError",
            "IntegrityError",
            "InternalError",
            "ProgrammingError",
            "NotSupportedError",
        ],
    )
    def test_exception_exported(self, name: str) -> None:
        import excel_dbapi

        cls = getattr(excel_dbapi, name)
        assert cls is not None
        # Must be importable as a class
        assert isinstance(cls, type)

    def test_exception_hierarchy(self) -> None:
        """PEP 249 exception inheritance must be correct."""
        import excel_dbapi

        assert issubclass(excel_dbapi.Warning, Exception)
        assert issubclass(excel_dbapi.Error, Exception)
        assert issubclass(excel_dbapi.InterfaceError, excel_dbapi.Error)
        assert issubclass(excel_dbapi.DatabaseError, excel_dbapi.Error)
        assert issubclass(excel_dbapi.DataError, excel_dbapi.DatabaseError)
        assert issubclass(excel_dbapi.OperationalError, excel_dbapi.DatabaseError)
        assert issubclass(excel_dbapi.IntegrityError, excel_dbapi.DatabaseError)
        assert issubclass(excel_dbapi.InternalError, excel_dbapi.DatabaseError)
        assert issubclass(excel_dbapi.ProgrammingError, excel_dbapi.DatabaseError)
        assert issubclass(excel_dbapi.NotSupportedError, excel_dbapi.DatabaseError)

    def test_connect_function(self) -> None:
        import excel_dbapi

        assert callable(excel_dbapi.connect)

    def test_all_contains_exceptions(self) -> None:
        import excel_dbapi

        for name in [
            "Error",
            "Warning",
            "InterfaceError",
            "DatabaseError",
            "DataError",
            "OperationalError",
            "IntegrityError",
            "InternalError",
            "ProgrammingError",
            "NotSupportedError",
        ]:
            assert name in excel_dbapi.__all__, f"{name} missing from __all__"


class TestCursorPEP249Stubs:
    """PEP 249 requires setinputsizes / setoutputsize on Cursor."""

    @pytest.fixture
    def cursor(self, tmp_path):
        from openpyxl import Workbook

        path = str(tmp_path / "test.xlsx")
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Sheet1"
        ws.append(["id", "name"])
        wb.save(path)

        from excel_dbapi.connection import ExcelConnection

        conn = ExcelConnection(path)
        cur = conn.cursor()
        yield cur
        conn.close()

    def test_setinputsizes_exists_and_returns_none(self, cursor) -> None:
        result = cursor.setinputsizes([10, 20])
        assert result is None

    def test_setoutputsize_exists_and_returns_none(self, cursor) -> None:
        result = cursor.setoutputsize(1000)
        assert result is None

    def test_setoutputsize_with_column(self, cursor) -> None:
        result = cursor.setoutputsize(1000, column=0)
        assert result is None

    def test_setinputsizes_raises_on_closed_cursor(self, cursor) -> None:
        from excel_dbapi.exceptions import InterfaceError

        cursor.close()
        with pytest.raises(InterfaceError):
            cursor.setinputsizes([])

    def test_setoutputsize_raises_on_closed_cursor(self, cursor) -> None:
        from excel_dbapi.exceptions import InterfaceError

        cursor.close()
        with pytest.raises(InterfaceError):
            cursor.setoutputsize(100)


class TestWorkbookBackendContract:
    def test_raises_not_implemented(self) -> None:
        from typing import Any

        from excel_dbapi.engines.base import TableData, WorkbookBackend

        class Stub(WorkbookBackend):
            def load(self) -> None:
                pass

            def save(self) -> None:
                pass

            def snapshot(self) -> Any:
                return None

            def restore(self, snapshot: Any) -> None:
                pass

            def list_sheets(self) -> list[str]:
                return ["Sheet1"]

            def read_sheet(self, sheet_name: str) -> TableData:
                return TableData(headers=["id"], rows=[[1]])

            def write_sheet(self, sheet_name: str, data: TableData) -> None:
                pass

            def append_row(self, sheet_name: str, row: list[Any]) -> int:
                return 1

            def create_sheet(self, name: str, headers: list[str]) -> None:
                pass

            def drop_sheet(self, name: str) -> None:
                pass

        engine = Stub("dummy.xlsx")
        assert engine.list_sheets() == ["Sheet1"]


class TestConnectSanitizeFormulasParam:
    """connect() and ExcelConnection accept sanitize_formulas parameter."""

    @pytest.fixture
    def xlsx_path(self, tmp_path):
        from openpyxl import Workbook

        path = str(tmp_path / "test.xlsx")
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Sheet1"
        ws.append(["id", "name"])
        wb.save(path)
        return path

    def test_connect_default_sanitize(self, xlsx_path: str) -> None:
        import excel_dbapi

        conn = excel_dbapi.connect(xlsx_path)
        assert conn.engine.sanitize_formulas is True
        conn.close()

    def test_connect_disable_sanitize(self, xlsx_path: str) -> None:
        import excel_dbapi

        conn = excel_dbapi.connect(xlsx_path, sanitize_formulas=False)
        assert conn.engine.sanitize_formulas is False
        conn.close()

    def test_connection_default_sanitize(self, xlsx_path: str) -> None:
        from excel_dbapi.connection import ExcelConnection

        conn = ExcelConnection(xlsx_path)
        assert conn.engine.sanitize_formulas is True
        conn.close()

    def test_connection_disable_sanitize(self, xlsx_path: str) -> None:
        from excel_dbapi.connection import ExcelConnection

        conn = ExcelConnection(xlsx_path, sanitize_formulas=False)
        assert conn.engine.sanitize_formulas is False
        conn.close()
