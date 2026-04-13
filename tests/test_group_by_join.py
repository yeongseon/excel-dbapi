from pathlib import Path

import pytest
from openpyxl import Workbook

from excel_dbapi.connection import ExcelConnection
from excel_dbapi.exceptions import ProgrammingError


def _create_group_join_workbook(path: Path) -> None:
    workbook = Workbook()

    employees = workbook.active
    assert employees is not None
    employees.title = "employees"
    employees.append(["id", "dept", "region_id", "name"])
    employees.append([1, "Eng", 10, "Ann"])
    employees.append([2, "Eng", 10, "Ben"])
    employees.append([3, "Sales", 20, "Cam"])
    employees.append([4, "Sales", 30, "Dan"])
    employees.append([5, "HR", None, "Eve"])

    orders = workbook.create_sheet("orders")
    orders.append(["id", "emp_id", "amount"])
    orders.append([101, 1, 100])
    orders.append([102, 1, 150])
    orders.append([103, 2, 200])
    orders.append([104, 3, 50])
    orders.append([105, 3, None])
    orders.append([106, 4, 70])
    orders.append([107, 99, 999])

    regions = workbook.create_sheet("regions")
    regions.append(["id", "region_name"])
    regions.append([10, "North"])
    regions.append([20, "South"])
    regions.append([30, "West"])

    workbook.save(path)


def _run_query(path: Path, query: str) -> list[tuple[object, ...]]:
    with ExcelConnection(str(path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute(query)
        return cursor.fetchall()


def test_group_by_inner_join_count_star(tmp_path: Path) -> None:
    file_path = tmp_path / "group_by_join_count.xlsx"
    _create_group_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT t1.dept, COUNT(*) FROM employees t1 "
        "INNER JOIN orders t2 ON t1.id = t2.emp_id "
        "GROUP BY t1.dept ORDER BY t1.dept",
    )
    assert rows == [("Eng", 3), ("Sales", 3)]


def test_group_by_join_sum_qualified_column(tmp_path: Path) -> None:
    file_path = tmp_path / "group_by_join_sum.xlsx"
    _create_group_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT t1.dept, SUM(t2.amount) FROM employees t1 "
        "JOIN orders t2 ON t1.id = t2.emp_id "
        "GROUP BY t1.dept ORDER BY t1.dept",
    )
    assert rows == [("Eng", 450), ("Sales", 120)]


def test_group_by_join_having(tmp_path: Path) -> None:
    file_path = tmp_path / "group_by_join_having.xlsx"
    _create_group_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT t1.dept, COUNT(*) FROM employees t1 "
        "JOIN orders t2 ON t1.id = t2.emp_id "
        "GROUP BY t1.dept HAVING COUNT(*) > 2 ORDER BY t1.dept",
    )
    assert rows == [("Eng", 3), ("Sales", 3)]


def test_group_by_join_order_by_aggregate(tmp_path: Path) -> None:
    file_path = tmp_path / "group_by_join_order_by_aggregate.xlsx"
    _create_group_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT t1.dept, SUM(t2.amount) FROM employees t1 "
        "JOIN orders t2 ON t1.id = t2.emp_id "
        "GROUP BY t1.dept ORDER BY SUM(t2.amount) DESC",
    )
    assert rows == [("Eng", 450), ("Sales", 120)]


def test_group_by_join_order_by_qualified_column(tmp_path: Path) -> None:
    file_path = tmp_path / "group_by_join_order_by_qualified.xlsx"
    _create_group_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT t1.dept, COUNT(*) FROM employees t1 "
        "JOIN orders t2 ON t1.id = t2.emp_id "
        "GROUP BY t1.dept ORDER BY t1.dept DESC",
    )
    assert rows == [("Sales", 3), ("Eng", 3)]


def test_group_by_join_multiple_aggregates(tmp_path: Path) -> None:
    file_path = tmp_path / "group_by_join_multi_aggregates.xlsx"
    _create_group_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT t1.dept, COUNT(*), SUM(t2.amount), AVG(t2.amount) FROM employees t1 "
        "JOIN orders t2 ON t1.id = t2.emp_id "
        "GROUP BY t1.dept ORDER BY t1.dept",
    )
    assert rows == [("Eng", 3, 450, 150), ("Sales", 3, 120, 60)]


def test_group_by_join_multiple_group_columns(tmp_path: Path) -> None:
    file_path = tmp_path / "group_by_join_multi_group_cols.xlsx"
    _create_group_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT t1.dept, t1.region_id, COUNT(*) FROM employees t1 "
        "JOIN orders t2 ON t1.id = t2.emp_id "
        "GROUP BY t1.dept, t1.region_id ORDER BY t1.dept, t1.region_id",
    )
    assert rows == [("Eng", 10, 3), ("Sales", 20, 2), ("Sales", 30, 1)]


def test_group_by_left_join_includes_null_groups(tmp_path: Path) -> None:
    file_path = tmp_path / "group_by_left_join_nulls.xlsx"
    _create_group_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT t1.dept, COUNT(t2.amount), SUM(t2.amount) FROM employees t1 "
        "LEFT JOIN orders t2 ON t1.id = t2.emp_id "
        "GROUP BY t1.dept ORDER BY t1.dept",
    )
    assert rows == [("Eng", 3, 450), ("HR", 0, None), ("Sales", 2, 120)]


def test_three_table_join_group_by(tmp_path: Path) -> None:
    file_path = tmp_path / "group_by_three_table_join.xlsx"
    _create_group_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT t3.region_name, SUM(t2.amount) FROM employees t1 "
        "JOIN orders t2 ON t1.id = t2.emp_id "
        "JOIN regions t3 ON t1.region_id = t3.id "
        "GROUP BY t3.region_name ORDER BY t3.region_name",
    )
    assert rows == [("North", 450), ("South", 50), ("West", 70)]


def test_group_by_join_limit_offset(tmp_path: Path) -> None:
    file_path = tmp_path / "group_by_join_limit_offset.xlsx"
    _create_group_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT t1.dept, COUNT(*) FROM employees t1 "
        "JOIN orders t2 ON t1.id = t2.emp_id "
        "GROUP BY t1.dept ORDER BY t1.dept LIMIT 1 OFFSET 1",
    )
    assert rows == [("Sales", 3)]


def test_group_by_join_rejects_bare_group_by_column(tmp_path: Path) -> None:
    file_path = tmp_path / "group_by_join_reject_bare_group.xlsx"
    _create_group_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        with pytest.raises(
            ProgrammingError,
            match="GROUP BY in JOIN queries requires qualified column names",
        ):
            cursor.execute(
                "SELECT t1.dept, COUNT(*) FROM employees t1 "
                "JOIN orders t2 ON t1.id = t2.emp_id "
                "GROUP BY dept"
            )


def test_group_by_join_rejects_bare_aggregate_argument(tmp_path: Path) -> None:
    file_path = tmp_path / "group_by_join_reject_bare_aggregate.xlsx"
    _create_group_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        with pytest.raises(
            ProgrammingError,
            match=r"Aggregate arguments in JOIN queries must be qualified column names or \*",
        ):
            cursor.execute(
                "SELECT t1.dept, SUM(amount) FROM employees t1 "
                "JOIN orders t2 ON t1.id = t2.emp_id "
                "GROUP BY t1.dept"
            )


def test_group_by_join_rejects_select_star(tmp_path: Path) -> None:
    file_path = tmp_path / "group_by_join_reject_star.xlsx"
    _create_group_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        with pytest.raises(
            ProgrammingError,
            match=r"SELECT \* is not supported with GROUP BY or aggregate functions",
        ):
            cursor.execute(
                "SELECT * FROM employees t1 "
                "JOIN orders t2 ON t1.id = t2.emp_id "
                "GROUP BY t1.dept"
            )


def test_group_by_join_table_qualified_columns(tmp_path: Path) -> None:
    """Regression: GROUP BY using table name instead of alias must work."""
    file_path = tmp_path / "group_by_join_table_qualified.xlsx"
    _create_group_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT employees.dept, COUNT(*) FROM employees e "
        "INNER JOIN orders o ON e.id = o.emp_id "
        "GROUP BY employees.dept ORDER BY employees.dept",
    )
    assert rows == [("Eng", 3), ("Sales", 3)]


def test_group_by_join_table_qualified_aggregate_arg(tmp_path: Path) -> None:
    """Regression: Aggregate arg using table name instead of alias must work."""
    file_path = tmp_path / "group_by_join_table_qualified_agg.xlsx"
    _create_group_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT e.dept, SUM(orders.amount) FROM employees e "
        "JOIN orders o ON e.id = o.emp_id "
        "GROUP BY e.dept ORDER BY e.dept",
    )
    assert rows == [("Eng", 450), ("Sales", 120)]


def test_group_by_join_having_table_qualified(tmp_path: Path) -> None:
    """Regression: HAVING with table-qualified aggregate arg."""
    file_path = tmp_path / "group_by_join_having_table_qual.xlsx"
    _create_group_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT employees.dept, SUM(orders.amount) FROM employees e "
        "JOIN orders o ON e.id = o.emp_id "
        "GROUP BY employees.dept HAVING SUM(orders.amount) > 200 "
        "ORDER BY employees.dept",
    )
    assert rows == [("Eng", 450)]


def test_group_by_join_order_by_table_qualified(tmp_path: Path) -> None:
    """Regression: ORDER BY with table-qualified column in GROUP BY context."""
    file_path = tmp_path / "group_by_join_order_table_qual.xlsx"
    _create_group_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT employees.dept, COUNT(*) FROM employees e "
        "JOIN orders o ON e.id = o.emp_id "
        "GROUP BY employees.dept ORDER BY employees.dept DESC",
    )
    assert rows == [("Sales", 3), ("Eng", 3)]


def test_group_by_join_mixed_alias_and_table_name(tmp_path: Path) -> None:
    """Regression: Mix of alias-qualified and table-qualified refs."""
    file_path = tmp_path / "group_by_join_mixed.xlsx"
    _create_group_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT employees.dept, SUM(o.amount) FROM employees e "
        "JOIN orders o ON e.id = o.emp_id "
        "GROUP BY employees.dept ORDER BY SUM(o.amount) DESC",
    )
    assert rows == [("Eng", 450), ("Sales", 120)]
