"""Tests for Unicode bare identifiers and double-quoted column identifiers."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
import pytest

from excel_dbapi.connection import ExcelConnection
from excel_dbapi.engines.openpyxl.backend import OpenpyxlBackend
from excel_dbapi.engines.result import ExecutionResult
from excel_dbapi.executor import SharedExecutor
from excel_dbapi.parser import parse_sql


def _create_test_workbook(path: Path) -> None:
    """Create a workbook with Unicode and spaced column headers."""
    wb = Workbook()

    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id", "이름", "Full Name", "나이", "Department Name"])
    ws.append([1, "홍길동", "Alice Kim", 30, "Engineering"])
    ws.append([2, "김철수", "Bob Park", 25, "Sales"])
    ws.append([3, "이영희", "Carol Lee", 35, "Engineering"])

    emp = wb.create_sheet("emp")
    emp.append(["id", "이름", "부서"])
    emp.append([1, "홍길동", "개발"])
    emp.append([2, "김철수", "영업"])
    emp.append([3, "이영희", "개발"])
    emp.append([4, "박지민", "영업"])

    wb.save(path)


def _create_dml_workbook(path: Path) -> None:
    """Create a fresh workbook for DML tests (INSERT/UPDATE/DELETE)."""
    wb = Workbook()

    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id", "이름", "Full Name", "나이"])
    ws.append([1, "홍길동", "Alice Kim", 30])
    ws.append([2, "김철수", "Bob Park", 25])

    wb.save(path)


def _execute(path: Path, sql: str) -> ExecutionResult:
    engine = OpenpyxlBackend(str(path))
    parsed = parse_sql(sql)
    return SharedExecutor(engine).execute(parsed)


# ---------------------------------------------------------------------------
# A. Unicode Bare Identifiers — Parser
# ---------------------------------------------------------------------------


def test_parse_unicode_columns_korean() -> None:
    parsed = parse_sql("SELECT 이름, 나이 FROM Sheet1")
    assert parsed["columns"] == ["이름", "나이"]
    assert parsed["table"] == "Sheet1"


def test_parse_unicode_columns_cjk() -> None:
    parsed = parse_sql("SELECT 名前 FROM Sheet1")
    assert parsed["columns"] == ["名前"]


def test_parse_unicode_mixed_ascii() -> None:
    parsed = parse_sql("SELECT id, 이름 FROM Sheet1")
    assert parsed["columns"] == ["id", "이름"]


def test_parse_unicode_where() -> None:
    parsed = parse_sql("SELECT * FROM Sheet1 WHERE 이름 = '홍길동'")
    assert parsed["where"]["conditions"][0]["column"] == "이름"
    assert parsed["where"]["conditions"][0]["value"] == "홍길동"


def test_parse_unicode_order_by() -> None:
    parsed = parse_sql("SELECT 이름 FROM Sheet1 ORDER BY 나이 DESC")
    assert parsed["order_by"][0]["column"] == "나이"
    assert parsed["order_by"][0]["direction"] == "DESC"


def test_parse_unicode_group_by() -> None:
    parsed = parse_sql("SELECT 부서, COUNT(*) FROM Sheet1 GROUP BY 부서")
    assert parsed["group_by"] == ["부서"]


def test_parse_unicode_aggregate() -> None:
    parsed = parse_sql("SELECT SUM(나이) FROM Sheet1")
    assert parsed["columns"] == [{'type': 'aggregate', 'func': 'SUM', 'arg': '나이'}]


def test_parse_unicode_insert() -> None:
    parsed = parse_sql("INSERT INTO Sheet1 (이름, 나이) VALUES ('홍길동', 30)")
    assert parsed["action"] == "INSERT"
    assert parsed["columns"] == ["이름", "나이"]


def test_parse_unicode_update() -> None:
    parsed = parse_sql("UPDATE Sheet1 SET 이름 = '김철수' WHERE 나이 = 30")
    assert parsed["action"] == "UPDATE"
    assert parsed["set"][0]["column"] == "이름"
    assert parsed["set"][0]["value"] == {'type': 'literal', 'value': '김철수'}
    assert parsed["where"]["conditions"][0]["column"] == "나이"


def test_parse_unicode_delete() -> None:
    parsed = parse_sql("DELETE FROM Sheet1 WHERE 이름 = '홍길동'")
    assert parsed["action"] == "DELETE"
    assert parsed["where"]["conditions"][0]["column"] == "이름"
    assert parsed["where"]["conditions"][0]["value"] == "홍길동"


# ---------------------------------------------------------------------------
# A. Unicode Bare Identifiers — Executor (end-to-end)
# ---------------------------------------------------------------------------


def test_exec_unicode_select(tmp_path: Path) -> None:
    path = tmp_path / "test.xlsx"
    _create_test_workbook(path)
    result = _execute(path, "SELECT 이름 FROM Sheet1 WHERE 나이 > 25")
    assert result.rows == [("홍길동",), ("이영희",)]


def test_exec_unicode_mixed_select(tmp_path: Path) -> None:
    path = tmp_path / "test.xlsx"
    _create_test_workbook(path)
    result = _execute(path, "SELECT id, 이름 FROM Sheet1 ORDER BY id")
    assert result.rows == [(1, "홍길동"), (2, "김철수"), (3, "이영희")]


def test_exec_unicode_group_by(tmp_path: Path) -> None:
    path = tmp_path / "test.xlsx"
    _create_test_workbook(path)
    result = _execute(
        path,
        "SELECT 부서, COUNT(*) FROM emp GROUP BY 부서 ORDER BY 부서",
    )
    assert sorted(result.rows) == [("개발", 2), ("영업", 2)]


def test_exec_unicode_aggregate(tmp_path: Path) -> None:
    path = tmp_path / "test.xlsx"
    _create_test_workbook(path)
    result = _execute(path, "SELECT SUM(나이) FROM Sheet1")
    assert result.rows == [(90,)]


def test_exec_unicode_insert(tmp_path: Path) -> None:
    path = tmp_path / "test.xlsx"
    _create_dml_workbook(path)
    engine = OpenpyxlBackend(str(path))
    executor = SharedExecutor(engine)
    executor.execute(parse_sql("INSERT INTO Sheet1 (id, 이름, 나이) VALUES (3, '이영희', 35)"))
    result = executor.execute(parse_sql("SELECT 이름 FROM Sheet1 WHERE id = 3"))
    assert result.rows == [("이영희",)]

def test_exec_unicode_update(tmp_path: Path) -> None:
    path = tmp_path / "test.xlsx"
    _create_dml_workbook(path)
    engine = OpenpyxlBackend(str(path))
    executor = SharedExecutor(engine)
    executor.execute(parse_sql("UPDATE Sheet1 SET 이름 = '박지민' WHERE id = 1"))
    result = executor.execute(parse_sql("SELECT 이름 FROM Sheet1 WHERE id = 1"))
    assert result.rows == [("박지민",)]

def test_exec_unicode_delete(tmp_path: Path) -> None:
    path = tmp_path / "test.xlsx"
    _create_dml_workbook(path)
    engine = OpenpyxlBackend(str(path))
    executor = SharedExecutor(engine)
    executor.execute(parse_sql("DELETE FROM Sheet1 WHERE 이름 = '홍길동'"))
    result = executor.execute(parse_sql("SELECT id FROM Sheet1"))
    assert result.rows == [(2,)]

# ---------------------------------------------------------------------------
# B. Double-Quoted Column Identifiers — Parser
# ---------------------------------------------------------------------------


def test_parse_quoted_column_spaces() -> None:
    parsed = parse_sql('SELECT "Full Name" FROM Sheet1')
    assert parsed["columns"] == ["Full Name"]


def test_parse_quoted_column_special_chars() -> None:
    parsed = parse_sql('SELECT "col-1" FROM Sheet1')
    assert parsed["columns"] == ["col-1"]


def test_parse_quoted_column_korean() -> None:
    parsed = parse_sql('SELECT "이름" FROM Sheet1')
    assert parsed["columns"] == ["이름"]


def test_parse_quoted_where() -> None:
    parsed = parse_sql("SELECT * FROM Sheet1 WHERE \"Full Name\" = 'Alice Kim'")
    cond = parsed["where"]["conditions"][0]
    assert cond["column"] == "Full Name"
    assert cond["value"] == "Alice Kim"


def test_parse_quoted_order_by() -> None:
    parsed = parse_sql('SELECT "Full Name" FROM Sheet1 ORDER BY "Full Name"')
    assert parsed["order_by"][0]["column"] == "Full Name"


def test_parse_quoted_group_by() -> None:
    parsed = parse_sql(
        'SELECT "Department Name", COUNT(*) FROM Sheet1 '
        'GROUP BY "Department Name"'
    )
    assert parsed["group_by"] == ["Department Name"]


def test_parse_quoted_having() -> None:
    parsed = parse_sql(
        'SELECT "부서", COUNT(*) FROM emp '
        'GROUP BY "부서" HAVING COUNT(*) > 1'
    )
    assert parsed["group_by"] == ["부서"]


def test_parse_quoted_alias() -> None:
    parsed = parse_sql('SELECT "Full Name" AS name FROM Sheet1')
    col = parsed["columns"][0]
    assert col == {'type': 'alias', 'alias': 'name', 'expression': 'Full Name'}


# ---------------------------------------------------------------------------
# B. Double-Quoted Column Identifiers — Executor (end-to-end)
# ---------------------------------------------------------------------------


def test_exec_quoted_select(tmp_path: Path) -> None:
    path = tmp_path / "test.xlsx"
    _create_test_workbook(path)
    result = _execute(path, 'SELECT "Full Name" FROM Sheet1 ORDER BY id')
    assert result.rows == [("Alice Kim",), ("Bob Park",), ("Carol Lee",)]


def test_exec_quoted_where(tmp_path: Path) -> None:
    path = tmp_path / "test.xlsx"
    _create_test_workbook(path)
    result = _execute(
        path,
        "SELECT id FROM Sheet1 WHERE \"Full Name\" = 'Alice Kim'",
    )
    assert result.rows == [(1,)]


def test_exec_quoted_order_by(tmp_path: Path) -> None:
    path = tmp_path / "test.xlsx"
    _create_test_workbook(path)
    result = _execute(
        path,
        'SELECT "Full Name" FROM Sheet1 ORDER BY "Full Name"',
    )
    assert result.rows == [("Alice Kim",), ("Bob Park",), ("Carol Lee",)]


def test_exec_quoted_group_by(tmp_path: Path) -> None:
    path = tmp_path / "test.xlsx"
    _create_test_workbook(path)
    result = _execute(
        path,
        'SELECT "Department Name", COUNT(*) FROM Sheet1 '
        'GROUP BY "Department Name" ORDER BY "Department Name"',
    )
    assert result.rows == [("Engineering", 2), ("Sales", 1)]


def test_exec_quoted_having(tmp_path: Path) -> None:
    path = tmp_path / "test.xlsx"
    _create_test_workbook(path)
    result = _execute(
        path,
        'SELECT "부서", COUNT(*) FROM emp '
        'GROUP BY "부서" HAVING COUNT(*) > 1',
    )
    # Both 개발 and 영업 have 2 rows
    assert len(result.rows) == 2


def test_exec_quoted_alias(tmp_path: Path) -> None:
    path = tmp_path / "test.xlsx"
    _create_test_workbook(path)
    result = _execute(path, 'SELECT "Full Name" AS name FROM Sheet1 WHERE id = 1')
    assert result.rows == [("Alice Kim",)]
    assert result.description[0][0] == "name"


# ---------------------------------------------------------------------------
# C. Qualified Quoted Identifiers — Parser
# ---------------------------------------------------------------------------


def test_parse_qualified_quoted_column() -> None:
    parsed = parse_sql('SELECT Sheet1."Full Name" FROM Sheet1')
    assert parsed["columns"] == [{'type': 'column', 'source': 'Sheet1', 'name': 'Full Name'}]


def test_parse_qualified_quoted_table_and_column() -> None:
    parsed = parse_sql('SELECT "Sheet1"."Full Name" FROM Sheet1')
    assert parsed["columns"] == [{'type': 'column', 'source': 'Sheet1', 'name': 'Full Name'}]


# ---------------------------------------------------------------------------
# C. Qualified Quoted Identifiers — Executor
# ---------------------------------------------------------------------------


def test_exec_qualified_quoted_column(tmp_path: Path) -> None:
    path = tmp_path / "test.xlsx"
    _create_test_workbook(path)
    result = _execute(path, 'SELECT Sheet1."Full Name" FROM Sheet1 WHERE id = 1')
    assert result.rows == [("Alice Kim",)]


# ---------------------------------------------------------------------------
# D. Mixed Quoting — Parser
# ---------------------------------------------------------------------------


def test_parse_mixed_quoted_unquoted() -> None:
    parsed = parse_sql('SELECT "Full Name", id FROM Sheet1')
    assert "Full Name" in parsed["columns"]
    assert "id" in parsed["columns"]


def test_parse_mixed_where_is_not_null() -> None:
    parsed = parse_sql(
        'SELECT id, "Full Name" FROM Sheet1 WHERE "Full Name" IS NOT NULL'
    )
    cond = parsed["where"]["conditions"][0]
    assert cond["column"] == "Full Name"
    assert cond["operator"] == "IS NOT"


# ---------------------------------------------------------------------------
# D. Mixed Quoting — Executor
# ---------------------------------------------------------------------------


def test_exec_mixed_quoted_unquoted(tmp_path: Path) -> None:
    path = tmp_path / "test.xlsx"
    _create_test_workbook(path)
    result = _execute(path, 'SELECT "Full Name", id FROM Sheet1 ORDER BY id')
    assert result.rows == [
        ("Alice Kim", 1),
        ("Bob Park", 2),
        ("Carol Lee", 3),
    ]


# ---------------------------------------------------------------------------
# E. CTE with Quoted Names — Parser
# ---------------------------------------------------------------------------


def test_parse_cte_quoted_name() -> None:
    parsed = parse_sql(
        'WITH "my cte" AS (SELECT id FROM Sheet1) '
        'SELECT * FROM "my cte"'
    )
    assert parsed["action"] == "SELECT"
    assert len(parsed.get("ctes", [])) == 1
    assert parsed["ctes"][0]["name"] == "my cte"


# ---------------------------------------------------------------------------
# F. INSERT/UPDATE/DELETE with Quoted Columns — Parser
# ---------------------------------------------------------------------------


def test_parse_insert_quoted_columns() -> None:
    parsed = parse_sql(
        "INSERT INTO Sheet1 (\"Full Name\", \"나이\") VALUES ('Alice', 30)"
    )
    assert parsed["action"] == "INSERT"
    assert parsed["columns"] == ["Full Name", "나이"]


def test_parse_update_quoted_column() -> None:
    parsed = parse_sql(
        "UPDATE Sheet1 SET \"Full Name\" = 'Bob' WHERE id = 1"
    )
    assert parsed["action"] == "UPDATE"
    assert parsed["set"][0]["column"] == "Full Name"


def test_parse_delete_quoted_where() -> None:
    parsed = parse_sql(
        "DELETE FROM Sheet1 WHERE \"Full Name\" = 'Alice Kim'"
    )
    assert parsed["action"] == "DELETE"
    assert parsed["where"]["conditions"][0]["column"] == "Full Name"


# ---------------------------------------------------------------------------
# F. INSERT/UPDATE/DELETE with Quoted Columns — Executor
# ---------------------------------------------------------------------------


def test_exec_insert_quoted_columns(tmp_path: Path) -> None:
    path = tmp_path / "test.xlsx"
    _create_dml_workbook(path)
    engine = OpenpyxlBackend(str(path))
    executor = SharedExecutor(engine)
    executor.execute(
        parse_sql(
            "INSERT INTO Sheet1 (id, \"Full Name\", 나이) VALUES (3, 'Carol', 35)"
        )
    )
    result = executor.execute(parse_sql('SELECT "Full Name" FROM Sheet1 WHERE id = 3'))
    assert result.rows == [("Carol",)]


def test_exec_update_quoted_column(tmp_path: Path) -> None:
    path = tmp_path / "test.xlsx"
    _create_dml_workbook(path)
    engine = OpenpyxlBackend(str(path))
    executor = SharedExecutor(engine)
    executor.execute(parse_sql("UPDATE Sheet1 SET \"Full Name\" = 'Updated' WHERE id = 1"))
    result = executor.execute(parse_sql('SELECT "Full Name" FROM Sheet1 WHERE id = 1'))
    assert result.rows == [("Updated",)]


def test_exec_delete_quoted_where(tmp_path: Path) -> None:
    path = tmp_path / "test.xlsx"
    _create_dml_workbook(path)
    engine = OpenpyxlBackend(str(path))
    executor = SharedExecutor(engine)
    executor.execute(parse_sql("DELETE FROM Sheet1 WHERE \"Full Name\" = 'Alice Kim'"))
    result = executor.execute(parse_sql("SELECT id FROM Sheet1"))
    assert result.rows == [(2,)]


# ---------------------------------------------------------------------------
# G. Edge Cases — Parser
# ---------------------------------------------------------------------------


def test_parse_embedded_double_quote() -> None:
    parsed = parse_sql('SELECT "col""name" FROM Sheet1')
    assert parsed["columns"] == ['col"name']


def test_single_quote_remains_literal() -> None:
    parsed = parse_sql("SELECT * FROM Sheet1 WHERE name = 'Alice'")
    cond = parsed["where"]["conditions"][0]
    assert cond["value"] == "Alice"
    assert cond["column"] == "name"


def test_long_unicode_identifier() -> None:
    long_name = "コラム" * 50  # 150 chars of CJK
    parsed = parse_sql(f"SELECT {long_name} FROM Sheet1")
    assert parsed["columns"] == [long_name]


# ---------------------------------------------------------------------------
# G. Edge Cases — Executor
# ---------------------------------------------------------------------------


def test_exec_quoted_korean_column(tmp_path: Path) -> None:
    """Quoted Korean column should work the same as unquoted."""
    path = tmp_path / "test.xlsx"
    _create_test_workbook(path)
    result = _execute(path, 'SELECT "이름" FROM Sheet1 WHERE id = 1')
    assert result.rows == [("홍길동",)]


def test_exec_unicode_order_by(tmp_path: Path) -> None:
    path = tmp_path / "test.xlsx"
    _create_test_workbook(path)
    result = _execute(path, "SELECT 이름 FROM Sheet1 ORDER BY 나이 DESC")
    assert result.rows == [("이영희",), ("홍길동",), ("김철수",)]


def test_exec_unicode_between(tmp_path: Path) -> None:
    path = tmp_path / "test.xlsx"
    _create_test_workbook(path)
    result = _execute(
        path,
        "SELECT 이름 FROM Sheet1 WHERE 나이 BETWEEN 25 AND 30 ORDER BY id",
    )
    assert result.rows == [("홍길동",), ("김철수",)]



def _create_r10_workbook(
    path: Path,
    *,
    headers: list[object],
    rows: list[list[object]],
    sheet_name: str,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = sheet_name
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()

def test_ascii_identifiers_work_end_to_end(tmp_path: Path) -> None:
    file_path = tmp_path / "ascii-identifiers.xlsx"
    _create_r10_workbook(
        file_path,
        headers=["user_id", "full_name"],
        rows=[[1, "Alice"], [2, "Bob"]],
        sheet_name="users",
    )

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT user_id, full_name FROM users ORDER BY user_id")
        assert cursor.fetchall() == [(1, "Alice"), (2, "Bob")]

@pytest.mark.xfail(
    reason=(
        "Quoted identifiers are not yet supported for table/column resolution; "
        "double quotes are currently parsed as string literals"
    ),
    strict=False,
)
def test_spaced_identifiers_quoted_are_currently_not_supported(tmp_path: Path) -> None:
    file_path = tmp_path / "spaced-identifiers.xlsx"
    _create_r10_workbook(
        file_path,
        headers=["id", "full name"],
        rows=[[1, "Alice"]],
        sheet_name="People Sheet",
    )

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT "full name" FROM "People Sheet"')
        assert cursor.fetchall() == [("Alice",)]



def _create_round13_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "t"
    sheet.append(["a", "b", "val", "grp"])
    sheet.append([1, 1, 6, "x"])
    sheet.append([2, 3, 5, "x"])
    sheet.append([1, 1, 7, "y"])
    sheet.append([5, 5, 1, "y"])
    workbook.save(path)

def test_parser_accepts_quoted_table_identifier_for_create_table() -> None:
    parsed = parse_sql('CREATE TABLE "Sales 2024" (id INTEGER, amount REAL)')
    assert parsed["table"] == "Sales 2024"

def test_select_from_quoted_table_name_with_space(tmp_path: Path) -> None:
    file_path = tmp_path / "round13_quoted_table_select.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Sales 2024"
    sheet.append(["id", "amount"])
    sheet.append([1, 100])
    workbook.save(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM "Sales 2024"')
        assert cursor.fetchall() == [(1,)]
