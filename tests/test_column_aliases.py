from pathlib import Path

import pytest
from openpyxl import Workbook

from excel_dbapi.engines.openpyxl.backend import OpenpyxlBackend
from excel_dbapi.executor import SharedExecutor
from excel_dbapi.parser import parse_sql


def _create_users_alias_workbook(path: Path) -> None:
    workbook = Workbook()
    users = workbook.active
    assert users is not None
    users.title = "users"
    users.append(["id", "name", "age", "department"])
    users.append([1, "Charlie", 40, "ops"])
    users.append([2, "Alice", 30, "eng"])
    users.append([3, "Bob", 20, "eng"])
    users.append([4, "Dora", 25, "sales"])
    users.append([5, "Evan", 28, "ops"])
    users.append([6, "Finn", 31, "eng"])
    workbook.save(path)


def _create_join_alias_workbook(path: Path) -> None:
    workbook = Workbook()
    t1 = workbook.active
    assert t1 is not None
    t1.title = "t1"
    t1.append(["id", "name"])
    t1.append([1, "Alice"])
    t1.append([2, "Bob"])
    t1.append([3, "Cara"])

    t2 = workbook.create_sheet("t2")
    t2.append(["user_id", "enabled"])
    t2.append([1, True])
    t2.append([2, False])
    workbook.save(path)


def test_parse_select_with_explicit_alias() -> None:
    parsed = parse_sql("SELECT name AS n FROM users")
    assert parsed["columns"] == [
        {"type": "alias", "alias": "n", "expression": "name"}
    ]


def test_parse_select_with_aggregate_alias() -> None:
    parsed = parse_sql("SELECT COUNT(*) AS total FROM users")
    assert parsed["columns"] == [
        {
            "type": "alias",
            "alias": "total",
            "expression": {"type": "aggregate", "func": "COUNT", "arg": "*"},
        }
    ]


def test_parse_join_select_with_qualified_alias() -> None:
    parsed = parse_sql(
        "SELECT t1.name AS user_name FROM t1 JOIN t2 ON t1.id = t2.user_id"
    )
    assert parsed["columns"] == [
        {
            "type": "alias",
            "alias": "user_name",
            "expression": {"type": "column", "source": "t1", "name": "name"},
        }
    ]


def test_parse_select_with_implicit_alias() -> None:
    parsed = parse_sql("SELECT name n FROM users")
    assert parsed["columns"] == [
        {"type": "alias", "alias": "n", "expression": "name"}
    ]


def test_parse_rejects_alias_on_wildcard() -> None:
    with pytest.raises(ValueError):
        parse_sql("SELECT * AS x FROM users")


def test_parse_rejects_reserved_alias() -> None:
    with pytest.raises(ValueError):
        parse_sql("SELECT name AS FROM users")


def test_parse_select_with_mixed_alias_and_bare_column() -> None:
    parsed = parse_sql("SELECT name AS n, age FROM users")
    assert parsed["columns"] == [
        {"type": "alias", "alias": "n", "expression": "name"},
        "age",
    ]


def test_parse_join_order_by_alias() -> None:
    parsed = parse_sql(
        "SELECT t1.name AS user_name FROM t1 JOIN t2 ON t1.id = t2.user_id ORDER BY user_name"
    )
    assert parsed["order_by"] == {"column": "user_name", "direction": "ASC"}


def test_executor_select_alias_description(tmp_path: Path) -> None:
    file_path = tmp_path / "alias_select.xlsx"
    _create_users_alias_workbook(file_path)

    parsed = parse_sql("SELECT name AS n FROM users ORDER BY id ASC")
    results = SharedExecutor(OpenpyxlBackend(str(file_path))).execute(parsed)

    assert [col[0] for col in results.description] == ["n"]
    assert results.rows == [
        ("Charlie",),
        ("Alice",),
        ("Bob",),
        ("Dora",),
        ("Evan",),
        ("Finn",),
    ]


def test_executor_aggregate_alias_description(tmp_path: Path) -> None:
    file_path = tmp_path / "alias_aggregate.xlsx"
    _create_users_alias_workbook(file_path)

    parsed = parse_sql("SELECT COUNT(*) AS total FROM users")
    results = SharedExecutor(OpenpyxlBackend(str(file_path))).execute(parsed)

    assert [col[0] for col in results.description] == ["total"]
    assert results.rows == [(6,)]


def test_executor_order_by_alias_ascending(tmp_path: Path) -> None:
    file_path = tmp_path / "alias_order_asc.xlsx"
    _create_users_alias_workbook(file_path)

    parsed = parse_sql("SELECT name AS n FROM users ORDER BY n")
    results = SharedExecutor(OpenpyxlBackend(str(file_path))).execute(parsed)

    assert results.rows == [
        ("Alice",),
        ("Bob",),
        ("Charlie",),
        ("Dora",),
        ("Evan",),
        ("Finn",),
    ]


def test_executor_order_by_alias_descending(tmp_path: Path) -> None:
    file_path = tmp_path / "alias_order_desc.xlsx"
    _create_users_alias_workbook(file_path)

    parsed = parse_sql("SELECT name AS n FROM users ORDER BY n DESC")
    results = SharedExecutor(OpenpyxlBackend(str(file_path))).execute(parsed)

    assert results.rows == [
        ("Finn",),
        ("Evan",),
        ("Dora",),
        ("Charlie",),
        ("Bob",),
        ("Alice",),
    ]


def test_executor_join_alias_description(tmp_path: Path) -> None:
    file_path = tmp_path / "alias_join.xlsx"
    _create_join_alias_workbook(file_path)

    parsed = parse_sql("SELECT t1.name AS user_name FROM t1 JOIN t2 ON t1.id = t2.user_id")
    results = SharedExecutor(OpenpyxlBackend(str(file_path))).execute(parsed)

    assert [col[0] for col in results.description] == ["user_name"]
    assert results.rows == [("Alice",), ("Bob",)]


def test_executor_group_by_aggregate_alias_with_order_by_alias(tmp_path: Path) -> None:
    file_path = tmp_path / "alias_group_order.xlsx"
    _create_users_alias_workbook(file_path)

    parsed = parse_sql(
        "SELECT department, COUNT(*) AS cnt FROM users GROUP BY department ORDER BY cnt"
    )
    results = SharedExecutor(OpenpyxlBackend(str(file_path))).execute(parsed)

    assert [col[0] for col in results.description] == ["department", "cnt"]
    assert results.rows == [("sales", 1), ("ops", 2), ("eng", 3)]


def test_executor_select_mixed_alias_and_non_alias_columns(tmp_path: Path) -> None:
    file_path = tmp_path / "alias_mixed_columns.xlsx"
    _create_users_alias_workbook(file_path)

    parsed = parse_sql("SELECT name AS n, age FROM users ORDER BY id ASC")
    results = SharedExecutor(OpenpyxlBackend(str(file_path))).execute(parsed)

    assert [col[0] for col in results.description] == ["n", "age"]
    assert results.rows == [
        ("Charlie", 40),
        ("Alice", 30),
        ("Bob", 20),
        ("Dora", 25),
        ("Evan", 28),
        ("Finn", 31),
    ]




def test_executor_alias_shadows_real_column_name(tmp_path: Path) -> None:
    """Alias 'age' shadows the real 'age' column; ORDER BY uses the alias target (name)."""
    file_path = tmp_path / "alias_shadow.xlsx"
    _create_users_alias_workbook(file_path)

    parsed = parse_sql("SELECT name AS age FROM users ORDER BY age")
    results = SharedExecutor(OpenpyxlBackend(str(file_path))).execute(parsed)

    assert [col[0] for col in results.description] == ["age"]
    # ORDER BY age resolves to alias target (name), so sorted alphabetically by name
    assert results.rows == [
        ("Alice",),
        ("Bob",),
        ("Charlie",),
        ("Dora",),
        ("Evan",),
        ("Finn",),
    ]


def test_executor_mixed_alias_and_bare_order_by(tmp_path: Path) -> None:
    """ORDER BY with alias column + non-alias column in multi-column sort."""
    file_path = tmp_path / "alias_mixed_order.xlsx"
    _create_users_alias_workbook(file_path)

    parsed = parse_sql(
        "SELECT department AS dept, name FROM users ORDER BY dept ASC, name DESC"
    )
    results = SharedExecutor(OpenpyxlBackend(str(file_path))).execute(parsed)

    assert [col[0] for col in results.description] == ["dept", "name"]
    # eng: Finn(31), Charlie(err-- wait, Charlie is ops), Bob, Alice
    # Actually: eng has Alice(30), Bob(20), Finn(31) -> name DESC: Finn, Bob, Alice
    # ops has Charlie(40), Evan(28) -> name DESC: Evan, Charlie
    # sales has Dora(25)
    assert results.rows == [
        ("eng", "Finn"),
        ("eng", "Bob"),
        ("eng", "Alice"),
        ("ops", "Evan"),
        ("ops", "Charlie"),
        ("sales", "Dora"),
    ]


def test_executor_single_row_aggregate_order_by_alias(tmp_path: Path) -> None:
    """Single-row aggregate (no GROUP BY) with ORDER BY alias — trivial but must not crash."""
    file_path = tmp_path / "alias_agg_order.xlsx"
    _create_users_alias_workbook(file_path)

    parsed = parse_sql("SELECT COUNT(*) AS total FROM users ORDER BY total")
    results = SharedExecutor(OpenpyxlBackend(str(file_path))).execute(parsed)

    assert [col[0] for col in results.description] == ["total"]
    assert results.rows == [(6,)]
