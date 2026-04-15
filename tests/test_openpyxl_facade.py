"""Tests for the excel_dbapi.openpyxl facade module.

Verifies that all re-exported symbols are importable and refer to the
correct underlying openpyxl classes/functions.
"""

from pathlib import Path

from openpyxl import Workbook
import openpyxl.comments
import openpyxl.styles
import openpyxl.utils
import openpyxl.utils.exceptions
import openpyxl.workbook.workbook
import openpyxl.worksheet.datavalidation
import openpyxl.worksheet.worksheet
import pytest

from excel_dbapi import openpyxl as facade
from excel_dbapi.engines.base import TableData
from excel_dbapi.engines.openpyxl.backend import OpenpyxlBackend
from excel_dbapi.exceptions import DatabaseError


class TestFacadeReexports:
    """Every symbol in __all__ resolves to the real openpyxl object."""

    def test_workbook_is_openpyxl_workbook(self):
        assert facade.Workbook is openpyxl.Workbook

    def test_load_workbook_is_openpyxl_load_workbook(self):
        assert facade.load_workbook is openpyxl.load_workbook

    def test_font(self):
        assert facade.Font is openpyxl.styles.Font

    def test_pattern_fill(self):
        assert facade.PatternFill is openpyxl.styles.PatternFill

    def test_border(self):
        assert facade.Border is openpyxl.styles.Border

    def test_side(self):
        assert facade.Side is openpyxl.styles.Side

    def test_alignment(self):
        assert facade.Alignment is openpyxl.styles.Alignment

    def test_comment(self):
        assert facade.Comment is openpyxl.comments.Comment

    def test_data_validation(self):
        assert facade.DataValidation is openpyxl.worksheet.datavalidation.DataValidation

    def test_get_column_letter(self):
        assert facade.get_column_letter is openpyxl.utils.get_column_letter

    def test_invalid_file_exception(self):
        assert (
            facade.InvalidFileException
            is openpyxl.utils.exceptions.InvalidFileException
        )

    def test_worksheet(self):
        assert facade.Worksheet is openpyxl.worksheet.worksheet.Worksheet


class TestFacadeAll:
    """__all__ is complete and consistent."""

    EXPECTED_SYMBOLS = {
        "Workbook",
        "load_workbook",
        "Alignment",
        "Border",
        "Font",
        "PatternFill",
        "Side",
        "Comment",
        "DataValidation",
        "get_column_letter",
        "InvalidFileException",
        "Worksheet",
    }

    def test_all_contains_expected_symbols(self):
        assert set(facade.__all__) == self.EXPECTED_SYMBOLS

    @pytest.mark.parametrize("name", sorted(EXPECTED_SYMBOLS))
    def test_each_symbol_is_accessible(self, name: str):
        obj = getattr(facade, name)
        assert obj is not None


class TestFacadeFunctionality:
    """Facade objects work identically to direct openpyxl usage."""

    def test_create_workbook(self):
        wb = facade.Workbook()
        assert wb.sheetnames == ["Sheet"]
        wb.close()

    def test_font_instantiation(self):
        font = facade.Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        assert font.bold is True
        assert font.name == "Calibri"

    def test_pattern_fill_instantiation(self):
        fill = facade.PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        assert fill.fill_type == "solid"

    def test_get_column_letter_returns_correct_value(self):
        assert facade.get_column_letter(1) == "A"
        assert facade.get_column_letter(26) == "Z"
        assert facade.get_column_letter(27) == "AA"

    def test_data_validation_creation(self):
        dv = facade.DataValidation(type="list", formula1='"A,B,C"')
        assert dv.type == "list"

    def test_comment_creation(self):
        comment = facade.Comment("test note", "author")
        assert comment.text == "test note"

    def test_invalid_file_exception_is_catchable(self):
        with pytest.raises(facade.InvalidFileException):
            raise facade.InvalidFileException("test")

    def test_border_creation(self):
        border = facade.Border(
            left=facade.Side(style="thin", color="000000"),
            right=facade.Side(style="thin", color="000000"),
        )
        assert border.left is not None
        assert border.left.style == "thin"

    def test_alignment_creation(self):
        alignment = facade.Alignment(horizontal="center", vertical="center")
        assert alignment.horizontal == "center"



def _xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id", "name"])
    ws.append([1, "Alice"])
    wb.save(path)

def test_openpyxl_backend_error_paths_and_execute_wrappers(tmp_path: Path) -> None:
    file_path = tmp_path / "openpyxl.xlsx"
    _xlsx(file_path)
    backend = OpenpyxlBackend(str(file_path), create=False)

    with pytest.raises(DatabaseError, match="not found"):
        backend.read_sheet("Missing")
    with pytest.raises(DatabaseError, match="not found"):
        backend.write_sheet("Missing", TableData(headers=["a"], rows=[]))
    with pytest.raises(DatabaseError, match="not found"):
        backend.append_row("Missing", [1])
    with pytest.raises(DatabaseError, match="already exists"):
        backend.create_sheet("Sheet1", ["a"])
    with pytest.raises(DatabaseError, match="not found"):
        backend.drop_sheet("Missing")

    backend.workbook = None
    with pytest.raises(DatabaseError, match="not loaded"):
        backend.snapshot()
    with pytest.raises(DatabaseError, match="not loaded"):
        backend.get_workbook()
    with pytest.raises(DatabaseError, match="not loaded"):
        backend.create_sheet("X", ["a"])
    with pytest.raises(DatabaseError, match="not loaded"):
        backend.drop_sheet("Sheet1")

    backend2 = OpenpyxlBackend(str(file_path), create=False)
    result1 = backend2.execute("SELECT * FROM Sheet1")
    result2 = backend2.execute_with_params("SELECT * FROM Sheet1 WHERE id = ?", (1,))
    assert result1.rowcount >= 1
    assert result2.rowcount == 1
