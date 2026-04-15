from pathlib import Path
from unittest.mock import PropertyMock, patch

from openpyxl import Workbook
import pytest

from excel_dbapi.connection import ExcelConnection
from excel_dbapi.exceptions import DatabaseError, NotSupportedError, ProgrammingError
from excel_dbapi.parser import parse_sql
from excel_dbapi.reflection import METADATA_SHEET, read_table_metadata, write_table_metadata


def _create_people_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "people"
    sheet.append(["id", "name"])
    sheet.append([1, "Alice"])
    sheet.append([2, "Bob"])
    workbook.save(path)


def _create_single_column_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "onecol"
    sheet.append(["id"])
    sheet.append([1])
    workbook.save(path)


def _query(path: Path, sql: str) -> tuple[list[tuple[object, ...]], list[str]]:
    with ExcelConnection(str(path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute(sql)
        rows = cursor.fetchall()
        description = [str(item[0]) for item in cursor.description or []]
        return rows, description


def test_parser_add_column() -> None:
    parsed = parse_sql("ALTER TABLE t ADD COLUMN age TEXT")
    assert parsed == {
        "action": "ALTER",
        "table": "t",
        "operation": "ADD_COLUMN",
        "column": "age",
        "type_name": "TEXT",
        "params": None,
    }


def test_parser_drop_column() -> None:
    parsed = parse_sql("ALTER TABLE t DROP COLUMN age")
    assert parsed == {
        "action": "ALTER",
        "table": "t",
        "operation": "DROP_COLUMN",
        "column": "age",
        "params": None,
    }


def test_parser_rename_column() -> None:
    parsed = parse_sql("ALTER TABLE t RENAME COLUMN old_name TO new_name")
    assert parsed == {
        "action": "ALTER",
        "table": "t",
        "operation": "RENAME_COLUMN",
        "old_column": "old_name",
        "new_column": "new_name",
        "params": None,
    }


def test_parser_add_column_type_normalization() -> None:
    parsed = parse_sql("ALTER TABLE t ADD COLUMN score FLOAT")
    assert parsed["type_name"] == "REAL"


def test_parser_error_invalid_type() -> None:
    with pytest.raises(DatabaseError, match="Unsupported ALTER TABLE column type"):
        parse_sql("ALTER TABLE t ADD COLUMN age BLOB")


def test_parser_error_missing_column_keyword() -> None:
    with pytest.raises(DatabaseError, match="Invalid ALTER TABLE format"):
        parse_sql("ALTER TABLE t ADD age INTEGER")


def test_add_column_basic(tmp_path: Path) -> None:
    file_path = tmp_path / "alter_add_basic.xlsx"
    _create_people_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("ALTER TABLE people ADD COLUMN age INTEGER")

    rows, description = _query(
        file_path, "SELECT id, name, age FROM people ORDER BY id"
    )
    assert description == ["id", "name", "age"]
    assert rows == [(1, "Alice", None), (2, "Bob", None)]


def test_add_column_duplicate_error(tmp_path: Path) -> None:
    file_path = tmp_path / "alter_add_dup.xlsx"
    _create_people_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="already exists"):
            cursor.execute("ALTER TABLE people ADD COLUMN name TEXT")


def test_add_column_nonexistent_table_error(tmp_path: Path) -> None:
    file_path = tmp_path / "alter_add_missing_table.xlsx"
    _create_people_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        with pytest.raises(
            ProgrammingError, match="Sheet 'missing' not found in Excel"
        ):
            cursor.execute("ALTER TABLE missing ADD COLUMN age INTEGER")


def test_drop_column_basic(tmp_path: Path) -> None:
    file_path = tmp_path / "alter_drop_basic.xlsx"
    _create_people_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("ALTER TABLE people DROP COLUMN name")

    rows, description = _query(file_path, "SELECT id FROM people ORDER BY id")
    assert description == ["id"]
    assert rows == [(1,), (2,)]


def test_drop_column_nonexistent_error(tmp_path: Path) -> None:
    file_path = tmp_path / "alter_drop_missing_col.xlsx"
    _create_people_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="Column 'age' not found"):
            cursor.execute("ALTER TABLE people DROP COLUMN age")


def test_drop_column_last_column_error(tmp_path: Path) -> None:
    file_path = tmp_path / "alter_drop_last_col.xlsx"
    _create_single_column_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="Cannot drop the only column"):
            cursor.execute("ALTER TABLE onecol DROP COLUMN id")


def test_rename_column_basic(tmp_path: Path) -> None:
    file_path = tmp_path / "alter_rename_basic.xlsx"
    _create_people_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("ALTER TABLE people RENAME COLUMN name TO full_name")

    rows, description = _query(
        file_path, "SELECT id, full_name FROM people ORDER BY id"
    )
    assert description == ["id", "full_name"]
    assert rows == [(1, "Alice"), (2, "Bob")]


def test_rename_column_nonexistent_error(tmp_path: Path) -> None:
    file_path = tmp_path / "alter_rename_missing_col.xlsx"
    _create_people_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="Column 'age' not found"):
            cursor.execute("ALTER TABLE people RENAME COLUMN age TO years")


def test_rename_column_duplicate_target_error(tmp_path: Path) -> None:
    file_path = tmp_path / "alter_rename_dup_target.xlsx"
    _create_people_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="Column 'id' already exists"):
            cursor.execute("ALTER TABLE people RENAME COLUMN name TO id")


def test_alter_then_select_insert(tmp_path: Path) -> None:
    file_path = tmp_path / "alter_then_insert.xlsx"
    _create_people_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("ALTER TABLE people ADD COLUMN age INTEGER")
        cursor.execute("INSERT INTO people (id, name, age) VALUES (3, 'Cara', 29)")

    rows, description = _query(
        file_path, "SELECT id, name, age FROM people ORDER BY id"
    )
    assert description == ["id", "name", "age"]
    assert rows == [(1, "Alice", None), (2, "Bob", None), (3, "Cara", 29)]


def test_alter_multiple_operations(tmp_path: Path) -> None:
    file_path = tmp_path / "alter_multiple_ops.xlsx"
    _create_people_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("ALTER TABLE people ADD COLUMN age INTEGER")
        cursor.execute("ALTER TABLE people RENAME COLUMN name TO full_name")
        cursor.execute("ALTER TABLE people DROP COLUMN age")
        cursor.execute("INSERT INTO people (id, full_name) VALUES (3, 'Cara')")

    rows, description = _query(
        file_path, "SELECT id, full_name FROM people ORDER BY id"
    )
    assert description == ["id", "full_name"]
    assert rows == [(1, "Alice"), (2, "Bob"), (3, "Cara")]


def test_add_column_reflection_type(tmp_path: Path) -> None:
    file_path = tmp_path / "alter_reflect.xlsx"
    _create_people_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("ALTER TABLE people ADD COLUMN score INTEGER")


    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        metadata = read_table_metadata(conn, "people")
        assert metadata is not None
        col_names = [entry["name"] for entry in metadata]
        assert "score" in col_names
        score_col = next(entry for entry in metadata if entry["name"] == "score")
        assert score_col["type_name"] == "INTEGER"


def test_alter_with_manual_commit(tmp_path: Path) -> None:
    file_path = tmp_path / "alter_manual_commit.xlsx"
    _create_people_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=False) as conn:
        cursor = conn.cursor()
        cursor.execute("ALTER TABLE people ADD COLUMN email TEXT")
        conn.commit()

    rows, description = _query(
        file_path, "SELECT id, name, email FROM people ORDER BY id"
    )
    assert description == ["id", "name", "email"]
    assert rows == [(1, "Alice", None), (2, "Bob", None)]


def test_alter_readonly_blocked(tmp_path: Path) -> None:
    file_path = tmp_path / "alter_readonly.xlsx"
    _create_people_workbook(file_path)



    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        with patch.object(type(conn.engine), "readonly", new_callable=PropertyMock, return_value=True):
            with pytest.raises(NotSupportedError, match="ALTER.*not supported.*read-only"):
                cursor.execute("ALTER TABLE people ADD COLUMN email TEXT")

def test_add_column_invalid_type_rejected_at_execute(tmp_path: Path) -> None:
    file_path = tmp_path / "alter_invalid_type.xlsx"
    _create_people_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        with pytest.raises(
            ProgrammingError, match="Unsupported ALTER TABLE column type"
        ):
            cursor.execute("ALTER TABLE people ADD COLUMN payload BLOB")



def _create_r7_workbook(
    path: Path, headers: list[object], rows: list[list[object]]
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Sheet1"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()

def test_ddl_updates_reflection_metadata(tmp_path: Path) -> None:
    file_path = tmp_path / "ddl-metadata.xlsx"
    _create_r7_workbook(file_path, ["seed"], [[1]])

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()

        cursor.execute("CREATE TABLE people (id INTEGER, name TEXT)")
        metadata = read_table_metadata(conn, "people")
        assert metadata is not None
        assert [entry["name"] for entry in metadata] == ["id", "name"]
        assert [entry["type_name"] for entry in metadata] == ["INTEGER", "TEXT"]

        cursor.execute("ALTER TABLE people ADD COLUMN email TEXT")
        metadata = read_table_metadata(conn, "people")
        assert metadata is not None
        assert [entry["name"] for entry in metadata] == ["id", "name", "email"]
        assert [entry["type_name"] for entry in metadata] == [
            "INTEGER",
            "TEXT",
            "TEXT",
        ]

        cursor.execute("ALTER TABLE people RENAME COLUMN name TO full_name")
        metadata = read_table_metadata(conn, "people")
        assert metadata is not None
        assert [entry["name"] for entry in metadata] == ["id", "full_name", "email"]
        assert [entry["type_name"] for entry in metadata] == [
            "INTEGER",
            "TEXT",
            "TEXT",
        ]

        cursor.execute("ALTER TABLE people DROP COLUMN email")
        metadata = read_table_metadata(conn, "people")
        assert metadata is not None
        assert [entry["name"] for entry in metadata] == ["id", "full_name"]

        cursor.execute("DROP TABLE people")
        assert read_table_metadata(conn, "people") is None

def test_drop_table_guard_ignores_metadata_sheet(tmp_path: Path) -> None:
    file_path = tmp_path / "drop-last-user-sheet.xlsx"
    _create_r7_workbook(file_path, ["id"], [[1]])

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        write_table_metadata(
            conn,
            "Sheet1",
            [
                {
                    "name": "id",
                    "type_name": "INTEGER",
                    "nullable": False,
                    "primary_key": True,
                }
            ],
        )
        assert METADATA_SHEET in conn.engine.list_sheets()

        cursor = conn.cursor()
        with pytest.raises(
            ProgrammingError, match="Cannot drop the only remaining sheet"
        ):
            cursor.execute("DROP TABLE Sheet1")



def _create_r9_workbook(
    path: Path,
    *,
    headers: list[object] | None = None,
    rows: list[list[object]] | None = None,
    sheet_name: str = "Sheet1",
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = sheet_name
    if headers is not None:
        sheet.append(headers)
    for row in rows or []:
        sheet.append(row)
    workbook.save(path)
    workbook.close()

def test_execute_create_table_rejects_malformed_definitions(tmp_path: Path) -> None:
    file_path = tmp_path / "create-malformed.xlsx"
    _create_r9_workbook(file_path, headers=["seed"], rows=[[1]])

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="Table name is required"):
            cursor.execute("CREATE TABLE (id INTEGER)")
        with pytest.raises(
            ProgrammingError,
            match="Malformed column definitions: empty column definition found",
        ):
            cursor.execute("CREATE TABLE t (id INTEGER,, name TEXT)")

def test_execute_create_table_rejects_trailing_comma(tmp_path: Path) -> None:
    file_path = tmp_path / "create-trailing-comma.xlsx"
    _create_r9_workbook(file_path, headers=["seed"], rows=[[1]])

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="empty column definition"):
            cursor.execute("CREATE TABLE people (id INTEGER, name TEXT,)")

@pytest.mark.parametrize(
    "sql",
    [
        "CREATE TABLE __excel_meta__ (id INTEGER)",
        "DROP TABLE __excel_meta__",
        "ALTER TABLE __excel_meta__ ADD COLUMN x TEXT",
    ],
)
def test_reserved_metadata_sheet_rejects_ddl(tmp_path: Path, sql: str) -> None:
    file_path = tmp_path / "reserved-ddl.xlsx"
    _create_r9_workbook(file_path, headers=["id"], rows=[[1]])

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        with pytest.raises(
            ProgrammingError,
            match="Cannot perform DDL on reserved metadata table '__excel_meta__'",
        ):
            cursor.execute(sql)

def test_alter_drop_column_is_case_insensitive(tmp_path: Path) -> None:
    file_path = tmp_path / "alter-drop-case-insensitive.xlsx"
    _create_r9_workbook(
        file_path, headers=["id", "Name"], rows=[[1, "Alice"]], sheet_name="people"
    )

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("ALTER TABLE people DROP COLUMN name")
        cursor.execute("SELECT * FROM people")
        assert [str(col[0]) for col in (cursor.description or [])] == ["id"]
        assert cursor.fetchall() == [(1,)]

def test_alter_rename_column_is_case_insensitive(tmp_path: Path) -> None:
    file_path = tmp_path / "alter-rename-case-insensitive.xlsx"
    _create_r9_workbook(
        file_path, headers=["id", "Name"], rows=[[1, "Alice"]], sheet_name="people"
    )

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("ALTER TABLE people RENAME COLUMN name TO full_name")
        cursor.execute("SELECT id, full_name FROM people")
        assert cursor.fetchall() == [(1, "Alice")]
