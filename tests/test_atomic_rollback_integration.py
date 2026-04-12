from pathlib import Path

import openpyxl
import pytest

from excel_dbapi.connection import ExcelConnection


@pytest.fixture
def workbook_path(tmp_path: Path) -> Path:
    path = tmp_path / "atomic.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id", "name"])
    ws.append([1, "Alice"])
    wb.save(path)
    wb.close()
    return path


def test_rollback_restores_snapshot_without_touching_file(workbook_path: Path):
    with ExcelConnection(str(workbook_path), autocommit=False) as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE Sheet1 SET name = 'Bob' WHERE id = 1")
        conn.rollback()
        cursor.execute("SELECT name FROM Sheet1 WHERE id = 1")
        assert cursor.fetchone() == ("Alice",)

    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    assert wb["Sheet1"].cell(row=2, column=2).value == "Alice"
    wb.close()


def test_atomic_save_preserves_original_file_on_replace_failure(
    workbook_path: Path, monkeypatch
):
    original_bytes = workbook_path.read_bytes()

    def _raise(*_args, **_kwargs):
        raise OSError("replace failed")

    monkeypatch.setattr("excel_dbapi.engines.openpyxl.backend.os.replace", _raise)

    with ExcelConnection(str(workbook_path), autocommit=False) as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE Sheet1 SET name = 'Bob' WHERE id = 1")
        with pytest.raises(OSError):
            conn.commit()

    assert workbook_path.read_bytes() == original_bytes
