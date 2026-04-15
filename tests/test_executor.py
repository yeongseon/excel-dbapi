from pathlib import Path
from typing import Any, cast
import datetime

from openpyxl import Workbook
import pandas as pd
import pytest

from excel_dbapi.connection import ExcelConnection
from excel_dbapi.engines.base import TableData
from excel_dbapi.engines.openpyxl.backend import OpenpyxlBackend
from excel_dbapi.engines.pandas.backend import PandasBackend
from excel_dbapi.exceptions import DatabaseError
from excel_dbapi.executor import SharedExecutor
from excel_dbapi.parser import parse_sql
import excel_dbapi.reflection as reflection


def test_executor_select():
    engine = OpenpyxlBackend("tests/data/sample.xlsx")
    parsed = parse_sql("SELECT * FROM Sheet1")
    results = SharedExecutor(engine).execute(parsed)
    assert isinstance(results.rows, list)
    assert isinstance(results.rows[0], tuple)


def test_executor_select_with_where():
    engine = OpenpyxlBackend("tests/data/sample.xlsx")
    parsed = parse_sql("SELECT * FROM Sheet1 WHERE id = 1")
    results = SharedExecutor(engine).execute(parsed)

    assert isinstance(results.rows, list)
    assert len(results.rows) == 1
    assert results.rows[0][0] == 1


def _create_select_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Sheet1"
    sheet.append(["id", "name", "score", "tags"])
    sheet.append([1, "A", 10, None])
    sheet.append([2, "A", 10, None])
    sheet.append([3, "B", None, None])
    sheet.append([4, "C", 30, "x"])
    sheet.append([5, "B", None, None])
    workbook.save(path)


def _create_empty_select_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Sheet1"
    sheet.append(["id", "name", "score"])
    workbook.save(path)


def _create_users_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "users"
    sheet.append(["id", "name", "age"])
    sheet.append([1, "Alice", 30])
    sheet.append([2, "Bob", 25])
    sheet.append([3, "Alice", 35])
    sheet.append([4, "Charlie", 40])
    workbook.save(path)


def _create_users_admins_workbook(path: Path) -> None:
    workbook = Workbook()

    users = workbook.active
    assert users is not None
    users.title = "users"
    users.append(["id", "name"])
    users.append([1, "Alice"])
    users.append([2, "Bob"])
    users.append([3, "Charlie"])

    admins = workbook.create_sheet("admins")
    admins.append(["id", "role"])
    admins.append([1, "admin"])
    admins.append([3, "editor"])

    workbook.save(path)


def test_executor_distinct_removes_duplicates_and_preserves_order(tmp_path: Path):
    file_path = tmp_path / "distinct_order.xlsx"
    _create_select_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql("SELECT DISTINCT name, score FROM Sheet1 ORDER BY name ASC")
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [("A", 10), ("B", None), ("C", 30)]


def test_executor_distinct_with_where_order_by_and_limit(tmp_path: Path):
    file_path = tmp_path / "distinct_where_limit.xlsx"
    _create_select_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT DISTINCT name FROM Sheet1 WHERE score IS NULL ORDER BY name ASC LIMIT 1"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [("B",)]


def test_executor_distinct_on_select_star(tmp_path: Path):
    file_path = tmp_path / "distinct_star.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Sheet1"
    sheet.append(["id", "name"])
    sheet.append([1, "A"])
    sheet.append([1, "A"])
    sheet.append([2, "B"])
    workbook.save(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql("SELECT DISTINCT * FROM Sheet1 ORDER BY id ASC")
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(1, "A"), (2, "B")]


def test_executor_offset_variants(tmp_path: Path):
    file_path = tmp_path / "offset_cases.xlsx"
    _create_select_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))

    parsed = parse_sql("SELECT id FROM Sheet1 ORDER BY id ASC OFFSET 2")
    results = SharedExecutor(engine).execute(parsed)
    assert results.rows == [(3,), (4,), (5,)]

    parsed = parse_sql("SELECT id FROM Sheet1 ORDER BY id ASC LIMIT 2 OFFSET 2")
    results = SharedExecutor(engine).execute(parsed)
    assert results.rows == [(3,), (4,)]

    parsed = parse_sql("SELECT id FROM Sheet1 ORDER BY id ASC OFFSET 99")
    results = SharedExecutor(engine).execute(parsed)
    assert results.rows == []

    parsed = parse_sql("SELECT id FROM Sheet1 ORDER BY id ASC OFFSET 0")
    results = SharedExecutor(engine).execute(parsed)
    assert results.rows == [(1,), (2,), (3,), (4,), (5,)]


def test_executor_offset_with_where_and_distinct_limit(tmp_path: Path):
    file_path = tmp_path / "offset_where_distinct.xlsx"
    _create_select_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT DISTINCT name FROM Sheet1 WHERE id >= 2 ORDER BY name ASC LIMIT 2 OFFSET 1"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [("B",), ("C",)]


def test_executor_distinct_applied_before_limit(tmp_path: Path):
    file_path = tmp_path / "distinct_before_limit.xlsx"
    _create_select_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql("SELECT DISTINCT name FROM Sheet1 ORDER BY name ASC LIMIT 2")
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [("A",), ("B",)]


def test_executor_distinct_rejects_order_by_non_selected_column(tmp_path: Path):
    file_path = tmp_path / "distinct_order_by_non_selected.xlsx"
    _create_select_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql("SELECT DISTINCT name FROM Sheet1 ORDER BY id ASC")
    with pytest.raises(DatabaseError, match="ORDER BY columns must appear in SELECT list when using DISTINCT",):
        SharedExecutor(engine).execute(parsed)


def test_executor_aggregate_count_star(tmp_path: Path):
    file_path = tmp_path / "aggregate_count.xlsx"
    _create_select_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql("SELECT COUNT(*) FROM Sheet1")
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(5,)]
    assert results.description[0][0] == "COUNT(*)"


def test_executor_aggregate_sum_avg_min_max(tmp_path: Path):
    file_path = tmp_path / "aggregate_numeric.xlsx"
    _create_select_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT SUM(score), AVG(score), MIN(score), MAX(score) FROM Sheet1"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(50.0, 50.0 / 3.0, 10.0, 30.0)]


def test_executor_aggregate_min_max_text_values(tmp_path: Path):
    file_path = tmp_path / "aggregate_text_min_max.xlsx"
    _create_select_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql("SELECT MIN(name), MAX(name) FROM Sheet1")
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [
        (
            "A",
            "C",
        )
    ]


def test_executor_aggregate_min_max_date_values(tmp_path: Path):
    file_path = tmp_path / "aggregate_date_min_max.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Sheet1"
    sheet.append(["id", "event_date"])
    sheet.append([1, datetime.date(2024, 2, 1)])
    sheet.append([2, datetime.date(2024, 1, 15)])
    sheet.append([3, datetime.date(2024, 3, 10)])
    workbook.save(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql("SELECT MIN(event_date), MAX(event_date) FROM Sheet1")
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [
        (datetime.datetime(2024, 1, 15, 0, 0), datetime.datetime(2024, 3, 10, 0, 0))
    ]


def test_executor_group_by_count(tmp_path: Path):
    file_path = tmp_path / "group_by_count.xlsx"
    _create_select_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT name, COUNT(*) FROM Sheet1 GROUP BY name ORDER BY name ASC"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [("A", 2), ("B", 2), ("C", 1)]


def test_executor_group_by_having_sum(tmp_path: Path):
    file_path = tmp_path / "group_by_having.xlsx"
    _create_select_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT name, SUM(score) FROM Sheet1 GROUP BY name HAVING SUM(score) > 15 ORDER BY name ASC"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [("A", 20.0), ("C", 30.0)]


def test_executor_aggregate_empty_table(tmp_path: Path):
    file_path = tmp_path / "aggregate_empty.xlsx"
    _create_empty_select_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT COUNT(*), SUM(score), AVG(score), MIN(score), MAX(score) FROM Sheet1"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(0, None, None, None, None)]


def test_executor_count_column_excludes_nulls(tmp_path: Path):
    file_path = tmp_path / "aggregate_count_nulls.xlsx"
    _create_select_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql("SELECT COUNT(*), COUNT(score) FROM Sheet1")
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(5, 3)]


def test_executor_group_by_with_order_limit_offset(tmp_path: Path):
    file_path = tmp_path / "group_by_order_limit_offset.xlsx"
    _create_select_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT name, COUNT(*) FROM Sheet1 GROUP BY name ORDER BY name DESC LIMIT 1 OFFSET 1"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [("B", 2)]


def test_executor_distinct_with_group_by(tmp_path: Path):
    file_path = tmp_path / "group_by_distinct.xlsx"
    _create_select_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT DISTINCT name FROM Sheet1 GROUP BY name ORDER BY name ASC"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [("A",), ("B",), ("C",)]


def test_executor_group_by_having_count_not_in_select(tmp_path: Path):
    file_path = tmp_path / "users_having_count_not_selected.xlsx"
    _create_users_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql("SELECT name FROM users GROUP BY name HAVING COUNT(*) > 1")
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [("Alice",)]


def test_executor_group_by_having_group_column_not_in_select(tmp_path: Path):
    file_path = tmp_path / "users_having_group_column_not_selected.xlsx"
    _create_users_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql("SELECT COUNT(*) FROM users GROUP BY name HAVING name = 'Alice'")
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(2,)]


def test_having_rejects_non_grouped_column(tmp_path: Path):
    file_path = tmp_path / "users_having_non_grouped_column.xlsx"
    _create_users_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql("SELECT COUNT(*) FROM users GROUP BY name HAVING age > 30")

    with pytest.raises(DatabaseError, match="in HAVING must be a GROUP BY column or aggregate function",):
        SharedExecutor(engine).execute(parsed)


def test_executor_group_by_order_by_group_key_not_in_select(tmp_path: Path):
    file_path = tmp_path / "users_order_by_group_key_not_selected.xlsx"
    _create_users_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql("SELECT COUNT(*) FROM users GROUP BY name ORDER BY name")
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(2,), (1,), (1,)]


def test_having_rejects_aggregate_with_expression_arg(tmp_path: Path):
    file_path = tmp_path / "users_having_expression_aggregate.xlsx"
    _create_users_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql("SELECT COUNT(*) FROM users GROUP BY name HAVING SUM(age+1) > 1")

    with pytest.raises(DatabaseError, match="must be a GROUP BY column or aggregate function"):
        SharedExecutor(engine).execute(parsed)


def test_order_by_rejects_aggregate_with_expression_arg(tmp_path: Path):
    file_path = tmp_path / "users_order_by_expression_aggregate.xlsx"
    _create_users_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    with pytest.raises(DatabaseError, match="Unsupported aggregate expression"):
        parsed = parse_sql(
            "SELECT name, COUNT(*) FROM users GROUP BY name ORDER BY SUM(age+1)"
        )
        SharedExecutor(engine).execute(parsed)


def test_subquery_in_where(tmp_path: Path):
    file_path = tmp_path / "subquery_in_where.xlsx"
    _create_users_admins_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql("SELECT id, name FROM users WHERE id IN (SELECT id FROM admins)")
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(1, "Alice"), (3, "Charlie")]


def test_subquery_returns_empty(tmp_path: Path):
    file_path = tmp_path / "subquery_empty.xlsx"
    _create_users_admins_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT id, name FROM users WHERE id IN (SELECT id FROM admins WHERE role = 'root')"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == []


def test_subquery_with_where(tmp_path: Path):
    file_path = tmp_path / "subquery_with_where.xlsx"
    _create_users_admins_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT id, name FROM users WHERE id IN (SELECT id FROM admins WHERE role = 'admin')"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(1, "Alice")]


def test_subquery_reexecution_safe(tmp_path: Path):
    file_path = tmp_path / "subquery_reexec.xlsx"
    _create_users_admins_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql("SELECT id, name FROM users WHERE id IN (SELECT id FROM admins)")
    executor = SharedExecutor(engine)

    result1 = executor.execute(parsed)
    result2 = executor.execute(parsed)

    assert result1.rows == result2.rows == [(1, "Alice"), (3, "Charlie")]


# --- JOIN executor tests ---


def test_executor_inner_join(tmp_path: Path):
    file_path = tmp_path / "inner_join.xlsx"
    _create_users_admins_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT a.id, a.name, b.role FROM users a INNER JOIN admins b ON a.id = b.id"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(1, "Alice", "admin"), (3, "Charlie", "editor")]
    assert len(results.description) == 3
    assert results.description[0][0] == "a.id"
    assert results.description[1][0] == "a.name"
    assert results.description[2][0] == "b.role"


def test_executor_left_join(tmp_path: Path):
    file_path = tmp_path / "left_join.xlsx"
    _create_users_admins_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT a.id, a.name, b.role FROM users a LEFT JOIN admins b ON a.id = b.id"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [
        (1, "Alice", "admin"),
        (2, "Bob", None),
        (3, "Charlie", "editor"),
    ]


def test_executor_join_with_where(tmp_path: Path):
    file_path = tmp_path / "join_where.xlsx"
    _create_users_admins_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT a.id, a.name, b.role FROM users a INNER JOIN admins b ON a.id = b.id WHERE b.role = 'admin'"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(1, "Alice", "admin")]


def test_executor_join_with_order_by(tmp_path: Path):
    file_path = tmp_path / "join_order.xlsx"
    _create_users_admins_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT a.id, b.role FROM users a INNER JOIN admins b ON a.id = b.id ORDER BY a.id DESC"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(3, "editor"), (1, "admin")]


def test_executor_join_with_limit_offset(tmp_path: Path):
    file_path = tmp_path / "join_limit.xlsx"
    _create_users_admins_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT a.id, a.name FROM users a LEFT JOIN admins b ON a.id = b.id LIMIT 2 OFFSET 1"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(2, "Bob"), (3, "Charlie")]


def test_executor_left_join_no_match(tmp_path: Path):
    """LEFT JOIN where right side has no matches at all -> all right columns None."""
    file_path = tmp_path / "left_join_no_match.xlsx"
    workbook = Workbook()
    users = workbook.active
    assert users is not None
    users.title = "users"
    users.append(["id", "name"])
    users.append([10, "Xander"])
    users.append([20, "Yara"])

    admins = workbook.create_sheet("admins")
    admins.append(["id", "role"])
    admins.append([1, "admin"])
    workbook.save(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT a.id, a.name, b.role FROM users a LEFT JOIN admins b ON a.id = b.id"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(10, "Xander", None), (20, "Yara", None)]


def test_executor_join_multi_condition_on(tmp_path: Path):
    """ON with AND (two equality conditions)."""
    file_path = tmp_path / "join_multi_on.xlsx"
    workbook = Workbook()
    left = workbook.active
    assert left is not None
    left.title = "employees"
    left.append(["dept", "grade", "name"])
    left.append(["eng", "senior", "Alice"])
    left.append(["eng", "junior", "Bob"])
    left.append(["hr", "senior", "Charlie"])

    right = workbook.create_sheet("salaries")
    right.append(["dept", "grade", "salary"])
    right.append(["eng", "senior", 150])
    right.append(["eng", "junior", 100])
    right.append(["hr", "senior", 120])
    workbook.save(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT a.name, b.salary FROM employees a INNER JOIN salaries b "
        "ON a.dept = b.dept AND a.grade = b.grade ORDER BY b.salary DESC"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [
        ("Alice", 150),
        ("Charlie", 120),
        ("Bob", 100),
    ]


def test_executor_join_table_name_qualifiers(tmp_path: Path):
    """Use table names (not aliases) as qualifiers."""
    file_path = tmp_path / "join_table_names.xlsx"
    _create_users_admins_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT users.id, admins.role FROM users INNER JOIN admins ON users.id = admins.id"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(1, "admin"), (3, "editor")]


def test_executor_join_missing_right_sheet(tmp_path: Path):
    """Joining a non-existent sheet raises ValueError."""
    file_path = tmp_path / "join_missing.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "users"
    sheet.append(["id"])
    sheet.append([1])
    workbook.save(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT a.id, b.id FROM users a INNER JOIN nonexistent b ON a.id = b.id"
    )
    with pytest.raises(DatabaseError, match="not found"):
        SharedExecutor(engine).execute(parsed)


def test_executor_join_null_keys_do_not_match(tmp_path: Path):
    """NULL join keys must not match per SQL standard: NULL != NULL."""
    file_path = tmp_path / "join_null_keys.xlsx"
    workbook = Workbook()
    left = workbook.active
    assert left is not None
    left.title = "left_t"
    left.append(["id", "name"])
    left.append([1, "Alice"])
    left.append([None, "Bob"])  # NULL key
    left.append([3, "Charlie"])

    right = workbook.create_sheet("right_t")
    right.append(["id", "role"])
    right.append([1, "admin"])
    right.append([None, "ghost"])  # NULL key
    right.append([3, "editor"])
    workbook.save(file_path)

    engine = OpenpyxlBackend(str(file_path))
    # INNER JOIN: NULL keys should NOT match
    parsed = parse_sql(
        "SELECT a.id, a.name, b.role FROM left_t a INNER JOIN right_t b ON a.id = b.id"
    )
    results = SharedExecutor(engine).execute(parsed)
    # Bob should NOT appear — NULL != NULL
    assert results.rows == [(1, "Alice", "admin"), (3, "Charlie", "editor")]


def test_executor_left_join_null_keys_unmatched(tmp_path: Path):
    """LEFT JOIN with NULL keys: left row with NULL key gets NULL-filled right columns."""
    file_path = tmp_path / "left_join_null_keys.xlsx"
    workbook = Workbook()
    left = workbook.active
    assert left is not None
    left.title = "left_t"
    left.append(["id", "name"])
    left.append([1, "Alice"])
    left.append([None, "Bob"])

    right = workbook.create_sheet("right_t")
    right.append(["id", "role"])
    right.append([1, "admin"])
    right.append([None, "ghost"])
    workbook.save(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT a.id, a.name, b.role FROM left_t a LEFT JOIN right_t b ON a.id = b.id"
    )
    results = SharedExecutor(engine).execute(parsed)
    # Bob (NULL key) has no match → right columns are None
    assert results.rows == [(1, "Alice", "admin"), (None, "Bob", None)]


def test_executor_join_numeric_string_coercion(tmp_path: Path):
    """Join key '1' (string) should match 1 (int) via numeric coercion."""
    file_path = tmp_path / "join_coercion.xlsx"
    workbook = Workbook()
    left = workbook.active
    assert left is not None
    left.title = "left_t"
    left.append(["id", "name"])
    left.append(["1", "Alice"])  # string '1'
    left.append(["2", "Bob"])

    right = workbook.create_sheet("right_t")
    right.append(["id", "role"])
    right.append([1, "admin"])  # int 1
    right.append([2, "editor"])
    workbook.save(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT a.id, a.name, b.role FROM left_t a INNER JOIN right_t b ON a.id = b.id"
    )
    results = SharedExecutor(engine).execute(parsed)
    assert len(results.rows) == 2
    assert results.rows[0][1] == "Alice"
    assert results.rows[0][2] == "admin"


def test_executor_join_unknown_select_column(tmp_path: Path):
    """SELECT referencing non-existent column in JOIN query raises ValueError."""
    file_path = tmp_path / "join_bad_col.xlsx"
    workbook = Workbook()
    left = workbook.active
    assert left is not None
    left.title = "users"
    left.append(["id", "name"])
    left.append([1, "Alice"])

    right = workbook.create_sheet("admins")
    right.append(["id", "role"])
    right.append([1, "admin"])
    workbook.save(file_path)

    engine = OpenpyxlBackend(str(file_path))
    # a.nonexistent doesn't exist in users sheet
    parsed = parse_sql(
        "SELECT a.nonexistent FROM users a INNER JOIN admins b ON a.id = b.id"
    )
    with pytest.raises(DatabaseError, match="Unknown column"):
        SharedExecutor(engine).execute(parsed)


def test_executor_join_unknown_on_column(tmp_path: Path):
    """ON referencing non-existent column raises ValueError."""
    file_path = tmp_path / "join_bad_on.xlsx"
    workbook = Workbook()
    left = workbook.active
    assert left is not None
    left.title = "users"
    left.append(["id", "name"])
    left.append([1, "Alice"])

    right = workbook.create_sheet("admins")
    right.append(["id", "role"])
    right.append([1, "admin"])
    workbook.save(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT a.id, b.id FROM users a INNER JOIN admins b ON a.id = b.nonexistent"
    )
    with pytest.raises(DatabaseError, match="Unknown column"):
        SharedExecutor(engine).execute(parsed)


def test_executor_join_with_as_alias(tmp_path: Path):
    """JOIN with AS keyword in alias (SQLAlchemy compatibility)."""
    file_path = tmp_path / "join_as_alias.xlsx"
    workbook = Workbook()
    left = workbook.active
    assert left is not None
    left.title = "users"
    left.append(["id", "name"])
    left.append([1, "Alice"])
    left.append([2, "Bob"])

    right = workbook.create_sheet("admins")
    right.append(["id", "role"])
    right.append([1, "admin"])
    workbook.save(file_path)

    engine = OpenpyxlBackend(str(file_path))
    # Use 'AS' keyword in both FROM and JOIN aliases
    parsed = parse_sql(
        "SELECT a.id, a.name, b.role FROM users AS a INNER JOIN admins AS b ON a.id = b.id"
    )
    results = SharedExecutor(engine).execute(parsed)
    assert results.rows == [(1, "Alice", "admin")]



class _NonTransactionalBackend:
    supports_transactions = False
    readonly = False

    def __init__(self) -> None:
        self._sheets = {
            "people": TableData(headers=["id", "name"], rows=[[1, "Alice"]])
        }

    def list_sheets(self) -> list[str]:
        return list(self._sheets)

    def read_sheet(self, name: str) -> TableData:
        return self._sheets[name]

    def write_sheet(self, name: str, data: TableData) -> None:
        self._sheets[name] = data

def test_non_transactional_metadata_read_failure_skips_lossy_rewrite(
    monkeypatch: pytest.MonkeyPatch,
) -> None:

    backend = _NonTransactionalBackend()
    executor = SharedExecutor(cast(Any, backend), connection=object())
    write_calls = 0

    def _boom_read(*_: object, **__: object) -> list[dict[str, object]]:
        raise RuntimeError("metadata read failed")

    def _track_write(*_: object, **__: object) -> None:
        nonlocal write_calls
        write_calls += 1

    monkeypatch.setattr(reflection, "read_table_metadata", _boom_read)
    monkeypatch.setattr(reflection, "write_table_metadata", _track_write)

    with pytest.warns(UserWarning, match="skipping metadata update to avoid data loss"):
        executor.execute(parse_sql("ALTER TABLE people DROP COLUMN name"))

    assert write_calls == 0



def _create_round11_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Sheet1"
    sheet.append(["id", "name"])
    sheet.append([1, "Alice"])
    sheet.append([2, "Bob"])

    table = workbook.create_sheet("t")
    table.append(["id", "a", "b", "c"])
    table.append([1, 10, 0, None])
    table.append([2, "alice", 0, None])

    workbook.save(path)

def test_update_set_supports_column_arithmetic_cast_and_function(
    tmp_path: Path,
) -> None:
    file_path = tmp_path / "round11_update_expressions.xlsx"
    _create_round11_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()

        cursor.execute("UPDATE t SET b = a WHERE id = 1")
        cursor.execute("UPDATE t SET b = a + 1 WHERE id = 1")
        cursor.execute("UPDATE t SET b = CAST(a AS TEXT) WHERE id = 1")
        cursor.execute("UPDATE t SET c = UPPER(a) WHERE id = 2")

        cursor.execute("SELECT b, c FROM t WHERE id = 1")
        assert cursor.fetchone() == ("10", None)
        cursor.execute("SELECT b, c FROM t WHERE id = 2")
        assert cursor.fetchone() == (0, "ALICE")



def _create_round12_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "t"
    sheet.append(["id", "txt", "col"])
    sheet.append([1, "alpha", 1])
    sheet.append([2, "beta", 2])
    sheet.append([3, "gamma", 3])
    workbook.save(path)

def test_update_where_detection_ignores_where_inside_string_literal(
    tmp_path: Path,
) -> None:
    file_path = tmp_path / "round12_update_where_string.xlsx"
    _create_round12_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE t SET txt = 'value WHERE other' WHERE id = 1")
        cursor.execute("SELECT txt FROM t WHERE id = 1")
        assert cursor.fetchone() == ("value WHERE other",)



def _create_round13_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "t"
    sheet.append(["a", "b", "val", "grp"])
    sheet.append([1, 1, 6, "x"])
    sheet.append([2, 3, 5, "x"])
    sheet.append([1, 1, 7, "y"])
    sheet.append([5, 5, 1, "y"])
    workbook.save(path)

def test_where_column_to_column_comparison_works_on_rhs_identifier(
    tmp_path: Path,
) -> None:
    file_path = tmp_path / "round13_where_column_rhs.xlsx"
    _create_round13_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT a, b FROM t WHERE a = b ORDER BY a")
        assert cursor.fetchall() == [(1, 1), (1, 1), (5, 5)]

def test_where_reversed_operands_resolve_rhs_identifier_as_column(
    tmp_path: Path,
) -> None:
    file_path = tmp_path / "round13_where_reversed_operands.xlsx"
    _create_round13_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT a FROM t WHERE 1 = a ORDER BY b")
        assert cursor.fetchall() == [(1,), (1,)]

def test_where_quoted_literal_still_treated_as_literal(tmp_path: Path) -> None:
    file_path = tmp_path / "round13_where_quoted_literal.xlsx"
    _create_round13_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT a FROM t WHERE a = 'b'")
        assert cursor.fetchall() == []

def test_having_aggregate_collected_from_rhs(tmp_path: Path) -> None:
    file_path = tmp_path / "round13_having_rhs_aggregate.xlsx"
    _create_round13_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT grp, SUM(val) FROM t GROUP BY grp HAVING 10 < SUM(val) ORDER BY grp"
        )
        assert cursor.fetchall() == [("x", 11.0)]

def test_having_aggregate_collection_handles_and_tree(tmp_path: Path) -> None:
    file_path = tmp_path / "round13_having_and_tree.xlsx"
    _create_round13_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT grp FROM t GROUP BY grp HAVING SUM(a) > 0 AND COUNT(*) > 1 ORDER BY grp"
        )
        assert cursor.fetchall() == [("x",), ("y",)]

def test_having_aggregate_collection_handles_not_wrapper(tmp_path: Path) -> None:
    file_path = tmp_path / "round13_having_not_wrapper.xlsx"
    _create_round13_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT grp FROM t GROUP BY grp HAVING NOT (SUM(val) > 10) ORDER BY grp"
        )
        assert cursor.fetchall() == [("y",)]



def _create_r16_workbook(
    path: Path, sheet: str, headers: list[object], rows: list[list[object]]
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = sheet
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    workbook.close()

def test_executor_resolves_unicode_sheet_names_with_casefold(tmp_path: Path) -> None:
    file_path = tmp_path / "unicode-sheet-casefold.xlsx"
    _create_r16_workbook(file_path, "Straße", ["id"], [[1]])

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM STRASSE")
        assert cursor.fetchall() == [(1,)]



def _create_people_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Sheet1"
    sheet.append(["id", "Name", "phrase"])
    sheet.append([1, "Alice", "Stra\u00dfe"])
    sheet.append([2, "Bob", "Road"])
    workbook.save(path)
    workbook.close()

def test_select_update_insert_column_resolution_is_case_insensitive(
    tmp_path: Path,
) -> None:
    file_path = tmp_path / "column-casefold.xlsx"
    _create_people_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()

        cursor.execute("SELECT name FROM Sheet1 ORDER BY id")
        assert cursor.fetchall() == [("Alice",), ("Bob",)]

        cursor.execute("UPDATE Sheet1 SET name = 'Carol' WHERE Name = 'Alice'")
        assert cursor.rowcount == 1
        cursor.execute("SELECT Name FROM Sheet1 WHERE id = 1")
        assert cursor.fetchall() == [("Carol",)]

        cursor.execute("INSERT INTO Sheet1 (id, name, phrase) VALUES (3, 'Dana', 'x')")
        cursor.execute("SELECT Name FROM Sheet1 WHERE id = 3")
        assert cursor.fetchall() == [("Dana",)]



@pytest.fixture
def tmp_xlsx(tmp_path):
    """Create a minimal xlsx file with a Sheet1 containing headers and one row."""
    path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id", "name", "score"])
    ws.append([1, "Alice", 90])
    ws.append([2, "Bob", 80])
    ws.append([3, None, 70])  # Row with NULL name
    wb.save(str(path))
    wb.close()
    return str(path)

@pytest.fixture
def tmp_xlsx_path(tmp_path):
    """Return a path (but don't create the file) — for testing create=True / missing file."""
    return str(tmp_path / "missing.xlsx")

class TestAndOrPrecedence:
    def test_and_binds_tighter_than_or_openpyxl(self, tmp_xlsx):
        """WHERE a = 1 OR b = 'Bob' AND score = 80 should be a = 1 OR (b = 'Bob' AND score = 80)."""
        conn = ExcelConnection(tmp_xlsx, engine="openpyxl")
        cur = conn.cursor()
        # id=1 (Alice,90), id=2 (Bob,80), id=3 (None,70)
        # With correct precedence: id=1 OR (name='Bob' AND score=80)
        # Should match id=1 (Alice) and id=2 (Bob)
        # With wrong precedence (left-to-right): (id=1 OR name='Bob') AND score=80
        # Would only match id=2 (Bob)
        cur.execute("SELECT * FROM Sheet1 WHERE id = 1 OR name = 'Bob' AND score = 80")
        rows = cur.fetchall()
        assert len(rows) == 2, (
            f"Expected 2 rows (AND before OR), got {len(rows)}: {rows}"
        )
        conn.close()

    def test_and_binds_tighter_than_or_pandas(self, tmp_xlsx):
        """Same test for pandas engine."""
        conn = ExcelConnection(tmp_xlsx, engine="pandas")
        cur = conn.cursor()
        cur.execute("SELECT * FROM Sheet1 WHERE id = 1 OR name = 'Bob' AND score = 80")
        rows = cur.fetchall()
        assert len(rows) == 2, (
            f"Expected 2 rows (AND before OR), got {len(rows)}: {rows}"
        )
        conn.close()

    def test_all_and_still_works(self, tmp_xlsx):
        """All AND conditions should still work correctly."""
        conn = ExcelConnection(tmp_xlsx, engine="openpyxl")
        cur = conn.cursor()
        cur.execute("SELECT * FROM Sheet1 WHERE id = 2 AND name = 'Bob' AND score = 80")
        rows = cur.fetchall()
        assert len(rows) == 1
        conn.close()

    def test_all_or_still_works(self, tmp_xlsx):
        """All OR conditions should still work correctly."""
        conn = ExcelConnection(tmp_xlsx, engine="openpyxl")
        cur = conn.cursor()
        cur.execute("SELECT * FROM Sheet1 WHERE id = 1 OR id = 2 OR id = 3")
        rows = cur.fetchall()
        assert len(rows) == 3
        conn.close()

class TestNullHandling:
    def test_is_null_openpyxl(self, tmp_xlsx):
        """IS NULL should match rows where column is None."""
        conn = ExcelConnection(tmp_xlsx, engine="openpyxl")
        cur = conn.cursor()
        cur.execute("SELECT * FROM Sheet1 WHERE name IS NULL")
        rows = cur.fetchall()
        assert len(rows) == 1
        assert rows[0][0] == 3  # id=3 has None name
        conn.close()

    def test_is_not_null_openpyxl(self, tmp_xlsx):
        """IS NOT NULL should match rows where column is not None."""
        conn = ExcelConnection(tmp_xlsx, engine="openpyxl")
        cur = conn.cursor()
        cur.execute("SELECT * FROM Sheet1 WHERE name IS NOT NULL")
        rows = cur.fetchall()
        assert len(rows) == 2
        conn.close()

    def test_is_null_pandas(self, tmp_xlsx):
        """IS NULL should work with pandas engine too."""
        conn = ExcelConnection(tmp_xlsx, engine="pandas")
        cur = conn.cursor()
        cur.execute("SELECT * FROM Sheet1 WHERE name IS NULL")
        rows = cur.fetchall()
        # pandas reads None from xlsx as NaN, which is also null
        assert len(rows) == 1
        conn.close()

    def test_is_not_null_pandas(self, tmp_xlsx):
        conn = ExcelConnection(tmp_xlsx, engine="pandas")
        cur = conn.cursor()
        cur.execute("SELECT * FROM Sheet1 WHERE name IS NOT NULL")
        rows = cur.fetchall()
        assert len(rows) == 2
        conn.close()

    def test_equality_with_null_returns_false(self, tmp_xlsx):
        """col = NULL should not match anything (SQL semantics)."""
        conn = ExcelConnection(tmp_xlsx, engine="openpyxl")
        cur = conn.cursor()
        cur.execute("SELECT * FROM Sheet1 WHERE name = NULL")
        rows = cur.fetchall()
        assert len(rows) == 0, "Equality comparison with NULL should return no rows"
        conn.close()

    def test_parse_is_null(self):

        parsed = parse_sql("SELECT * FROM Sheet1 WHERE name IS NULL")
        cond = parsed["where"]["conditions"][0]
        assert cond["column"] == "name"
        assert cond["operator"] == "IS"
        assert cond["value"] is None

    def test_parse_is_not_null(self):

        parsed = parse_sql("SELECT * FROM Sheet1 WHERE name IS NOT NULL")
        cond = parsed["where"]["conditions"][0]
        assert cond["column"] == "name"
        assert cond["operator"] == "IS NOT"
        assert cond["value"] is None



def test_pandas_backend_error_paths_and_execute_wrappers(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    file_path = tmp_path / "pandas.xlsx"
    pd.DataFrame([{"id": 1, "name": "Alice"}]).to_excel(
        file_path, index=False, sheet_name="Sheet1"
    )
    backend = PandasBackend(str(file_path), create=False)

    with pytest.raises(DatabaseError, match="not found"):
        backend.read_sheet("Missing")
    with pytest.raises(DatabaseError, match="not found"):
        backend.write_sheet("Missing", TableData(headers=["a"], rows=[]))
    with pytest.raises(DatabaseError, match="not found"):
        backend.append_row("Missing", [1])
    with pytest.raises(DatabaseError, match="already exists"):
        backend.create_sheet("Sheet1", ["a"])
    with pytest.raises(DatabaseError, match="not found"):
        backend.drop_sheet("Missing")

    result1 = backend.execute("SELECT * FROM Sheet1")
    result2 = backend.execute_with_params("SELECT * FROM Sheet1 WHERE id = ?", (1,))
    assert result1.rowcount >= 1
    assert result2.rowcount == 1

    created_temp: dict[str, str] = {}

    def fail_replace(src: str, dst: str) -> None:
        del dst
        created_temp["path"] = src
        raise OSError("forced replace failure")

    monkeypatch.setattr("excel_dbapi.engines.pandas.backend.os.replace", fail_replace)
    with pytest.raises(OSError, match="forced replace failure"):
        backend.save()
    temp_path = Path(created_temp["path"])
    assert not temp_path.exists()
