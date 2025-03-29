import pytest

from excel_dbapi.connection import ExcelConnection
from excel_dbapi.exceptions import OperationalError
from excel_dbapi.table import ExcelTable


def test_table_fetch_all(tmp_path):
    from openpyxl import Workbook

    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Age"])
    ws.append(["Alice", 25])
    ws.append(["Bob", 30])
    wb.save(file_path)

    with ExcelConnection(file_path) as conn:
        with ExcelTable(conn, "Sheet") as table:
            data = table.fetch_all()
            assert data == [["Name", "Age"], ["Alice", 25], ["Bob", 30]]


def test_table_fetch_row(tmp_path):
    from openpyxl import Workbook

    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Age"])
    ws.append(["Alice", 25])
    wb.save(file_path)

    with ExcelConnection(file_path) as conn:
        with ExcelTable(conn, "Sheet") as table:
            row = table.fetch_row(2)
            assert row == ["Alice", 25]


def test_table_invalid_sheet(tmp_path):
    from openpyxl import Workbook

    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    wb.save(file_path)

    with ExcelConnection(file_path) as conn:
        with pytest.raises(OperationalError):
            with ExcelTable(conn, "Nonexistent"):
                pass
