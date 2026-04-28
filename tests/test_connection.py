import os
import stat
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook

from excel_dbapi import ExcelConnection, connect
from excel_dbapi.connection import _resolve_engine_and_location
from excel_dbapi.engines.base import _normalize_headers
from excel_dbapi.exceptions import (
    BackendOperationError,
    DataError,
    DatabaseError,
    InterfaceError,
    NotSupportedError,
    OperationalError,
    ProgrammingError,
)


def test_connection_open_and_close():
    conn = ExcelConnection("tests/data/sample.xlsx")
    assert conn.closed is False
    conn.close()
    assert conn.closed is True


def test_connection_cursor():
    conn = ExcelConnection("tests/data/sample.xlsx")
    cursor = conn.cursor()
    assert cursor is not None
    conn.close()
    with pytest.raises(InterfaceError):
        conn.cursor()


def test_rollback_autocommit_raises():
    with ExcelConnection("tests/data/sample.xlsx", autocommit=True) as conn:
        with pytest.raises(NotSupportedError):
            conn.rollback()


@pytest.mark.parametrize(
    ("raised", "expected"),
    [
        (ValueError("bad query"), ProgrammingError),
        (NotImplementedError("not supported"), NotSupportedError),
        (RuntimeError("boom"), DatabaseError),
    ],
)
def test_connection_execute_maps_exceptions(raised, expected):
    conn = ExcelConnection("tests/data/sample.xlsx")

    def _raise(*args, **kwargs):
        del args, kwargs
        raise raised

    conn._executor.execute_with_params = _raise

    with pytest.raises(expected):
        conn.execute("SELECT * FROM Sheet1")

    conn.close()


def test_nonexistent_file_raises_operational_error():
    """Issue 1: FileNotFoundError should be wrapped as OperationalError."""
    with pytest.raises(OperationalError):
        ExcelConnection("nonexistent_file.xlsx")


def test_bad_graph_dsn_raises_operational_error():
    """Issue 1: Invalid DSN scheme should be rejected as OperationalError."""
    with pytest.raises(OperationalError, match="Unsupported DSN scheme"):
        ExcelConnection("bad://dsn")


def test_corrupt_file_raises_operational_error(tmp_path):
    """BadZipFile from corrupt .xlsx must be wrapped as OperationalError."""
    bad_file = tmp_path / "corrupt.xlsx"
    bad_file.write_bytes(b"not a real xlsx file")

    with pytest.raises(OperationalError):
        ExcelConnection(str(bad_file))



def _create_r7_workbook(
    path: Path, headers: list[object], rows: list[list[object]]
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Sheet1"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()

def test_pandas_backend_rejects_blank_headers(tmp_path: Path) -> None:
    file_path = tmp_path / "pandas-unnamed.xlsx"
    _create_r7_workbook(file_path, ["id", None], [[1, "Alice"]])

    with pytest.raises(DataError, match="Empty or None header"):
        ExcelConnection(str(file_path), engine="pandas")

def test_pandas_backend_rejects_duplicate_headers(tmp_path: Path) -> None:
    file_path = tmp_path / "pandas-duplicate.xlsx"
    _create_r7_workbook(file_path, ["id", "id"], [[1, 2]])

    with pytest.raises(DataError, match="Duplicate header"):
        ExcelConnection(str(file_path), engine="pandas")

def test_pandas_backend_rejects_data_only_false(tmp_path: Path) -> None:
    file_path = tmp_path / "pandas-data-only.xlsx"
    _create_r7_workbook(file_path, ["id"], [[1]])

    with pytest.raises(NotSupportedError, match="does not support data_only=False"):
        ExcelConnection(str(file_path), engine="pandas", data_only=False)



def _create_people_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Sheet1"
    sheet.append(["id", "Name", "phrase"])
    sheet.append([1, "Alice", "Stra\u00dfe"])
    sheet.append([2, "Bob", "Road"])
    workbook.save(path)
    workbook.close()

def test_commit_wraps_non_dbapi_backend_exceptions(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    file_path = tmp_path / "commit-wrap.xlsx"
    _create_people_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:

        def _boom_save() -> None:
            raise RuntimeError("save failed")

        monkeypatch.setattr(conn.engine, "save", _boom_save)
        with pytest.raises(OperationalError, match="save failed"):
            conn.commit()

def test_close_wraps_non_dbapi_backend_exceptions(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    file_path = tmp_path / "close-wrap.xlsx"
    _create_people_workbook(file_path)

    conn = ExcelConnection(str(file_path), engine="openpyxl", autocommit=True)

    def _boom_close() -> None:
        raise RuntimeError("close failed")

    monkeypatch.setattr(conn.engine, "close", _boom_close)
    with pytest.raises(OperationalError, match="close failed"):
        conn.close()



def _make_xlsx(
    path: Path,
    sheet: str = "users",
    headers: list[str] | None = None,
    rows: list[list[Any]] | None = None,
) -> str:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet
    ws.append(headers or ["id", "name"])
    for row in rows or []:
        ws.append(row)
    fpath = str(path)
    wb.save(fpath)
    wb.close()
    return fpath

class TestConnectionExecuteAutocommit:
    """Fix 1: connection.execute() must save when autocommit=True."""

    def test_connection_execute_insert_persists(self, tmp_path: Path) -> None:
        """connection.execute('INSERT ...') persists data when autocommit=True."""
        fpath = _make_xlsx(tmp_path / "test.xlsx")
        with ExcelConnection(fpath, engine="openpyxl", autocommit=True) as conn:
            conn.execute("INSERT INTO users VALUES (1, 'Alice')")

        # Re-open and verify data was persisted
        with ExcelConnection(fpath, engine="openpyxl") as conn:
            result = conn.execute("SELECT * FROM users")
            assert len(result.rows) == 1
            assert result.rows[0] == (1, "Alice")

    def test_connection_execute_update_persists(self, tmp_path: Path) -> None:
        """connection.execute('UPDATE ...') persists data."""
        fpath = _make_xlsx(tmp_path / "test.xlsx", rows=[[1, "Alice"]])
        with ExcelConnection(fpath, engine="openpyxl", autocommit=True) as conn:
            conn.execute("UPDATE users SET name = 'Bob' WHERE id = 1")

        with ExcelConnection(fpath, engine="openpyxl") as conn:
            result = conn.execute("SELECT * FROM users WHERE id = 1")
            assert result.rows[0] == (1, "Bob")

    def test_connection_execute_delete_persists(self, tmp_path: Path) -> None:
        """connection.execute('DELETE ...') persists data."""
        fpath = _make_xlsx(tmp_path / "test.xlsx", rows=[[1, "Alice"], [2, "Bob"]])
        with ExcelConnection(fpath, engine="openpyxl", autocommit=True) as conn:
            conn.execute("DELETE FROM users WHERE id = 1")

        with ExcelConnection(fpath, engine="openpyxl") as conn:
            result = conn.execute("SELECT * FROM users")
            assert len(result.rows) == 1

    def test_connection_execute_no_save_without_autocommit(
        self, tmp_path: Path
    ) -> None:
        """connection.execute() does NOT save when autocommit=False."""
        fpath = _make_xlsx(tmp_path / "test.xlsx")
        with ExcelConnection(fpath, engine="openpyxl", autocommit=False) as conn:
            conn.execute("INSERT INTO users VALUES (1, 'Alice')")
            # Not committed — rollback
            conn.rollback()

        with ExcelConnection(fpath, engine="openpyxl") as conn:
            result = conn.execute("SELECT * FROM users")
            assert len(result.rows) == 0

    def test_connection_execute_select_no_save(self, tmp_path: Path) -> None:
        """SELECT via connection.execute() does NOT trigger save."""
        fpath = _make_xlsx(tmp_path / "test.xlsx", rows=[[1, "Alice"]])
        mtime_before = os.path.getmtime(fpath)
        import time; time.sleep(0.05)  # ensure measurable mtime gap if save occurs
        with ExcelConnection(fpath, engine="openpyxl", autocommit=True) as conn:
            result = conn.execute("SELECT * FROM users")
            assert len(result.rows) == 1
        mtime_after = os.path.getmtime(fpath)
        assert mtime_after == mtime_before, "SELECT should not trigger save"
    def test_executemany_saves_once_not_per_row(self, tmp_path: Path) -> None:
        """executemany() should save only once at end, not per row."""
        fpath = _make_xlsx(tmp_path / "test.xlsx")
        save_count = 0
        with ExcelConnection(fpath, engine="openpyxl", autocommit=True) as conn:
            original_save = conn.engine.save

            def counting_save() -> None:
                nonlocal save_count
                save_count += 1
                original_save()

            conn.engine.save = counting_save  # type: ignore[assignment]
            cursor = conn.cursor()
            cursor.executemany(
                "INSERT INTO users VALUES (?, ?)",
                [(1, "Alice"), (2, "Bob"), (3, "Charlie")],
            )
        assert save_count == 1, f"Expected 1 save, got {save_count}"

    def test_connection_execute_create_table_persists(self, tmp_path: Path) -> None:
        """connection.execute('CREATE TABLE ...') persists."""
        fpath = _make_xlsx(tmp_path / "test.xlsx")
        with ExcelConnection(fpath, engine="openpyxl", autocommit=True) as conn:
            conn.execute("CREATE TABLE orders (id INTEGER, product TEXT)")

        with ExcelConnection(fpath, engine="openpyxl") as conn:
            result = conn.execute("SELECT * FROM orders")
            assert result.rows == []

class TestDSNAutoEngineSelection:
    """Fix 2: engine default=None allows DSN-based auto-detection."""

    def test_connect_defaults_to_openpyxl_for_local_files(self, tmp_path: Path) -> None:
        """connect() with local file still uses openpyxl when engine=None."""
        fpath = _make_xlsx(tmp_path / "test.xlsx")
        conn = connect(fpath, autocommit=True)
        assert conn.engine_name == "openpyxl"
        conn.close()

    def test_connect_explicit_engine_still_works(self, tmp_path: Path) -> None:
        """Explicit engine='openpyxl' still works."""
        fpath = _make_xlsx(tmp_path / "test.xlsx")
        conn = connect(fpath, engine="openpyxl", autocommit=True)
        assert conn.engine_name == "openpyxl"
        conn.close()

    def test_dsn_engine_auto_detection(self) -> None:
        """msgraph:// DSN auto-detects 'graph' engine (would fail with
        engine='openpyxl' default due to mismatch error)."""

        engine, location = _resolve_engine_and_location(
            "msgraph://drives/fake/items/fake", None
        )
        assert engine == "graph"
        assert location == "msgraph://drives/fake/items/fake"

    def test_dsn_engine_mismatch_raises(self) -> None:
        """Explicit engine conflicting with DSN raises BackendOperationError."""

        with pytest.raises(BackendOperationError, match="Engine mismatch"):
            _resolve_engine_and_location("msgraph://drives/fake/items/fake", "openpyxl")

    @pytest.mark.parametrize(
        ("dsn", "expected_location"),
        [
            (
                "sharepoint://sites/team/drives/drive-1/items/item-1",
                "sharepoint://sites/team/drives/drive-1/items/item-1",
            ),
            (
                "onedrive://me/drive/items/item-2",
                "onedrive://me/drive/items/item-2",
            ),
        ],
    )
    def test_extended_graph_schemes_auto_detected(
        self, dsn: str, expected_location: str
    ) -> None:

        engine, location = _resolve_engine_and_location(dsn, None)
        assert engine == "graph"
        assert location == expected_location

class TestHeaderNormalization:
    """Fix 4: _normalize_headers validates and coerces headers."""

    def test_valid_headers(self) -> None:
        result = _normalize_headers(["id", "name", "age"])
        assert result == ["id", "name", "age"]

    def test_numeric_headers_coerced_to_string(self) -> None:
        result = _normalize_headers([1, 2, 3])
        assert result == ["1", "2", "3"]

    def test_none_header_raises_data_error(self) -> None:
        with pytest.raises(DataError, match="Empty or None header at column index 1"):
            _normalize_headers(["id", None, "age"])

    def test_empty_string_header_raises_data_error(self) -> None:
        with pytest.raises(DataError, match="Empty or None header"):
            _normalize_headers(["id", "", "age"])

    def test_whitespace_only_header_raises_data_error(self) -> None:
        with pytest.raises(DataError, match="Empty or None header"):
            _normalize_headers(["id", "   ", "age"])

    def test_duplicate_headers_raises_data_error(self) -> None:
        with pytest.raises(DataError, match="Duplicate header"):
            _normalize_headers(["id", "name", "id"])

    def test_case_insensitive_duplicate_raises_data_error(self) -> None:
        with pytest.raises(DataError, match="Duplicate header"):
            _normalize_headers(["Name", "age", "name"])

    def test_duplicate_headers_in_xlsx(self, tmp_path: Path) -> None:
        """Reading a sheet with duplicate headers raises DataError."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "bad"
        ws.append(["id", "name", "id"])
        ws.append([1, "Alice", 2])
        fpath = str(tmp_path / "dup.xlsx")
        wb.save(fpath)
        wb.close()

        with ExcelConnection(fpath, engine="openpyxl") as conn:
            with pytest.raises(DataError, match="Duplicate header"):
                conn.execute("SELECT * FROM bad")

    def test_empty_header_in_xlsx(self, tmp_path: Path) -> None:
        """Reading a sheet with empty/None header raises DataError."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "bad"
        ws.append(["id", None, "age"])
        ws.append([1, "Alice", 25])
        fpath = str(tmp_path / "empty.xlsx")
        wb.save(fpath)
        wb.close()

        with ExcelConnection(fpath, engine="openpyxl") as conn:
            with pytest.raises(DataError, match="Empty or None header"):
                conn.execute("SELECT * FROM bad")

class TestHeaderWhitespaceTrimming:
    """Regression: _normalize_headers strips leading/trailing whitespace."""

    def test_whitespace_trimmed_from_headers(self) -> None:
        """Headers with leading/trailing whitespace are trimmed."""
        result = _normalize_headers(["  id  ", " name ", "age"])
        assert result == ["id", "name", "age"]

    def test_trimmed_duplicates_detected(self) -> None:
        """After trimming, duplicate headers are detected."""
        with pytest.raises(DataError, match="Duplicate header"):
            _normalize_headers(["  id  ", "id", "name"])



@pytest.fixture
def tmp_xlsx(tmp_path):
    """Create a minimal xlsx file with a Sheet1 containing headers and one row."""
    path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id", "name", "score"])
    ws.append([1, "Alice", 90])
    ws.append([2, "Bob", 80])
    ws.append([3, None, 70])  # Row with NULL name
    wb.save(str(path))
    wb.close()
    return str(path)

@pytest.fixture
def tmp_xlsx_path(tmp_path):
    """Return a path (but don't create the file) — for testing create=True / missing file."""
    return str(tmp_path / "missing.xlsx")

class TestPathCanonicalization:
    def test_dotdot_in_path_is_resolved(self, tmp_path, tmp_xlsx):
        # Paths with '..' are canonicalized, not rejected (library, not sandbox)
        # Create the file first, then reference it with '..'
        subdir = tmp_path / "subdir"
        subdir.mkdir()
        path_with_dotdot = str(subdir / ".." / "test.xlsx")
        conn = ExcelConnection(path_with_dotdot)
        # The stored path should be the resolved canonical form
        assert ".." not in conn.file_path
        conn.close()

    def test_absolute_path_accepted(self, tmp_xlsx):
        conn = ExcelConnection(tmp_xlsx)
        assert conn.closed is False
        conn.close()

    def test_resolved_path_stored(self, tmp_xlsx):
        conn = ExcelConnection(tmp_xlsx)
        assert os.path.isabs(conn.file_path)
        conn.close()

    def test_tilde_expanded(self, tmp_xlsx):
        # expanduser should work (though may not change anything in test env)
        conn = ExcelConnection(tmp_xlsx)
        assert "~" not in conn.file_path
        conn.close()

class TestTempFilePermissions:
    def test_openpyxl_save_creates_restricted_temp(self, tmp_xlsx):
        """After save, the target file should exist and be writable by owner."""
        conn = ExcelConnection(tmp_xlsx, autocommit=False)
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO Sheet1 (id, name, score) VALUES (?, ?, ?)", (10, "Test", 100)
        )
        conn.commit()
        conn.close()
        # File should still be readable
        assert os.path.exists(tmp_xlsx)
        mode = os.stat(tmp_xlsx).st_mode
        # Owner should have read+write
        assert mode & stat.S_IRUSR
        assert mode & stat.S_IWUSR

    def test_pandas_save_creates_restricted_temp(self, tmp_xlsx):
        """Same test for pandas engine."""
        conn = ExcelConnection(tmp_xlsx, engine="pandas", autocommit=False)
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO Sheet1 (id, name, score) VALUES (?, ?, ?)", (10, "Test", 100)
        )
        conn.commit()
        conn.close()
        assert os.path.exists(tmp_xlsx)

class TestAutocommitSnapshot:
    def test_autocommit_write_updates_snapshot(self, tmp_xlsx):
        """After an autocommit write, switching to manual mode and rolling back
        should NOT undo the autocommitted data."""
        conn = ExcelConnection(tmp_xlsx, autocommit=True)
        cur = conn.cursor()

        # Autocommit write — should save AND update snapshot
        cur.execute(
            "INSERT INTO Sheet1 (id, name, score) VALUES (?, ?, ?)", (99, "Zara", 100)
        )

        # Switch to manual mode
        conn.autocommit = False

        # Rollback should restore to post-autocommit state (not pre-autocommit)
        conn.rollback()

        # The autocommitted row should still be present
        cur.execute("SELECT * FROM Sheet1 WHERE id = 99")
        rows = cur.fetchall()
        assert len(rows) == 1, (
            "Autocommitted row should survive rollback after switching to manual mode"
        )
        conn.close()

    def test_executemany_autocommit_updates_snapshot(self, tmp_xlsx):
        """executemany with autocommit should also update snapshot."""
        conn = ExcelConnection(tmp_xlsx, autocommit=True)
        cur = conn.cursor()

        cur.executemany(
            "INSERT INTO Sheet1 (id, name, score) VALUES (?, ?, ?)",
            [(10, "X", 50), (11, "Y", 60)],
        )

        conn.autocommit = False
        conn.rollback()

        cur.execute("SELECT * FROM Sheet1 WHERE id = 10")
        assert len(cur.fetchall()) == 1
        cur.execute("SELECT * FROM Sheet1 WHERE id = 11")
        assert len(cur.fetchall()) == 1
        conn.close()

    def test_autocommit_write_persisted_on_disk_after_rollback(self, tmp_xlsx):
        """GH#13 exact reproduction: autocommit write must survive rollback AND
        be present on disk (not just in memory). Verifies the file is persisted."""
        conn = ExcelConnection(tmp_xlsx, autocommit=True)
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO Sheet1 (id, name, score) VALUES (?, ?, ?)", (42, "Persist", 95)
        )

        # Toggle autocommit off and rollback
        conn.autocommit = False
        conn.rollback()
        conn.close()

        # Re-open the file fresh — verify data is on disk
        conn2 = ExcelConnection(tmp_xlsx, autocommit=True)
        cur2 = conn2.cursor()
        cur2.execute("SELECT * FROM Sheet1 WHERE id = 42")
        rows = cur2.fetchall()
        assert len(rows) == 1, "Autocommitted row must be on disk even after rollback"
        assert rows[0][1] == "Persist"
        conn2.close()

    def test_autocommit_data_only_false_updates_snapshot(self, tmp_xlsx):
        """Snapshot is updated after autocommit write even with data_only=False."""
        conn = ExcelConnection(tmp_xlsx, autocommit=True, data_only=False)
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO Sheet1 (id, name, score) VALUES (?, ?, ?)", (77, "NoCache", 80)
        )
        conn.autocommit = False
        conn.rollback()
        cur.execute("SELECT * FROM Sheet1 WHERE id = 77")
        assert len(cur.fetchall()) == 1, (
            "Autocommitted row should survive rollback with data_only=False"
        )
        conn.close()

    def test_multiple_autocommit_writes_snapshot_stays_current(self, tmp_xlsx):
        """Two consecutive autocommit writes: rollback after switching to manual
        should keep both rows (snapshot updated after each save)."""
        conn = ExcelConnection(tmp_xlsx, autocommit=True)
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO Sheet1 (id, name, score) VALUES (?, ?, ?)", (50, "First", 10)
        )
        cur.execute(
            "INSERT INTO Sheet1 (id, name, score) VALUES (?, ?, ?)", (51, "Second", 20)
        )
        conn.autocommit = False
        conn.rollback()
        cur.execute("SELECT * FROM Sheet1 WHERE id IN (50, 51)")
        rows = cur.fetchall()
        assert len(rows) == 2, (
            "Both autocommitted rows should survive rollback"
        )
        conn.close()

class TestExceptionTypes:
    def test_unsupported_engine_raises_operational_error(self, tmp_xlsx):
        with pytest.raises(NotSupportedError, match="Unsupported engine"):
            ExcelConnection(tmp_xlsx, engine="sqlite")

    def test_missing_file_raises_operational_error(self, tmp_xlsx_path):
        with pytest.raises(OperationalError, match="not found"):
            ExcelConnection(tmp_xlsx_path)

    def test_missing_file_with_create_succeeds(self, tmp_xlsx_path):
        conn = ExcelConnection(tmp_xlsx_path, create=True)
        assert conn.closed is False
        assert os.path.exists(tmp_xlsx_path)
        conn.close()

    def test_existing_file_without_create_succeeds(self, tmp_xlsx):
        conn = ExcelConnection(tmp_xlsx)
        assert conn.closed is False
        conn.close()

    def test_bad_engine_checked_before_file_existence(self, tmp_xlsx_path):
        """Engine validation should happen before file existence check."""
        # Both conditions fail: bad engine + missing file
        # Should get OperationalError about engine, not about file
        with pytest.raises(NotSupportedError, match="Unsupported engine"):
            ExcelConnection(tmp_xlsx_path, engine="nonexistent")



def _xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id", "name"])
    ws.append([1, "Alice"])
    wb.save(path)

def test_connection_str_and_repr(tmp_path: Path) -> None:
    file_path = tmp_path / "repr.xlsx"
    _xlsx(file_path)
    conn = ExcelConnection(str(file_path), engine="openpyxl")
    assert "ExcelConnection" in str(conn)
    assert repr(conn) == str(conn)
    conn.close()
