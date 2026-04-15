from pathlib import Path

import pytest
from excel_dbapi.exceptions import DatabaseError
from openpyxl import Workbook

from excel_dbapi.connection import ExcelConnection
from excel_dbapi.parser import parse_sql


def _create_case_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "t"
    sheet.append(["id", "score", "tier", "label", "points", "status", "x"])
    sheet.append([1, 15, "gold", "x", 2, "active", 5])
    sheet.append([2, 8, "silver", None, 3, "inactive", -2])
    sheet.append([3, None, "bronze", "z", 4, "inactive", 0])
    sheet.append([4, 12, "gold", "w", 0, "active", 7])
    workbook.save(path)


def _query(
    path: Path,
    sql: str,
    params: tuple[object, ...] | None = None,
) -> tuple[list[tuple[object, ...]], list[str]]:
    with ExcelConnection(str(path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute(sql, params)
        rows = cursor.fetchall()
        description = [str(item[0]) for item in cursor.description or []]
        return rows, description


def _execute(path: Path, sql: str, params: tuple[object, ...] | None = None) -> int:
    with ExcelConnection(str(path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(sql, params)
        return cursor.rowcount


def test_parser_builds_searched_case_ast() -> None:
    parsed = parse_sql(
        "SELECT CASE WHEN score > 10 THEN 'high' ELSE 'low' END AS bucket FROM t"
    )
    alias_expr = parsed["columns"][0]
    assert alias_expr["type"] == "alias"
    case_expr = alias_expr["expression"]
    assert case_expr["type"] == "case"
    assert case_expr["mode"] == "searched"
    assert case_expr["value"] is None
    assert case_expr["else"]["type"] == "literal"


def test_parser_builds_simple_case_ast() -> None:
    parsed = parse_sql("SELECT CASE tier WHEN 'gold' THEN 1 ELSE 2 END FROM t")
    case_expr = parsed["columns"][0]
    assert case_expr["type"] == "case"
    assert case_expr["mode"] == "simple"
    assert case_expr["value"] == "tier"
    assert case_expr["whens"][0]["match"]["type"] == "literal"


def test_parser_builds_case_in_update_set_and_where_operand() -> None:
    parsed = parse_sql(
        "UPDATE t "
        "SET label = CASE WHEN score > 10 THEN 'a' ELSE 'b' END "
        "WHERE CASE WHEN score > 10 THEN 'a' ELSE 'b' END = 'a'"
    )
    assert parsed["set"][0]["value"]["type"] == "case"
    where_condition = parsed["where"]["conditions"][0]
    assert where_condition["column"]["type"] == "case"


def test_searched_case_select_with_alias(tmp_path: Path) -> None:
    file_path = tmp_path / "case_searched_alias.xlsx"
    _create_case_workbook(file_path)

    rows, description = _query(
        file_path,
        "SELECT id, CASE WHEN score >= 10 THEN 'high' WHEN score >= 5 THEN 'mid' ELSE 'low' END AS band "
        "FROM t ORDER BY id",
    )
    assert description == ["id", "band"]
    assert rows == [(1, "high"), (2, "mid"), (3, "low"), (4, "high")]


def test_simple_case_select_and_implicit_alias(tmp_path: Path) -> None:
    file_path = tmp_path / "case_simple_alias.xlsx"
    _create_case_workbook(file_path)

    rows, description = _query(
        file_path,
        "SELECT CASE tier WHEN 'gold' THEN 1 WHEN 'silver' THEN 2 ELSE 3 END rank FROM t ORDER BY id",
    )
    assert description == ["rank"]
    assert rows == [(1,), (2,), (3,), (1,)]


def test_case_update_set_value(tmp_path: Path) -> None:
    file_path = tmp_path / "case_update.xlsx"
    _create_case_workbook(file_path)

    rowcount = _execute(
        file_path,
        "UPDATE t SET label = CASE WHEN score >= 10 THEN 'keep' ELSE 'drop' END",
    )
    assert rowcount == 4

    rows, _ = _query(file_path, "SELECT id, label FROM t ORDER BY id")
    assert rows == [(1, "keep"), (2, "drop"), (3, "drop"), (4, "keep")]


def test_case_null_and_no_else_behavior(tmp_path: Path) -> None:
    file_path = tmp_path / "case_null.xlsx"
    _create_case_workbook(file_path)

    rows, _ = _query(
        file_path,
        "SELECT id, CASE WHEN score IS NULL THEN 'missing' ELSE NULL END AS note FROM t ORDER BY id",
    )
    assert rows == [(1, None), (2, None), (3, "missing"), (4, None)]

    rows, _ = _query(
        file_path,
        "SELECT CASE WHEN score > 100 THEN 'x' END AS missing_case FROM t WHERE id = 1",
    )
    assert rows == [(None,)]


def test_case_with_arithmetic_and_nested_case(tmp_path: Path) -> None:
    file_path = tmp_path / "case_arithmetic_nested.xlsx"
    _create_case_workbook(file_path)

    rows, _ = _query(
        file_path,
        "SELECT id, CASE WHEN score >= 10 THEN score + points ELSE points * 2 END AS calc FROM t ORDER BY id",
    )
    assert rows == [(1, 17.0), (2, 6.0), (3, 8.0), (4, 12.0)]

    rows, _ = _query(
        file_path,
        "SELECT id, "
        "CASE WHEN score >= 10 THEN CASE WHEN points > 1 THEN 'A' ELSE 'B' END ELSE 'C' END AS nested_value "
        "FROM t ORDER BY id",
    )
    assert rows == [(1, "A"), (2, "C"), (3, "C"), (4, "B")]


def test_case_in_where_operand(tmp_path: Path) -> None:
    file_path = tmp_path / "case_where_operand.xlsx"
    _create_case_workbook(file_path)

    rows, _ = _query(
        file_path,
        "SELECT id FROM t "
        "WHERE CASE WHEN score > 10 THEN 'a' ELSE 'b' END = 'a' "
        "ORDER BY id",
    )
    assert rows == [(1,), (4,)]


def test_case_order_by_case_expression(tmp_path: Path) -> None:
    file_path = tmp_path / "case_order_by_case_asc.xlsx"
    _create_case_workbook(file_path)

    rows, _ = _query(
        file_path,
        "SELECT * FROM t ORDER BY CASE WHEN status = 'active' THEN 0 ELSE 1 END ASC",
    )
    assert [row[0] for row in rows] == [1, 4, 2, 3]


def test_case_order_by_case_desc(tmp_path: Path) -> None:
    file_path = tmp_path / "case_order_by_case_desc.xlsx"
    _create_case_workbook(file_path)

    rows, _ = _query(
        file_path,
        "SELECT * FROM t ORDER BY CASE WHEN status = 'active' THEN 0 ELSE 1 END DESC",
    )
    assert [row[0] for row in rows] == [2, 3, 1, 4]


def test_case_arithmetic_addition(tmp_path: Path) -> None:
    file_path = tmp_path / "case_arithmetic_addition.xlsx"
    _create_case_workbook(file_path)

    rows, description = _query(
        file_path,
        "SELECT CASE WHEN x > 0 THEN x ELSE 0 END + 10 AS result FROM t ORDER BY id",
    )
    assert description == ["result"]
    assert rows == [(15.0,), (10.0,), (10.0,), (17.0,)]


def test_case_arithmetic_multiplication(tmp_path: Path) -> None:
    file_path = tmp_path / "case_arithmetic_multiplication.xlsx"
    _create_case_workbook(file_path)

    rows, _ = _query(
        file_path,
        "SELECT CASE WHEN x > 0 THEN x ELSE 0 END * 2 FROM t ORDER BY id",
    )
    assert rows == [(10.0,), (0.0,), (0.0,), (14.0,)]


def test_case_in_complex_arithmetic(tmp_path: Path) -> None:
    file_path = tmp_path / "case_complex_arithmetic.xlsx"
    _create_case_workbook(file_path)

    rows, _ = _query(
        file_path,
        "SELECT 100 + CASE WHEN status = 'active' THEN 50 ELSE 0 END FROM t ORDER BY id",
    )
    assert rows == [(150.0,), (100.0,), (100.0,), (150.0,)]


def test_case_parser_error_missing_end() -> None:
    with pytest.raises(DatabaseError, match="missing END"):
        parse_sql("SELECT CASE WHEN x = 1 THEN 2 FROM t")


def test_case_parser_error_missing_when() -> None:
    with pytest.raises(DatabaseError, match="missing WHEN"):
        parse_sql("SELECT CASE THEN 1 END FROM t")


def test_case_parameter_binding_select_and_update(tmp_path: Path) -> None:
    file_path = tmp_path / "case_params.xlsx"
    _create_case_workbook(file_path)

    rows, _ = _query(
        file_path,
        "SELECT CASE WHEN score >= ? THEN ? ELSE ? END AS bucket FROM t WHERE id = ?",
        (10, "high", "low", 2),
    )
    assert rows == [("low",)]

    rowcount = _execute(
        file_path,
        "UPDATE t SET label = CASE WHEN score >= ? THEN ? ELSE ? END WHERE id = ?",
        (10, "high", "low", 1),
    )
    assert rowcount == 1

    rows, _ = _query(file_path, "SELECT label FROM t WHERE id = 1")
    assert rows == [("high",)]


def test_case_order_by_case_with_parameters(tmp_path: Path) -> None:
    file_path = tmp_path / "case_order_by_case_params.xlsx"
    _create_case_workbook(file_path)

    rows, _ = _query(
        file_path,
        "SELECT id, status FROM t "
        "ORDER BY CASE WHEN status = ? THEN ? WHEN status = ? THEN ? ELSE ? END ASC",
        ("active", 0, "pending", 1, 2),
    )
    assert rows == [(1, "active"), (4, "active"), (2, "inactive"), (3, "inactive")]


def test_case_order_by_case_with_join(tmp_path: Path) -> None:
    file_path = tmp_path / "case_order_by_case_join.xlsx"

    workbook = Workbook()
    sheet_employees = workbook.active
    assert sheet_employees is not None
    sheet_employees.title = "employees"
    sheet_employees.append(["id", "name", "dept_id"])
    sheet_employees.append([1, "Alice", 10])
    sheet_employees.append([2, "Bob", 20])
    sheet_employees.append([3, "Carol", 10])

    sheet_departments = workbook.create_sheet("departments")
    sheet_departments.append(["id", "dept_name"])
    sheet_departments.append([10, "Engineering"])
    sheet_departments.append([20, "Sales"])
    workbook.save(file_path)

    rows, _ = _query(
        file_path,
        "SELECT employees.name, departments.dept_name, "
        "CASE WHEN departments.dept_name = 'Sales' THEN 0 ELSE 1 END AS sort_key "
        "FROM employees "
        "INNER JOIN departments ON employees.dept_id = departments.id "
        "ORDER BY sort_key ASC",
    )
    assert rows[0] == ("Bob", "Sales", 0)
    assert len(rows) == 3


def test_case_order_by_case_with_group_by(tmp_path: Path) -> None:
    file_path = tmp_path / "case_order_by_case_group_by.xlsx"

    workbook = Workbook()
    sheet_sales = workbook.active
    assert sheet_sales is not None
    sheet_sales.title = "sales"
    sheet_sales.append(["region", "amount"])
    sheet_sales.append(["East", 100])
    sheet_sales.append(["West", 200])
    sheet_sales.append(["East", 150])
    sheet_sales.append(["West", 50])
    workbook.save(file_path)

    rows, _ = _query(
        file_path,
        "SELECT region, SUM(amount) FROM sales GROUP BY region "
        "ORDER BY CASE WHEN region = 'West' THEN 0 ELSE 1 END ASC",
    )
    assert rows[0] == ("West", 250)
    assert rows[1] == ("East", 250)
