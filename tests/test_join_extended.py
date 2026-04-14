from pathlib import Path

import pytest
from openpyxl import Workbook

from excel_dbapi.connection import ExcelConnection
from excel_dbapi.exceptions import ProgrammingError


def _create_join_workbook(path: Path) -> None:
    workbook = Workbook()

    t1 = workbook.active
    assert t1 is not None
    t1.title = "t1"
    t1.append(["id", "val1"])
    t1.append([1, "a1"])
    t1.append([2, "a2"])
    t1.append([4, "a4"])

    t2 = workbook.create_sheet("t2")
    t2.append(["id", "val2"])
    t2.append([1, "b1"])
    t2.append([2, "b2"])
    t2.append([3, "b3"])

    t3 = workbook.create_sheet("t3")
    t3.append(["id", "val3"])
    t3.append([1, "c1"])
    t3.append([3, "c3"])
    t3.append([4, "c4"])

    t4 = workbook.create_sheet("t4")
    t4.append(["id", "val4"])
    t4.append([1, "d1"])
    t4.append([2, "d2"])
    t4.append([4, "d4"])

    workbook.save(path)


def _run_query(path: Path, query: str) -> list[tuple[object, ...]]:
    with ExcelConnection(str(path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute(query)
        return cursor.fetchall()


def test_right_join_basic_matching_rows(tmp_path: Path) -> None:
    file_path = tmp_path / "right_join_basic.xlsx"
    _create_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, b.val2 FROM t1 a RIGHT JOIN t2 b ON a.id = b.id WHERE a.id IS NOT NULL ORDER BY b.id",
    )
    assert rows == [(1, "b1"), (2, "b2")]


def test_right_join_non_matching_right_rows_have_null_left(tmp_path: Path) -> None:
    file_path = tmp_path / "right_join_null_left.xlsx"
    _create_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, b.id FROM t1 a RIGHT JOIN t2 b ON a.id = b.id ORDER BY b.id",
    )
    assert rows == [(1, 1), (2, 2), (None, 3)]


def test_right_outer_join_syntax(tmp_path: Path) -> None:
    file_path = tmp_path / "right_outer_join.xlsx"
    _create_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, b.id FROM t1 a RIGHT OUTER JOIN t2 b ON a.id = b.id ORDER BY b.id",
    )
    assert rows == [(1, 1), (2, 2), (None, 3)]


def test_right_join_with_where_clause(tmp_path: Path) -> None:
    file_path = tmp_path / "right_join_where.xlsx"
    _create_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, b.id FROM t1 a RIGHT JOIN t2 b ON a.id = b.id WHERE b.id >= 2 ORDER BY b.id",
    )
    assert rows == [(2, 2), (None, 3)]


def test_right_join_with_order_by(tmp_path: Path) -> None:
    file_path = tmp_path / "right_join_order.xlsx"
    _create_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, b.id FROM t1 a RIGHT JOIN t2 b ON a.id = b.id ORDER BY b.id DESC",
    )
    assert rows == [(None, 3), (2, 2), (1, 1)]


def test_chained_inner_join_three_tables(tmp_path: Path) -> None:
    file_path = tmp_path / "chained_inner_three.xlsx"
    _create_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, b.val2, c.val3 FROM t1 a JOIN t2 b ON a.id = b.id JOIN t3 c ON b.id = c.id ORDER BY a.id",
    )
    assert rows == [(1, "b1", "c1")]


def test_chained_left_then_inner_join(tmp_path: Path) -> None:
    file_path = tmp_path / "chained_left_inner.xlsx"
    _create_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, b.val2, c.val3 FROM t1 a LEFT JOIN t2 b ON a.id = b.id JOIN t3 c ON b.id = c.id ORDER BY a.id",
    )
    assert rows == [(1, "b1", "c1")]


def test_chained_inner_then_left_join(tmp_path: Path) -> None:
    file_path = tmp_path / "chained_inner_left.xlsx"
    _create_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, c.val3 FROM t1 a JOIN t2 b ON a.id = b.id LEFT JOIN t3 c ON b.id = c.id ORDER BY a.id",
    )
    assert rows == [(1, "c1"), (2, None)]


def test_chained_three_joins_four_tables(tmp_path: Path) -> None:
    file_path = tmp_path / "chained_three_joins.xlsx"
    _create_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, d.val4 FROM t1 a JOIN t2 b ON a.id = b.id JOIN t3 c ON b.id = c.id JOIN t4 d ON c.id = d.id",
    )
    assert rows == [(1, "d1")]


def test_chained_join_where_across_multiple_tables(tmp_path: Path) -> None:
    file_path = tmp_path / "chained_where.xlsx"
    _create_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id FROM t1 a JOIN t2 b ON a.id = b.id LEFT JOIN t3 c ON b.id = c.id WHERE c.val3 IS NULL AND b.val2 = 'b2'",
    )
    assert rows == [(2,)]


def test_chained_join_with_order_by(tmp_path: Path) -> None:
    file_path = tmp_path / "chained_order_by.xlsx"
    _create_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, c.val3 FROM t1 a JOIN t2 b ON a.id = b.id LEFT JOIN t3 c ON b.id = c.id ORDER BY a.id DESC",
    )
    assert rows == [(2, None), (1, "c1")]


def test_chained_right_then_left_join(tmp_path: Path) -> None:
    file_path = tmp_path / "chained_right_left.xlsx"
    _create_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, b.id, c.val3 FROM t1 a RIGHT JOIN t2 b ON a.id = b.id LEFT JOIN t3 c ON b.id = c.id ORDER BY b.id",
    )
    assert rows == [(1, 1, "c1"), (2, 2, None), (None, 3, "c3")]


def test_right_join_without_on_raises_error(tmp_path: Path) -> None:
    file_path = tmp_path / "right_join_no_on.xlsx"
    _create_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="JOIN requires ON condition"):
            cursor.execute("SELECT a.id FROM t1 a RIGHT JOIN t2 b")


def test_chained_join_duplicate_alias_raises_error(tmp_path: Path) -> None:
    file_path = tmp_path / "chained_duplicate_alias.xlsx"
    _create_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="Ambiguous table reference"):
            cursor.execute(
                "SELECT a.id FROM t1 a JOIN t2 b ON a.id = b.id JOIN t3 b ON b.id = b.id"
            )


def test_chained_join_on_references_first_table(tmp_path: Path) -> None:
    """Third JOIN's ON clause references the first table, not the immediately previous one."""
    file_path = tmp_path / "chained_on_first.xlsx"
    _create_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        # t3 ON clause references a.id (first table), not b.id (immediate previous)
        cursor.execute(
            "SELECT a.id, b.val2, c.val3 FROM t1 a "
            "JOIN t2 b ON a.id = b.id "
            "LEFT JOIN t3 c ON a.id = c.id "
            "ORDER BY a.id"
        )
        rows = cursor.fetchall()
        # t1 has ids 1,2,4; t2 has 1,2,3; inner join gives 1,2
        # t3 has 1,3,4; left join on a.id=c.id for ids 1,2 gives:
        # id=1: c.val3='c1'; id=2: c.val3=None (no match)
        assert len(rows) == 2
        assert rows[0] == (1, "b1", "c1")
        assert rows[1] == (2, "b2", None)


def test_select_star_inner_join(tmp_path: Path) -> None:
    file_path = tmp_path / "select_star_inner_join.xlsx"
    _create_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM t1 a JOIN t2 b ON a.id = b.id ORDER BY a.id")
        rows = cursor.fetchall()
        description = cursor.description

    assert rows == [(1, "a1", 1, "b1"), (2, "a2", 2, "b2")]
    assert description is not None
    assert [col[0] for col in description] == ["a.id", "a.val1", "b.id", "b.val2"]


def test_select_star_left_join(tmp_path: Path) -> None:
    file_path = tmp_path / "select_star_left_join.xlsx"
    _create_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM t1 a LEFT JOIN t2 b ON a.id = b.id ORDER BY a.id")
        rows = cursor.fetchall()
        description = cursor.description

    assert rows == [(1, "a1", 1, "b1"), (2, "a2", 2, "b2"), (4, "a4", None, None)]
    assert description is not None
    assert [col[0] for col in description] == ["a.id", "a.val1", "b.id", "b.val2"]


def test_select_star_right_join(tmp_path: Path) -> None:
    file_path = tmp_path / "select_star_right_join.xlsx"
    _create_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT * FROM t1 a RIGHT JOIN t2 b ON a.id = b.id ORDER BY b.id"
        )
        rows = cursor.fetchall()
        description = cursor.description

    assert rows == [(1, "a1", 1, "b1"), (2, "a2", 2, "b2"), (None, None, 3, "b3")]
    assert description is not None
    assert [col[0] for col in description] == ["a.id", "a.val1", "b.id", "b.val2"]


def test_select_star_chained_join(tmp_path: Path) -> None:
    file_path = tmp_path / "select_star_chained_join.xlsx"
    _create_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT * FROM t1 a JOIN t2 b ON a.id = b.id JOIN t3 c ON b.id = c.id"
        )
        rows = cursor.fetchall()
        description = cursor.description

    assert rows == [(1, "a1", 1, "b1", 1, "c1")]
    assert description is not None
    assert [col[0] for col in description] == [
        "a.id",
        "a.val1",
        "b.id",
        "b.val2",
        "c.id",
        "c.val3",
    ]


def test_select_star_with_where(tmp_path: Path) -> None:
    file_path = tmp_path / "select_star_with_where.xlsx"
    _create_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT * FROM t1 a JOIN t2 b ON a.id = b.id WHERE a.val1 = 'a1'"
        )
        rows = cursor.fetchall()
        description = cursor.description

    assert rows == [(1, "a1", 1, "b1")]
    assert description is not None
    assert [col[0] for col in description] == ["a.id", "a.val1", "b.id", "b.val2"]


def test_select_star_with_order_by(tmp_path: Path) -> None:
    file_path = tmp_path / "select_star_with_order_by.xlsx"
    _create_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM t1 a JOIN t2 b ON a.id = b.id ORDER BY a.id DESC")
        rows = cursor.fetchall()
        description = cursor.description

    assert rows == [(2, "a2", 2, "b2"), (1, "a1", 1, "b1")]
    assert description is not None
    assert [col[0] for col in description] == ["a.id", "a.val1", "b.id", "b.val2"]


def test_select_star_with_limit_offset(tmp_path: Path) -> None:
    file_path = tmp_path / "select_star_with_limit_offset.xlsx"
    _create_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT * FROM t1 a LEFT JOIN t2 b ON a.id = b.id ORDER BY a.id LIMIT 2 OFFSET 1"
        )
        rows = cursor.fetchall()
        description = cursor.description

    assert rows == [(2, "a2", 2, "b2"), (4, "a4", None, None)]
    assert description is not None
    assert [col[0] for col in description] == ["a.id", "a.val1", "b.id", "b.val2"]


def test_select_star_description_columns(tmp_path: Path) -> None:
    file_path = tmp_path / "select_star_description_columns.xlsx"
    _create_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM t1 a JOIN t2 b ON a.id = b.id")
        _ = cursor.fetchall()
        description = cursor.description

    assert description is not None
    assert [col[0] for col in description] == ["a.id", "a.val1", "b.id", "b.val2"]


def test_select_star_empty_result_has_description(tmp_path: Path) -> None:
    file_path = tmp_path / "select_star_empty_result.xlsx"
    _create_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM t1 a JOIN t2 b ON a.id = b.id WHERE a.id = 999")
        rows = cursor.fetchall()
        description = cursor.description

    assert rows == []
    assert description is not None
    assert len(description) == 4
    assert [col[0] for col in description] == ["a.id", "a.val1", "b.id", "b.val2"]


def test_select_star_duplicate_column_names(tmp_path: Path) -> None:
    file_path = tmp_path / "select_star_duplicate_column_names.xlsx"
    _create_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM t1 a JOIN t2 b ON a.id = b.id ORDER BY a.id")
        rows = cursor.fetchall()
        description = cursor.description

    assert rows == [(1, "a1", 1, "b1"), (2, "a2", 2, "b2")]
    assert description is not None
    assert [col[0] for col in description] == ["a.id", "a.val1", "b.id", "b.val2"]


def test_select_star_chained_left_right_join(tmp_path: Path) -> None:
    file_path = tmp_path / "select_star_chained_left_right_join.xlsx"
    _create_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT * FROM t1 a LEFT JOIN t2 b ON a.id = b.id RIGHT JOIN t3 c ON b.id = c.id ORDER BY c.id"
        )
        rows = cursor.fetchall()
        description = cursor.description

    assert rows == [
        (1, "a1", 1, "b1", 1, "c1"),
        (None, None, None, None, 3, "c3"),
        (None, None, None, None, 4, "c4"),
    ]
    assert description is not None
    assert [col[0] for col in description] == [
        "a.id",
        "a.val1",
        "b.id",
        "b.val2",
        "c.id",
        "c.val3",
    ]


def test_select_star_mixed_with_columns_rejected(tmp_path: Path) -> None:
    """SELECT *, a.id with JOIN is rejected at parse time."""
    file_path = tmp_path / "mixed_star.xlsx"
    _create_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="SELECT \\* cannot be mixed"):
            cursor.execute("SELECT *, a.id FROM t1 a JOIN t2 b ON a.id = b.id")


def test_select_star_mixed_trailing_rejected(tmp_path: Path) -> None:
    """SELECT a.id, * with JOIN is rejected at parse time."""
    file_path = tmp_path / "mixed_star_trailing.xlsx"
    _create_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="SELECT \\* cannot be mixed"):
            cursor.execute("SELECT a.id, * FROM t1 a JOIN t2 b ON a.id = b.id")


def test_select_table_dot_star_rejected(tmp_path: Path) -> None:
    """SELECT a.* with JOIN is rejected (only bare * is supported)."""
    file_path = tmp_path / "table_dot_star.xlsx"
    _create_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError):
            cursor.execute("SELECT a.* FROM t1 a JOIN t2 b ON a.id = b.id")


def test_select_star_with_alias_and_no_alias_description(tmp_path: Path) -> None:
    """SELECT * FROM t1 a JOIN t2 ON ... uses alias for t1, bare name for t2."""
    file_path = tmp_path / "mixed_alias_desc.xlsx"
    _create_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM t1 a JOIN t2 ON a.id = t2.id")
        rows = cursor.fetchall()
        description = cursor.description

    # a is alias for t1, t2 has no alias — description uses ref names
    assert description is not None
    assert [col[0] for col in description] == [
        "a.id",
        "a.val1",
        "t2.id",
        "t2.val2",
    ]
    assert len(rows) == 2  # ids 1 and 2 match


def test_full_outer_join_basic(tmp_path: Path) -> None:
    file_path = tmp_path / "full_outer_join_basic.xlsx"
    _create_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, a.val1, b.id, b.val2 "
        "FROM t1 a FULL OUTER JOIN t2 b ON a.id = b.id "
        "ORDER BY a.id",
    )
    assert rows == [
        (1, "a1", 1, "b1"),
        (2, "a2", 2, "b2"),
        (4, "a4", None, None),
        (None, None, 3, "b3"),
    ]


def test_full_join_without_outer_keyword(tmp_path: Path) -> None:
    file_path = tmp_path / "full_join_no_outer_keyword.xlsx"
    _create_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, b.id FROM t1 a FULL JOIN t2 b ON a.id = b.id ORDER BY a.id",
    )
    assert rows == [(1, 1), (2, 2), (4, None), (None, 3)]


def test_full_outer_join_all_match(tmp_path: Path) -> None:
    file_path = tmp_path / "full_outer_join_all_match.xlsx"
    _create_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, a.val1, d.id, d.val4 "
        "FROM t1 a FULL OUTER JOIN t4 d ON a.id = d.id "
        "ORDER BY a.id",
    )
    assert rows == [(1, "a1", 1, "d1"), (2, "a2", 2, "d2"), (4, "a4", 4, "d4")]


def test_full_outer_join_no_match(tmp_path: Path) -> None:
    file_path = tmp_path / "full_outer_join_no_match.xlsx"
    workbook = Workbook()

    t1 = workbook.active
    assert t1 is not None
    t1.title = "t1"
    t1.append(["id", "val1"])
    t1.append([1, "a1"])
    t1.append([2, "a2"])

    t2 = workbook.create_sheet("t2")
    t2.append(["id", "val2"])
    t2.append([3, "b3"])
    t2.append([4, "b4"])
    workbook.save(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, a.val1, b.id, b.val2 FROM t1 a FULL JOIN t2 b ON a.id = b.id ORDER BY a.id",
    )
    assert rows == [
        (1, "a1", None, None),
        (2, "a2", None, None),
        (None, None, 3, "b3"),
        (None, None, 4, "b4"),
    ]


def test_full_outer_join_with_where(tmp_path: Path) -> None:
    file_path = tmp_path / "full_outer_join_with_where.xlsx"
    _create_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, b.id "
        "FROM t1 a FULL JOIN t2 b ON a.id = b.id "
        "WHERE b.id IS NULL OR a.id = 1 "
        "ORDER BY a.id",
    )
    assert rows == [(1, 1), (4, None)]


def test_full_outer_join_with_select_star(tmp_path: Path) -> None:
    file_path = tmp_path / "full_outer_join_select_star.xlsx"
    _create_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM t1 a FULL JOIN t2 b ON a.id = b.id ORDER BY a.id")
        rows = cursor.fetchall()
        description = cursor.description

    assert rows == [
        (1, "a1", 1, "b1"),
        (2, "a2", 2, "b2"),
        (4, "a4", None, None),
        (None, None, 3, "b3"),
    ]
    assert description is not None
    assert [col[0] for col in description] == ["a.id", "a.val1", "b.id", "b.val2"]


def test_full_outer_join_chained(tmp_path: Path) -> None:
    file_path = tmp_path / "full_outer_join_chained.xlsx"
    _create_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, b.id, c.val3 "
        "FROM t1 a FULL JOIN t2 b ON a.id = b.id "
        "INNER JOIN t3 c ON b.id = c.id "
        "ORDER BY c.id",
    )
    assert rows == [(1, 1, "c1"), (None, 3, "c3")]


def test_full_outer_join_description(tmp_path: Path) -> None:
    file_path = tmp_path / "full_outer_join_description.xlsx"
    _create_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT a.id, a.val1, b.id, b.val2 FROM t1 a FULL JOIN t2 b ON a.id = b.id"
        )
        _ = cursor.fetchall()
        description = cursor.description

    assert description is not None
    assert [col[0] for col in description] == ["a.id", "a.val1", "b.id", "b.val2"]


def test_full_outer_join_duplicate_keys(tmp_path: Path) -> None:
    file_path = tmp_path / "full_outer_join_duplicate_keys.xlsx"
    workbook = Workbook()

    t1 = workbook.active
    assert t1 is not None
    t1.title = "t1"
    t1.append(["id", "val1"])
    t1.append([1, "a1"])
    t1.append([1, "a2"])
    t1.append([3, "a3"])

    t2 = workbook.create_sheet("t2")
    t2.append(["id", "val2"])
    t2.append([1, "b1"])
    t2.append([1, "b2"])
    t2.append([2, "b2"])
    workbook.save(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, a.val1, b.id, b.val2 FROM t1 a FULL JOIN t2 b ON a.id = b.id",
    )
    assert len(rows) == 6
    assert rows.count((1, "a1", 1, "b1")) == 1
    assert rows.count((1, "a1", 1, "b2")) == 1
    assert rows.count((1, "a2", 1, "b1")) == 1
    assert rows.count((1, "a2", 1, "b2")) == 1
    assert rows.count((3, "a3", None, None)) == 1
    assert rows.count((None, None, 2, "b2")) == 1


def test_full_outer_join_null_keys(tmp_path: Path) -> None:
    file_path = tmp_path / "full_outer_join_null_keys.xlsx"
    workbook = Workbook()

    t1 = workbook.active
    assert t1 is not None
    t1.title = "t1"
    t1.append(["id", "val1"])
    t1.append([1, "a1"])
    t1.append([None, "a_null"])

    t2 = workbook.create_sheet("t2")
    t2.append(["id", "val2"])
    t2.append([1, "b1"])
    t2.append([None, "b_null"])
    workbook.save(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, a.val1, b.id, b.val2 FROM t1 a FULL JOIN t2 b ON a.id = b.id",
    )
    assert len(rows) == 3
    assert rows.count((1, "a1", 1, "b1")) == 1
    assert rows.count((None, "a_null", None, None)) == 1
    assert rows.count((None, None, None, "b_null")) == 1


def test_full_outer_join_empty_left(tmp_path: Path) -> None:
    file_path = tmp_path / "full_outer_join_empty_left.xlsx"
    workbook = Workbook()

    t1 = workbook.active
    assert t1 is not None
    t1.title = "t1"
    t1.append(["id", "val1"])

    t2 = workbook.create_sheet("t2")
    t2.append(["id", "val2"])
    t2.append([1, "b1"])
    t2.append([2, "b2"])
    workbook.save(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, a.val1, b.id, b.val2 FROM t1 a FULL JOIN t2 b ON a.id = b.id ORDER BY b.id",
    )
    assert rows == [(None, None, 1, "b1"), (None, None, 2, "b2")]


def test_cross_join_basic(tmp_path: Path) -> None:
    file_path = tmp_path / "cross_join_basic.xlsx"
    _create_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, a.val1, b.id, b.val2 FROM t1 a CROSS JOIN t2 b ORDER BY a.id, b.id",
    )
    assert len(rows) == 9
    assert rows[0] == (1, "a1", 1, "b1")
    assert rows[-1] == (4, "a4", 3, "b3")


def test_cross_join_with_where(tmp_path: Path) -> None:
    file_path = tmp_path / "cross_join_with_where.xlsx"
    _create_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, b.id FROM t1 a CROSS JOIN t2 b "
        "WHERE a.id = 1 AND b.id = 1 OR a.id = 2 AND b.id = 2 "
        "ORDER BY a.id",
    )
    assert rows == [(1, 1), (2, 2)]


def test_cross_join_with_select_star(tmp_path: Path) -> None:
    file_path = tmp_path / "cross_join_select_star.xlsx"
    _create_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM t1 a CROSS JOIN t2 b ORDER BY a.id, b.id")
        rows = cursor.fetchall()
        description = cursor.description

    assert len(rows) == 9
    assert rows[0] == (1, "a1", 1, "b1")
    assert rows[-1] == (4, "a4", 3, "b3")
    assert description is not None
    assert [col[0] for col in description] == ["a.id", "a.val1", "b.id", "b.val2"]


def test_cross_join_empty_table(tmp_path: Path) -> None:
    file_path = tmp_path / "cross_join_empty_table.xlsx"
    workbook = Workbook()

    t1 = workbook.active
    assert t1 is not None
    t1.title = "t1"
    t1.append(["id", "val1"])
    t1.append([1, "a1"])

    t2 = workbook.create_sheet("t2")
    t2.append(["id", "val2"])
    workbook.save(file_path)

    rows = _run_query(file_path, "SELECT a.id, b.id FROM t1 a CROSS JOIN t2 b")
    assert rows == []


def test_cross_join_single_row(tmp_path: Path) -> None:
    file_path = tmp_path / "cross_join_single_row.xlsx"
    workbook = Workbook()

    t1 = workbook.active
    assert t1 is not None
    t1.title = "t1"
    t1.append(["id", "val1"])
    t1.append([1, "a1"])

    t2 = workbook.create_sheet("t2")
    t2.append(["id", "val2"])
    t2.append([1, "b1"])
    t2.append([2, "b2"])
    t2.append([3, "b3"])
    workbook.save(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, b.id FROM t1 a CROSS JOIN t2 b ORDER BY b.id",
    )
    assert rows == [(1, 1), (1, 2), (1, 3)]


def test_cross_join_chained(tmp_path: Path) -> None:
    file_path = tmp_path / "cross_join_chained.xlsx"
    _create_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, b.id, c.val3 "
        "FROM t1 a CROSS JOIN t2 b "
        "INNER JOIN t3 c ON b.id = c.id "
        "ORDER BY a.id, b.id",
    )
    assert rows == [
        (1, 1, "c1"),
        (1, 3, "c3"),
        (2, 1, "c1"),
        (2, 3, "c3"),
        (4, 1, "c1"),
        (4, 3, "c3"),
    ]


def test_cross_join_rejects_on_clause(tmp_path: Path) -> None:
    file_path = tmp_path / "cross_join_rejects_on_clause.xlsx"
    _create_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        with pytest.raises(
            ProgrammingError, match="CROSS JOIN does not accept ON condition"
        ):
            cursor.execute("SELECT a.id FROM t1 a CROSS JOIN t2 b ON a.id = b.id")


def test_cross_join_on_rejected(tmp_path: Path) -> None:
    file_path = tmp_path / "cross_join_on_rejected.xlsx"
    _create_join_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        with pytest.raises(
            ProgrammingError, match="CROSS JOIN does not accept ON condition"
        ):
            cursor.execute(
                "SELECT a.id FROM t1 a CROSS JOIN t2 b ON a.id = b.id WHERE a.id = b.id"
            )


def _create_non_equi_join_workbook(path: Path) -> None:
    workbook = Workbook()

    a_sheet = workbook.active
    assert a_sheet is not None
    a_sheet.title = "a"
    a_sheet.append(["id", "name", "x"])
    a_sheet.append([1, "alpha", 10])
    a_sheet.append([2, "beta", 20])
    a_sheet.append([3, "shared", 30])

    b_sheet = workbook.create_sheet("b")
    b_sheet.append(["id", "name", "y"])
    b_sheet.append([0, "zero", 5])
    b_sheet.append([1, "uno", 5])
    b_sheet.append([2, "other", 20])
    b_sheet.append([4, "shared", 15])

    workbook.save(path)


def test_join_on_greater_than_condition(tmp_path: Path) -> None:
    file_path = tmp_path / "join_on_greater_than.xlsx"
    _create_non_equi_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, b.id FROM a JOIN b ON a.id > b.id ORDER BY a.id, b.id",
    )
    assert rows == [(1, 0), (2, 0), (2, 1), (3, 0), (3, 1), (3, 2)]


def test_join_on_not_equal_condition(tmp_path: Path) -> None:
    file_path = tmp_path / "join_on_not_equal.xlsx"
    _create_non_equi_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, b.id FROM a JOIN b ON a.x != b.y ORDER BY a.id, b.id",
    )
    assert rows == [
        (1, 0),
        (1, 1),
        (1, 2),
        (1, 4),
        (2, 0),
        (2, 1),
        (2, 4),
        (3, 0),
        (3, 1),
        (3, 2),
        (3, 4),
    ]


def test_join_on_or_condition(tmp_path: Path) -> None:
    file_path = tmp_path / "join_on_or_condition.xlsx"
    _create_non_equi_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, b.id FROM a JOIN b ON a.id = b.id OR a.name = b.name ORDER BY a.id, b.id",
    )
    assert rows == [(1, 1), (2, 2), (3, 4)]


def test_join_on_mixed_and_or_with_parentheses(tmp_path: Path) -> None:
    file_path = tmp_path / "join_on_mixed_and_or.xlsx"
    _create_non_equi_join_workbook(file_path)

    rows = _run_query(
        file_path,
        "SELECT a.id, b.id FROM a JOIN b ON (a.id = b.id AND a.x > b.y) OR a.name = b.name ORDER BY a.id, b.id",
    )
    assert rows == [(1, 1), (3, 4)]
