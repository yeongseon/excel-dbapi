"""Tests for openpyxl write_sheet preserving cell formatting (issue #91)."""

import openpyxl
from openpyxl.styles import Font, PatternFill

from excel_dbapi.connection import ExcelConnection


class TestWriteSheetPreservesFormatting:
    """write_sheet must not destroy cell formatting."""

    def test_update_preserves_font(self, tmp_path):
        """Font formatting must survive UPDATE operations."""
        file = tmp_path / "test.xlsx"
        # Create workbook with formatting
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "t"
        ws.append(["id", "name"])
        ws.append([1, "Alice"])
        ws.append([2, "Bob"])
        # Apply bold font to header
        bold = Font(bold=True)
        ws["A1"].font = bold
        ws["B1"].font = bold
        # Apply red fill to data cells
        red_fill = PatternFill(start_color="FF0000", fill_type="solid")
        ws["A2"].fill = red_fill
        wb.save(str(file))
        wb.close()

        # Update via excel-dbapi
        with ExcelConnection(str(file), autocommit=True) as conn:
            cur = conn.cursor()
            cur.execute("UPDATE t SET name = 'Ann' WHERE id = 1")

        # Verify formatting preserved
        wb = openpyxl.load_workbook(str(file))
        ws = wb.active
        assert ws["A1"].font.bold is True
        assert ws["B1"].font.bold is True
        assert ws["A2"].fill.start_color.rgb == "00FF0000"
        # Verify data updated
        assert ws["B2"].value == "Ann"
        wb.close()

    def test_delete_preserves_remaining_formatting(self, tmp_path):
        """Formatting on non-deleted rows must survive DELETE."""
        file = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "t"
        ws.append(["id", "name"])
        ws.append([1, "Alice"])
        ws.append([2, "Bob"])
        bold = Font(bold=True)
        ws["A1"].font = bold
        wb.save(str(file))
        wb.close()

        with ExcelConnection(str(file), autocommit=True) as conn:
            cur = conn.cursor()
            cur.execute("DELETE FROM t WHERE id = 2")

        wb = openpyxl.load_workbook(str(file))
        ws = wb.active
        assert ws["A1"].font.bold is True
        assert ws["A2"].value == 1
        assert ws["B2"].value == "Alice"
        # Row 3 should be gone
        assert ws["A3"].value is None
        wb.close()

    def test_insert_with_existing_formatting(self, tmp_path):
        """INSERT must not destroy existing formatting."""
        file = tmp_path / "test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "t"
        ws.append(["id", "name"])
        ws.append([1, "Alice"])
        bold = Font(bold=True)
        ws["A1"].font = bold
        wb.save(str(file))
        wb.close()

        with ExcelConnection(str(file), autocommit=True) as conn:
            cur = conn.cursor()
            cur.execute("INSERT INTO t VALUES (2, 'Bob')")

        wb = openpyxl.load_workbook(str(file))
        ws = wb.active
        assert ws["A1"].font.bold is True
        assert ws["A3"].value == 2
        wb.close()
