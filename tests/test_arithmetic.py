from pathlib import Path

import pytest
from openpyxl import Workbook

from excel_dbapi.connection import ExcelConnection
from excel_dbapi.exceptions import ProgrammingError


def _create_arithmetic_workbook(path: Path) -> None:
    workbook = Workbook()

    sheet = workbook.active
    assert sheet is not None
    sheet.title = "t"
    sheet.append(
        [
            "id",
            "price",
            "quantity",
            "qty",
            "tax",
            "discount",
            "total",
            "count",
            "a",
            "b",
            "c",
            "label",
        ]
    )
    sheet.append([1, 10, 2, 2, 1.5, 1, 20, 4, 1, 2, 3, "x"])
    sheet.append([2, 5, 3, 0, 0.5, None, 15, 5, 2, 3, 4, "y"])
    sheet.append([3, None, 4, 1, 1.0, 2, 8, 2, 1, None, 5, "z"])

    t1 = workbook.create_sheet("t1")
    t1.append(["id", "price"])
    t1.append([1, 10])
    t1.append([2, 5])

    t2 = workbook.create_sheet("t2")
    t2.append(["id", "qty"])
    t2.append([1, 3])
    t2.append([2, 4])

    workbook.save(path)


def _run_query(path: Path, query: str) -> tuple[list[tuple[object, ...]], list[str]]:
    with ExcelConnection(str(path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute(query)
        rows = cursor.fetchall()
        description = [str(col[0]) for col in cursor.description or []]
        return rows, description


def test_basic_arithmetic_operations(tmp_path: Path) -> None:
    file_path = tmp_path / "arithmetic_basic.xlsx"
    _create_arithmetic_workbook(file_path)

    rows, _ = _run_query(file_path, "SELECT price * quantity FROM t ORDER BY id")
    assert rows == [(20.0,), (15.0,), (None,)]

    rows, _ = _run_query(file_path, "SELECT price + tax FROM t ORDER BY id")
    assert rows == [(11.5,), (5.5,), (None,)]

    rows, _ = _run_query(file_path, "SELECT total - discount FROM t ORDER BY id")
    assert rows == [(19.0,), (None,), (6.0,)]

    rows, _ = _run_query(file_path, "SELECT total / count FROM t ORDER BY id")
    assert rows == [(5.0,), (3.0,), (4.0,)]

    rows, _ = _run_query(file_path, "SELECT -discount FROM t ORDER BY id")
    assert rows == [(-1.0,), (None,), (-2.0,)]


def test_arithmetic_precedence_and_parentheses(tmp_path: Path) -> None:
    file_path = tmp_path / "arithmetic_precedence.xlsx"
    _create_arithmetic_workbook(file_path)

    rows, _ = _run_query(file_path, "SELECT a + b * c FROM t WHERE id = 1")
    assert rows == [(7.0,)]

    rows, _ = _run_query(file_path, "SELECT (a + b) * c FROM t WHERE id = 1")
    assert rows == [(9.0,)]


def test_arithmetic_literals_and_aliases(tmp_path: Path) -> None:
    file_path = tmp_path / "arithmetic_literals_aliases.xlsx"
    _create_arithmetic_workbook(file_path)

    rows, description = _run_query(
        file_path,
        "SELECT price * 1.1, price + 100, price * 2 AS doubled FROM t WHERE id = 1",
    )
    assert rows == [(11.0, 110.0, 20.0)]
    assert description == ["price * 1.1", "price + 100", "doubled"]


def test_arithmetic_alias_projection_and_description(tmp_path: Path) -> None:
    file_path = tmp_path / "arithmetic_alias_description.xlsx"
    _create_arithmetic_workbook(file_path)

    rows, description = _run_query(
        file_path,
        "SELECT price * quantity AS total, (a + b) * 2 AS result FROM t WHERE id = 1",
    )
    assert rows == [(20.0, 6.0)]
    assert description == ["total", "result"]


def test_arithmetic_join_qualified_columns(tmp_path: Path) -> None:
    file_path = tmp_path / "arithmetic_join.xlsx"
    _create_arithmetic_workbook(file_path)

    rows, description = _run_query(
        file_path,
        "SELECT t1.price * t2.qty FROM t1 INNER JOIN t2 ON t1.id = t2.id ORDER BY t1.id",
    )
    assert rows == [(30.0,), (20.0,)]
    assert description == ["t1.price * t2.qty"]


def test_arithmetic_null_propagation(tmp_path: Path) -> None:
    file_path = tmp_path / "arithmetic_null_propagation.xlsx"
    _create_arithmetic_workbook(file_path)

    rows, _ = _run_query(file_path, "SELECT price * quantity FROM t WHERE id = 3")
    assert rows == [(None,)]


def test_arithmetic_division_by_zero_raises_programming_error(tmp_path: Path) -> None:
    file_path = tmp_path / "arithmetic_div_zero.xlsx"
    _create_arithmetic_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="Division by zero"):
            cursor.execute("SELECT total / qty FROM t WHERE id = 2")


def test_arithmetic_non_numeric_raises_programming_error(tmp_path: Path) -> None:
    file_path = tmp_path / "arithmetic_non_numeric.xlsx"
    _create_arithmetic_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="numeric operands"):
            cursor.execute("SELECT label + price FROM t WHERE id = 1")


def test_order_by_arithmetic_alias(tmp_path: Path) -> None:
    file_path = tmp_path / "arithmetic_order_by_alias.xlsx"
    _create_arithmetic_workbook(file_path)

    rows, description = _run_query(
        file_path,
        "SELECT price * qty AS total FROM t ORDER BY total",
    )
    assert description == ["total"]
    assert rows == [(0.0,), (20.0,), (None,)]


def test_arithmetic_with_where_clause(tmp_path: Path) -> None:
    file_path = tmp_path / "arithmetic_where.xlsx"
    _create_arithmetic_workbook(file_path)

    rows, _ = _run_query(
        file_path,
        "SELECT price * qty AS total FROM t WHERE price > 0 ORDER BY id",
    )
    assert rows == [(20.0,), (0.0,)]


def test_backward_compatibility_non_expression_select_forms(tmp_path: Path) -> None:
    file_path = tmp_path / "arithmetic_backward_compat.xlsx"
    _create_arithmetic_workbook(file_path)

    rows, description = _run_query(file_path, "SELECT price FROM t WHERE id = 1")
    assert rows == [(10,)]
    assert description == ["price"]

    rows, description = _run_query(file_path, "SELECT price AS p FROM t WHERE id = 1")
    assert rows == [(10,)]
    assert description == ["p"]

    rows, description = _run_query(file_path, "SELECT COUNT(*) FROM t")
    assert rows == [(3,)]
    assert description == ["COUNT(*)"]

    rows, description = _run_query(file_path, "SELECT * FROM t ORDER BY id")
    assert len(rows) == 3
    assert description[0] == "id"


def test_aggregate_with_arithmetic_arg_rejected(tmp_path: Path) -> None:
    """SUM(price * qty) and COUNT(price + 1) must be rejected by the parser."""
    file_path = tmp_path / "arithmetic_agg_reject.xlsx"
    _create_arithmetic_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError):
            cursor.execute("SELECT SUM(price * qty) FROM t")

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError):
            cursor.execute("SELECT COUNT(price + 1) FROM t")


def test_arithmetic_alias_order_by_with_limit_offset(tmp_path: Path) -> None:
    """Arithmetic alias + ORDER BY + LIMIT/OFFSET regression test."""
    file_path = tmp_path / "arithmetic_order_limit.xlsx"
    _create_arithmetic_workbook(file_path)

    rows, description = _run_query(
        file_path,
        "SELECT price * qty AS total FROM t ORDER BY total LIMIT 2",
    )
    assert description == ["total"]
    assert len(rows) == 2
    assert rows == [(0.0,), (20.0,)]

    rows, description = _run_query(
        file_path,
        "SELECT price * qty AS total FROM t ORDER BY total LIMIT 2 OFFSET 1",
    )
    assert description == ["total"]
    assert len(rows) == 2
    assert rows == [(20.0,), (None,)]
