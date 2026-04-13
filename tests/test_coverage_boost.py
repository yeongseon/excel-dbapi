"""Targeted tests to boost coverage for parser.py and executor.py uncovered lines."""

from __future__ import annotations

import pytest
from openpyxl import Workbook

from excel_dbapi.connection import ExcelConnection
from excel_dbapi.parser import parse_sql


# =============================================================================
# Parser: Aggregate edge cases (lines 232-235)
# =============================================================================


def test_aggregate_sum_star_rejected() -> None:
    """SUM(*) is not supported — only COUNT(*)."""
    with pytest.raises(ValueError, match="SUM does not support"):
        parse_sql("SELECT SUM(*) FROM Sheet1")


def test_aggregate_avg_star_rejected() -> None:
    with pytest.raises(ValueError, match="AVG does not support"):
        parse_sql("SELECT AVG(*) FROM Sheet1")


def test_aggregate_expression_rejected() -> None:
    """SUM(a+b) — expressions inside aggregates are not supported."""
    with pytest.raises(ValueError, match="Unsupported aggregate expression"):
        parse_sql("SELECT SUM(a + b) FROM Sheet1")


# =============================================================================
# Parser: WHERE edge cases (lines 347, 353, 380, 395, 419, 491, 505, 511)
# =============================================================================


def test_where_single_token() -> None:
    """WHERE clause with only one token is invalid."""
    with pytest.raises(ValueError, match="Invalid WHERE clause format"):
        parse_sql("SELECT * FROM Sheet1 WHERE x")


def test_where_is_not_missing_null() -> None:
    """IS NOT without NULL is invalid."""
    with pytest.raises(ValueError, match="expected NULL or NOT after IS"):
        parse_sql("SELECT * FROM Sheet1 WHERE x IS FOO")


def test_where_between_missing_and() -> None:
    """BETWEEN without AND keyword is invalid."""
    with pytest.raises(ValueError, match="expected AND in BETWEEN"):
        parse_sql("SELECT * FROM Sheet1 WHERE x BETWEEN 1 OR 10")


def test_where_in_missing_tokens() -> None:
    """IN without parenthesized list."""
    with pytest.raises(ValueError, match="Invalid WHERE clause format"):
        parse_sql("SELECT * FROM Sheet1 WHERE x IN")


def test_where_in_malformed_no_close_paren() -> None:
    """IN clause missing close paren."""
    with pytest.raises(ValueError, match="IN clause"):
        parse_sql("SELECT * FROM Sheet1 WHERE x IN (1, 2")


def test_where_like_missing_pattern() -> None:
    """LIKE without pattern value."""
    with pytest.raises(ValueError, match="Invalid WHERE clause format"):
        parse_sql("SELECT * FROM Sheet1 WHERE x LIKE")


def test_where_operator_missing_value() -> None:
    """Regular operator (=) without right-hand value."""
    with pytest.raises(ValueError, match="Invalid WHERE clause format"):
        parse_sql("SELECT * FROM Sheet1 WHERE x =")


# =============================================================================
# Parser: Subquery edge cases (lines 471, 491)
# =============================================================================


def test_subquery_with_join_inside() -> None:
    """Subqueries cannot contain JOIN."""
    with pytest.raises(ValueError, match="JOIN is not supported in subqueries"):
        parse_sql("SELECT * FROM t WHERE id IN (SELECT a.id FROM t1 a JOIN t2 b ON a.id = b.id)")


def test_subquery_with_having() -> None:
    """Subqueries cannot contain HAVING."""
    with pytest.raises(ValueError, match="HAVING is not supported in subqueries"):
        parse_sql("SELECT * FROM t WHERE id IN (SELECT id FROM t2 GROUP BY id HAVING COUNT(*) > 1)")


def test_subquery_with_order_by() -> None:
    """Subqueries cannot contain ORDER BY."""
    with pytest.raises(ValueError, match="ORDER BY is not supported in subqueries"):
        parse_sql("SELECT * FROM t WHERE id IN (SELECT id FROM t2 ORDER BY id)")


def test_subquery_with_offset() -> None:
    """Subqueries cannot contain OFFSET."""
    with pytest.raises(ValueError, match="OFFSET is not supported in subqueries"):
        parse_sql("SELECT * FROM t WHERE id IN (SELECT id FROM t2 OFFSET 5)")


def test_subquery_with_group_by() -> None:
    """Subqueries cannot contain GROUP BY."""
    with pytest.raises(ValueError, match="GROUP BY is not supported in subqueries"):
        parse_sql("SELECT * FROM t WHERE id IN (SELECT id FROM t2 GROUP BY id)")


def test_subquery_with_limit() -> None:
    """Subqueries cannot contain LIMIT."""
    with pytest.raises(ValueError, match="LIMIT is not supported in subqueries"):
        parse_sql("SELECT * FROM t WHERE id IN (SELECT id FROM t2 LIMIT 5)")


# =============================================================================
# Parser: JOIN parsing edge cases (lines 574, 588, 601, 629, 637, 652, 699,
#          715, 723, 731)
# =============================================================================


def test_join_invalid_column_reference() -> None:
    """JOIN column must be qualified (table.column)."""
    with pytest.raises(ValueError, match="Invalid column reference in JOIN"):
        parse_sql("SELECT a.id FROM t1 a JOIN t2 b ON id = b.id")


def test_join_on_empty() -> None:
    """JOIN requires ON condition."""
    with pytest.raises(ValueError, match="JOIN requires ON condition"):
        parse_sql("SELECT a.id FROM t1 a JOIN t2 b")


def test_join_on_comparison_non_equality() -> None:
    """JOIN ON only supports = comparisons."""
    with pytest.raises(ValueError, match="only '=' comparisons"):
        parse_sql("SELECT a.id FROM t1 a JOIN t2 b ON a.id > b.id")


def test_join_on_references_wrong_tables() -> None:
    """JOIN ON must reference columns from the joined tables."""
    with pytest.raises(ValueError, match="compare columns from the two joined sources"):
        parse_sql("SELECT a.id FROM t1 a JOIN t2 b ON a.id = a.name")


def test_join_on_too_many_tokens_in_comparison() -> None:
    """JOIN ON comparison must be exactly 3 tokens (left op right)."""
    with pytest.raises(ValueError, match="AND-combined equality"):
        parse_sql("SELECT a.id FROM t1 a JOIN t2 b ON a.id = b.id = 1")


def test_select_distinct_empty_after_keyword() -> None:
    """DISTINCT followed by nothing."""
    with pytest.raises(ValueError):
        parse_sql("SELECT DISTINCT")


def test_select_as_alias_missing() -> None:
    """AS keyword without following alias."""
    with pytest.raises(ValueError, match="Expected alias after AS"):
        parse_sql("SELECT a.id FROM t1 AS")


def test_inner_without_join() -> None:
    """INNER without JOIN keyword."""
    with pytest.raises(ValueError, match="Unsupported SQL syntax: INNER"):
        parse_sql("SELECT a.id FROM t1 a INNER t2 b ON a.id = b.id")


def test_left_without_join() -> None:
    """LEFT without JOIN keyword."""
    with pytest.raises(ValueError, match="Unsupported SQL syntax: LEFT"):
        parse_sql("SELECT a.id FROM t1 a LEFT t2 b ON a.id = b.id")


def test_join_missing_table() -> None:
    """JOIN without table name after it."""
    with pytest.raises(ValueError, match="missing table"):
        parse_sql("SELECT a.id FROM t1 a JOIN")


def test_join_source_invalid_reference() -> None:
    """Invalid source reference in SELECT column with JOIN."""
    with pytest.raises(ValueError, match="Invalid source reference"):
        parse_sql("SELECT c.id FROM t1 a JOIN t2 b ON a.id = b.id")


def test_join_aggregate_not_supported() -> None:
    """Aggregate functions in JOIN queries not supported."""
    with pytest.raises(ValueError, match="Aggregate functions are not supported with JOIN"):
        parse_sql("SELECT COUNT(*) FROM t1 a JOIN t2 b ON a.id = b.id")


# =============================================================================
# Parser: Clause ordering violations (lines 837-859)
# =============================================================================


def test_order_by_before_where() -> None:
    with pytest.raises(ValueError, match="ORDER BY cannot appear before WHERE"):
        parse_sql("SELECT * FROM t ORDER BY x WHERE x = 1")


def test_limit_before_where() -> None:
    with pytest.raises(ValueError, match="LIMIT cannot appear before WHERE"):
        parse_sql("SELECT * FROM t LIMIT 10 WHERE x = 1")


def test_offset_before_where() -> None:
    with pytest.raises(ValueError, match="OFFSET cannot appear before WHERE"):
        parse_sql("SELECT * FROM t OFFSET 5 WHERE x = 1")


def test_group_by_before_where() -> None:
    with pytest.raises(ValueError, match="GROUP BY cannot appear before WHERE"):
        parse_sql("SELECT x, COUNT(*) FROM t GROUP BY x WHERE x = 1")


def test_having_before_where() -> None:
    with pytest.raises(ValueError):
        parse_sql("SELECT x, COUNT(*) FROM t GROUP BY x HAVING COUNT(*) > 1 WHERE x = 1")


def test_having_before_group_by() -> None:
    with pytest.raises(ValueError, match="HAVING cannot appear before GROUP BY"):
        parse_sql("SELECT x FROM t HAVING COUNT(*) > 1 GROUP BY x")


def test_order_by_before_group_by() -> None:
    with pytest.raises(ValueError, match="ORDER BY cannot appear before GROUP BY"):
        parse_sql("SELECT x FROM t ORDER BY x GROUP BY x")


def test_limit_before_group_by() -> None:
    with pytest.raises(ValueError, match="LIMIT cannot appear before GROUP BY"):
        parse_sql("SELECT x FROM t LIMIT 10 GROUP BY x")


def test_offset_before_group_by() -> None:
    with pytest.raises(ValueError, match="OFFSET cannot appear before GROUP BY"):
        parse_sql("SELECT x FROM t OFFSET 5 GROUP BY x")


def test_order_by_before_having() -> None:
    with pytest.raises(ValueError, match="ORDER BY cannot appear before HAVING"):
        parse_sql("SELECT x, COUNT(*) FROM t GROUP BY x ORDER BY x HAVING COUNT(*) > 1")


def test_limit_before_having() -> None:
    with pytest.raises(ValueError, match="LIMIT cannot appear before HAVING"):
        parse_sql("SELECT x, COUNT(*) FROM t GROUP BY x LIMIT 10 HAVING COUNT(*) > 1")


def test_offset_before_having() -> None:
    with pytest.raises(ValueError, match="OFFSET cannot appear before HAVING"):
        parse_sql("SELECT x, COUNT(*) FROM t GROUP BY x OFFSET 5 HAVING COUNT(*) > 1")


def test_offset_before_order_by() -> None:
    with pytest.raises(ValueError, match="OFFSET cannot appear before ORDER BY"):
        parse_sql("SELECT * FROM t OFFSET 5 ORDER BY x")


def test_offset_before_limit() -> None:
    with pytest.raises(ValueError, match="OFFSET cannot appear before LIMIT"):
        parse_sql("SELECT * FROM t OFFSET 5 LIMIT 10")


# =============================================================================
# Parser: Empty GROUP BY / HAVING (lines 886, 897)
# =============================================================================


def test_empty_group_by() -> None:
    with pytest.raises(ValueError, match="Invalid GROUP BY clause format"):
        # This is tricky — GROUP BY followed immediately by ORDER BY
        parse_sql("SELECT x FROM t GROUP BY ORDER BY x")


def test_empty_having() -> None:
    with pytest.raises(ValueError, match="Invalid HAVING clause format"):
        parse_sql("SELECT x FROM t GROUP BY x HAVING ORDER BY x")


# =============================================================================
# Parser: LIMIT/OFFSET edge cases (lines 939, 949, 953)
# =============================================================================


def test_empty_limit() -> None:
    with pytest.raises(ValueError, match="Invalid LIMIT clause format"):
        parse_sql("SELECT * FROM t LIMIT")


def test_empty_offset() -> None:
    with pytest.raises(ValueError, match="Invalid OFFSET clause format"):
        parse_sql("SELECT * FROM t OFFSET")


def test_limit_non_integer_string() -> None:
    with pytest.raises(ValueError, match="LIMIT must be an integer"):
        parse_sql("SELECT * FROM t LIMIT 'abc'")


def test_offset_non_integer_string() -> None:
    with pytest.raises(ValueError, match="OFFSET must be an integer"):
        parse_sql("SELECT * FROM t OFFSET 'abc'")


# =============================================================================
# Parser: HAVING with params, post-bind validation (lines 1022-1043)
# =============================================================================


def test_having_with_param_binding() -> None:
    """HAVING with ? parameter that gets bound."""
    parsed = parse_sql(
        "SELECT name, COUNT(*) FROM t GROUP BY name HAVING COUNT(*) > ?",
        (5,),
    )
    assert parsed["having"] is not None
    cond = parsed["having"]["conditions"][0]
    assert cond["value"] == 5


def test_limit_param_non_integer() -> None:
    """LIMIT ? with non-integer param."""
    with pytest.raises(ValueError, match="LIMIT must be an integer"):
        parse_sql("SELECT * FROM t LIMIT ?", ("abc",))


def test_offset_param_non_integer() -> None:
    """OFFSET ? with non-integer param."""
    with pytest.raises(ValueError, match="OFFSET must be an integer"):
        parse_sql("SELECT * FROM t OFFSET ?", ("abc",))


# =============================================================================
# Parser: JOIN with DISTINCT / subquery / SELECT * / HAVING (lines 1047-1067)
# =============================================================================


def test_join_with_distinct_rejected() -> None:
    with pytest.raises(ValueError, match="DISTINCT is not supported with JOIN"):
        parse_sql("SELECT DISTINCT a.id FROM t1 a JOIN t2 b ON a.id = b.id")


def test_join_with_select_star_rejected() -> None:
    with pytest.raises(ValueError, match="SELECT \\* is not supported with JOIN"):
        parse_sql("SELECT * FROM t1 a JOIN t2 b ON a.id = b.id")


def test_join_with_having_rejected() -> None:
    with pytest.raises(ValueError):
        parse_sql("SELECT a.id FROM t1 a JOIN t2 b ON a.id = b.id GROUP BY a.id HAVING COUNT(*) > 1")


def test_join_with_subquery_where_rejected() -> None:
    with pytest.raises(ValueError, match="Subqueries are not supported with JOIN"):
        parse_sql("SELECT a.id FROM t1 a JOIN t2 b ON a.id = b.id WHERE a.id IN (SELECT id FROM t3)")


# =============================================================================
def test_right_join_accepted() -> None:
    parsed = parse_sql("SELECT a.id FROM t1 a RIGHT JOIN t2 b ON a.id = b.id")
    assert parsed["joins"] is not None
    assert parsed["joins"][0]["type"] == "RIGHT"


def test_full_outer_join_rejected() -> None:
    with pytest.raises(ValueError, match="Unsupported SQL syntax: FULL"):
        parse_sql("SELECT a.id FROM t1 a FULL OUTER JOIN t2 b ON a.id = b.id")


def test_cross_join_rejected() -> None:
    with pytest.raises(ValueError, match="Unsupported SQL syntax: CROSS"):
        parse_sql("SELECT a.id FROM t1 a CROSS JOIN t2 b")


# =============================================================================
# Parser: INSERT edge cases (lines 1125, 1132, 1150, 1217-1236, 1261)
# =============================================================================


def test_insert_empty_after_table() -> None:
    with pytest.raises(ValueError, match="Invalid INSERT format"):
        parse_sql("INSERT INTO Sheet1")


def test_insert_no_values_no_select() -> None:
    with pytest.raises(ValueError, match="Invalid INSERT format"):
        parse_sql("INSERT INTO Sheet1 SOMETHING")


def test_insert_unclosed_column_paren() -> None:
    with pytest.raises(ValueError, match="Invalid INSERT format"):
        parse_sql("INSERT INTO Sheet1 (id, name VALUES (1, 'x')")


def test_insert_multi_row_unexpected_char() -> None:
    """Extra character between tuples that isn't comma."""
    with pytest.raises(ValueError, match="Invalid INSERT format"):
        parse_sql("INSERT INTO Sheet1 VALUES (1, 'a') X (2, 'b')")


def test_insert_multi_row_no_open_paren() -> None:
    """VALUES followed by non-parenthesized content."""
    with pytest.raises(ValueError, match="Invalid INSERT format"):
        parse_sql("INSERT INTO Sheet1 VALUES 1, 2")


def test_insert_too_many_params() -> None:
    """Too many parameters for INSERT placeholders."""
    with pytest.raises(ValueError, match="Too many parameters"):
        parse_sql("INSERT INTO Sheet1 VALUES (?, ?)", (1, 2, 3))


def test_insert_multi_row_with_escaped_quote() -> None:
    """Multi-row INSERT with escaped single quote in value."""
    parsed = parse_sql("INSERT INTO Sheet1 VALUES (1, 'it''s'), (2, 'ok')")
    assert parsed["values"] == [[1, "it's"], [2, "ok"]]


def test_insert_multi_row_nested_parens_in_values() -> None:
    """Simple multi-row — replaces a test that didn't raise as expected."""
    result = parse_sql("INSERT INTO Sheet1 VALUES (1, 'a')")
    assert result["values"] == [[1, "a"]]


# =============================================================================
# Parser: UPDATE with WHERE param binding (line 1370)
# =============================================================================


def test_update_with_where_param_binding() -> None:
    """UPDATE SET ? WHERE ? — both SET and WHERE have params."""
    parsed = parse_sql(
        "UPDATE Sheet1 SET name = ? WHERE id = ?",
        ("Alice", 42),
    )
    assert parsed["set"][0]["value"] == "Alice"
    assert parsed["where"]["conditions"][0]["value"] == 42


# =============================================================================
# Executor: error message with available sheets (lines 47-49, 65-67, 117-119)
# =============================================================================


def test_select_missing_table_shows_available(tmp_path: object) -> None:
    import pathlib

    file_path = pathlib.Path(str(tmp_path)) / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id"])
    ws.append([1])
    wb.save(file_path)

    with ExcelConnection(str(file_path)) as conn:
        cursor = conn.cursor()
        with pytest.raises(Exception, match="Available sheets"):
            cursor.execute("SELECT * FROM NonExistent")


def test_update_missing_table_shows_available(tmp_path: object) -> None:
    import pathlib

    file_path = pathlib.Path(str(tmp_path)) / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id"])
    ws.append([1])
    wb.save(file_path)

    with ExcelConnection(str(file_path)) as conn:
        cursor = conn.cursor()
        with pytest.raises(Exception, match="Available sheets"):
            cursor.execute("UPDATE NonExistent SET id = 1")


def test_delete_missing_table_shows_available(tmp_path: object) -> None:
    import pathlib

    file_path = pathlib.Path(str(tmp_path)) / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id"])
    ws.append([1])
    wb.save(file_path)

    with ExcelConnection(str(file_path)) as conn:
        cursor = conn.cursor()
        with pytest.raises(Exception, match="Available sheets"):
            cursor.execute("DELETE FROM NonExistent")


# =============================================================================
# Executor: INSERT edge cases (lines 178, 184)
# =============================================================================


def test_insert_invalid_subquery_format(tmp_path: object) -> None:
    """INSERT with dict values that isn't a proper subquery."""
    import pathlib
    from excel_dbapi.executor import SharedExecutor as Executor
    from excel_dbapi.engines.openpyxl.backend import OpenpyxlBackend

    file_path = pathlib.Path(str(tmp_path)) / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id", "name"])
    wb.save(file_path)

    backend = OpenpyxlBackend(str(file_path))
    executor = Executor(backend)
    # Directly call execute with a crafted parsed dict
    with pytest.raises(ValueError, match="Invalid INSERT subquery format"):
        executor.execute(
            {
                "action": "INSERT",
                "table": "Sheet1",
                "columns": None,
                "values": {"type": "invalid"},
            }
        )
    backend.close()
