"""Tests for NOT operator, parenthesized WHERE, and NOT IN/LIKE/BETWEEN.

Phase 10 of the SQL feature implementation.
"""

from pathlib import Path

import pytest
from openpyxl import Workbook

from excel_dbapi.connection import ExcelConnection
from excel_dbapi.parser import _parse_where_expression


# ---------------------------------------------------------------------------
# Fixture helpers
# ---------------------------------------------------------------------------


def _create_basic_workbook(path: Path) -> None:
    """id(int), name(str), score(int), grade(str)."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "data"
    ws.append(["id", "name", "score", "grade"])
    ws.append([1, "Alice", 85, "A"])
    ws.append([2, "Bob", 72, "B"])
    ws.append([3, "Charlie", 91, "A"])
    ws.append([4, "Diana", 68, "C"])
    ws.append([5, "Eve", 45, "F"])
    wb.save(path)


def _create_two_sheet_workbook(path: Path) -> None:
    """Two sheets: 'users' and 'scores'."""
    wb = Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "users"
    ws1.append(["id", "name"])
    ws1.append([1, "Alice"])
    ws1.append([2, "Bob"])
    ws1.append([3, "Charlie"])

    ws2 = wb.create_sheet("scores")
    ws2.append(["user_id", "score"])
    ws2.append([1, 85])
    ws2.append([2, 72])
    ws2.append([4, 99])  # user_id=4 not in users
    wb.save(path)


# ===================================================================
# Parser-level tests (unit tests on parse tree structure)
# ===================================================================


class TestParserNotOperator:
    """Unary NOT operator in WHERE expressions."""

    def test_not_simple_condition(self) -> None:
        result = _parse_where_expression(
            "NOT x = 1", params=None, bind_params=False
        )
        assert "conditions" in result
        cond = result["conditions"][0]
        assert cond["type"] == "not"
        assert cond["operand"]["column"] == "x"
        assert cond["operand"]["operator"] == "="

    def test_not_greater_than(self) -> None:
        result = _parse_where_expression(
            "NOT x > 5", params=None, bind_params=False
        )
        cond = result["conditions"][0]
        assert cond["type"] == "not"
        assert cond["operand"]["operator"] == ">"
        assert cond["operand"]["value"] == 5

    def test_not_is_null(self) -> None:
        result = _parse_where_expression(
            "NOT x IS NULL", params=None, bind_params=False
        )
        cond = result["conditions"][0]
        assert cond["type"] == "not"
        assert cond["operand"]["operator"] == "IS"
        assert cond["operand"]["value"] is None

    def test_not_with_and(self) -> None:
        result = _parse_where_expression(
            "NOT x = 1 AND y = 2", params=None, bind_params=False
        )
        # AND binds tighter, so: (NOT x=1) AND (y=2)
        assert len(result["conditions"]) == 2
        assert result["conditions"][0]["type"] == "not"
        assert result["conjunctions"] == ["AND"]

    def test_not_parenthesized_group(self) -> None:
        result = _parse_where_expression(
            "NOT (x = 1 OR y = 2)", params=None, bind_params=False
        )
        cond = result["conditions"][0]
        assert cond["type"] == "not"
        operand = cond["operand"]
        assert operand["type"] == "compound"
        assert len(operand["conditions"]) == 2
        assert operand["conjunctions"] == ["OR"]

    def test_double_not(self) -> None:
        result = _parse_where_expression(
            "NOT NOT x = 1", params=None, bind_params=False
        )
        cond = result["conditions"][0]
        assert cond["type"] == "not"
        inner = cond["operand"]
        assert inner["type"] == "not"
        assert inner["operand"]["column"] == "x"


class TestParserNotInLikeBetween:
    """NOT IN, NOT LIKE, NOT BETWEEN as operator-level negation."""

    def test_not_in_literals(self) -> None:
        result = _parse_where_expression(
            "name NOT IN ('Alice', 'Bob')", params=None, bind_params=False
        )
        cond = result["conditions"][0]
        assert cond["operator"] == "NOT IN"
        assert cond["column"] == "name"
        assert cond["value"] == ("Alice", "Bob")

    def test_not_like(self) -> None:
        result = _parse_where_expression(
            "name NOT LIKE 'A%'", params=None, bind_params=False
        )
        cond = result["conditions"][0]
        assert cond["operator"] == "NOT LIKE"
        assert cond["value"] == "A%"

    def test_not_between(self) -> None:
        result = _parse_where_expression(
            "score NOT BETWEEN 70 AND 90", params=None, bind_params=False
        )
        cond = result["conditions"][0]
        assert cond["operator"] == "NOT BETWEEN"
        assert cond["value"] == (70, 90)

    def test_not_in_with_and(self) -> None:
        result = _parse_where_expression(
            "name NOT IN ('Alice') AND score > 50",
            params=None,
            bind_params=False,
        )
        assert len(result["conditions"]) == 2
        assert result["conditions"][0]["operator"] == "NOT IN"
        assert result["conjunctions"] == ["AND"]


class TestParserParenthesizedWhere:
    """Parenthesized expressions in WHERE clause."""

    def test_simple_parenthesized(self) -> None:
        result = _parse_where_expression(
            "(x = 1)", params=None, bind_params=False
        )
        cond = result["conditions"][0]
        # Simple (x = 1) is semantically x = 1 — compound wrapper is unwrapped
        assert cond["column"] == "x"
        assert cond["operator"] == "="

    def test_or_with_parenthesized_and(self) -> None:
        result = _parse_where_expression(
            "(x = 1 OR y = 2) AND z = 3", params=None, bind_params=False
        )
        assert len(result["conditions"]) == 2
        group = result["conditions"][0]
        assert group.get("type") == "compound"
        assert group["conjunctions"] == ["OR"]

    def test_and_with_parenthesized_or(self) -> None:
        result = _parse_where_expression(
            "x = 1 AND (y = 2 OR z = 3)", params=None, bind_params=False
        )
        assert len(result["conditions"]) == 2
        assert result["conjunctions"] == ["AND"]
        group = result["conditions"][1]
        assert group.get("type") == "compound"
        assert group["conjunctions"] == ["OR"]

    def test_two_parenthesized_groups(self) -> None:
        result = _parse_where_expression(
            "(a = 1 AND b = 2) OR (c = 3 AND d = 4)",
            params=None,
            bind_params=False,
        )
        assert result["conjunctions"] == ["OR"]
        left = result["conditions"][0]
        right = result["conditions"][1]
        assert left.get("type") == "compound"
        assert right.get("type") == "compound"
        assert left["conjunctions"] == ["AND"]
        assert right["conjunctions"] == ["AND"]

    def test_nested_parentheses(self) -> None:
        result = _parse_where_expression(
            "((x = 1))", params=None, bind_params=False
        )
        # Double parens unwrap to simple atom at top level
        cond = result["conditions"][0]
        assert cond["column"] == "x"

    def test_parenthesized_with_not_in(self) -> None:
        result = _parse_where_expression(
            "(x NOT IN (1, 2)) AND y = 3",
            params=None,
            bind_params=False,
        )
        group = result["conditions"][0]
        assert group.get("type") == "compound"
        inner = group["conditions"][0]
        assert inner["operator"] == "NOT IN"


class TestParserParameterBinding:
    """Parameter binding works with new tree nodes."""

    def test_not_with_placeholder(self) -> None:
        result = _parse_where_expression("NOT x = ?", params=(42,))
        cond = result["conditions"][0]
        assert cond["type"] == "not"
        assert cond["operand"]["value"] == 42

    def test_parenthesized_with_placeholders(self) -> None:
        result = _parse_where_expression(
            "(x = ? OR y = ?) AND z = ?", params=(1, 2, 3)
        )
        group = result["conditions"][0]
        assert group["conditions"][0]["value"] == 1
        assert group["conditions"][1]["value"] == 2
        assert result["conditions"][1]["value"] == 3

    def test_not_in_with_placeholders(self) -> None:
        result = _parse_where_expression(
            "name NOT IN (?, ?)", params=("Alice", "Bob")
        )
        cond = result["conditions"][0]
        assert cond["operator"] == "NOT IN"
        assert cond["value"] == ("Alice", "Bob")

    def test_not_between_with_placeholders(self) -> None:
        result = _parse_where_expression(
            "score NOT BETWEEN ? AND ?", params=(70, 90)
        )
        cond = result["conditions"][0]
        assert cond["operator"] == "NOT BETWEEN"
        assert cond["value"] == (70, 90)

    def test_not_like_with_placeholder(self) -> None:
        result = _parse_where_expression(
            "name NOT LIKE ?", params=("A%",)
        )
        cond = result["conditions"][0]
        assert cond["operator"] == "NOT LIKE"
        assert cond["value"] == "A%"


# ===================================================================
# Regression tests (existing patterns still work)
# ===================================================================


class TestParserRegression:
    """Ensure existing WHERE patterns remain unchanged."""

    def test_simple_and(self) -> None:
        result = _parse_where_expression(
            "x = 1 AND y = 2", params=None, bind_params=False
        )
        assert result["conjunctions"] == ["AND"]
        assert len(result["conditions"]) == 2

    def test_simple_or(self) -> None:
        result = _parse_where_expression(
            "x = 1 OR y = 2", params=None, bind_params=False
        )
        assert result["conjunctions"] == ["OR"]

    def test_and_or_precedence(self) -> None:
        result = _parse_where_expression(
            "x = 1 AND y = 2 OR z = 3", params=None, bind_params=False
        )
        # AND binds tighter: (x=1 AND y=2) OR z=3
        # New parser: top-level OR, left side is AND compound
        assert result["conjunctions"] == ["OR"]

    def test_in_clause(self) -> None:
        result = _parse_where_expression(
            "x IN (1, 2, 3)", params=None, bind_params=False
        )
        cond = result["conditions"][0]
        assert cond["operator"] == "IN"
        assert cond["value"] == (1, 2, 3)

    def test_between_clause(self) -> None:
        result = _parse_where_expression(
            "x BETWEEN 1 AND 10", params=None, bind_params=False
        )
        cond = result["conditions"][0]
        assert cond["operator"] == "BETWEEN"
        assert cond["value"] == (1, 10)

    def test_like_clause(self) -> None:
        result = _parse_where_expression(
            "x LIKE 'A%'", params=None, bind_params=False
        )
        cond = result["conditions"][0]
        assert cond["operator"] == "LIKE"

    def test_is_null(self) -> None:
        result = _parse_where_expression(
            "x IS NULL", params=None, bind_params=False
        )
        cond = result["conditions"][0]
        assert cond["operator"] == "IS"
        assert cond["value"] is None

    def test_is_not_null(self) -> None:
        result = _parse_where_expression(
            "x IS NOT NULL", params=None, bind_params=False
        )
        cond = result["conditions"][0]
        assert cond["operator"] == "IS NOT"
        assert cond["value"] is None


# ===================================================================
# E2E tests (full SQL against actual Excel files)
# ===================================================================


class TestE2ENotOperator:
    """NOT operator end-to-end with openpyxl backend."""

    def test_not_equals(self, tmp_path: Path) -> None:
        f = tmp_path / "test.xlsx"
        _create_basic_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute("SELECT name FROM data WHERE NOT grade = 'A'")
            names = sorted(r[0] for r in cur.fetchall())
            assert names == ["Bob", "Diana", "Eve"]

    def test_not_greater_than(self, tmp_path: Path) -> None:
        f = tmp_path / "test.xlsx"
        _create_basic_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute("SELECT name FROM data WHERE NOT score > 80")
            names = sorted(r[0] for r in cur.fetchall())
            assert names == ["Bob", "Diana", "Eve"]

    def test_not_is_null(self, tmp_path: Path) -> None:
        f = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "data"
        ws.append(["id", "val"])
        ws.append([1, "x"])
        ws.append([2, None])
        ws.append([3, "z"])
        wb.save(f)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute("SELECT id FROM data WHERE NOT val IS NULL")
            ids = sorted(r[0] for r in cur.fetchall())
            assert ids == [1, 3]

    def test_not_parenthesized(self, tmp_path: Path) -> None:
        f = tmp_path / "test.xlsx"
        _create_basic_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT name FROM data WHERE NOT (grade = 'A' OR grade = 'B')"
            )
            names = sorted(r[0] for r in cur.fetchall())
            assert names == ["Diana", "Eve"]


class TestE2ENotInLikeBetween:
    """NOT IN, NOT LIKE, NOT BETWEEN end-to-end."""

    def test_not_in_literals(self, tmp_path: Path) -> None:
        f = tmp_path / "test.xlsx"
        _create_basic_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT name FROM data WHERE name NOT IN ('Alice', 'Bob')"
            )
            names = sorted(r[0] for r in cur.fetchall())
            assert names == ["Charlie", "Diana", "Eve"]

    def test_not_in_with_params(self, tmp_path: Path) -> None:
        f = tmp_path / "test.xlsx"
        _create_basic_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT name FROM data WHERE name NOT IN (?, ?)",
                ("Alice", "Bob"),
            )
            names = sorted(r[0] for r in cur.fetchall())
            assert names == ["Charlie", "Diana", "Eve"]

    def test_not_in_subquery(self, tmp_path: Path) -> None:
        f = tmp_path / "test.xlsx"
        _create_two_sheet_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT name FROM users WHERE id NOT IN "
                "(SELECT user_id FROM scores)"
            )
            names = [r[0] for r in cur.fetchall()]
            # users id=3 (Charlie) is not in scores
            assert names == ["Charlie"]

    def test_not_like(self, tmp_path: Path) -> None:
        f = tmp_path / "test.xlsx"
        _create_basic_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute("SELECT name FROM data WHERE name NOT LIKE 'A%'")
            names = sorted(r[0] for r in cur.fetchall())
            assert names == ["Bob", "Charlie", "Diana", "Eve"]

    def test_not_like_with_param(self, tmp_path: Path) -> None:
        f = tmp_path / "test.xlsx"
        _create_basic_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT name FROM data WHERE name NOT LIKE ?", ("A%",)
            )
            names = sorted(r[0] for r in cur.fetchall())
            assert names == ["Bob", "Charlie", "Diana", "Eve"]

    def test_not_between(self, tmp_path: Path) -> None:
        f = tmp_path / "test.xlsx"
        _create_basic_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT name FROM data WHERE score NOT BETWEEN 70 AND 90"
            )
            names = sorted(r[0] for r in cur.fetchall())
            # 85 is between, 72 is between, 91 NOT between, 68 NOT between, 45 NOT between
            assert names == ["Charlie", "Diana", "Eve"]

    def test_not_between_with_params(self, tmp_path: Path) -> None:
        f = tmp_path / "test.xlsx"
        _create_basic_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT name FROM data WHERE score NOT BETWEEN ? AND ?",
                (70, 90),
            )
            names = sorted(r[0] for r in cur.fetchall())
            assert names == ["Charlie", "Diana", "Eve"]


class TestE2EParenthesizedWhere:
    """Parenthesized WHERE expressions end-to-end."""

    def test_or_parenthesized_with_and(self, tmp_path: Path) -> None:
        f = tmp_path / "test.xlsx"
        _create_basic_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            # (grade A or B) AND score > 80 → Alice(85,A), Charlie(91,A)
            cur.execute(
                "SELECT name FROM data WHERE (grade = 'A' OR grade = 'B') AND score > 80"
            )
            names = sorted(r[0] for r in cur.fetchall())
            assert names == ["Alice", "Charlie"]

    def test_and_parenthesized_with_or(self, tmp_path: Path) -> None:
        f = tmp_path / "test.xlsx"
        _create_basic_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            # grade=F OR (grade=A AND score>90) → Eve(45,F), Charlie(91,A)
            cur.execute(
                "SELECT name FROM data WHERE grade = 'F' OR (grade = 'A' AND score > 90)"
            )
            names = sorted(r[0] for r in cur.fetchall())
            assert names == ["Charlie", "Eve"]

    def test_two_parenthesized_groups(self, tmp_path: Path) -> None:
        f = tmp_path / "test.xlsx"
        _create_basic_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            # (grade=A AND score>90) OR (grade=F AND score<50)
            # → Charlie(91,A) OR Eve(45,F)
            cur.execute(
                "SELECT name FROM data WHERE "
                "(grade = 'A' AND score > 90) OR (grade = 'F' AND score < 50)"
            )
            names = sorted(r[0] for r in cur.fetchall())
            assert names == ["Charlie", "Eve"]

    def test_simple_parenthesized(self, tmp_path: Path) -> None:
        f = tmp_path / "test.xlsx"
        _create_basic_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute("SELECT name FROM data WHERE (grade = 'A')")
            names = sorted(r[0] for r in cur.fetchall())
            assert names == ["Alice", "Charlie"]

    def test_parenthesized_with_params(self, tmp_path: Path) -> None:
        f = tmp_path / "test.xlsx"
        _create_basic_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT name FROM data WHERE (grade = ? OR grade = ?) AND score > ?",
                ("A", "B", 80),
            )
            names = sorted(r[0] for r in cur.fetchall())
            assert names == ["Alice", "Charlie"]


class TestE2EComplexCombinations:
    """Complex combinations of NOT, parentheses, and negated operators."""

    def test_not_with_not_in(self, tmp_path: Path) -> None:
        f = tmp_path / "test.xlsx"
        _create_basic_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            # NOT (grade=A AND name NOT IN ('Alice'))
            # → NOT (grade=A AND name not Alice) = NOT (Charlie)
            # → everyone except Charlie: Alice, Bob, Diana, Eve
            cur.execute(
                "SELECT name FROM data WHERE "
                "NOT (grade = 'A' AND name NOT IN ('Alice'))"
            )
            names = sorted(r[0] for r in cur.fetchall())
            assert names == ["Alice", "Bob", "Diana", "Eve"]

    def test_parenthesized_or_with_not_like_and_not_between(
        self, tmp_path: Path
    ) -> None:
        f = tmp_path / "test.xlsx"
        _create_basic_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            # (name NOT LIKE 'A%' OR name NOT LIKE 'B%') AND score NOT BETWEEN 70 AND 90
            # name NOT LIKE 'A%': Bob, Charlie, Diana, Eve
            # name NOT LIKE 'B%': Alice, Charlie, Diana, Eve
            # OR → all 5 (everyone matches at least one)
            # AND score NOT BETWEEN 70 AND 90: Charlie(91), Diana(68), Eve(45)
            cur.execute(
                "SELECT name FROM data WHERE "
                "(name NOT LIKE 'A%' OR name NOT LIKE 'B%') "
                "AND score NOT BETWEEN 70 AND 90"
            )
            names = sorted(r[0] for r in cur.fetchall())
            assert names == ["Charlie", "Diana", "Eve"]

    def test_update_with_not_in(self, tmp_path: Path) -> None:
        f = tmp_path / "test.xlsx"
        _create_basic_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl", autocommit=True) as conn:
            cur = conn.cursor()
            cur.execute(
                "UPDATE data SET grade = 'X' WHERE name NOT IN ('Alice', 'Bob')"
            )
            cur.execute("SELECT name, grade FROM data WHERE grade = 'X'")
            names = sorted(r[0] for r in cur.fetchall())
            assert names == ["Charlie", "Diana", "Eve"]

    def test_delete_with_not_between(self, tmp_path: Path) -> None:
        f = tmp_path / "test.xlsx"
        _create_basic_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl", autocommit=True) as conn:
            cur = conn.cursor()
            cur.execute(
                "DELETE FROM data WHERE score NOT BETWEEN 70 AND 92"
            )
            cur.execute("SELECT name FROM data")
            names = sorted(r[0] for r in cur.fetchall())
            # Kept: Alice(85), Bob(72), Charlie(91) — all between 70-92
            assert names == ["Alice", "Bob", "Charlie"]

    def test_delete_with_not_like(self, tmp_path: Path) -> None:
        f = tmp_path / "test.xlsx"
        _create_basic_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl", autocommit=True) as conn:
            cur = conn.cursor()
            cur.execute("DELETE FROM data WHERE name NOT LIKE 'A%'")
            cur.execute("SELECT name FROM data")
            names = sorted(r[0] for r in cur.fetchall())
            assert names == ["Alice"]

    def test_parenthesized_where_with_update(self, tmp_path: Path) -> None:
        f = tmp_path / "test.xlsx"
        _create_basic_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl", autocommit=True) as conn:
            cur = conn.cursor()
            cur.execute(
                "UPDATE data SET grade = 'S' "
                "WHERE (grade = 'A' AND score > 90) OR grade = 'F'"
            )
            cur.execute("SELECT name FROM data WHERE grade = 'S'")
            names = sorted(r[0] for r in cur.fetchall())
            assert names == ["Charlie", "Eve"]


# ===================================================================
# Regression tests for recursive AST walkers (Oracle review fix)
# ===================================================================


class TestJoinWithNotAndParens:
    """JOIN queries with NOT, parenthesized WHERE, and NOT IN subqueries."""

    @staticmethod
    def _create_join_workbook(path: Path) -> None:
        """Two sheets: users(id, name, dept) and orders(id, user_id, amount)."""
        wb = Workbook()
        ws1 = wb.active
        assert ws1 is not None
        ws1.title = "users"
        ws1.append(["id", "name", "dept"])
        ws1.append([1, "Alice", "eng"])
        ws1.append([2, "Bob", "hr"])
        ws1.append([3, "Charlie", "eng"])
        ws1.append([4, "Diana", "sales"])

        ws2 = wb.create_sheet("orders")
        ws2.append(["id", "user_id", "amount"])
        ws2.append([10, 1, 100])
        ws2.append([11, 2, 200])
        ws2.append([12, 1, 150])
        ws2.append([13, 3, 300])
        wb.save(path)

    def test_join_where_not_operator(self, tmp_path: Path) -> None:
        """JOIN ... WHERE NOT users.dept = 'hr' should work."""
        f = tmp_path / "test.xlsx"
        self._create_join_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT users.name, orders.amount FROM users "
                "JOIN orders ON users.id = orders.user_id "
                "WHERE NOT users.dept = 'hr'"
            )
            rows = cur.fetchall()
            names = sorted(r[0] for r in rows)
            assert "Bob" not in names
            assert "Alice" in names

    def test_join_where_parenthesized_and_or(self, tmp_path: Path) -> None:
        """JOIN ... WHERE (a AND b) OR c with parenthesized groups."""
        f = tmp_path / "test.xlsx"
        self._create_join_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT users.name, orders.amount FROM users "
                "JOIN orders ON users.id = orders.user_id "
                "WHERE (users.dept = 'eng' AND orders.amount > 100) "
                "OR users.dept = 'hr'"
            )
            rows = cur.fetchall()
            names = sorted(r[0] for r in rows)
            # eng with amount > 100: Alice(150), Charlie(300)
            # hr: Bob(200)
            assert names == ["Alice", "Bob", "Charlie"]


class TestNestedSubqueryResolution:
    """Subqueries nested inside NOT/compound nodes must be resolved."""

    @staticmethod
    def _create_subquery_workbook(path: Path) -> None:
        """users(id, name) and scores(user_id, score)."""
        wb = Workbook()
        ws1 = wb.active
        assert ws1 is not None
        ws1.title = "users"
        ws1.append(["id", "name"])
        ws1.append([1, "Alice"])
        ws1.append([2, "Bob"])
        ws1.append([3, "Charlie"])
        ws1.append([4, "Diana"])

        ws2 = wb.create_sheet("scores")
        ws2.append(["user_id", "score"])
        ws2.append([1, 85])
        ws2.append([2, 72])
        wb.save(path)

    def test_not_in_subquery_nested_in_or(self, tmp_path: Path) -> None:
        """name = 'X' OR (id NOT IN (SELECT ...)) must resolve subquery."""
        f = tmp_path / "test.xlsx"
        self._create_subquery_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT name FROM users WHERE name = 'Alice' OR "
                "(id NOT IN (SELECT user_id FROM scores))"
            )
            rows = cur.fetchall()
            names = sorted(r[0] for r in rows)
            # Alice matches first condition
            # Charlie(3) and Diana(4) match NOT IN subquery
            assert names == ["Alice", "Charlie", "Diana"]

    def test_not_wrapping_in_subquery(self, tmp_path: Path) -> None:
        """NOT (id IN (SELECT ...)) must resolve subquery."""
        f = tmp_path / "test.xlsx"
        self._create_subquery_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT name FROM users WHERE NOT (id IN "
                "(SELECT user_id FROM scores))"
            )
            rows = cur.fetchall()
            names = sorted(r[0] for r in rows)
            # users 1,2 are in scores; 3,4 are NOT
            assert names == ["Charlie", "Diana"]

    def test_is_subquery_condition_nested(self, tmp_path: Path) -> None:
        """_is_subquery_condition must detect subqueries inside NOT/compound."""
        from excel_dbapi.parser import _is_subquery_condition

        # Flat subquery (already worked before)
        where_flat: dict[str, object] = {
            "conditions": [
                {"column": "id", "operator": "IN", "value": {"type": "subquery", "query": "SELECT 1"}}
            ],
            "conjunctions": [],
        }
        assert _is_subquery_condition(where_flat) is True  # type: ignore[arg-type]

        # Nested in NOT
        where_not: dict[str, object] = {
            "conditions": [
                {
                    "type": "not",
                    "operand": {
                        "column": "id",
                        "operator": "NOT IN",
                        "value": {"type": "subquery", "query": "SELECT 1"},
                    },
                }
            ],
            "conjunctions": [],
        }
        assert _is_subquery_condition(where_not) is True  # type: ignore[arg-type]

        # Nested in compound
        where_compound: dict[str, object] = {
            "conditions": [
                {
                    "type": "compound",
                    "conditions": [
                        {"column": "id", "operator": "IN", "value": {"type": "subquery", "query": "SELECT 1"}}
                    ],
                    "conjunctions": [],
                }
            ],
            "conjunctions": [],
        }
        assert _is_subquery_condition(where_compound) is True  # type: ignore[arg-type]

        # No subquery
        where_none: dict[str, object] = {
            "conditions": [{"column": "id", "operator": "=", "value": 1}],
            "conjunctions": [],
        }
        assert _is_subquery_condition(where_none) is False  # type: ignore[arg-type]


class TestPrecedenceGroupedNodes:
    """Precedence-generated nested groups (no 'type' key) must be walked recursively.

    When `a OR b AND c` is parsed, it becomes:
      {conditions: [a, {conditions: [b, c], conjunctions: ['AND']}], conjunctions: ['OR']}
    The inner group has no 'type': 'compound' — it's implicit precedence grouping.
    """

    @staticmethod
    def _create_workbook(path: Path) -> None:
        """users(id, name) and scores(user_id, score)."""
        wb = Workbook()
        ws1 = wb.active
        assert ws1 is not None
        ws1.title = "users"
        ws1.append(["id", "name"])
        ws1.append([1, "Alice"])
        ws1.append([2, "Bob"])
        ws1.append([3, "Charlie"])
        ws1.append([4, "Diana"])

        ws2 = wb.create_sheet("scores")
        ws2.append(["user_id", "score"])
        ws2.append([1, 85])
        ws2.append([2, 72])
        wb.save(path)

    def test_non_parenthesized_subquery_in_or_and(self, tmp_path: Path) -> None:
        """name = 'Alice' OR id NOT IN (SELECT ...) AND id > 0 must resolve subquery."""
        f = tmp_path / "test.xlsx"
        self._create_workbook(f)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT name FROM users WHERE name = 'Alice' OR "
                "id NOT IN (SELECT user_id FROM scores) AND id > 0"
            )
            rows = cur.fetchall()
            names = sorted(r[0] for r in rows)
            # Alice matches first condition
            # id NOT IN (1,2) AND id > 0: Charlie(3) and Diana(4)
            assert names == ["Alice", "Charlie", "Diana"]

    def test_is_subquery_condition_precedence_group(self) -> None:
        """_is_subquery_condition must detect subqueries in precedence groups (no 'type')."""
        from excel_dbapi.parser import _is_subquery_condition

        # Precedence-grouped: a OR (b_with_subquery AND c)
        where: dict[str, object] = {
            "conditions": [
                {"column": "name", "operator": "=", "value": "Alice"},
                {
                    "conditions": [
                        {
                            "column": "id",
                            "operator": "NOT IN",
                            "value": {"type": "subquery", "query": "SELECT user_id FROM scores"},
                        },
                        {"column": "id", "operator": ">", "value": 0},
                    ],
                    "conjunctions": ["AND"],
                },
            ],
            "conjunctions": ["OR"],
        }
        assert _is_subquery_condition(where) is True  # type: ignore[arg-type]

    def test_join_validation_precedence_group(self, tmp_path: Path) -> None:
        """JOIN WHERE with OR/AND precedence groups must validate all column refs."""
        f = tmp_path / "join_test.xlsx"
        wb = Workbook()
        ws1 = wb.active
        assert ws1 is not None
        ws1.title = "users"
        ws1.append(["id", "name"])
        ws1.append([1, "Alice"])
        ws1.append([2, "Bob"])

        ws2 = wb.create_sheet("orders")
        ws2.append(["id", "user_id", "amount"])
        ws2.append([10, 1, 100])
        ws2.append([11, 2, 200])
        wb.save(f)

        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            # Valid: all refs are qualified and belong to join sources
            cur.execute(
                "SELECT users.name, orders.amount FROM users "
                "JOIN orders ON users.id = orders.user_id "
                "WHERE users.name = 'Alice' OR orders.amount > 100 AND users.id > 0"
            )
            rows = cur.fetchall()
            names = sorted(r[0] for r in rows)
            # Alice: matches first cond. Bob: amount=200 > 100 AND id=2 > 0
            assert names == ["Alice", "Bob"]

    def test_join_rejects_subquery_in_precedence_group(self, tmp_path: Path) -> None:
        f = tmp_path / "join_sub.xlsx"
        wb = Workbook()
        ws1 = wb.active
        assert ws1 is not None
        ws1.title = "users"
        ws1.append(["id", "name"])
        ws1.append([1, "Alice"])

        ws2 = wb.create_sheet("orders")
        ws2.append(["id", "user_id", "amount"])
        ws2.append([10, 1, 100])
        wb.save(f)

        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute(
                "SELECT users.name FROM users "
                "JOIN orders ON users.id = orders.user_id "
                "WHERE users.name = 'Alice' OR "
                "users.id IN (SELECT id FROM users) AND orders.amount > 0"
            )
            rows = cur.fetchall()
            assert rows == [("Alice",)]


class TestNotInNullSemantics:
    """SQL-standard NULL handling for IN/NOT IN operators.

    Per SQL spec: ``x NOT IN (1, NULL)`` yields UNKNOWN when x != 1,
    which is treated as FALSE in WHERE clauses.
    """

    def test_not_in_with_null_candidate(self, tmp_path: Path) -> None:
        """NOT IN with NULL candidate should return empty when no match."""
        f = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "data"
        ws.append(["id", "val"])
        ws.append([1, 10])
        ws.append([2, 20])
        ws.append([3, 30])
        wb.save(f)

        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            # val NOT IN (10, NULL): 10 matches → excluded; 20, 30 → UNKNOWN → excluded
            cur.execute("SELECT id FROM data WHERE val NOT IN (10, NULL)")
            assert cur.fetchall() == []

    def test_not_in_without_null(self, tmp_path: Path) -> None:
        """NOT IN without NULL works normally."""
        f = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "data"
        ws.append(["id", "val"])
        ws.append([1, 10])
        ws.append([2, 20])
        ws.append([3, 30])
        wb.save(f)

        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute("SELECT id FROM data WHERE val NOT IN (10, 20)")
            assert cur.fetchall() == [(3,)]

    def test_not_in_only_null(self, tmp_path: Path) -> None:
        """NOT IN (NULL) → UNKNOWN for all rows → no rows returned."""
        f = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "data"
        ws.append(["id", "val"])
        ws.append([1, 10])
        ws.append([2, 20])
        wb.save(f)

        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute("SELECT id FROM data WHERE val NOT IN (NULL)")
            assert cur.fetchall() == []

    def test_in_with_null_candidate(self, tmp_path: Path) -> None:
        """IN with NULL candidate should find matches, skip NULLs."""
        f = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "data"
        ws.append(["id", "val"])
        ws.append([1, 10])
        ws.append([2, 20])
        ws.append([3, 30])
        wb.save(f)

        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            # val IN (10, NULL): 10 matches → included; 20, 30 don't match → excluded
            cur.execute("SELECT id FROM data WHERE val IN (10, NULL)")
            assert cur.fetchall() == [(1,)]


class TestThreeValuedLogic:
    """SQL three-valued logic (TRUE / FALSE / UNKNOWN) in WHERE clauses.

    Per SQL spec, comparisons involving NULL yield UNKNOWN, and:
    - NOT UNKNOWN = UNKNOWN
    - TRUE AND UNKNOWN = UNKNOWN
    - FALSE AND UNKNOWN = FALSE
    - TRUE OR UNKNOWN = TRUE
    - FALSE OR UNKNOWN = UNKNOWN
    All UNKNOWN results in WHERE are treated as FALSE (row excluded).
    """

    def _make_workbook(self, tmp_path: Path) -> Path:
        f = tmp_path / "test.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "data"
        ws.append(["id", "x"])
        ws.append([1, None])
        ws.append([2, 10])
        ws.append([3, None])
        wb.save(f)
        return f

    def test_not_equals_null(self, tmp_path: Path) -> None:
        """NOT (x = NULL) → NOT UNKNOWN = UNKNOWN → no rows."""
        f = self._make_workbook(tmp_path)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute("SELECT id FROM data WHERE NOT (x = NULL)")
            assert cur.fetchall() == []

    def test_not_in_null_subexpression(self, tmp_path: Path) -> None:
        """NOT (x IN (NULL)) → NOT UNKNOWN = UNKNOWN → no rows."""
        f = self._make_workbook(tmp_path)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute("SELECT id FROM data WHERE NOT (x IN (NULL))")
            assert cur.fetchall() == []

    def test_null_equals_null(self, tmp_path: Path) -> None:
        """x = NULL → UNKNOWN → no rows (use IS NULL instead)."""
        f = self._make_workbook(tmp_path)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute("SELECT id FROM data WHERE x = NULL")
            assert cur.fetchall() == []

    def test_not_null_comparison(self, tmp_path: Path) -> None:
        """NOT (x > NULL) → NOT UNKNOWN = UNKNOWN → no rows."""
        f = self._make_workbook(tmp_path)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute("SELECT id FROM data WHERE NOT (x > NULL)")
            assert cur.fetchall() == []

    def test_and_with_unknown(self, tmp_path: Path) -> None:
        """TRUE AND UNKNOWN = UNKNOWN → excludes row.
        For id=2: (x = 10) is TRUE, but (x = NULL) is UNKNOWN → AND → UNKNOWN.
        """
        f = self._make_workbook(tmp_path)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute("SELECT id FROM data WHERE x = 10 AND x = NULL")
            assert cur.fetchall() == []

    def test_or_with_unknown(self, tmp_path: Path) -> None:
        """TRUE OR UNKNOWN = TRUE.
        For id=2: (x = 10) is TRUE, (x = NULL) is UNKNOWN → OR → TRUE.
        """
        f = self._make_workbook(tmp_path)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute("SELECT id FROM data WHERE x = 10 OR x = NULL")
            assert cur.fetchall() == [(2,)]

    def test_false_and_unknown(self, tmp_path: Path) -> None:
        """FALSE AND UNKNOWN = FALSE → still excluded.
        For id=2: (x = 999) is FALSE, (x = NULL) is UNKNOWN → AND → FALSE.
        """
        f = self._make_workbook(tmp_path)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute("SELECT id FROM data WHERE x = 999 AND x = NULL")
            assert cur.fetchall() == []

    def test_is_null_unaffected(self, tmp_path: Path) -> None:
        """IS NULL always returns TRUE/FALSE, never UNKNOWN."""
        f = self._make_workbook(tmp_path)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute("SELECT id FROM data WHERE x IS NULL")
            assert cur.fetchall() == [(1,), (3,)]

    def test_not_is_null_unaffected(self, tmp_path: Path) -> None:
        """NOT (x IS NULL) is fine — IS NULL always returns bool, not UNKNOWN."""
        f = self._make_workbook(tmp_path)
        with ExcelConnection(str(f), engine="openpyxl") as conn:
            cur = conn.cursor()
            cur.execute("SELECT id FROM data WHERE NOT (x IS NULL)")
            assert cur.fetchall() == [(2,)]
