from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from excel_dbapi.connection import ExcelConnection


def _create_sample_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["id", "name"])
    ws.append([1, "Alice"])
    wb.save(path)


def _create_multi_row_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["id", "name"])
    ws.append([1, "Alice"])
    ws.append([2, "Bob"])
    ws.append([3, "Cara"])
    wb.save(path)


def test_openpyxl_insert_and_executemany(tmp_path: Path) -> None:
    file_path = tmp_path / "sample.xlsx"
    _create_sample_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("INSERT INTO Sheet1 (id, name) VALUES (2, 'Bob')")
        assert cursor.rowcount == 1

        cursor.executemany(
            "INSERT INTO Sheet1 (id, name) VALUES (?, ?)",
            [(3, "Cora"), (4, "Dane")],
        )
        assert cursor.rowcount == 2

    wb = load_workbook(file_path, data_only=True)
    rows = list(wb["Sheet1"].iter_rows(values_only=True))
    assert rows[-3:] == [(2, "Bob"), (3, "Cora"), (4, "Dane")]


def test_openpyxl_executemany_rollback_on_error(tmp_path: Path) -> None:
    file_path = tmp_path / "sample.xlsx"
    _create_sample_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=False) as conn:
        cursor = conn.cursor()
        with pytest.raises(Exception):
            cursor.executemany(
                "INSERT INTO Sheet1 (id, name) VALUES (?, ?)",
                [(2, "Bob"), (3,)],
            )
        conn.rollback()

    wb = load_workbook(file_path, data_only=True)
    rows = list(wb["Sheet1"].iter_rows(values_only=True))
    assert rows == [("id", "name"), (1, "Alice")]


def test_openpyxl_create_and_drop_table(tmp_path: Path) -> None:
    file_path = tmp_path / "sample.xlsx"
    _create_sample_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("CREATE TABLE NewSheet (col1, col2)")
        cursor.execute("DROP TABLE NewSheet")

    wb = load_workbook(file_path, data_only=True)
    assert "NewSheet" not in wb.sheetnames


def test_pandas_insert_and_create(tmp_path: Path) -> None:
    file_path = tmp_path / "sample.xlsx"
    df = pd.DataFrame([{"id": 1, "name": "Alice"}])
    df.to_excel(file_path, index=False, sheet_name="Sheet1")

    with ExcelConnection(str(file_path), engine="pandas", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("INSERT INTO Sheet1 (id, name) VALUES (2, 'Bob')")
        cursor.execute("CREATE TABLE Extra (col1, col2)")

    data = pd.read_excel(file_path, sheet_name=None)
    assert len(data["Sheet1"]) == 2
    assert set(data["Extra"].columns) == {"col1", "col2"}


def test_openpyxl_update_delete_and_rollback(tmp_path: Path) -> None:
    file_path = tmp_path / "sample.xlsx"
    _create_sample_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=False) as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE Sheet1 SET name = 'Ann' WHERE id = 1")
        assert cursor.rowcount == 1
        cursor.execute("DELETE FROM Sheet1 WHERE id = 1")
        assert cursor.rowcount == 1
        conn.rollback()

    wb = load_workbook(file_path, data_only=True)
    rows = list(wb["Sheet1"].iter_rows(values_only=True))
    assert rows[1] == (1, "Alice")


def test_openpyxl_update_and_delete_all(tmp_path: Path) -> None:
    file_path = tmp_path / "sample.xlsx"
    _create_sample_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE Sheet1 SET name = 'All'")
        assert cursor.rowcount == 1
        cursor.execute("DELETE FROM Sheet1")
        assert cursor.rowcount == 1

    wb = load_workbook(file_path, data_only=True)
    rows = list(wb["Sheet1"].iter_rows(values_only=True))
    assert rows == [("id", "name")]


def test_pandas_update_and_delete(tmp_path: Path) -> None:
    file_path = tmp_path / "sample.xlsx"
    df = pd.DataFrame([{"id": 1, "name": "Alice"}, {"id": 2, "name": "Bob"}])
    df.to_excel(file_path, index=False, sheet_name="Sheet1")

    with ExcelConnection(str(file_path), engine="pandas", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE Sheet1 SET name = 'Ann' WHERE id = 2")
        assert cursor.rowcount == 1
        cursor.execute("DELETE FROM Sheet1 WHERE id = 1")
        assert cursor.rowcount == 1

    data = pd.read_excel(file_path, sheet_name=None)
    assert list(data["Sheet1"]["name"]) == ["Ann"]


def test_select_order_limit_with_where_openpyxl(tmp_path: Path) -> None:
    file_path = tmp_path / "sample.xlsx"
    _create_multi_row_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, name FROM Sheet1 WHERE id >= 2 ORDER BY id DESC LIMIT 1"
        )
        results = cursor.fetchall()
        assert results == [(3, "Cara")]


def test_select_order_limit_with_where_pandas(tmp_path: Path) -> None:
    file_path = tmp_path / "sample.xlsx"
    df = pd.DataFrame(
        [
            {"id": 1, "name": "Alice"},
            {"id": 2, "name": "Bob"},
            {"id": 3, "name": "Cara"},
        ]
    )
    df.to_excel(file_path, index=False, sheet_name="Sheet1")

    with ExcelConnection(str(file_path), engine="pandas", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, name FROM Sheet1 WHERE id >= 2 ORDER BY id DESC LIMIT 1"
        )
        results = cursor.fetchall()
        assert results == [(3, "Cara")]



# ── Multi-row INSERT & INSERT...SELECT tests ──


def test_multi_row_insert_two_rows(tmp_path: Path) -> None:
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id", "name"])
    ws.append([1, "Alice"])
    wb.save(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("INSERT INTO Sheet1 VALUES (2, 'Bob'), (3, 'Carol')")
        assert cursor.rowcount == 2

    wb = load_workbook(file_path, data_only=True)
    rows = list(wb["Sheet1"].iter_rows(values_only=True))
    assert len(rows) == 4  # header + 3 data rows
    assert rows[2] == (2, "Bob")
    assert rows[3] == (3, "Carol")


def test_multi_row_insert_five_rows(tmp_path: Path) -> None:
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id", "name"])
    wb.save(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO Sheet1 VALUES "
            "(1, 'A'), (2, 'B'), (3, 'C'), (4, 'D'), (5, 'E')"
        )
        assert cursor.rowcount == 5

    wb = load_workbook(file_path, data_only=True)
    rows = list(wb["Sheet1"].iter_rows(values_only=True))
    assert len(rows) == 6  # header + 5 data rows
    assert rows[5] == (5, "E")


def test_multi_row_insert_with_columns(tmp_path: Path) -> None:
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id", "name", "extra"])
    wb.save(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO Sheet1 (id, name) VALUES (1, 'Alice'), (2, 'Bob')"
        )
        assert cursor.rowcount == 2

    wb = load_workbook(file_path, data_only=True)
    rows = list(wb["Sheet1"].iter_rows(values_only=True))
    assert len(rows) == 3  # header + 2 data rows
    assert rows[1] == (1, "Alice", None)
    assert rows[2] == (2, "Bob", None)


def test_multi_row_insert_with_params(tmp_path: Path) -> None:
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id", "name"])
    wb.save(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO Sheet1 VALUES (?, ?), (?, ?)",
            (1, "Alice", 2, "Bob"),
        )
        assert cursor.rowcount == 2

    wb = load_workbook(file_path, data_only=True)
    rows = list(wb["Sheet1"].iter_rows(values_only=True))
    assert rows[1] == (1, "Alice")
    assert rows[2] == (2, "Bob")


def test_insert_select_cross_sheet(tmp_path: Path) -> None:
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Source"
    ws.append(["id", "name"])
    ws.append([1, "Alice"])
    ws.append([2, "Bob"])
    ws2 = wb.create_sheet("Target")
    ws2.append(["id", "name"])
    wb.save(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("INSERT INTO Target SELECT id, name FROM Source")
        assert cursor.rowcount == 2

    wb = load_workbook(file_path, data_only=True)
    rows = list(wb["Target"].iter_rows(values_only=True))
    assert len(rows) == 3  # header + 2 data rows
    assert rows[1] == (1, "Alice")
    assert rows[2] == (2, "Bob")


def test_insert_select_with_where(tmp_path: Path) -> None:
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Source"
    ws.append(["id", "name"])
    ws.append([1, "Alice"])
    ws.append([2, "Bob"])
    ws.append([3, "Carol"])
    ws2 = wb.create_sheet("Target")
    ws2.append(["id", "name"])
    wb.save(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO Target SELECT id, name FROM Source WHERE id >= 2"
        )
        assert cursor.rowcount == 2

    wb = load_workbook(file_path, data_only=True)
    rows = list(wb["Target"].iter_rows(values_only=True))
    assert len(rows) == 3  # header + 2 data rows
    assert rows[1] == (2, "Bob")
    assert rows[2] == (3, "Carol")


def test_insert_select_empty_result(tmp_path: Path) -> None:
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Source"
    ws.append(["id", "name"])
    ws.append([1, "Alice"])
    ws2 = wb.create_sheet("Target")
    ws2.append(["id", "name"])
    wb.save(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO Target SELECT id, name FROM Source WHERE id > 999"
        )
        assert cursor.rowcount == 0

    wb = load_workbook(file_path, data_only=True)
    rows = list(wb["Target"].iter_rows(values_only=True))
    assert len(rows) == 1  # header only


def test_multi_row_insert_column_count_mismatch(tmp_path: Path) -> None:
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id", "name"])
    wb.save(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        with pytest.raises(Exception, match="count"):
            cursor.execute("INSERT INTO Sheet1 VALUES (1, 'Alice'), (2)")


def test_insert_select_same_sheet(tmp_path: Path) -> None:
    """INSERT...SELECT from the same sheet duplicates rows."""
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id", "name"])
    ws.append([1, "Alice"])
    wb.save(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("INSERT INTO Sheet1 SELECT id, name FROM Sheet1 WHERE id = 1")
        assert cursor.rowcount == 1

    wb = load_workbook(file_path, data_only=True)
    rows = list(wb["Sheet1"].iter_rows(values_only=True))
    assert len(rows) == 3  # header + original + duplicate
    assert rows[1] == (1, "Alice")
    assert rows[2] == (1, "Alice")


def test_multi_row_insert_atomicity_on_failure(tmp_path: Path) -> None:
    """Multi-row INSERT must not leave partial rows when a later row fails validation."""
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id", "name"])
    ws.append([1, "Existing"])
    wb.save(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        # Row 2 has wrong column count — should fail atomically
        with pytest.raises(Exception, match="count"):
            cursor.execute(
                "INSERT INTO Sheet1 VALUES (2, 'Good'), (3)"
            )

    # Verify NO partial rows were inserted
    wb = load_workbook(file_path, data_only=True)
    rows = list(wb["Sheet1"].iter_rows(values_only=True))
    assert len(rows) == 2  # header + original only
    assert rows[1] == (1, "Existing")


def test_insert_select_zero_rows_column_mismatch(tmp_path: Path) -> None:
    """INSERT...SELECT with zero rows must still validate column count."""
    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Target"
    ws.append(["id", "name"])
    wb.save(file_path)

    # Create Source with 3 columns (mismatched with Target's 2)
    wb = load_workbook(file_path)
    ws2 = wb.create_sheet("Source")
    ws2.append(["a", "b", "c"])
    wb.save(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        # SELECT returns 0 rows but has 3 columns vs Target's 2 — must fail
        with pytest.raises(Exception, match="mismatch"):
            cursor.execute(
                "INSERT INTO Target SELECT a, b, c FROM Source WHERE a = 999"
            )
