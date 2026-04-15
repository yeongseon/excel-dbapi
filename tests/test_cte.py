from pathlib import Path

import pytest
from excel_dbapi.exceptions import DatabaseError
from openpyxl import Workbook

from excel_dbapi.engines.openpyxl.backend import OpenpyxlBackend
from excel_dbapi.executor import SharedExecutor
from excel_dbapi.parser import parse_sql


def _create_cte_workbook(path: Path) -> None:
    workbook = Workbook()

    users = workbook.active
    assert users is not None
    users.title = "users"
    users.append(["id", "name", "age", "dept_id"])
    users.append([1, "Alice", 30, 10])
    users.append([2, "Bob", 19, 20])
    users.append([3, "Cara", 27, 10])
    users.append([4, "Duke", 45, 30])

    emp = workbook.create_sheet("emp")
    emp.append(["id", "dept"])
    for row_id in range(1, 8):
        emp.append([row_id, "Eng"])
    for row_id in range(8, 12):
        emp.append([row_id, "Sales"])

    dept = workbook.create_sheet("dept")
    dept.append(["id", "name"])
    dept.append([10, "Engineering"])
    dept.append([20, "Sales"])
    dept.append([30, "Operations"])

    workbook.save(path)


def test_basic_cte_select(tmp_path: Path) -> None:
    file_path = tmp_path / "cte_basic.xlsx"
    _create_cte_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "WITH t AS (SELECT id, name FROM users WHERE age > 20) "
        "SELECT id, name FROM t ORDER BY id"
    )

    assert parsed["action"] == "SELECT"
    assert len(parsed.get("ctes", [])) == 1

    result = SharedExecutor(engine).execute(parsed)
    assert result.rows == [(1, "Alice"), (3, "Cara"), (4, "Duke")]


def test_cte_with_aggregation(tmp_path: Path) -> None:
    file_path = tmp_path / "cte_agg.xlsx"
    _create_cte_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "WITH summary AS ("
        "SELECT dept, COUNT(*) AS cnt FROM emp GROUP BY dept"
        ") "
        "SELECT dept, cnt FROM summary WHERE cnt > 5"
    )
    result = SharedExecutor(engine).execute(parsed)

    assert result.rows == [("Eng", 7)]


def test_multiple_ctes_join(tmp_path: Path) -> None:
    file_path = tmp_path / "cte_multi.xlsx"
    _create_cte_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "WITH a AS (SELECT id, dept_id FROM users WHERE age > 20), "
        "b AS (SELECT id, name FROM dept) "
        "SELECT a.id, b.name "
        "FROM a JOIN b ON a.dept_id = b.id "
        "ORDER BY a.id"
    )
    result = SharedExecutor(engine).execute(parsed)

    assert result.rows == [(1, "Engineering"), (3, "Engineering"), (4, "Operations")]


def test_cte_referenced_multiple_times(tmp_path: Path) -> None:
    file_path = tmp_path / "cte_reuse.xlsx"
    _create_cte_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "WITH adults AS (SELECT id, dept_id FROM users WHERE age > 20), "
        "adults_ids AS (SELECT id, dept_id FROM adults) "
        "SELECT a.id, i.id "
        "FROM adults AS a JOIN adults_ids AS i ON a.dept_id = i.dept_id "
        "ORDER BY a.id, i.id"
    )
    result = SharedExecutor(engine).execute(parsed)

    assert result.rows == [(1, 1), (1, 3), (3, 1), (3, 3), (4, 4)]


def test_recursive_cte_rejected() -> None:
    with pytest.raises(DatabaseError, match="Recursive CTEs are not supported"):
        parse_sql("WITH t AS (SELECT id FROM t) SELECT * FROM t")
