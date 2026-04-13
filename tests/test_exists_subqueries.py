from pathlib import Path

import pytest
from openpyxl import Workbook

from excel_dbapi.connection import ExcelConnection
from excel_dbapi.engines.openpyxl.backend import OpenpyxlBackend
from excel_dbapi.exceptions import ProgrammingError
from excel_dbapi.executor import SharedExecutor
from excel_dbapi.parser import parse_sql


def _create_workbook(path: Path) -> None:
    workbook = Workbook()

    users = workbook.active
    assert users is not None
    users.title = "users"
    users.append(["id", "name", "dept_id"])
    users.append([1, "Alice", 10])
    users.append([2, "Bob", 10])
    users.append([3, "Carol", 20])
    users.append([4, "Dan", 30])

    admins = workbook.create_sheet("admins")
    admins.append(["id", "role", "dept_id"])
    admins.append([1, "owner", 10])
    admins.append([3, "editor", 20])

    departments = workbook.create_sheet("departments")
    departments.append(["id", "active"])
    departments.append([10, 1])
    departments.append([20, 1])
    departments.append([30, 0])

    workbook.save(path)


def _select(path: Path, query: str) -> list[tuple[object, ...]]:
    with ExcelConnection(str(path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute(query)
        return cursor.fetchall()


def test_exists_basic_uncorrelated(tmp_path: Path) -> None:
    file_path = tmp_path / "exists_basic.xlsx"
    _create_workbook(file_path)

    rows = _select(
        file_path,
        "SELECT id FROM users WHERE EXISTS (SELECT id FROM admins) ORDER BY id",
    )
    assert rows == [(1,), (2,), (3,), (4,)]


def test_parser_exists_and_not_exists_nodes() -> None:
    parsed_exists = parse_sql(
        "SELECT * FROM users u WHERE EXISTS (SELECT id FROM admins a WHERE a.id = u.id)"
    )
    exists_condition = parsed_exists["where"]["conditions"][0]
    assert exists_condition["type"] == "exists"
    assert exists_condition["correlated"] is True
    assert exists_condition["outer_refs"] == ["u.id"]

    parsed_not_exists = parse_sql(
        "SELECT * FROM users u WHERE NOT EXISTS (SELECT id FROM admins a WHERE a.id = u.id)"
    )
    not_condition = parsed_not_exists["where"]["conditions"][0]
    assert not_condition["type"] == "not"
    assert not_condition["operand"]["type"] == "exists"


def test_parser_scalar_subquery_node_metadata() -> None:
    parsed = parse_sql("SELECT id FROM users WHERE id = (SELECT MAX(id) FROM admins)")
    value = parsed["where"]["conditions"][0]["value"]
    assert value["type"] == "subquery"
    assert value["mode"] == "scalar"
    assert value["correlated"] is False
    assert value["outer_refs"] == []


def test_not_exists_correlated(tmp_path: Path) -> None:
    file_path = tmp_path / "not_exists.xlsx"
    _create_workbook(file_path)

    rows = _select(
        file_path,
        "SELECT id FROM users u "
        "WHERE NOT EXISTS (SELECT id FROM admins a WHERE a.id = u.id) "
        "ORDER BY id",
    )
    assert rows == [(2,), (4,)]


def test_exists_correlated(tmp_path: Path) -> None:
    file_path = tmp_path / "exists_correlated.xlsx"
    _create_workbook(file_path)

    rows = _select(
        file_path,
        "SELECT id FROM users u "
        "WHERE EXISTS (SELECT id FROM admins a WHERE a.id = u.id) "
        "ORDER BY id",
    )
    assert rows == [(1,), (3,)]


def test_scalar_subquery_in_select(tmp_path: Path) -> None:
    file_path = tmp_path / "scalar_select.xlsx"
    _create_workbook(file_path)

    rows = _select(
        file_path,
        "SELECT id, (SELECT MAX(id) FROM admins) AS max_admin_id FROM users ORDER BY id",
    )
    assert rows == [(1, 3.0), (2, 3.0), (3, 3.0), (4, 3.0)]


def test_scalar_subquery_in_where_comparison(tmp_path: Path) -> None:
    file_path = tmp_path / "scalar_where.xlsx"
    _create_workbook(file_path)

    rows = _select(
        file_path,
        "SELECT id FROM users WHERE id = (SELECT MAX(id) FROM admins)",
    )
    assert rows == [(3,)]


def test_scalar_subquery_null_result(tmp_path: Path) -> None:
    file_path = tmp_path / "scalar_null.xlsx"
    _create_workbook(file_path)

    rows = _select(
        file_path,
        "SELECT id FROM users WHERE id = (SELECT id FROM admins WHERE role = 'missing')",
    )
    assert rows == []


def test_scalar_subquery_multiple_rows_raises(tmp_path: Path) -> None:
    file_path = tmp_path / "scalar_error.xlsx"
    _create_workbook(file_path)

    with pytest.raises(ProgrammingError, match="Scalar subquery returned more than one row"):
        _select(file_path, "SELECT id FROM users WHERE id = (SELECT id FROM admins)")


def test_exists_false_on_empty_subquery(tmp_path: Path) -> None:
    file_path = tmp_path / "exists_empty.xlsx"
    _create_workbook(file_path)

    rows = _select(
        file_path,
        "SELECT id FROM users WHERE EXISTS (SELECT id FROM admins WHERE role = 'missing')",
    )
    assert rows == []


def test_nested_exists(tmp_path: Path) -> None:
    file_path = tmp_path / "nested_exists.xlsx"
    _create_workbook(file_path)

    rows = _select(
        file_path,
        "SELECT id FROM users u "
        "WHERE EXISTS ("
        "SELECT id FROM admins a "
        "WHERE a.id = u.id "
        "AND EXISTS (SELECT id FROM departments d WHERE d.id = a.dept_id AND d.active = 1)"
        ") ORDER BY id",
    )
    assert rows == [(1,), (3,)]


def test_exists_with_join_in_outer_query(tmp_path: Path) -> None:
    file_path = tmp_path / "exists_join_outer.xlsx"
    _create_workbook(file_path)

    rows = _select(
        file_path,
        "SELECT u.id FROM users u "
        "LEFT JOIN departments d ON u.dept_id = d.id "
        "WHERE EXISTS (SELECT id FROM admins a WHERE a.id = u.id) "
        "ORDER BY u.id",
    )
    assert rows == [(1,), (3,)]


def test_non_correlated_scalar_subquery_is_cached_once(tmp_path: Path) -> None:
    file_path = tmp_path / "scalar_cache.xlsx"
    _create_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    executor = SharedExecutor(engine)
    parsed = parse_sql(
        "SELECT id, (SELECT MAX(id) FROM admins) AS max_admin_id FROM users ORDER BY id"
    )
    result = executor.execute(parsed)

    assert result.rows == [(1, 3.0), (2, 3.0), (3, 3.0), (4, 3.0)]
    assert len(executor._subquery_cache) == 1
