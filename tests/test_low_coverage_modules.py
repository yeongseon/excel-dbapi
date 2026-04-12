from __future__ import annotations

import sys
import types
from pathlib import Path
from typing import Any, cast

import httpx
import pandas as pd
import pytest
from openpyxl import Workbook

from excel_dbapi.connection import ExcelConnection
from excel_dbapi.cursor import ExcelCursor
from excel_dbapi.engines.base import TableData, WorkbookBackend
from excel_dbapi.engines.graph.auth import (
    _has_get_token_with_args,
    normalize_token_provider,
)
from excel_dbapi.engines.graph.client import GraphClient
from excel_dbapi.engines.graph.session import WorkbookSession
from excel_dbapi.engines.openpyxl.backend import OpenpyxlBackend
from excel_dbapi.engines.pandas.backend import PandasBackend
from excel_dbapi.engines.result import ExecutionResult
from excel_dbapi.exceptions import NotSupportedError, OperationalError
from excel_dbapi.executor import SharedExecutor


class MemoryBackend(WorkbookBackend):
    def __init__(self, sheets: dict[str, TableData], readonly: bool = False) -> None:
        super().__init__("memory.xlsx")
        self._sheets = sheets
        self.readonly = readonly

    def load(self) -> None:
        return None

    def save(self) -> None:
        return None

    def snapshot(self) -> dict[str, TableData]:
        return self._sheets

    def restore(self, snapshot: Any) -> None:
        self._sheets = snapshot

    def list_sheets(self) -> list[str]:
        return list(self._sheets.keys())

    def read_sheet(self, sheet_name: str) -> TableData:
        if sheet_name not in self._sheets:
            raise ValueError(f"Sheet '{sheet_name}' not found in Excel")
        return self._sheets[sheet_name]

    def write_sheet(self, sheet_name: str, data: TableData) -> None:
        self._sheets[sheet_name] = data

    def append_row(self, sheet_name: str, row: list[Any]) -> int:
        self._sheets[sheet_name].rows.append(row)
        return len(self._sheets[sheet_name].rows) + 1

    def create_sheet(self, name: str, headers: list[str]) -> None:
        self._sheets[name] = TableData(headers=headers, rows=[])

    def drop_sheet(self, name: str) -> None:
        del self._sheets[name]


def _xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id", "name"])
    ws.append([1, "Alice"])
    wb.save(path)


def test_executor_error_paths_and_utility_paths() -> None:
    backend = MemoryBackend(
        {
            "Users": TableData(headers=["id", "name", "extra"], rows=[[1, "A"]]),
            "Empty": TableData(headers=[], rows=[]),
        }
    )
    executor = SharedExecutor(backend)

    with pytest.raises(ValueError, match="Sheet 'Missing' not found"):
        executor.execute(
            {
                "action": "UPDATE",
                "table": "Missing",
                "set": [{"column": "name", "value": "B"}],
                "where": None,
            }
        )

    update_empty = executor.execute(
        {
            "action": "UPDATE",
            "table": "Empty",
            "set": [{"column": "name", "value": "B"}],
            "where": None,
        }
    )
    assert update_empty.rowcount == 0

    with pytest.raises(ValueError, match="Unknown column"):
        executor.execute(
            {
                "action": "UPDATE",
                "table": "Users",
                "set": [{"column": "missing", "value": "B"}],
                "where": None,
            }
        )

    updated = executor.execute(
        {
            "action": "UPDATE",
            "table": "Users",
            "set": [{"column": "extra", "value": "Z"}],
            "where": {
                "conditions": [{"column": "id", "operator": "=", "value": 1}],
                "conjunctions": [],
            },
        }
    )
    assert updated.rowcount == 1
    assert backend.read_sheet("Users").rows[0] == [1, "A", "Z"]

    with pytest.raises(ValueError, match="Sheet 'Missing' not found"):
        executor.execute({"action": "DELETE", "table": "Missing", "where": None})

    delete_empty = executor.execute(
        {"action": "DELETE", "table": "Empty", "where": None}
    )
    assert delete_empty.rowcount == 0

    with pytest.raises(ValueError, match="Sheet 'Missing' not found"):
        executor.execute(
            {
                "action": "INSERT",
                "table": "Missing",
                "columns": None,
                "values": [1],
            }
        )

    with pytest.raises(ValueError, match="without headers"):
        executor.execute(
            {
                "action": "INSERT",
                "table": "Empty",
                "columns": None,
                "values": [1],
            }
        )

    with pytest.raises(ValueError, match="header count"):
        executor.execute(
            {
                "action": "INSERT",
                "table": "Users",
                "columns": None,
                "values": [1],
            }
        )

    with pytest.raises(ValueError, match="Unknown column"):
        executor.execute(
            {
                "action": "INSERT",
                "table": "Users",
                "columns": ["missing"],
                "values": [1],
            }
        )

    assert (
        executor._matches_where(
            {"id": 1}, {"column": "id", "operator": "=", "value": 1}
        )
        is True
    )
    assert (
        executor._evaluate_condition(
            {"id": None}, {"column": "id", "operator": "BETWEEN", "value": (1, 2)}
        )
        is False
    )
    assert (
        executor._evaluate_condition(
            {"id": 5}, {"column": "id", "operator": "BETWEEN", "value": (None, 9)}
        )
        is False
    )
    assert (
        executor._evaluate_condition(
            {"id": 1}, {"column": "id", "operator": "<", "value": 2}
        )
        is True
    )
    with pytest.raises(NotImplementedError, match="Unsupported operator"):
        executor._evaluate_condition(
            {"id": 1}, {"column": "id", "operator": "~~~", "value": 2}
        )
    assert executor._sort_key(None) == (1, "")
    assert executor._to_number(True) is None
    assert executor._to_number({"x": 1}) is None

    with pytest.raises(ValueError, match="Unsupported action"):
        executor.execute({"action": "BOOM", "table": "Users"})


def test_cursor_paths_for_executemany_and_fetch() -> None:
    class FakeEngine:
        def __init__(self) -> None:
            self.restored: Any = None
            self.saved = False

        def snapshot(self) -> str:
            return "snap"

        def restore(self, snapshot: Any) -> None:
            self.restored = snapshot

        def save(self) -> None:
            self.saved = True

    class FakeConnection:
        def __init__(self) -> None:
            self.closed = False
            self.autocommit = False
            self.engine = FakeEngine()
            self._snapshot = None

        def execute(
            self, query: str, params: tuple[Any, ...] | None = None
        ) -> ExecutionResult:
            raise NotImplementedError("not supported")

    cursor = ExcelCursor(FakeConnection())
    with pytest.raises(NotSupportedError):
        cursor.executemany("SELECT 1", [(1,)])
    assert cursor.fetchone() is None
    assert cursor.fetchmany(0) == []


def test_connection_str_and_repr(tmp_path: Path) -> None:
    file_path = tmp_path / "repr.xlsx"
    _xlsx(file_path)
    conn = ExcelConnection(str(file_path), engine="openpyxl")
    assert "ExcelConnection" in str(conn)
    assert repr(conn) == str(conn)
    conn.close()


def test_base_abstract_method_bodies_are_executable() -> None:
    backend = MemoryBackend({"T": TableData(headers=["id"], rows=[[1]])})
    WorkbookBackend.load(backend)
    WorkbookBackend.save(backend)
    WorkbookBackend.snapshot(backend)
    WorkbookBackend.restore(backend, None)
    WorkbookBackend.list_sheets(backend)
    WorkbookBackend.read_sheet(backend, "T")
    WorkbookBackend.write_sheet(backend, "T", TableData(headers=["id"], rows=[]))
    WorkbookBackend.append_row(backend, "T", [2])
    WorkbookBackend.create_sheet(backend, "U", ["id"])
    WorkbookBackend.drop_sheet(backend, "U")


def test_openpyxl_backend_error_paths_and_execute_wrappers(tmp_path: Path) -> None:
    file_path = tmp_path / "openpyxl.xlsx"
    _xlsx(file_path)
    backend = OpenpyxlBackend(str(file_path), create=False)

    with pytest.raises(ValueError, match="not found"):
        backend.read_sheet("Missing")
    with pytest.raises(ValueError, match="not found"):
        backend.write_sheet("Missing", TableData(headers=["a"], rows=[]))
    with pytest.raises(ValueError, match="not found"):
        backend.append_row("Missing", [1])
    with pytest.raises(ValueError, match="already exists"):
        backend.create_sheet("Sheet1", ["a"])
    with pytest.raises(ValueError, match="not found"):
        backend.drop_sheet("Missing")

    backend.workbook = None
    with pytest.raises(ValueError, match="not loaded"):
        backend.snapshot()
    with pytest.raises(ValueError, match="not loaded"):
        backend.get_workbook()
    with pytest.raises(ValueError, match="not loaded"):
        backend.create_sheet("X", ["a"])
    with pytest.raises(ValueError, match="not loaded"):
        backend.drop_sheet("Sheet1")

    backend2 = OpenpyxlBackend(str(file_path), create=False)
    result1 = backend2.execute("SELECT * FROM Sheet1")
    result2 = backend2.execute_with_params("SELECT * FROM Sheet1 WHERE id = ?", (1,))
    assert result1.rowcount >= 1
    assert result2.rowcount == 1


def test_pandas_backend_error_paths_and_execute_wrappers(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    file_path = tmp_path / "pandas.xlsx"
    pd.DataFrame([{"id": 1, "name": "Alice"}]).to_excel(
        file_path, index=False, sheet_name="Sheet1"
    )
    backend = PandasBackend(str(file_path), create=False)

    with pytest.raises(ValueError, match="not found"):
        backend.read_sheet("Missing")
    with pytest.raises(ValueError, match="not found"):
        backend.write_sheet("Missing", TableData(headers=["a"], rows=[]))
    with pytest.raises(ValueError, match="not found"):
        backend.append_row("Missing", [1])
    with pytest.raises(ValueError, match="already exists"):
        backend.create_sheet("Sheet1", ["a"])
    with pytest.raises(ValueError, match="not found"):
        backend.drop_sheet("Missing")

    result1 = backend.execute("SELECT * FROM Sheet1")
    result2 = backend.execute_with_params("SELECT * FROM Sheet1 WHERE id = ?", (1,))
    assert result1.rowcount >= 1
    assert result2.rowcount == 1

    created_temp: dict[str, str] = {}

    def fail_replace(src: str, dst: str) -> None:
        del dst
        created_temp["path"] = src
        raise OSError("forced replace failure")

    monkeypatch.setattr("excel_dbapi.engines.pandas.backend.os.replace", fail_replace)
    with pytest.raises(OSError, match="forced replace failure"):
        backend.save()
    temp_path = Path(created_temp["path"])
    assert not temp_path.exists()


def test_graph_client_transport_error_and_session_property() -> None:
    class BrokenTransport(httpx.BaseTransport):
        def handle_request(self, request: httpx.Request) -> httpx.Response:
            raise httpx.ConnectError("boom", request=request)

    client = GraphClient(normalize_token_provider("tok"), transport=BrokenTransport())
    assert client.session_id is None
    client.session_id = "sid"
    assert client.session_id == "sid"
    with pytest.raises(OperationalError, match="Graph API request failed"):
        client.get("/x")
    client.close()


def test_graph_session_reopen_and_close_swallow_errors() -> None:
    class FakeClient:
        def __init__(self) -> None:
            self.session_id: str | None = "sid-1"
            self.calls = 0

        def post(self, path: str, json: dict[str, Any]) -> Any:
            self.calls += 1
            if path.endswith("closeSession"):
                raise RuntimeError("close fail")
            return types.SimpleNamespace(json=lambda: {"id": "sid-2"})

    class FakeLocator:
        item_path = "/drives/d/items/i"

    client = FakeClient()
    session = WorkbookSession(
        cast(Any, client), cast(Any, FakeLocator()), persist_changes=False
    )
    session._open = True
    session.reopen()
    assert session.is_open is True
    session.close()
    assert session.is_open is False


def test_graph_auth_additional_paths(monkeypatch: pytest.MonkeyPatch) -> None:
    class BadSignatureObj:
        def get_token(self) -> str:
            return "x"

    def explode_signature(obj: Any) -> Any:
        raise TypeError("no signature")

    monkeypatch.setattr("inspect.signature", explode_signature)
    assert _has_get_token_with_args(BadSignatureObj()) is False

    module_azure = types.ModuleType("azure")
    module_identity = types.ModuleType("azure.identity")

    class DefaultAzureCredential:
        def get_token(self, scope: str) -> Any:
            return types.SimpleNamespace(token=f"azure:{scope}")

    setattr(module_identity, "DefaultAzureCredential", DefaultAzureCredential)
    setattr(module_azure, "identity", module_identity)
    monkeypatch.setitem(sys.modules, "azure", module_azure)
    monkeypatch.setitem(sys.modules, "azure.identity", module_identity)
    provider = normalize_token_provider(None)
    assert provider.get_token().startswith("azure:")
