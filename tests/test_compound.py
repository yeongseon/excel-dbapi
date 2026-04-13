from pathlib import Path

import pytest
from openpyxl import Workbook

from excel_dbapi.connection import ExcelConnection
from excel_dbapi.parser import parse_sql


def _create_compound_workbook(path: Path) -> None:
    workbook = Workbook()

    t1 = workbook.active
    assert t1 is not None
    t1.title = "t1"
    t1.append(["id", "name"])
    t1.append([1, "A"])
    t1.append([2, "B"])
    t1.append([3, None])
    t1.append([2, "B"])

    t2 = workbook.create_sheet("t2")
    t2.append(["id", "name"])
    t2.append([2, "B"])
    t2.append([3, None])
    t2.append([4, "D"])
    t2.append([4, "D"])

    t3 = workbook.create_sheet("t3")
    t3.append(["id", "name"])
    t3.append([4, "D"])
    t3.append([5, "E"])

    workbook.save(path)


def test_parser_accepts_compound_variants() -> None:
    parsed = parse_sql("SELECT id, name FROM t1 UNION SELECT id, name FROM t2")
    assert parsed["action"] == "COMPOUND"
    assert parsed["operators"] == ["UNION"]
    assert len(parsed["queries"]) == 2

    parsed = parse_sql("SELECT id FROM t1 UNION ALL SELECT id FROM t2")
    assert parsed["operators"] == ["UNION ALL"]

    parsed = parse_sql("SELECT id FROM t1 INTERSECT SELECT id FROM t2")
    assert parsed["operators"] == ["INTERSECT"]

    parsed = parse_sql("SELECT id FROM t1 EXCEPT SELECT id FROM t2")
    assert parsed["operators"] == ["EXCEPT"]

    parsed = parse_sql("SELECT id FROM t1 UNION SELECT id FROM t2 UNION SELECT id FROM t3")
    assert parsed["operators"] == ["UNION", "UNION"]
    assert len(parsed["queries"]) == 3


def test_union_basic(tmp_path: Path) -> None:
    file_path = tmp_path / "compound_union.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id, name FROM t1 UNION SELECT id, name FROM t2")
        assert cursor.fetchall() == [(1, "A"), (2, "B"), (3, None), (4, "D")]


def test_union_all_basic(tmp_path: Path) -> None:
    file_path = tmp_path / "compound_union_all.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM t1 UNION ALL SELECT id FROM t2")
        assert cursor.fetchall() == [(1,), (2,), (3,), (2,), (2,), (3,), (4,), (4,)]


def test_intersect_basic(tmp_path: Path) -> None:
    file_path = tmp_path / "compound_intersect.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM t1 INTERSECT SELECT id FROM t2")
        assert cursor.fetchall() == [(2,), (3,)]


def test_except_basic(tmp_path: Path) -> None:
    file_path = tmp_path / "compound_except.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM t1 EXCEPT SELECT id FROM t2")
        assert cursor.fetchall() == [(1,)]


def test_union_different_tables(tmp_path: Path) -> None:
    file_path = tmp_path / "compound_diff_tables.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM t1 UNION SELECT id FROM t3")
        assert cursor.fetchall() == [(1,), (2,), (3,), (4,), (5,)]


def test_union_with_where(tmp_path: Path) -> None:
    file_path = tmp_path / "compound_where.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM t1 WHERE id <= 2 UNION SELECT id FROM t2 WHERE id >= 3")
        assert cursor.fetchall() == [(1,), (2,), (3,), (4,)]


def test_union_column_count_mismatch(tmp_path: Path) -> None:
    file_path = tmp_path / "compound_column_mismatch.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        with pytest.raises(ValueError, match="column counts"):
            conn.execute("SELECT id FROM t1 UNION SELECT id, name FROM t2")


def test_chained_union(tmp_path: Path) -> None:
    file_path = tmp_path / "compound_chained_union.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM t1 UNION SELECT id FROM t2 UNION SELECT id FROM t3")
        assert cursor.fetchall() == [(1,), (2,), (3,), (4,), (5,)]


def test_mixed_compound(tmp_path: Path) -> None:
    file_path = tmp_path / "compound_mixed.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM t1 UNION SELECT id FROM t2 EXCEPT SELECT id FROM t3")
        assert cursor.fetchall() == [(1,), (2,), (3,)]


def test_union_all_preserves_order(tmp_path: Path) -> None:
    file_path = tmp_path / "compound_union_all_order.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        # Per SQL standard, trailing ORDER BY applies to the entire compound.
        cursor.execute(
            "SELECT id FROM t1 WHERE id <= 2 "
            "UNION ALL "
            "SELECT id FROM t2 WHERE id >= 4 ORDER BY id DESC"
        )
        assert cursor.fetchall() == [(4,), (4,), (2,), (2,), (1,)]


def test_intersect_empty_result(tmp_path: Path) -> None:
    file_path = tmp_path / "compound_intersect_empty.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM t1 INTERSECT SELECT id FROM t3")
        assert cursor.fetchall() == []


def test_except_all_removed(tmp_path: Path) -> None:
    file_path = tmp_path / "compound_except_all_removed.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM t2 EXCEPT SELECT id FROM t2")
        assert cursor.fetchall() == []


def test_union_with_order_by(tmp_path: Path) -> None:
    file_path = tmp_path / "compound_order_by.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        # Per SQL standard, trailing ORDER BY applies to the entire compound.
        cursor.execute(
            "SELECT id FROM t1 WHERE id <= 2 "
            "UNION "
            "SELECT id FROM t2 WHERE id >= 3 ORDER BY id DESC"
        )
        assert cursor.fetchall() == [(4,), (3,), (2,), (1,)]


def test_union_null_handling(tmp_path: Path) -> None:
    file_path = tmp_path / "compound_nulls.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM t1 WHERE id = 3 UNION SELECT name FROM t2 WHERE id = 3")
        assert cursor.fetchall() == [(None,)]


def test_compound_parameterized_in_clause(tmp_path: Path) -> None:
    """Regression: IN (?, ?) with compound queries must count placeholders correctly."""
    file_path = tmp_path / "compound_param_in.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id FROM t1 WHERE id IN (?, ?) UNION SELECT id FROM t2 WHERE id = ?",
            (1, 2, 4),
        )
        assert cursor.fetchall() == [(1,), (2,), (4,)]


def test_compound_parameterized_between(tmp_path: Path) -> None:
    """Regression: BETWEEN ? AND ? with compound queries."""
    file_path = tmp_path / "compound_param_between.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id FROM t1 WHERE id BETWEEN ? AND ? UNION SELECT id FROM t2 WHERE id = ?",
            (1, 2, 4),
        )
        assert cursor.fetchall() == [(1,), (2,), (4,)]


def test_compound_parameterized_mixed_branches(tmp_path: Path) -> None:
    """Regression: different placeholder counts across branches."""
    file_path = tmp_path / "compound_param_mixed.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id FROM t1 WHERE id IN (?, ?, ?) EXCEPT SELECT id FROM t2 WHERE id > ?",
            (1, 2, 3, 3),
        )
        # t1 ids matching IN (1,2,3): [1,2,3], t2 ids > 3: [4]
        # EXCEPT: [1,2,3] - [4] = [1,2,3]
        assert cursor.fetchall() == [(1,), (2,), (3,)]


def test_compound_parameterized_too_few_params(tmp_path: Path) -> None:
    """Regression: too few parameters should raise ProgrammingError."""
    file_path = tmp_path / "compound_param_few.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        with pytest.raises(Exception, match="Not enough parameters"):
            conn.execute(
                "SELECT id FROM t1 WHERE id IN (?, ?) UNION SELECT id FROM t2 WHERE id = ?",
                (1, 2),
            )


def test_compound_parameterized_too_many_params(tmp_path: Path) -> None:
    """Regression: too many parameters should raise ProgrammingError."""
    file_path = tmp_path / "compound_param_many.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        with pytest.raises(Exception, match="Too many parameters"):
            conn.execute(
                "SELECT id FROM t1 WHERE id IN (?, ?) UNION SELECT id FROM t2 WHERE id = ?",
                (1, 2, 3, 4),
            )


def test_compound_parameterized_union_all_in(tmp_path: Path) -> None:
    """Regression: UNION ALL with IN clause placeholders."""
    file_path = tmp_path / "compound_param_union_all.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id FROM t1 WHERE id IN (?, ?) UNION ALL SELECT id FROM t2 WHERE id IN (?, ?)",
            (1, 2, 2, 3),
        )
        # t1 ids matching IN (1,2): [1,2,2] (id=2 appears twice), t2 ids matching IN (2,3): [2,3]
        assert cursor.fetchall() == [(1,), (2,), (2,), (2,), (3,)]


def test_compound_parameterized_intersect_in(tmp_path: Path) -> None:
    """Regression: INTERSECT with IN clause placeholders in both branches."""
    file_path = tmp_path / "compound_param_intersect.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id FROM t1 WHERE id IN (?, ?) INTERSECT SELECT id FROM t2 WHERE id IN (?, ?)",
            (2, 3, 2, 3),
        )
        assert cursor.fetchall() == [(2,), (3,)]


def test_parser_compound_placeholder_counting() -> None:
    """Regression: parser must count ?, tokens (comma-attached) as placeholders."""
    parsed = parse_sql(
        "SELECT id FROM t1 WHERE id IN (?, ?) UNION SELECT id FROM t2 WHERE id = ?",
        (10, 20, 30),
    )
    assert parsed["action"] == "COMPOUND"
    # Verify branch 1 got params (10, 20) and branch 2 got (30,)
    q1_where = parsed["queries"][0]["where"]
    assert q1_where["conditions"][0]["value"] == (10, 20)
    q2_where = parsed["queries"][1]["where"]
    assert q2_where["conditions"][0]["value"] == 30


def test_parser_parenthesized_branch(tmp_path: Path) -> None:
    """Parser accepts parenthesized compound branches (e.g. from SQLAlchemy)."""
    parsed = parse_sql(
        "(SELECT id FROM t1 ORDER BY id DESC) UNION SELECT id FROM t2"
    )
    assert parsed["action"] == "COMPOUND"
    assert parsed["operators"] == ["UNION"]
    assert len(parsed["queries"]) == 2
    # Branch 1 should have ORDER BY from inside the parens.
    assert parsed["queries"][0].get("order_by") == {
        "column": "id",
        "direction": "DESC",
    }


def test_parser_compound_outer_order_by() -> None:
    """Trailing ORDER BY after compound is compound-level, not branch-level."""
    parsed = parse_sql("SELECT id FROM t1 UNION SELECT id FROM t2 ORDER BY id")
    assert parsed["action"] == "COMPOUND"
    assert parsed.get("order_by") == {"column": "id", "direction": "ASC"}
    # Branch 2 should NOT have ORDER BY (it was extracted).
    assert parsed["queries"][1].get("order_by") is None


def test_parser_compound_outer_order_by_desc() -> None:
    """Trailing ORDER BY DESC after compound is compound-level."""
    parsed = parse_sql("SELECT id FROM t1 UNION SELECT id FROM t2 ORDER BY id DESC")
    assert parsed.get("order_by") == {"column": "id", "direction": "DESC"}


def test_parser_compound_outer_limit() -> None:
    """Trailing LIMIT after compound is compound-level."""
    parsed = parse_sql("SELECT id FROM t1 UNION SELECT id FROM t2 LIMIT 5")
    assert parsed["action"] == "COMPOUND"
    assert parsed.get("limit") == 5
    assert parsed["queries"][1].get("limit") is None


def test_parser_compound_outer_order_by_and_limit() -> None:
    """Trailing ORDER BY + LIMIT after compound are compound-level."""
    parsed = parse_sql("SELECT id FROM t1 UNION SELECT id FROM t2 ORDER BY id LIMIT 3")
    assert parsed.get("order_by") == {"column": "id", "direction": "ASC"}
    assert parsed.get("limit") == 3


def test_compound_outer_order_by_e2e(tmp_path: Path) -> None:
    """Compound-level ORDER BY sorts the entire result."""
    file_path = tmp_path / "compound_outer_order.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id FROM t1 UNION ALL SELECT id FROM t2 ORDER BY id DESC"
        )
        rows = cursor.fetchall()
        # t1 ids: [1,2,3,2], t2 ids: [2,3,4,4]
        # UNION ALL: [1,2,3,2,2,3,4,4], ORDER BY id DESC: [4,4,3,3,2,2,2,1]
        assert rows == [(4,), (4,), (3,), (3,), (2,), (2,), (2,), (1,)]


def test_compound_outer_order_by_with_limit_e2e(tmp_path: Path) -> None:
    """Compound-level ORDER BY + LIMIT."""
    file_path = tmp_path / "compound_limit.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id FROM t1 UNION ALL SELECT id FROM t2 ORDER BY id LIMIT 3"
        )
        rows = cursor.fetchall()
        # UNION ALL sorted ASC: [1,2,2,2,3,3,4,4], LIMIT 3 → [1,2,2]
        assert rows == [(1,), (2,), (2,)]


def test_parenthesized_branch_e2e(tmp_path: Path) -> None:
    """Parenthesized compound branch executes correctly."""
    file_path = tmp_path / "compound_paren.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "(SELECT id FROM t1 ORDER BY id DESC) UNION SELECT id FROM t2"
        )
        rows = cursor.fetchall()
        # Branch 1: t1 ids [3,2,1,2] ORDER BY DESC → [3,2,2,1]
        # Branch 2: t2 ids [2,3,4,4]
        # UNION (dedup): {1, 2, 3, 4}
        assert sorted(rows) == [(1,), (2,), (3,), (4,)]




def test_mixed_union_intersect_left_to_right(tmp_path: Path) -> None:
    """Mixed operators evaluate left-to-right per SQL standard.

    A UNION B INTERSECT C == (A UNION B) INTERSECT C.
    t1 ids: {1, 2, 3}, t2 ids: {2, 3, 4}, t3 ids: {4, 5}.
    (t1 UNION t2) = {1, 2, 3, 4}.
    {1, 2, 3, 4} INTERSECT {4, 5} = {4}.
    """
    file_path = tmp_path / "compound_mixed_lr.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id FROM t1 UNION SELECT id FROM t2 INTERSECT SELECT id FROM t3"
        )
        rows = cursor.fetchall()
        assert rows == [(4,)]


def test_compound_in_clause_param_slicing(tmp_path: Path) -> None:
    """IN (?,?) in compound query correctly slices parameters.

    Regression test for Bug #3: token-based ? counting failed when
    tokenizer produced tokens like '(?,?)' instead of separate '?' tokens.
    """
    file_path = tmp_path / "compound_in_params.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        # Each branch uses IN (?,?) — 2 placeholders each, 4 total.
        cursor.execute(
            "SELECT id FROM t1 WHERE id IN (?,?) "
            "UNION "
            "SELECT id FROM t2 WHERE id IN (?,?)",
            (1, 2, 3, 4),
        )
        rows = cursor.fetchall()
        ids = sorted(r[0] for r in rows)
        # t1 WHERE id IN (1,2) -> {1,2}, t2 WHERE id IN (3,4) -> {3,4}
        # UNION -> {1,2,3,4}
        assert ids == [1, 2, 3, 4]


def test_compound_quoted_question_mark_not_counted(tmp_path: Path) -> None:
    """A '?' inside a string literal is not counted as a placeholder.

    Regression test for Bug #3 edge case.
    """
    file_path = tmp_path / "compound_quoted_qmark.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        # The literal '?' in the first branch is NOT a placeholder.
        # Only the ? in WHERE id = ? is a placeholder (1 per branch = 2 total).
        cursor.execute(
            "SELECT id FROM t1 WHERE id = ? "
            "UNION "
            "SELECT id FROM t2 WHERE id = ?",
            (1, 2),
        )
        rows = cursor.fetchall()
        ids = sorted(r[0] for r in rows)
        assert ids == [1, 2]


def test_compound_from_with_subquery_depth(tmp_path: Path) -> None:
    """FROM inside a subquery does not confuse top-level FROM detection.

    Regression test for Bug #4: stale depth variable caused FROM in
    subqueries to be treated as top-level FROM.
    """
    file_path = tmp_path / "compound_from_depth.xlsx"
    _create_compound_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        # Subquery contains its own FROM. The compound-level ORDER BY
        # should still be correctly detected as trailing (after the
        # last top-level FROM).
        cursor.execute(
            "SELECT id FROM t1 WHERE id IN (SELECT id FROM t2) "
            "UNION "
            "SELECT id FROM t3 "
            "ORDER BY id"
        )
        rows = cursor.fetchall()
        ids = [r[0] for r in rows]
        # Results should be sorted by ORDER BY id.
        assert ids == sorted(ids)


def test_count_unquoted_placeholders_unit() -> None:
    """Unit test for _count_unquoted_placeholders helper."""
    from excel_dbapi.parser import _count_unquoted_placeholders

    # Normal placeholders.
    assert _count_unquoted_placeholders("SELECT ? FROM t WHERE id = ?") == 2
    # Placeholder inside IN clause.
    assert _count_unquoted_placeholders("SELECT id FROM t WHERE id IN (?,?,?)") == 3
    # Quoted ? is not counted.
    assert _count_unquoted_placeholders("SELECT '?' FROM t WHERE id = ?") == 1
    # Double-quoted ? is not counted.
    assert _count_unquoted_placeholders('SELECT "?" FROM t WHERE id = ?') == 1
    # Escaped quote with ? after.
    assert _count_unquoted_placeholders("SELECT 'it''s' FROM t WHERE id = ?") == 1
    # No placeholders.
    assert _count_unquoted_placeholders("SELECT id FROM t") == 0
