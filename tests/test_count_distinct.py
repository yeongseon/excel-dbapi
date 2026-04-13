from pathlib import Path

import pytest
from openpyxl import Workbook

from excel_dbapi.engines.openpyxl.backend import OpenpyxlBackend
from excel_dbapi.executor import SharedExecutor
from excel_dbapi.parser import parse_sql


def _create_users_dept_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "users"
    sheet.append(["id", "name", "dept"])
    sheet.append([1, "A", "eng"])
    sheet.append([2, "B", "eng"])
    sheet.append([3, "C", "sales"])
    sheet.append([4, "D", "eng"])
    sheet.append([5, None, "sales"])
    sheet.append([6, "E", None])
    workbook.save(path)


def _create_team_dept_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "data"
    sheet.append(["id", "team", "dept"])
    sheet.append([1, "team1", "eng"])
    sheet.append([2, "team1", "eng"])
    sheet.append([3, "team1", "sales"])
    sheet.append([4, "team2", "ops"])
    sheet.append([5, "team2", "ops"])
    workbook.save(path)


def _create_all_none_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "all_none"
    sheet.append(["id", "dept"])
    sheet.append([1, None])
    sheet.append([2, None])
    sheet.append([3, None])
    workbook.save(path)


def test_parse_count_distinct() -> None:
    parsed = parse_sql("SELECT COUNT(DISTINCT dept) FROM users")
    assert parsed["columns"] == [
        {"type": "aggregate", "func": "COUNT", "arg": "dept", "distinct": True}
    ]


def test_parse_count_distinct_qualified_supported() -> None:
    parsed = parse_sql("SELECT COUNT(DISTINCT t1.dept) FROM users AS t1")
    assert parsed["columns"] == [
        {"type": "aggregate", "func": "COUNT", "arg": "t1.dept", "distinct": True}
    ]


def test_parse_count_without_distinct() -> None:
    parsed = parse_sql("SELECT COUNT(dept) FROM users")
    aggregate = parsed["columns"][0]
    assert aggregate == {"type": "aggregate", "func": "COUNT", "arg": "dept"}
    assert "distinct" not in aggregate


def test_parse_rejects_sum_distinct() -> None:
    with pytest.raises(ValueError):
        parse_sql("SELECT SUM(DISTINCT age) FROM users")


def test_parse_rejects_avg_distinct() -> None:
    with pytest.raises(ValueError):
        parse_sql("SELECT AVG(DISTINCT age) FROM users")


def test_parse_rejects_count_distinct_star() -> None:
    with pytest.raises(ValueError):
        parse_sql("SELECT COUNT(DISTINCT *) FROM users")


def test_parse_count_distinct_with_alias() -> None:
    parsed = parse_sql("SELECT COUNT(DISTINCT dept) AS unique_depts FROM users")
    assert parsed["columns"] == [
        {
            "type": "alias",
            "alias": "unique_depts",
            "expression": {
                "type": "aggregate",
                "func": "COUNT",
                "arg": "dept",
                "distinct": True,
            },
        }
    ]


def test_executor_count_distinct(tmp_path: Path) -> None:
    file_path = tmp_path / "count_distinct.xlsx"
    _create_users_dept_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql("SELECT COUNT(DISTINCT dept) FROM users")
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(2,)]


def test_executor_count_distinct_with_group_by(tmp_path: Path) -> None:
    file_path = tmp_path / "count_distinct_group_by.xlsx"
    _create_team_dept_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT team, COUNT(DISTINCT dept) FROM data GROUP BY team ORDER BY team ASC"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [("team1", 2), ("team2", 1)]


def test_executor_count_distinct_all_none(tmp_path: Path) -> None:
    file_path = tmp_path / "count_distinct_all_none.xlsx"
    _create_all_none_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql("SELECT COUNT(DISTINCT dept) FROM all_none")
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(0,)]


def test_executor_count_distinct_with_alias_description(tmp_path: Path) -> None:
    file_path = tmp_path / "count_distinct_alias.xlsx"
    _create_users_dept_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql("SELECT COUNT(DISTINCT dept) AS unique_depts FROM users")
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(2,)]
    assert results.description[0][0] == "unique_depts"


def test_executor_count_distinct_order_by(tmp_path: Path) -> None:
    file_path = tmp_path / "count_distinct_order_by.xlsx"
    _create_team_dept_workbook(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql(
        "SELECT team, COUNT(DISTINCT dept) FROM data GROUP BY team "
        "ORDER BY COUNT(DISTINCT dept) DESC, team ASC"
    )
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [("team1", 2), ("team2", 1)]


def test_executor_count_distinct_empty_table(tmp_path: Path) -> None:
    """COUNT(DISTINCT col) on an empty table returns 0."""
    file_path = tmp_path / "empty.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "empty"
    sheet.append(["id", "dept"])  # headers only, no data rows
    workbook.save(file_path)

    engine = OpenpyxlBackend(str(file_path))
    parsed = parse_sql("SELECT COUNT(DISTINCT dept) FROM empty")
    results = SharedExecutor(engine).execute(parsed)

    assert results.rows == [(0,)]


def test_aggregate_label_round_trip() -> None:
    """Verify _aggregate_label and _aggregate_spec_from_label round-trip for DISTINCT."""
    from excel_dbapi.executor import SharedExecutor

    executor = SharedExecutor.__new__(SharedExecutor)

    # Test DISTINCT label
    distinct_agg: dict[str, object] = {
        "type": "aggregate",
        "func": "COUNT",
        "arg": "dept",
        "distinct": True,
    }
    label = SharedExecutor._aggregate_label(executor, distinct_agg)
    assert label == "COUNT(DISTINCT dept)"
    spec = SharedExecutor._aggregate_spec_from_label(executor, label)
    assert spec is not None
    func, arg, is_distinct, filter_condition = spec
    assert func == "COUNT"
    assert arg == "dept"
    assert is_distinct is True
    assert filter_condition is None

    # Test non-DISTINCT label
    regular_agg: dict[str, object] = {
        "type": "aggregate",
        "func": "COUNT",
        "arg": "dept",
    }
    label2 = SharedExecutor._aggregate_label(executor, regular_agg)
    assert label2 == "COUNT(dept)"
    spec2 = SharedExecutor._aggregate_spec_from_label(executor, label2)
    assert spec2 is not None
    func2, arg2, is_distinct2, filter_condition2 = spec2
    assert func2 == "COUNT"
    assert arg2 == "dept"
    assert is_distinct2 is False
    assert filter_condition2 is None
