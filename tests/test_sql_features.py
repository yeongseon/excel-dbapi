from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from excel_dbapi.connection import ExcelConnection
from excel_dbapi.engine.parser import parse_sql


def _create_feature_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["id", "name", "score"])
    ws.append([1, "Alice", 10])
    ws.append([2, "Bob", 20])
    ws.append([3, "Cara", None])
    ws.append([4, "Dane", 15])
    wb.save(path)


def test_param_binding_count_mismatch() -> None:
    with pytest.raises(ValueError):
        parse_sql("INSERT INTO Sheet1 (id, name) VALUES (?, ?)")

    with pytest.raises(ValueError):
        parse_sql("INSERT INTO Sheet1 (id, name) VALUES (1, 2)", (1, 2, 3))


def test_limit_requires_integer() -> None:
    with pytest.raises(ValueError):
        parse_sql("SELECT * FROM Sheet1 LIMIT 'A'")


def test_openpyxl_where_and_or_and_order_limit(tmp_path: Path) -> None:
    file_path = tmp_path / "sample.xlsx"
    _create_feature_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, name FROM Sheet1 WHERE id >= 2 AND name != 'Cara' ORDER BY id DESC LIMIT 2"
        )
        results = cursor.fetchall()
        assert results == [(4, "Dane"), (2, "Bob")]

        cursor.execute(
            "SELECT id FROM Sheet1 WHERE id = 1 OR id = 3 ORDER BY id ASC"
        )
        assert cursor.fetchall() == [(1,), (3,)]


def test_pandas_where_and_or_and_order_limit(tmp_path: Path) -> None:
    file_path = tmp_path / "sample.xlsx"
    df = pd.DataFrame(
        [
            {"id": 1, "name": "Alice", "score": 10},
            {"id": 2, "name": "Bob", "score": 20},
            {"id": 3, "name": "Cara", "score": None},
            {"id": 4, "name": "Dane", "score": 15},
        ]
    )
    df.to_excel(file_path, index=False, sheet_name="Sheet1")

    with ExcelConnection(str(file_path), engine="pandas") as conn:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT id, name FROM Sheet1 WHERE id >= 2 AND name != 'Cara' ORDER BY id DESC LIMIT 2"
        )
        results = cursor.fetchall()
        assert results == [(4, "Dane"), (2, "Bob")]

        cursor.execute(
            "SELECT id FROM Sheet1 WHERE id = 1 OR id = 3 ORDER BY id ASC"
        )
        assert cursor.fetchall() == [(1,), (3,)]


def test_openpyxl_update_delete_rowcount_and_commit(tmp_path: Path) -> None:
    file_path = tmp_path / "sample.xlsx"
    _create_feature_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=False) as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE Sheet1 SET score = 99 WHERE id >= 2 AND id <= 3")
        assert cursor.rowcount == 2
        cursor.execute("DELETE FROM Sheet1 WHERE id = 1")
        assert cursor.rowcount == 1
        conn.commit()

    wb = load_workbook(file_path, data_only=True)
    rows = list(wb["Sheet1"].iter_rows(values_only=True))
    assert rows[1][0] == 2
    assert rows[1][2] == 99


def test_pandas_update_delete_rowcount(tmp_path: Path) -> None:
    file_path = tmp_path / "sample.xlsx"
    df = pd.DataFrame(
        [
            {"id": 1, "name": "Alice", "score": 10},
            {"id": 2, "name": "Bob", "score": 20},
            {"id": 3, "name": "Cara", "score": 30},
        ]
    )
    df.to_excel(file_path, index=False, sheet_name="Sheet1")

    with ExcelConnection(str(file_path), engine="pandas", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute("UPDATE Sheet1 SET score = 0 WHERE score >= 20")
        assert cursor.rowcount == 2
        cursor.execute("DELETE FROM Sheet1")
        assert cursor.rowcount == 3

    data = pd.read_excel(file_path, sheet_name=None)
    assert len(data["Sheet1"]) == 0
