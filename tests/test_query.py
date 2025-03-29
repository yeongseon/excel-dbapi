from excel_dbapi.connection import ExcelConnection
from excel_dbapi.table import ExcelTable


def test_query_all(tmp_path):
    from openpyxl import Workbook

    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Age", "City"])
    ws.append(["Alice", 25, "Seoul"])
    ws.append(["Bob", 30, "Tokyo"])
    wb.save(file_path)

    with ExcelConnection(file_path) as conn:
        with ExcelTable(conn, "Sheet") as table:
            result = table.query()
            assert result == [
                {"Name": "Alice", "Age": 25, "City": "Seoul"},
                {"Name": "Bob", "Age": 30, "City": "Tokyo"},
            ]


def test_query_columns(tmp_path):
    from openpyxl import Workbook

    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Age", "City"])
    ws.append(["Alice", 25, "Seoul"])
    wb.save(file_path)

    with ExcelConnection(file_path) as conn:
        with ExcelTable(conn, "Sheet") as table:
            result = table.query(columns=["Name"])
            assert result == [{"Name": "Alice"}]


def test_query_where(tmp_path):
    from openpyxl import Workbook

    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Age", "City"])
    ws.append(["Alice", 25, "Seoul"])
    ws.append(["Bob", 30, "Tokyo"])
    wb.save(file_path)

    with ExcelConnection(file_path) as conn:
        with ExcelTable(conn, "Sheet") as table:
            result = table.query(where=lambda row: row[1] > 25)
            assert result == [{"Name": "Bob", "Age": 30, "City": "Tokyo"}]


def test_query_limit(tmp_path):
    from openpyxl import Workbook

    file_path = tmp_path / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Age"])
    ws.append(["Alice", 25])
    ws.append(["Bob", 30])
    wb.save(file_path)

    with ExcelConnection(file_path) as conn:
        with ExcelTable(conn, "Sheet") as table:
            result = table.query(limit=1)
            assert result == [{"Name": "Alice", "Age": 25}]
