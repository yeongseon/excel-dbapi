from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from excel_dbapi.cli import main


@pytest.fixture
def sample_workbook(tmp_path: Path) -> Path:
    """Create a temporary workbook with sample data for testing."""
    file_path = tmp_path / "sample.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Sheet1"

    # Add headers and data
    sheet.append(["id", "name", "score"])
    sheet.append([1, "Alice", 95])
    sheet.append([2, "Bob", 87])

    workbook.save(file_path)
    workbook.close()

    return file_path


@pytest.fixture
def multi_sheet_workbook(tmp_path: Path) -> Path:
    """Create a temporary workbook with multiple sheets for testing."""
    file_path = tmp_path / "multi.xlsx"
    workbook = Workbook()

    # First sheet
    sheet1 = workbook.active
    assert sheet1 is not None
    sheet1.title = "Sheet1"
    sheet1.append(["id", "name", "score"])
    sheet1.append([1, "Alice", 95])
    sheet1.append([2, "Bob", 87])

    # Second sheet
    sheet2 = workbook.create_sheet("Sheet2")
    sheet2.append(["product_id", "product_name", "price"])
    sheet2.append([101, "Widget", 9.99])
    sheet2.append([102, "Gadget", 19.99])

    workbook.save(file_path)
    workbook.close()

    return file_path


class TestInspectCommand:
    """Tests for the inspect command."""

    def test_inspect_displays_workbook_info(
        self, sample_workbook: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        """Test that inspect shows workbook name, engine, and sheet info."""
        with pytest.raises(SystemExit) as exc_info:
            main(["inspect", str(sample_workbook)])
        assert exc_info.value.code == 0
        captured = capsys.readouterr()

        assert "Workbook: sample.xlsx" in captured.out
        assert "Engine: openpyxl" in captured.out
        assert "Sheet1" in captured.out
        assert "rows: 2" in captured.out
        assert "columns: 3" in captured.out
        assert "headers: id, name, score" in captured.out

    def test_inspect_with_multiple_sheets(
        self, multi_sheet_workbook: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        """Test that inspect shows all sheets in multi-sheet workbook."""
        with pytest.raises(SystemExit) as exc_info:
            main(["inspect", str(multi_sheet_workbook)])
        assert exc_info.value.code == 0
        captured = capsys.readouterr()

        assert "Workbook: multi.xlsx" in captured.out
        assert "Sheet1" in captured.out
        assert "Sheet2" in captured.out
        assert "product_id, product_name, price" in captured.out

    def test_inspect_nonexistent_file(self, capsys: pytest.CaptureFixture[str]) -> None:
        """Test that inspect raises SystemExit(1) for missing file."""
        with pytest.raises(SystemExit) as exc_info:
            main(["inspect", "/nonexistent/file.xlsx"])

        assert exc_info.value.code == 1
        captured = capsys.readouterr()
        assert "Error:" in captured.err


class TestTablesCommand:
    """Tests for the tables command."""

    def test_tables_lists_sheet_names(
        self, sample_workbook: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        """Test that tables command lists all sheet names."""
        with pytest.raises(SystemExit) as exc_info:
            main(["tables", str(sample_workbook)])
        assert exc_info.value.code == 0
        captured = capsys.readouterr()

        assert "Sheet1" in captured.out

    def test_tables_lists_multiple_sheets(
        self, multi_sheet_workbook: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        """Test that tables command lists all sheets in multi-sheet workbook."""
        with pytest.raises(SystemExit) as exc_info:
            main(["tables", str(multi_sheet_workbook)])
        assert exc_info.value.code == 0
        captured = capsys.readouterr()

        assert "Sheet1" in captured.out
        assert "Sheet2" in captured.out

    def test_tables_nonexistent_file(self, capsys: pytest.CaptureFixture[str]) -> None:
        """Test that tables raises SystemExit(1) for missing file."""
        with pytest.raises(SystemExit) as exc_info:
            main(["tables", "/nonexistent/file.xlsx"])

        assert exc_info.value.code == 1
        captured = capsys.readouterr()
        assert "Error:" in captured.err


class TestSchemaCommand:
    """Tests for the schema command."""

    def test_schema_all_sheets(
        self, sample_workbook: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        """Test that schema command shows all sheets with headers and row counts."""
        with pytest.raises(SystemExit) as exc_info:
            main(["schema", str(sample_workbook)])
        assert exc_info.value.code == 0
        captured = capsys.readouterr()

        assert "Sheet1:" in captured.out
        assert "rows: 2" in captured.out
        assert "headers: id, name, score" in captured.out

    def test_schema_specific_sheet(
        self, multi_sheet_workbook: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        """Test that schema command shows only the specified sheet."""
        with pytest.raises(SystemExit) as exc_info:
            main(["schema", str(multi_sheet_workbook), "Sheet2"])
        assert exc_info.value.code == 0
        captured = capsys.readouterr()

        assert "Sheet2:" in captured.out
        assert "rows: 2" in captured.out
        assert "headers: product_id, product_name, price" in captured.out
        # Ensure Sheet1 is not shown
        assert "Sheet1:" not in captured.out

    def test_schema_multiple_sheets_output(
        self, multi_sheet_workbook: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        """Test that schema shows all sheets when no sheet arg provided."""
        with pytest.raises(SystemExit) as exc_info:
            main(["schema", str(multi_sheet_workbook)])
        assert exc_info.value.code == 0
        captured = capsys.readouterr()

        assert "Sheet1:" in captured.out
        assert "Sheet2:" in captured.out

    def test_schema_nonexistent_file(self, capsys: pytest.CaptureFixture[str]) -> None:
        """Test that schema raises SystemExit(1) for missing file."""
        with pytest.raises(SystemExit) as exc_info:
            main(["schema", "/nonexistent/file.xlsx"])

        assert exc_info.value.code == 1
        captured = capsys.readouterr()
        assert "Error:" in captured.err


class TestQueryCommand:
    """Tests for the query command."""

    def test_query_select_displays_tabular_output(
        self, sample_workbook: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        """Test that query command displays SELECT results in tabular format."""
        with pytest.raises(SystemExit) as exc_info:
            main(["query", str(sample_workbook), "SELECT * FROM Sheet1"])
        assert exc_info.value.code == 0
        captured = capsys.readouterr()

        # Should have header row with column names
        assert "id" in captured.out
        assert "name" in captured.out
        assert "score" in captured.out
        # Should have data rows
        assert "Alice" in captured.out
        assert "Bob" in captured.out
        assert "95" in captured.out
        assert "87" in captured.out
        # Should show row count
        assert "2 row(s)" in captured.out

    def test_query_select_specific_columns(
        self, sample_workbook: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        """Test that query command correctly filters columns."""
        with pytest.raises(SystemExit) as exc_info:
            main(["query", str(sample_workbook), "SELECT name, score FROM Sheet1"])
        assert exc_info.value.code == 0
        captured = capsys.readouterr()

        assert "name" in captured.out
        assert "score" in captured.out
        assert "Alice" in captured.out
        # Verify projection: id column must NOT appear in header row
        header_line = captured.out.strip().split("\n")[0]
        assert "id" not in header_line.lower().split()
        assert "2 row(s)" in captured.out

    def test_query_select_with_where_clause(
        self, sample_workbook: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        """Test that query command respects WHERE clause."""
        with pytest.raises(SystemExit) as exc_info:
            main(
                ["query", str(sample_workbook), "SELECT * FROM Sheet1 WHERE score > 90"]
            )
        assert exc_info.value.code == 0
        captured = capsys.readouterr()

        assert "Alice" in captured.out
        assert "95" in captured.out
        assert "1 row(s)" in captured.out

    def test_query_select_empty_sheet(
        self, tmp_path: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        """SELECT on a sheet with headers but no data rows shows 0 row(s), not OK."""
        file_path = tmp_path / "empty.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "Sheet1"
        ws.append(["id", "name"])
        wb.save(file_path)
        wb.close()

        with pytest.raises(SystemExit) as exc_info:
            main(["query", str(file_path), "SELECT * FROM Sheet1"])
        assert exc_info.value.code == 0
        captured = capsys.readouterr()

        assert "0 row(s)" in captured.out
        assert "OK" not in captured.out

    def test_query_insert_single_row(
        self, sample_workbook: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        """Test that INSERT statement displays OK message."""
        with pytest.raises(SystemExit) as exc_info:
            main(
                [
                    "query",
                    str(sample_workbook),
                    "INSERT INTO Sheet1 VALUES (3, 'Carol', 92)",
                ]
            )
        assert exc_info.value.code == 0
        captured = capsys.readouterr()

        assert "OK" in captured.out
        assert "1 rows affected" in captured.out

    def test_query_insert_multiple_rows(
        self, sample_workbook: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        """Test that INSERT with multiple rows displays correct row count."""
        with pytest.raises(SystemExit) as exc_info:
            main(
                [
                    "query",
                    str(sample_workbook),
                    "INSERT INTO Sheet1 VALUES (3, 'Carol', 92), (4, 'David', 88)",
                ]
            )
        assert exc_info.value.code == 0
        captured = capsys.readouterr()

        assert "OK" in captured.out
        assert "2 rows affected" in captured.out

    def test_query_update_statement(
        self, sample_workbook: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        """Test that UPDATE statement displays OK message."""
        with pytest.raises(SystemExit) as exc_info:
            main(
                [
                    "query",
                    str(sample_workbook),
                    "UPDATE Sheet1 SET name = 'Alicia' WHERE id = 1",
                ]
            )
        assert exc_info.value.code == 0
        captured = capsys.readouterr()

        assert "OK" in captured.out

    def test_query_delete_statement(
        self, sample_workbook: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        """Test that DELETE statement displays OK message."""
        with pytest.raises(SystemExit) as exc_info:
            main(["query", str(sample_workbook), "DELETE FROM Sheet1 WHERE id = 1"])
        assert exc_info.value.code == 0
        captured = capsys.readouterr()

        assert "OK" in captured.out
        assert "1 rows affected" in captured.out

    def test_query_nonexistent_file(self, capsys: pytest.CaptureFixture[str]) -> None:
        """Test that query raises SystemExit(1) for missing file."""
        with pytest.raises(SystemExit) as exc_info:
            main(["query", "/nonexistent/file.xlsx", "SELECT * FROM Sheet1"])

        assert exc_info.value.code == 1
        captured = capsys.readouterr()
        assert "Error:" in captured.err

    def test_query_invalid_sql(
        self, sample_workbook: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        """Test that invalid SQL raises SystemExit(1)."""
        with pytest.raises(SystemExit) as exc_info:
            main(["query", str(sample_workbook), "INVALID SQL QUERY"])

        assert exc_info.value.code == 1
        captured = capsys.readouterr()
        assert "Error:" in captured.err


class TestErrorHandling:
    """Tests for error handling and edge cases."""

    def test_missing_file_path_argument(
        self, capsys: pytest.CaptureFixture[str]
    ) -> None:
        """Test that missing file path argument raises SystemExit."""
        with pytest.raises(SystemExit):
            main(["inspect"])

    def test_missing_sql_argument(
        self, sample_workbook: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        """Test that query without SQL argument raises SystemExit."""
        with pytest.raises(SystemExit):
            main(["query", str(sample_workbook)])

    def test_unknown_command(self, capsys: pytest.CaptureFixture[str]) -> None:
        """Test that unknown command raises SystemExit."""
        with pytest.raises(SystemExit):
            main(["unknown", "/some/file.xlsx"])


class TestEngineOption:
    """Tests for the --engine option."""

    def test_inspect_with_openpyxl_engine(
        self, sample_workbook: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        """Test that --engine option works with inspect command."""
        with pytest.raises(SystemExit) as exc_info:
            main(["inspect", str(sample_workbook), "--engine", "openpyxl"])
        assert exc_info.value.code == 0
        captured = capsys.readouterr()

        assert "Engine: openpyxl" in captured.out

    def test_query_with_openpyxl_engine(
        self, sample_workbook: Path, capsys: pytest.CaptureFixture[str]
    ) -> None:
        """Test that --engine option works with query command."""
        with pytest.raises(SystemExit) as exc_info:
            main(
                [
                    "query",
                    str(sample_workbook),
                    "SELECT * FROM Sheet1",
                    "--engine",
                    "openpyxl",
                ]
            )
        assert exc_info.value.code == 0
        captured = capsys.readouterr()

        assert "Alice" in captured.out
        assert "Bob" in captured.out
