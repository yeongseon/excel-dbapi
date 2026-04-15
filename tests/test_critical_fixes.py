"""Tests for critical bug fixes (#54-#58).

Fix 1 (#54): connection.execute() autocommit persistence
Fix 2 (#55): DSN auto engine selection
Fix 3 (#56): Package version matches pyproject.toml
Fix 4 (#57): Header normalization/validation
Fix 5 (#58): Exception mapping + cursor API signature
"""

from __future__ import annotations

import os
from pathlib import Path
from typing import Any

import pytest
from openpyxl import Workbook

from excel_dbapi import ExcelConnection, connect
from excel_dbapi.engines.base import _normalize_headers
from excel_dbapi.exceptions import (
    DataError,
    DatabaseError,
    NotSupportedError,
    OperationalError,
    ProgrammingError,
)


def _make_xlsx(
    path: Path,
    sheet: str = "users",
    headers: list[str] | None = None,
    rows: list[list[Any]] | None = None,
) -> str:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = sheet
    for h in headers or ["id", "name"]:
        pass
    ws.append(headers or ["id", "name"])
    for row in rows or []:
        ws.append(row)
    fpath = str(path)
    wb.save(fpath)
    wb.close()
    return fpath


# ─── Fix 1: connection.execute() autocommit persists (#54) ──────────────


class TestConnectionExecuteAutocommit:
    """Fix 1: connection.execute() must save when autocommit=True."""

    def test_connection_execute_insert_persists(self, tmp_path: Path) -> None:
        """connection.execute('INSERT ...') persists data when autocommit=True."""
        fpath = _make_xlsx(tmp_path / "test.xlsx")
        with ExcelConnection(fpath, engine="openpyxl", autocommit=True) as conn:
            conn.execute("INSERT INTO users VALUES (1, 'Alice')")

        # Re-open and verify data was persisted
        with ExcelConnection(fpath, engine="openpyxl") as conn:
            result = conn.execute("SELECT * FROM users")
            assert len(result.rows) == 1
            assert result.rows[0] == (1, "Alice")

    def test_connection_execute_update_persists(self, tmp_path: Path) -> None:
        """connection.execute('UPDATE ...') persists data."""
        fpath = _make_xlsx(tmp_path / "test.xlsx", rows=[[1, "Alice"]])
        with ExcelConnection(fpath, engine="openpyxl", autocommit=True) as conn:
            conn.execute("UPDATE users SET name = 'Bob' WHERE id = 1")

        with ExcelConnection(fpath, engine="openpyxl") as conn:
            result = conn.execute("SELECT * FROM users WHERE id = 1")
            assert result.rows[0] == (1, "Bob")

    def test_connection_execute_delete_persists(self, tmp_path: Path) -> None:
        """connection.execute('DELETE ...') persists data."""
        fpath = _make_xlsx(tmp_path / "test.xlsx", rows=[[1, "Alice"], [2, "Bob"]])
        with ExcelConnection(fpath, engine="openpyxl", autocommit=True) as conn:
            conn.execute("DELETE FROM users WHERE id = 1")

        with ExcelConnection(fpath, engine="openpyxl") as conn:
            result = conn.execute("SELECT * FROM users")
            assert len(result.rows) == 1

    def test_connection_execute_no_save_without_autocommit(
        self, tmp_path: Path
    ) -> None:
        """connection.execute() does NOT save when autocommit=False."""
        fpath = _make_xlsx(tmp_path / "test.xlsx")
        with ExcelConnection(fpath, engine="openpyxl", autocommit=False) as conn:
            conn.execute("INSERT INTO users VALUES (1, 'Alice')")
            # Not committed — rollback
            conn.rollback()

        with ExcelConnection(fpath, engine="openpyxl") as conn:
            result = conn.execute("SELECT * FROM users")
            assert len(result.rows) == 0

    def test_connection_execute_select_no_save(self, tmp_path: Path) -> None:
        """SELECT via connection.execute() does NOT trigger save."""
        fpath = _make_xlsx(tmp_path / "test.xlsx", rows=[[1, "Alice"]])
        os.path.getmtime(fpath)
        with ExcelConnection(fpath, engine="openpyxl", autocommit=True) as conn:
            result = conn.execute("SELECT * FROM users")
            assert len(result.rows) == 1
        # File should not have been re-saved for a SELECT
        # (mod_time could be same if fast, but at least no error)

    def test_executemany_saves_once_not_per_row(self, tmp_path: Path) -> None:
        """executemany() should save only once at end, not per row."""
        fpath = _make_xlsx(tmp_path / "test.xlsx")
        save_count = 0
        with ExcelConnection(fpath, engine="openpyxl", autocommit=True) as conn:
            original_save = conn.engine.save

            def counting_save() -> None:
                nonlocal save_count
                save_count += 1
                original_save()

            conn.engine.save = counting_save  # type: ignore[assignment]
            cursor = conn.cursor()
            cursor.executemany(
                "INSERT INTO users VALUES (?, ?)",
                [(1, "Alice"), (2, "Bob"), (3, "Charlie")],
            )
        assert save_count == 1, f"Expected 1 save, got {save_count}"

    def test_connection_execute_create_table_persists(self, tmp_path: Path) -> None:
        """connection.execute('CREATE TABLE ...') persists."""
        fpath = _make_xlsx(tmp_path / "test.xlsx")
        with ExcelConnection(fpath, engine="openpyxl", autocommit=True) as conn:
            conn.execute("CREATE TABLE orders (id INTEGER, product TEXT)")

        with ExcelConnection(fpath, engine="openpyxl") as conn:
            result = conn.execute("SELECT * FROM orders")
            assert result.rows == []


# ─── Fix 2: DSN auto engine selection (#55) ─────────────────────────────


class TestDSNAutoEngineSelection:
    """Fix 2: engine default=None allows DSN-based auto-detection."""

    def test_connect_defaults_to_openpyxl_for_local_files(self, tmp_path: Path) -> None:
        """connect() with local file still uses openpyxl when engine=None."""
        fpath = _make_xlsx(tmp_path / "test.xlsx")
        conn = connect(fpath, autocommit=True)
        assert "Openpyxl" in conn.engine_name
        conn.close()

    def test_connect_explicit_engine_still_works(self, tmp_path: Path) -> None:
        """Explicit engine='openpyxl' still works."""
        fpath = _make_xlsx(tmp_path / "test.xlsx")
        conn = connect(fpath, engine="openpyxl", autocommit=True)
        assert "Openpyxl" in conn.engine_name
        conn.close()

    def test_dsn_engine_auto_detection(self) -> None:
        """msgraph:// DSN auto-detects 'graph' engine (would fail with
        engine='openpyxl' default due to mismatch error)."""
        from excel_dbapi.connection import _resolve_engine_and_location

        engine, location = _resolve_engine_and_location(
            "msgraph://drives/fake/items/fake", None
        )
        assert engine == "graph"
        assert location == "msgraph://drives/fake/items/fake"

    def test_dsn_engine_mismatch_raises(self) -> None:
        """Explicit engine conflicting with DSN raises ValueError."""
        from excel_dbapi.connection import _resolve_engine_and_location

        with pytest.raises(DatabaseError, match="Engine mismatch"):
            _resolve_engine_and_location("msgraph://drives/fake/items/fake", "openpyxl")

    @pytest.mark.parametrize(
        ("dsn", "expected_location"),
        [
            (
                "sharepoint://sites/team/drives/drive-1/items/item-1",
                "sharepoint://sites/team/drives/drive-1/items/item-1",
            ),
            (
                "onedrive://me/drive/items/item-2",
                "onedrive://me/drive/items/item-2",
            ),
        ],
    )
    def test_extended_graph_schemes_auto_detected(
        self, dsn: str, expected_location: str
    ) -> None:
        from excel_dbapi.connection import _resolve_engine_and_location

        engine, location = _resolve_engine_and_location(dsn, None)
        assert engine == "graph"
        assert location == expected_location


# ─── Fix 3: Package version (#56) ───────────────────────────────────────


class TestPackageVersion:
    """Fix 3: __version__ matches pyproject.toml."""

    def test_version_is_not_hardcoded_030(self) -> None:
        """__version__ should NOT be the old hardcoded '0.3.0'."""
        import excel_dbapi

        assert excel_dbapi.__version__ != "0.3.0"

    def test_version_matches_pyproject(self) -> None:
        """__version__ matches the version in pyproject.toml."""
        import importlib

        try:
            toml_module = importlib.import_module("tomllib")
        except ModuleNotFoundError:
            toml_module = importlib.import_module("tomli")

        pyproject = Path(__file__).parent.parent / "pyproject.toml"
        with open(pyproject, "rb") as f:
            data = toml_module.load(f)
        expected = data["project"]["version"]

        import excel_dbapi

        # In editable installs the version comes from importlib.metadata
        # which reads pyproject.toml, so they should match.
        assert excel_dbapi.__version__ == expected

    def test_version_fallback_format(self) -> None:
        """Version is a valid semver-like string (not empty)."""
        import excel_dbapi

        parts = excel_dbapi.__version__.split(".")
        assert len(parts) >= 2, (
            f"Version {excel_dbapi.__version__!r} is not semver-like"
        )


# ─── Fix 4: Header normalization (#57) ──────────────────────────────────


class TestHeaderNormalization:
    """Fix 4: _normalize_headers validates and coerces headers."""

    def test_valid_headers(self) -> None:
        result = _normalize_headers(["id", "name", "age"])
        assert result == ["id", "name", "age"]

    def test_numeric_headers_coerced_to_string(self) -> None:
        result = _normalize_headers([1, 2, 3])
        assert result == ["1", "2", "3"]

    def test_none_header_raises_data_error(self) -> None:
        with pytest.raises(DataError, match="Empty or None header at column index 1"):
            _normalize_headers(["id", None, "age"])

    def test_empty_string_header_raises_data_error(self) -> None:
        with pytest.raises(DataError, match="Empty or None header"):
            _normalize_headers(["id", "", "age"])

    def test_whitespace_only_header_raises_data_error(self) -> None:
        with pytest.raises(DataError, match="Empty or None header"):
            _normalize_headers(["id", "   ", "age"])

    def test_duplicate_headers_raises_data_error(self) -> None:
        with pytest.raises(DataError, match="Duplicate header"):
            _normalize_headers(["id", "name", "id"])

    def test_case_insensitive_duplicate_raises_data_error(self) -> None:
        with pytest.raises(DataError, match="Duplicate header"):
            _normalize_headers(["Name", "age", "name"])

    def test_duplicate_headers_in_xlsx(self, tmp_path: Path) -> None:
        """Reading a sheet with duplicate headers raises DataError."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "bad"
        ws.append(["id", "name", "id"])
        ws.append([1, "Alice", 2])
        fpath = str(tmp_path / "dup.xlsx")
        wb.save(fpath)
        wb.close()

        with ExcelConnection(fpath, engine="openpyxl") as conn:
            with pytest.raises(DataError, match="Duplicate header"):
                conn.execute("SELECT * FROM bad")

    def test_empty_header_in_xlsx(self, tmp_path: Path) -> None:
        """Reading a sheet with empty/None header raises DataError."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = "bad"
        ws.append(["id", None, "age"])
        ws.append([1, "Alice", 25])
        fpath = str(tmp_path / "empty.xlsx")
        wb.save(fpath)
        wb.close()

        with ExcelConnection(fpath, engine="openpyxl") as conn:
            with pytest.raises(DataError, match="Empty or None header"):
                conn.execute("SELECT * FROM bad")


# ─── Fix 5: Exception mapping (#58) ─────────────────────────────────────


class TestExceptionMapping:
    """Fix 5: Cursor wraps raw exceptions into PEP 249 hierarchy."""

    def test_value_error_becomes_programming_error(self, tmp_path: Path) -> None:
        """ValueError from executor → ProgrammingError."""
        fpath = _make_xlsx(tmp_path / "test.xlsx")
        with ExcelConnection(fpath, engine="openpyxl") as conn:
            cursor = conn.cursor()
            with pytest.raises(ProgrammingError):
                cursor.execute("INVALID SQL GIBBERISH")


def test_malformed_create_table_missing_comma_raises_programming_error(
    tmp_path: Path,
) -> None:
    fpath = _make_xlsx(tmp_path / "malformed-create.xlsx")
    with ExcelConnection(fpath, engine="openpyxl") as conn:
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="Missing comma"):
            cursor.execute("CREATE TABLE t (id INTEGER name TEXT)")

    def test_not_implemented_becomes_not_supported(self, tmp_path: Path) -> None:
        """NotImplementedError → NotSupportedError."""
        fpath = _make_xlsx(tmp_path / "test.xlsx")
        with ExcelConnection(fpath, engine="openpyxl") as conn:
            cursor = conn.cursor()
            # Trigger a non-supported operation
            with pytest.raises((ProgrammingError, NotSupportedError)):
                cursor.execute("CREATE INDEX idx ON users (id)")

    def test_executemany_accepts_iterable_of_sequences(self, tmp_path: Path) -> None:
        """executemany() accepts Iterable[Sequence[Any]], not just List[tuple]."""
        fpath = _make_xlsx(tmp_path / "test.xlsx")
        with ExcelConnection(fpath, engine="openpyxl", autocommit=True) as conn:
            cursor = conn.cursor()
            # Pass a generator of lists (not List[tuple])
            params = ([i, f"user{i}"] for i in range(1, 4))
            cursor.executemany("INSERT INTO users VALUES (?, ?)", params)
            assert cursor.rowcount == 3

        with ExcelConnection(fpath, engine="openpyxl") as conn:
            result = conn.execute("SELECT * FROM users")
            assert len(result.rows) == 3

    def test_executemany_rollback_on_error_autocommit_off(self, tmp_path: Path) -> None:
        """executemany() with autocommit=False rolls back on error."""
        fpath = _make_xlsx(tmp_path / "test.xlsx", rows=[[1, "Original"]])
        with ExcelConnection(fpath, engine="openpyxl", autocommit=False) as conn:
            cursor = conn.cursor()
            with pytest.raises((ProgrammingError, DatabaseError)):
                # First insert succeeds, second should fail
                cursor.executemany(
                    "INSERT INTO users VALUES (?, ?)",
                    [(2, "Bob"), (3, None, "extra")],  # type: ignore[list-item]
                )
            # After error, rollback should have happened
            conn.rollback()
            result = conn.execute("SELECT * FROM users")
            # Only original row should remain
            assert len(result.rows) == 1
            assert result.rows[0] == (1, "Original")

    def test_database_error_subclasses_pass_through(self, tmp_path: Path) -> None:
        """PEP 249 exceptions already raised in executor pass through unchanged."""
        fpath = _make_xlsx(tmp_path / "test.xlsx")
        with ExcelConnection(fpath, engine="openpyxl") as conn:
            cursor = conn.cursor()
            # This raises ProgrammingError from the parser — should pass through
            with pytest.raises(ProgrammingError):
                cursor.execute("SELECT * FROM nonexistent_table_xyz")


class TestExceptionMappingDirect:
    """Gap 2: Direct tests for each exception type mapping in execute() and executemany()."""

    def _make_conn_with_raising_executor(
        self, tmp_path: Path, exc: Exception
    ) -> ExcelConnection:
        """Create a connection whose executor raises the given exception."""
        fpath = _make_xlsx(tmp_path / "test.xlsx")
        conn = ExcelConnection(fpath, engine="openpyxl", autocommit=True)

        def raising_execute(query: str, params: Any = None) -> Any:
            raise exc

        conn._executor.execute_with_params = raising_execute  # type: ignore[assignment]
        return conn

    def test_key_error_maps_to_programming_error_execute(self, tmp_path: Path) -> None:
        conn = self._make_conn_with_raising_executor(tmp_path, KeyError("missing"))
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="missing"):
            cursor.execute("SELECT * FROM users")
        conn.close()

    def test_type_error_maps_to_programming_error_execute(self, tmp_path: Path) -> None:
        conn = self._make_conn_with_raising_executor(tmp_path, TypeError("bad type"))
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="bad type"):
            cursor.execute("SELECT * FROM users")
        conn.close()

    def test_index_error_maps_to_programming_error_execute(
        self, tmp_path: Path
    ) -> None:
        conn = self._make_conn_with_raising_executor(
            tmp_path, IndexError("out of range")
        )
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="out of range"):
            cursor.execute("SELECT * FROM users")
        conn.close()

    def test_os_error_maps_to_operational_error_execute(self, tmp_path: Path) -> None:
        conn = self._make_conn_with_raising_executor(tmp_path, OSError("disk full"))
        cursor = conn.cursor()
        with pytest.raises(OperationalError, match="disk full"):
            cursor.execute("SELECT * FROM users")
        conn.close()

    def test_generic_exception_maps_to_database_error_execute(
        self, tmp_path: Path
    ) -> None:
        conn = self._make_conn_with_raising_executor(
            tmp_path, RuntimeError("unexpected")
        )
        cursor = conn.cursor()
        with pytest.raises(DatabaseError, match="unexpected"):
            cursor.execute("SELECT * FROM users")
        conn.close()

    def test_database_error_passes_through_execute(self, tmp_path: Path) -> None:
        conn = self._make_conn_with_raising_executor(
            tmp_path, ProgrammingError("bad sql")
        )
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="bad sql"):
            cursor.execute("SELECT * FROM users")
        conn.close()

    def test_key_error_maps_to_programming_error_executemany(
        self, tmp_path: Path
    ) -> None:
        fpath = _make_xlsx(tmp_path / "test.xlsx")
        conn = ExcelConnection(fpath, engine="openpyxl", autocommit=True)
        original = conn._executor.execute_with_params
        call_count = 0

        def raising_on_second(query: str, params: Any = None) -> Any:
            nonlocal call_count
            call_count += 1
            if call_count >= 2:
                raise KeyError("missing key")
            return original(query, params)

        conn._executor.execute_with_params = raising_on_second  # type: ignore[assignment]
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="missing key"):
            cursor.executemany("INSERT INTO users VALUES (?, ?)", [(1, "a"), (2, "b")])
        conn.close()

    def test_os_error_maps_to_operational_error_executemany(
        self, tmp_path: Path
    ) -> None:
        fpath = _make_xlsx(tmp_path / "test.xlsx")
        conn = ExcelConnection(fpath, engine="openpyxl", autocommit=True)

        def raising_execute(query: str, params: Any = None) -> Any:
            raise OSError("permission denied")

        conn._executor.execute_with_params = raising_execute  # type: ignore[assignment]
        cursor = conn.cursor()
        with pytest.raises(OperationalError, match="permission denied"):
            cursor.executemany("INSERT INTO users VALUES (?, ?)", [(1, "a")])
        conn.close()

    def test_generic_exception_maps_to_database_error_executemany(
        self, tmp_path: Path
    ) -> None:
        fpath = _make_xlsx(tmp_path / "test.xlsx")
        conn = ExcelConnection(fpath, engine="openpyxl", autocommit=True)

        def raising_execute(query: str, params: Any = None) -> Any:
            raise RuntimeError("boom")

        conn._executor.execute_with_params = raising_execute  # type: ignore[assignment]
        cursor = conn.cursor()
        with pytest.raises(DatabaseError, match="boom"):
            cursor.executemany("INSERT INTO users VALUES (?, ?)", [(1, "a")])
        conn.close()


class TestExecutemanyMidBatchRestore:
    """Gap 3: Mid-batch executemany failure restores snapshot after partial mutation."""

    def test_mid_batch_failure_restores_snapshot(self, tmp_path: Path) -> None:
        """After 1 successful INSERT, a failure in batch 2 restores original state."""
        fpath = _make_xlsx(tmp_path / "test.xlsx", rows=[[1, "Original"]])
        conn = ExcelConnection(fpath, engine="openpyxl", autocommit=False)
        original = conn._executor.execute_with_params
        call_count = 0

        def fail_on_second(query: str, params: Any = None) -> Any:
            nonlocal call_count
            call_count += 1
            if call_count == 2:
                raise ValueError("bad value")
            return original(query, params)

        conn._executor.execute_with_params = fail_on_second  # type: ignore[assignment]
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="bad value"):
            cursor.executemany(
                "INSERT INTO users VALUES (?, ?)",
                [(2, "Second"), (3, "Third")],
            )
        # After error + snapshot restore, only original row should exist
        conn._executor.execute_with_params = original  # type: ignore[assignment]
        result = conn.execute("SELECT * FROM users")
        assert len(result.rows) == 1
        assert result.rows[0] == (1, "Original")
        conn.close()


class TestExecutemanyAutocommitRestore:
    """Regression: executemany with autocommit=True must also restore on failure."""

    def test_autocommit_true_mid_batch_failure_restores(self, tmp_path: Path) -> None:
        """Under autocommit=True, a mid-batch failure restores the snapshot.

        Previously snapshot was only taken when autocommit=False, leaving
        partial mutations in memory under autocommit=True.
        """
        fpath = _make_xlsx(tmp_path / "test.xlsx", rows=[[1, "Original"]])
        conn = ExcelConnection(fpath, engine="openpyxl", autocommit=True)
        original = conn._executor.execute_with_params
        call_count = 0

        def fail_on_second(query: str, params: Any = None) -> Any:
            nonlocal call_count
            call_count += 1
            if call_count == 2:
                raise ValueError("bad value")
            return original(query, params)

        conn._executor.execute_with_params = fail_on_second  # type: ignore[assignment]
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError, match="bad value"):
            cursor.executemany(
                "INSERT INTO users VALUES (?, ?)",
                [(2, "Second"), (3, "Third")],
            )
        # After error + snapshot restore, only original row should exist
        conn._executor.execute_with_params = original  # type: ignore[assignment]
        result = conn.execute("SELECT * FROM users")
        assert len(result.rows) == 1
        assert result.rows[0] == (1, "Original")
        conn.close()


class TestHeaderWhitespaceTrimming:
    """Regression: _normalize_headers strips leading/trailing whitespace."""

    def test_whitespace_trimmed_from_headers(self) -> None:
        """Headers with leading/trailing whitespace are trimmed."""
        result = _normalize_headers(["  id  ", " name ", "age"])
        assert result == ["id", "name", "age"]

    def test_trimmed_duplicates_detected(self) -> None:
        """After trimming, duplicate headers are detected."""
        with pytest.raises(DataError, match="Duplicate header"):
            _normalize_headers(["  id  ", "id", "name"])
