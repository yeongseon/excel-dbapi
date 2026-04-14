"""Tests for API extensions: create flag, data_only control, workbook property."""

import os

import pytest
from openpyxl import Workbook, load_workbook

import excel_dbapi
from excel_dbapi.connection import ExcelConnection
from excel_dbapi.exceptions import NotSupportedError


@pytest.fixture
def tmp_xlsx(tmp_path):
    """Create a temporary Excel file with sample data."""
    path = str(tmp_path / "sample.xlsx")
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id", "name", "value"])
    ws.append([1, "Alice", 100])
    ws.append([2, "Bob", 200])
    wb.save(path)
    return path


@pytest.fixture
def tmp_xlsx_with_formula(tmp_path):
    """Create a temporary Excel file with a formula."""
    path = str(tmp_path / "formula.xlsx")
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["a", "b", "total"])
    ws["A2"] = 10
    ws["B2"] = 20
    ws["C2"] = "=A2+B2"
    wb.save(path)
    return path


@pytest.fixture
def nonexistent_path(tmp_path):
    """Return a path that does not exist."""
    return str(tmp_path / "nonexistent.xlsx")


# ==============================================================================
# Extension 1: create flag
# ==============================================================================


class TestCreateFlag:
    """Tests for the create=True/False parameter."""

    def test_create_true_new_file_openpyxl(self, nonexistent_path):
        """create=True with non-existent file should create a new workbook."""
        assert not os.path.exists(nonexistent_path)
        conn = ExcelConnection(nonexistent_path, engine="openpyxl", create=True)
        assert os.path.exists(nonexistent_path)
        conn.close()

    def test_create_true_existing_file_openpyxl(self, tmp_xlsx):
        """create=True with existing file should load normally."""
        conn = ExcelConnection(tmp_xlsx, engine="openpyxl", create=True)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM Sheet1")
        rows = cursor.fetchall()
        assert len(rows) == 2
        assert rows[0][1] == "Alice"
        conn.close()

    def test_create_false_nonexistent_file_raises(self, nonexistent_path):
        """create=False (default) with non-existent file should raise."""
        with pytest.raises(Exception):
            ExcelConnection(nonexistent_path, engine="openpyxl", create=False)

    def test_create_default_is_false(self, nonexistent_path):
        """Default create value should be False."""
        with pytest.raises(Exception):
            ExcelConnection(nonexistent_path, engine="openpyxl")

    def test_create_true_new_file_pandas(self, nonexistent_path):
        """create=True with non-existent file should work for pandas engine."""
        assert not os.path.exists(nonexistent_path)
        conn = ExcelConnection(nonexistent_path, engine="pandas", create=True)
        assert os.path.exists(nonexistent_path)
        conn.close()

    def test_create_true_zero_byte_file_pandas(self, tmp_path):
        file_path = tmp_path / "zero-byte.xlsx"
        file_path.write_bytes(b"")

        conn = ExcelConnection(str(file_path), engine="pandas", create=True)
        assert file_path.exists()
        assert file_path.stat().st_size > 0
        conn.close()

    def test_create_via_connect_function(self, nonexistent_path):
        """create=True should work via the module-level connect() function."""
        assert not os.path.exists(nonexistent_path)
        conn = excel_dbapi.connect(nonexistent_path, create=True)
        assert os.path.exists(nonexistent_path)
        conn.close()

    def test_create_then_insert_and_select(self, nonexistent_path):
        """After creating a new file, should be able to CREATE TABLE, INSERT, SELECT."""
        conn = ExcelConnection(nonexistent_path, engine="openpyxl", create=True)
        cursor = conn.cursor()
        cursor.execute("CREATE TABLE Users (id, name)")
        cursor.execute("INSERT INTO Users (id, name) VALUES (1, 'Alice')")
        conn.commit()
        cursor.execute("SELECT * FROM Users")
        rows = cursor.fetchall()
        assert len(rows) == 1
        assert rows[0][0] == 1
        assert rows[0][1] == "Alice"
        conn.close()


# ==============================================================================
# Extension 2: data_only control
# ==============================================================================


class TestDataOnlyControl:
    """Tests for the data_only parameter."""

    def test_data_only_true_default(self, tmp_xlsx_with_formula):
        """data_only=True (default) should read cached values, not formulas."""
        conn = ExcelConnection(tmp_xlsx_with_formula, engine="openpyxl", data_only=True)
        # With data_only=True, the formula cell returns its cached value (or None
        # if the workbook hasn't been opened by Excel to compute cached values).
        # Since we created with openpyxl (no cached values), it returns None.
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM Sheet1")
        rows = cursor.fetchall()
        # The formula cell should NOT contain the formula string
        assert rows[0][2] != "=A2+B2"
        conn.close()

    def test_data_only_false_reads_formulas(self, tmp_xlsx_with_formula):
        """data_only=False should read formula strings."""
        conn = ExcelConnection(
            tmp_xlsx_with_formula, engine="openpyxl", data_only=False
        )
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM Sheet1")
        rows = cursor.fetchall()
        # The formula cell should contain the formula string
        assert rows[0][2] == "=A2+B2"
        conn.close()

    def test_data_only_via_connect_function(self, tmp_xlsx_with_formula):
        """data_only should work via the module-level connect() function."""
        conn = excel_dbapi.connect(tmp_xlsx_with_formula, data_only=False)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM Sheet1")
        rows = cursor.fetchall()
        assert rows[0][2] == "=A2+B2"
        conn.close()

    def test_data_only_pandas_engine_accepted(self, tmp_xlsx):
        """pandas engine should accept data_only without error (ignored)."""
        conn = ExcelConnection(tmp_xlsx, engine="pandas", data_only=False)
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM Sheet1")
        rows = cursor.fetchall()
        assert len(rows) == 2
        conn.close()


# ==============================================================================
# Extension 3: workbook property
# ==============================================================================


class TestWorkbookProperty:
    """Tests for the connection.workbook property."""

    def test_workbook_returns_openpyxl_workbook(self, tmp_xlsx):
        """workbook property should return openpyxl Workbook for OpenpyxlEngine."""
        conn = ExcelConnection(tmp_xlsx, engine="openpyxl")
        wb = conn.workbook
        assert isinstance(wb, Workbook)
        assert "Sheet1" in wb.sheetnames
        conn.close()

    def test_workbook_raises_for_pandas_engine(self, tmp_xlsx):
        """workbook property should raise NotSupportedError for PandasEngine."""
        conn = ExcelConnection(tmp_xlsx, engine="pandas")
        with pytest.raises(NotSupportedError, match="does not expose a workbook"):
            _ = conn.workbook
        conn.close()

    def test_workbook_allows_direct_styling(self, tmp_xlsx):
        """Should be able to modify styling via workbook property."""
        conn = ExcelConnection(tmp_xlsx, engine="openpyxl")
        wb = conn.workbook
        ws = wb["Sheet1"]
        from openpyxl.styles import Font

        ws["A1"].font = Font(bold=True)
        conn.commit()
        conn.close()

        # Verify styling persisted
        wb2 = load_workbook(tmp_xlsx)
        assert wb2["Sheet1"]["A1"].font.bold is True
        wb2.close()

    def test_workbook_allows_data_validation(self, tmp_xlsx):
        """Should be able to add DataValidation via workbook property."""
        conn = ExcelConnection(tmp_xlsx, engine="openpyxl")
        wb = conn.workbook
        ws = wb["Sheet1"]
        from openpyxl.worksheet.datavalidation import DataValidation

        dv = DataValidation(type="list", formula1='"Yes,No"')
        dv.add("D1:D100")
        ws.add_data_validation(dv)
        conn.commit()
        conn.close()

        # Verify DV persisted
        wb2 = load_workbook(tmp_xlsx)
        assert len(wb2["Sheet1"].data_validations.dataValidation) > 0
        wb2.close()

    def test_workbook_via_connect_function(self, tmp_xlsx):
        """workbook property should work via module-level connect()."""
        conn = excel_dbapi.connect(tmp_xlsx)
        wb = conn.workbook
        assert isinstance(wb, Workbook)
        conn.close()
