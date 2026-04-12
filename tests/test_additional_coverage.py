from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from excel_dbapi.connection import ExcelConnection
from excel_dbapi.exceptions import ProgrammingError


def _create_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["id", "name"])
    ws.append([1, "Alice"])
    ws.append([2, "Bob"])
    wb.save(path)


def test_fetchmany_and_arraysize(tmp_path: Path) -> None:
    file_path = tmp_path / "sample.xlsx"
    _create_workbook(file_path)

    with ExcelConnection(str(file_path)) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id, name FROM Sheet1 ORDER BY id ASC")
        cursor.arraysize = 1
        first = cursor.fetchmany()
        assert first == [(1, "Alice")]
        second = cursor.fetchmany(2)
        assert second == [(2, "Bob")]


def test_executemany_autocommit(tmp_path: Path) -> None:
    file_path = tmp_path / "sample.xlsx"
    _create_workbook(file_path)

    with ExcelConnection(str(file_path), autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.executemany(
            "INSERT INTO Sheet1 (id, name) VALUES (?, ?)",
            [(3, "Cara"), (4, "Dane")],
        )
        assert cursor.rowcount == 2

    wb = load_workbook(file_path, data_only=True)
    rows = list(wb["Sheet1"].iter_rows(values_only=True))
    assert rows[-2:] == [(3, "Cara"), (4, "Dane")]


def test_select_with_params_and_limit(tmp_path: Path) -> None:
    file_path = tmp_path / "sample.xlsx"
    _create_workbook(file_path)

    with ExcelConnection(str(file_path)) as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM Sheet1 WHERE id >= ? LIMIT ?", (1, 1))
        assert cursor.fetchall() == [(1,)]


def test_pandas_insert_column_mismatch(tmp_path: Path) -> None:
    file_path = tmp_path / "sample.xlsx"
    df = pd.DataFrame([{"id": 1, "name": "Alice"}])
    df.to_excel(file_path, index=False, sheet_name="Sheet1")

    with ExcelConnection(str(file_path), engine="pandas") as conn:
        cursor = conn.cursor()
        with pytest.raises(ProgrammingError):
            cursor.execute("INSERT INTO Sheet1 (id) VALUES (1, 'A')")
