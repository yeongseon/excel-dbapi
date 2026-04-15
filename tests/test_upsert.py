from pathlib import Path

from openpyxl import Workbook
import pytest

from excel_dbapi.connection import ExcelConnection
from excel_dbapi.exceptions import DatabaseError, ProgrammingError
from excel_dbapi.parser import parse_sql


def _write_sheet(
    workbook: Workbook, name: str, headers: list[str], rows: list[list[object]]
) -> None:
    if workbook.sheetnames:
        sheet = workbook.active
        assert sheet is not None
        sheet.title = name
    else:
        sheet = workbook.create_sheet(name)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)


def _create_items_workbook(
    path: Path, rows: list[list[object]], *, include_source: bool = False
) -> None:
    workbook = Workbook()
    headers = ["id", "name", "age", "status", "code", "value"]
    _write_sheet(workbook, "items", headers, rows)
    if include_source:
        source = workbook.create_sheet("incoming")
        source.append(["id", "name", "age"])
        source.append([1, "FromSource", 99])
        source.append([2, "NewFromSource", 20])
    workbook.save(path)


def _fetch_rows(path: Path, query: str) -> list[tuple[object, ...]]:
    with ExcelConnection(str(path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute(query)
        return cursor.fetchall()


def test_parser_on_conflict_do_nothing() -> None:
    parsed = parse_sql(
        "INSERT INTO items (id, name) VALUES (1, 'Alice') ON CONFLICT (id) DO NOTHING"
    )
    assert parsed["on_conflict"] == {"target_columns": ["id"], "action": "NOTHING"}


def test_parser_on_conflict_do_update() -> None:
    parsed = parse_sql(
        "INSERT INTO items (id, name, age) VALUES (1, 'Alice', 30) "
        "ON CONFLICT (id) DO UPDATE SET name = excluded.name, age = 31"
    )
    on_conflict = parsed["on_conflict"]
    assert on_conflict["target_columns"] == ["id"]
    assert on_conflict["action"] == "UPDATE"
    first_value = on_conflict["set"][0]["value"]
    assert first_value["type"] == "column"
    assert first_value["source"] == "excluded"
    assert first_value["table"] == "excluded"
    assert first_value["name"] == "name"
    assert on_conflict["set"][1]["value"] == {"type": "literal", "value": 31}


@pytest.mark.parametrize(
    "sql",
    [
        "INSERT INTO items VALUES (1, 'A') ON CONFLICT (id) NOTHING",
        "INSERT INTO items VALUES (1, 'A') ON CONFLICT (id) DO MERGE",
        "INSERT INTO items VALUES (1, 'A') ON CONFLICT id DO NOTHING",
        "INSERT INTO items VALUES (1, 'A') ON CONFLICT (id) DO UPDATE name = 'x'",
    ],
)
def test_parser_on_conflict_invalid(sql: str) -> None:
    with pytest.raises(DatabaseError):
        parse_sql(sql)


def test_upsert_do_nothing_basic(tmp_path: Path) -> None:
    file_path = tmp_path / "upsert_nothing_basic.xlsx"
    _create_items_workbook(file_path, [[1, "Old", 30, "active", "A", 10]])

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO items (id, name, age, status, code, value) "
            "VALUES (1, 'New', 40, 'inactive', 'B', 11) "
            "ON CONFLICT (id) DO NOTHING"
        )
        assert cursor.rowcount == 0

    assert _fetch_rows(file_path, "SELECT id, name FROM items ORDER BY id") == [
        (1, "Old")
    ]


def test_upsert_do_nothing_no_conflict(tmp_path: Path) -> None:
    file_path = tmp_path / "upsert_nothing_no_conflict.xlsx"
    _create_items_workbook(file_path, [[1, "Old", 30, "active", "A", 10]])

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO items (id, name, age, status, code, value) "
            "VALUES (2, 'New', 20, 'active', 'B', 5) "
            "ON CONFLICT (id) DO NOTHING"
        )
        assert cursor.rowcount == 1

    assert _fetch_rows(file_path, "SELECT id, name FROM items ORDER BY id") == [
        (1, "Old"),
        (2, "New"),
    ]


def test_upsert_do_nothing_multi_row_mixed(tmp_path: Path) -> None:
    file_path = tmp_path / "upsert_nothing_mixed.xlsx"
    _create_items_workbook(
        file_path,
        [[1, "Old1", 30, "active", "A", 10], [3, "Old3", 35, "active", "C", 12]],
    )

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO items (id, name, age, status, code, value) VALUES "
            "(1, 'Ignored', 99, 'x', 'X', 1), "
            "(2, 'Inserted', 25, 'active', 'B', 7), "
            "(3, 'Ignored2', 98, 'x', 'Y', 2) "
            "ON CONFLICT (id) DO NOTHING"
        )
        assert cursor.rowcount == 1

    assert _fetch_rows(file_path, "SELECT id, name FROM items ORDER BY id") == [
        (1, "Old1"),
        (2, "Inserted"),
        (3, "Old3"),
    ]


def test_upsert_do_update_basic(tmp_path: Path) -> None:
    file_path = tmp_path / "upsert_update_basic.xlsx"
    _create_items_workbook(file_path, [[1, "Old", 30, "active", "A", 10]])

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO items (id, name, age, status, code, value) "
            "VALUES (1, 'New', 50, 'active', 'A', 99) "
            "ON CONFLICT (id) DO UPDATE SET age = 42"
        )
        assert cursor.rowcount == 1

    assert _fetch_rows(file_path, "SELECT id, name, age FROM items ORDER BY id") == [
        (1, "Old", 42)
    ]


def test_upsert_do_update_excluded_reference(tmp_path: Path) -> None:
    file_path = tmp_path / "upsert_update_excluded.xlsx"
    _create_items_workbook(file_path, [[1, "Old", 30, "active", "A", 10]])

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO items (id, name, age, status, code, value) "
            "VALUES (1, 'Incoming', 31, 'active', 'A', 10) "
            "ON CONFLICT (id) DO UPDATE SET name = excluded.name"
        )
        assert cursor.rowcount == 1

    assert _fetch_rows(file_path, "SELECT id, name FROM items ORDER BY id") == [
        (1, "Incoming")
    ]


def test_upsert_do_update_literal_values(tmp_path: Path) -> None:
    file_path = tmp_path / "upsert_update_literal.xlsx"
    _create_items_workbook(file_path, [[1, "Old", 30, "active", "A", 10]])

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO items (id, name, age, status, code, value) "
            "VALUES (1, 'Incoming', 31, 'inactive', 'A', 10) "
            "ON CONFLICT (id) DO UPDATE SET status = 'updated'"
        )
        assert cursor.rowcount == 1

    assert _fetch_rows(file_path, "SELECT id, status FROM items ORDER BY id") == [
        (1, "updated")
    ]


def test_upsert_do_update_multi_column_target(tmp_path: Path) -> None:
    file_path = tmp_path / "upsert_update_multi_target.xlsx"
    _create_items_workbook(
        file_path,
        [[1, "Old", 30, "active", "A", 10], [1, "Second", 31, "active", "B", 11]],
    )

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO items (id, name, age, status, code, value) "
            "VALUES (1, 'Incoming', 99, 'inactive', 'A', 20) "
            "ON CONFLICT (id, code) DO UPDATE SET age = 50"
        )
        assert cursor.rowcount == 1

    assert _fetch_rows(file_path, "SELECT id, code, age FROM items ORDER BY code") == [
        (1, "A", 50),
        (1, "B", 31),
    ]


def test_upsert_do_update_multi_row_mixed(tmp_path: Path) -> None:
    file_path = tmp_path / "upsert_update_mixed.xlsx"
    _create_items_workbook(
        file_path,
        [[1, "Old1", 30, "active", "A", 10], [3, "Old3", 33, "active", "C", 30]],
    )

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO items (id, name, age, status, code, value) VALUES "
            "(1, 'New1', 35, 'active', 'A', 10), "
            "(2, 'New2', 20, 'active', 'B', 20), "
            "(3, 'New3', 40, 'active', 'C', 30) "
            "ON CONFLICT (id) DO UPDATE SET name = excluded.name"
        )
        assert cursor.rowcount == 3

    assert _fetch_rows(file_path, "SELECT id, name FROM items ORDER BY id") == [
        (1, "New1"),
        (2, "New2"),
        (3, "New3"),
    ]


def test_upsert_do_update_all_rows_conflict(tmp_path: Path) -> None:
    file_path = tmp_path / "upsert_update_all_conflict.xlsx"
    _create_items_workbook(
        file_path, [[1, "A", 30, "active", "A", 10], [2, "B", 20, "active", "B", 20]]
    )

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO items (id, name, age, status, code, value) VALUES "
            "(1, 'A2', 30, 'active', 'A', 10), "
            "(2, 'B2', 20, 'active', 'B', 20) "
            "ON CONFLICT (id) DO UPDATE SET name = excluded.name"
        )
        assert cursor.rowcount == 2

    assert _fetch_rows(file_path, "SELECT id, name FROM items ORDER BY id") == [
        (1, "A2"),
        (2, "B2"),
    ]


def test_upsert_do_nothing_all_rows_conflict(tmp_path: Path) -> None:
    file_path = tmp_path / "upsert_nothing_all_conflict.xlsx"
    _create_items_workbook(
        file_path, [[1, "A", 30, "active", "A", 10], [2, "B", 20, "active", "B", 20]]
    )

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO items (id, name, age, status, code, value) VALUES "
            "(1, 'A2', 30, 'active', 'A', 10), "
            "(2, 'B2', 20, 'active', 'B', 20) "
            "ON CONFLICT (id) DO NOTHING"
        )
        assert cursor.rowcount == 0

    assert _fetch_rows(file_path, "SELECT id, name FROM items ORDER BY id") == [
        (1, "A"),
        (2, "B"),
    ]


def test_upsert_executor_invalid_conflict_column(tmp_path: Path) -> None:
    file_path = tmp_path / "upsert_invalid_column.xlsx"
    _create_items_workbook(file_path, [[1, "A", 30, "active", "A", 10]])

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        with pytest.raises(
            ProgrammingError, match="ON CONFLICT column 'missing' not found in headers"
        ):
            cursor.execute(
                "INSERT INTO items (id, name, age, status, code, value) "
                "VALUES (1, 'X', 30, 'active', 'A', 1) "
                "ON CONFLICT (missing) DO NOTHING"
            )


def test_upsert_parameterized_values_and_set(tmp_path: Path) -> None:
    file_path = tmp_path / "upsert_params.xlsx"
    _create_items_workbook(file_path, [[1, "Old", 30, "active", "A", 10]])

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO items (id, name, age, status, code, value) VALUES (?, ?, ?, ?, ?, ?) "
            "ON CONFLICT (id) DO UPDATE SET name = ?, age = ?",
            (1, "Incoming", 31, "active", "A", 10, "BoundName", 88),
        )
        assert cursor.rowcount == 1

    assert _fetch_rows(file_path, "SELECT id, name, age FROM items ORDER BY id") == [
        (1, "BoundName", 88)
    ]


def test_insert_select_with_on_conflict(tmp_path: Path) -> None:
    file_path = tmp_path / "upsert_insert_select.xlsx"
    _create_items_workbook(
        file_path, [[1, "Old", 30, "active", "A", 10]], include_source=True
    )

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO items (id, name, age) SELECT id, name, age FROM incoming "
            "ON CONFLICT (id) DO UPDATE SET name = excluded.name"
        )
        assert cursor.rowcount == 2

    assert _fetch_rows(file_path, "SELECT id, name, age FROM items ORDER BY id") == [
        (1, "FromSource", 30),
        (2, "NewFromSource", 20),
    ]


def test_upsert_on_empty_table(tmp_path: Path) -> None:
    file_path = tmp_path / "upsert_empty_table.xlsx"
    _create_items_workbook(file_path, [])

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO items (id, name, age, status, code, value) "
            "VALUES (1, 'First', 21, 'active', 'A', 1) "
            "ON CONFLICT (id) DO NOTHING"
        )
        assert cursor.rowcount == 1

    assert _fetch_rows(file_path, "SELECT id, name FROM items ORDER BY id") == [
        (1, "First")
    ]


def test_upsert_do_update_multiple_set_assignments(tmp_path: Path) -> None:
    file_path = tmp_path / "upsert_multi_set.xlsx"
    _create_items_workbook(file_path, [[1, "Old", 30, "active", "A", 10]])

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO items (id, name, age, status, code, value) "
            "VALUES (1, 'Incoming', 45, 'active', 'B', 99) "
            "ON CONFLICT (id) DO UPDATE SET "
            "name = excluded.name, status = 'updated', code = excluded.code"
        )
        assert cursor.rowcount == 1

    assert _fetch_rows(
        file_path, "SELECT id, name, status, code FROM items ORDER BY id"
    ) == [(1, "Incoming", "updated", "B")]


def test_upsert_do_update_arithmetic_expression(tmp_path: Path) -> None:
    file_path = tmp_path / "upsert_expression.xlsx"
    _create_items_workbook(file_path, [[1, "Old", 30, "active", "A", 10]])

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO items (id, name, age, status, code, value) "
            "VALUES (1, 'Incoming', 31, 'active', 'A', 5) "
            "ON CONFLICT (id) DO UPDATE SET value = excluded.value + 1"
        )
        assert cursor.rowcount == 1

    assert _fetch_rows(file_path, "SELECT id, value FROM items ORDER BY id") == [(1, 6)]


# ── Oracle Review: NULL Conflict Semantics Tests ──


def test_do_nothing_with_null_target_value(tmp_path: Path) -> None:
    """NULL target values should never match: SQL NULL != NULL semantics."""
    file_path = tmp_path / "test.xlsx"
    _create_items_workbook(
        file_path,
        [
            [None, "Alice", 30, "active", "A", 1],
            [2, "Bob", 25, "active", "B", 2],
        ],
    )

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO items (id, name, age, status, code, value) "
            "VALUES (NULL, 'NullInsert', 40, 'active', 'C', 3) "
            "ON CONFLICT (id) DO NOTHING"
        )
        # NULL in target should NOT conflict with existing NULL — row should be inserted
        assert cursor.rowcount == 1

    rows = _fetch_rows(file_path, "SELECT name FROM items ORDER BY name")
    assert rows == [("Alice",), ("Bob",), ("NullInsert",)]


def test_do_update_with_null_target_value(tmp_path: Path) -> None:
    """DO UPDATE with NULL targets: both NULL rows should be inserted, not conflict."""
    file_path = tmp_path / "test.xlsx"
    _create_items_workbook(
        file_path,
        [
            [None, "Existing", 30, "active", "A", 1],
        ],
    )

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO items (id, name, age, status, code, value) "
            "VALUES (NULL, 'NewNull', 40, 'active', 'B', 2) "
            "ON CONFLICT (id) DO UPDATE SET name = excluded.name"
        )
        # NULL target should NOT conflict — row should be inserted, not updated
        assert cursor.rowcount == 1

    rows = _fetch_rows(file_path, "SELECT name FROM items ORDER BY name")
    assert rows == [("Existing",), ("NewNull",)]


# ── Oracle Review: Parameter Binding Order Tests ──


def test_insert_select_on_conflict_with_placeholders_in_select_and_set(
    tmp_path: Path,
) -> None:
    """Parameter order: SELECT placeholders first, then SET placeholders."""
    file_path = tmp_path / "test.xlsx"
    _create_items_workbook(
        file_path,
        [[1, "Alice", 30, "active", "A", 10]],
        include_source=True,
    )

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO items (id, name, age, status, code, value) "
            "SELECT id, name, age, ?, ?, ? FROM incoming "
            "ON CONFLICT (id) DO UPDATE SET status = ?, code = ?",
            ("pending", "X", 0, "updated", "Y"),
        )
        # id=1 conflicts (exists), id=2 new — SELECT params are first 3, SET params are last 2
        assert cursor.rowcount == 2

    rows = _fetch_rows(
        file_path, "SELECT id, name, status, code FROM items ORDER BY id"
    )
    assert rows == [
        (1, "Alice", "updated", "Y"),  # conflicted, DO UPDATE applied
        (2, "NewFromSource", "pending", "X"),  # inserted with SELECT params
    ]


# ── Oracle Review: Sanitize Formulas on DO UPDATE Path ──


def test_do_update_with_sanitize_formulas(tmp_path: Path) -> None:
    """Formula values via excluded.col should be sanitized when sanitize_formulas=True."""
    file_path = tmp_path / "test.xlsx"
    _create_items_workbook(
        file_path,
        [
            [1, "Alice", 30, "active", "A", 10],
        ],
    )

    with ExcelConnection(
        str(file_path), engine="openpyxl", autocommit=True, sanitize_formulas=True
    ) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO items (id, name, age, status, code, value) "
            "VALUES (1, '=HYPERLINK(\"http://evil.com\")', 30, 'active', 'A', 10) "
            "ON CONFLICT (id) DO UPDATE SET name = excluded.name"
        )
        assert cursor.rowcount == 1

    rows = _fetch_rows(file_path, "SELECT id, name FROM items")
    # Formula should be sanitized (prefixed with single quote)
    assert rows == [(1, '\'=HYPERLINK("http://evil.com")')]


# ── Oracle Review Round 2: Composite NULL + Bare Identifier Tests ──


def test_do_nothing_with_null_in_composite_target(tmp_path: Path) -> None:
    """Composite ON CONFLICT (id, code) where one target col is NULL => no conflict."""
    file_path = tmp_path / "test.xlsx"
    _create_items_workbook(
        file_path,
        [
            [1, "Alice", 30, "active", None, 10],
        ],
    )

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO items (id, name, age, status, code, value) "
            "VALUES (1, 'NewAlice', 40, 'active', NULL, 99) "
            "ON CONFLICT (id, code) DO NOTHING"
        )
        # NULL in 'code' column on both sides => no conflict (NULL != NULL)
        assert cursor.rowcount == 1

    rows = _fetch_rows(file_path, "SELECT id, name FROM items ORDER BY name")
    assert rows == [(1, "Alice"), (1, "NewAlice")]


def test_do_update_with_null_in_composite_target(tmp_path: Path) -> None:
    """Composite ON CONFLICT (id, code) where existing row has NULL code => insert, not update."""
    file_path = tmp_path / "test.xlsx"
    _create_items_workbook(
        file_path,
        [
            [1, "Alice", 30, "active", None, 10],
            [2, "Bob", 25, "active", "B", 20],
        ],
    )

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        # Row 1: id=1, code=NULL incoming vs id=1, code=NULL existing => NO conflict
        # Row 2: id=2, code='B' incoming vs id=2, code='B' existing => CONFLICT => UPDATE
        cursor.execute(
            "INSERT INTO items (id, name, age, status, code, value) VALUES "
            "(1, 'NewAlice', 40, 'active', NULL, 99), "
            "(2, 'NewBob', 35, 'active', 'B', 25) "
            "ON CONFLICT (id, code) DO UPDATE SET name = excluded.name"
        )
        assert cursor.rowcount == 2  # 1 insert + 1 update

    rows = _fetch_rows(file_path, "SELECT id, name, code FROM items ORDER BY id, name")
    assert rows == [
        (1, "Alice", None),  # original — not updated (NULL != NULL)
        (1, "NewAlice", None),  # inserted (no conflict)
        (2, "NewBob", "B"),  # updated (conflict on id=2, code='B')
    ]


def test_do_update_bare_identifier_stores_literal(tmp_path: Path) -> None:
    """SET col = name stores the string 'name' as literal, NOT a column reference."""
    file_path = tmp_path / "test.xlsx"
    _create_items_workbook(
        file_path,
        [
            [1, "Alice", 30, "active", "A", 10],
        ],
    )

    with ExcelConnection(str(file_path), engine="openpyxl", autocommit=True) as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO items (id, name, age, status, code, value) "
            "VALUES (1, 'Incoming', 40, 'active', 'A', 99) "
            "ON CONFLICT (id) DO UPDATE SET status = hello"
        )
        assert cursor.rowcount == 1

    rows = _fetch_rows(file_path, "SELECT id, status FROM items")
    # 'hello' is a bare identifier — treated as string literal per SQL_SPEC.md
    assert rows == [(1, "hello")]



def _create_round11_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Sheet1"
    sheet.append(["id", "name"])
    sheet.append([1, "Alice"])
    sheet.append([2, "Bob"])

    table = workbook.create_sheet("t")
    table.append(["id", "a", "b", "c"])
    table.append([1, 10, 0, None])
    table.append([2, "alice", 0, None])

    workbook.save(path)

def test_upsert_update_set_supports_expression_nodes(tmp_path: Path) -> None:
    file_path = tmp_path / "round11_upsert_expressions.xlsx"
    _create_round11_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO t (id, a, b, c) VALUES (1, 9, 0, 'incoming') "
            "ON CONFLICT (id) DO UPDATE SET b = a + 1, c = CAST(excluded.a AS TEXT)"
        )

        cursor.execute("SELECT a, b, c FROM t WHERE id = 1")
        assert cursor.fetchone() == (10, 11, "9")
