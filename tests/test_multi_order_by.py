from pathlib import Path

import pytest
from openpyxl import Workbook

from excel_dbapi.engines.openpyxl.backend import OpenpyxlBackend
from excel_dbapi.executor import SharedExecutor
from excel_dbapi.parser import parse_sql


@pytest.fixture
def multi_order_xlsx(tmp_path: Path) -> str:
    path = str(tmp_path / "test.xlsx")
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "users"
    ws.append(["name", "age", "city"])
    ws.append(["Alice", 30, "NYC"])
    ws.append(["Bob", 25, "LA"])
    ws.append(["Charlie", 30, "Chicago"])
    ws.append(["Diana", 25, "NYC"])
    ws.append(["Eve", 30, "LA"])
    wb.save(path)
    return path


def test_parser_order_by_single_column_defaults_asc() -> None:
    parsed = parse_sql("SELECT * FROM users ORDER BY name")
    assert parsed["order_by"] == [{"column": "name", "direction": "ASC"}]


def test_parser_order_by_single_column_desc() -> None:
    parsed = parse_sql("SELECT * FROM users ORDER BY name DESC")
    assert parsed["order_by"] == [{"column": "name", "direction": "DESC"}]


def test_parser_order_by_multiple_columns() -> None:
    parsed = parse_sql("SELECT * FROM users ORDER BY name DESC, age ASC")
    assert parsed["order_by"] == [
        {"column": "name", "direction": "DESC"},
        {"column": "age", "direction": "ASC"},
    ]


def test_parser_order_by_multiple_columns_mixed_defaults() -> None:
    parsed = parse_sql("SELECT * FROM users ORDER BY name DESC, age ASC, city")
    assert parsed["order_by"] == [
        {"column": "name", "direction": "DESC"},
        {"column": "age", "direction": "ASC"},
        {"column": "city", "direction": "ASC"},
    ]


def test_parser_order_by_second_column_defaults_asc() -> None:
    parsed = parse_sql("SELECT * FROM users ORDER BY name DESC, age")
    assert parsed["order_by"] == [
        {"column": "name", "direction": "DESC"},
        {"column": "age", "direction": "ASC"},
    ]


def test_parser_order_by_empty_clause_raises() -> None:
    with pytest.raises(ValueError, match="Invalid ORDER BY clause format"):
        parse_sql("SELECT * FROM users ORDER BY")


def test_parser_order_by_invalid_direction_raises() -> None:
    with pytest.raises(ValueError, match="Invalid ORDER BY direction"):
        parse_sql("SELECT * FROM users ORDER BY name SIDEWAYS")


def test_parser_compound_trailing_multi_order_by() -> None:
    parsed = parse_sql(
        "SELECT a, b FROM t1 UNION SELECT a, b FROM t2 ORDER BY a ASC, b DESC LIMIT 5"
    )
    assert parsed["action"] == "COMPOUND"
    assert parsed.get("order_by") == [
        {"column": "a", "direction": "ASC"},
        {"column": "b", "direction": "DESC"},
    ]
    assert parsed.get("limit") == 5


def test_executor_simple_select_multi_order_by_age_then_name(
    multi_order_xlsx: str,
) -> None:
    engine = OpenpyxlBackend(multi_order_xlsx)
    parsed = parse_sql("SELECT name, age FROM users ORDER BY age ASC, name DESC")
    result = SharedExecutor(engine).execute(parsed)
    assert result.rows == [
        ("Diana", 25),
        ("Bob", 25),
        ("Eve", 30),
        ("Charlie", 30),
        ("Alice", 30),
    ]


def test_executor_simple_select_multi_order_by_name_then_age(
    multi_order_xlsx: str,
) -> None:
    engine = OpenpyxlBackend(multi_order_xlsx)
    parsed = parse_sql("SELECT name, age FROM users ORDER BY name ASC, age DESC")
    result = SharedExecutor(engine).execute(parsed)
    assert result.rows == [
        ("Alice", 30),
        ("Bob", 25),
        ("Charlie", 30),
        ("Diana", 25),
        ("Eve", 30),
    ]


def test_executor_join_multi_order_by(tmp_path: Path) -> None:
    path = str(tmp_path / "join_multi_order.xlsx")
    wb = Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "t1"
    ws1.append(["id", "name"])
    ws1.append([1, "Alice"])
    ws1.append([2, "Alice"])
    ws1.append([3, "Bob"])
    ws1.append([4, "Bob"])

    ws2 = wb.create_sheet("t2")
    ws2.append(["id", "amount"])
    ws2.append([1, 100])
    ws2.append([2, 150])
    ws2.append([3, 90])
    ws2.append([4, 110])
    wb.save(path)

    engine = OpenpyxlBackend(path)
    parsed = parse_sql(
        "SELECT t1.name, t2.amount FROM t1 INNER JOIN t2 ON t1.id = t2.id "
        "ORDER BY t1.name ASC, t2.amount DESC"
    )
    result = SharedExecutor(engine).execute(parsed)
    assert result.rows == [
        ("Alice", 150),
        ("Alice", 100),
        ("Bob", 110),
        ("Bob", 90),
    ]


def test_executor_group_by_multi_order_by_aggregate(tmp_path: Path) -> None:
    path = str(tmp_path / "group_multi_order.xlsx")
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "t"
    ws.append(["name"])
    ws.append(["Alice"])
    ws.append(["Bob"])
    ws.append(["Alice"])
    ws.append(["Cara"])
    ws.append(["Bob"])
    ws.append(["Bob"])
    wb.save(path)

    engine = OpenpyxlBackend(path)
    parsed = parse_sql(
        "SELECT name, COUNT(*) FROM t GROUP BY name ORDER BY COUNT(*) DESC, name ASC"
    )
    result = SharedExecutor(engine).execute(parsed)
    assert result.rows == [("Bob", 3), ("Alice", 2), ("Cara", 1)]


def test_executor_compound_multi_order_by(tmp_path: Path) -> None:
    path = str(tmp_path / "compound_multi_order.xlsx")
    wb = Workbook()
    ws1 = wb.active
    assert ws1 is not None
    ws1.title = "t1"
    ws1.append(["col1", "col2"])
    ws1.append([1, 10])
    ws1.append([1, 20])
    ws1.append([2, 5])

    ws2 = wb.create_sheet("t2")
    ws2.append(["col1", "col2"])
    ws2.append([1, 15])
    ws2.append([2, 7])
    ws2.append([2, 6])
    wb.save(path)

    engine = OpenpyxlBackend(path)
    parsed = parse_sql(
        "SELECT col1, col2 FROM t1 UNION ALL SELECT col1, col2 FROM t2 "
        "ORDER BY col1 ASC, col2 DESC"
    )
    result = SharedExecutor(engine).execute(parsed)
    assert result.rows == [(1, 20), (1, 15), (1, 10), (2, 7), (2, 6), (2, 5)]


def test_executor_unknown_column_in_second_order_position_raises(
    multi_order_xlsx: str,
) -> None:
    engine = OpenpyxlBackend(multi_order_xlsx)
    parsed = parse_sql("SELECT name, age FROM users ORDER BY age ASC, missing DESC")
    with pytest.raises(ValueError, match="Unknown column: missing"):
        SharedExecutor(engine).execute(parsed)


def test_executor_single_column_order_by_backward_compat_dict_shape(
    multi_order_xlsx: str,
) -> None:
    engine = OpenpyxlBackend(multi_order_xlsx)
    parsed = parse_sql("SELECT name, age FROM users")
    parsed["order_by"] = {"column": "age", "direction": "DESC"}
    result = SharedExecutor(engine).execute(parsed)
    assert result.rows == [
        ("Alice", 30),
        ("Charlie", 30),
        ("Eve", 30),
        ("Bob", 25),
        ("Diana", 25),
    ]
