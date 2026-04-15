from datetime import date
from pathlib import Path
from typing import Any
import pathlib

from openpyxl import Workbook
import pytest

from excel_dbapi.connection import ExcelConnection
from excel_dbapi.engines.base import TableData, WorkbookBackend
from excel_dbapi.engines.openpyxl.backend import OpenpyxlBackend
from excel_dbapi.engines.result import ExecutionResult
from excel_dbapi.exceptions import DatabaseError, ProgrammingError
from excel_dbapi.executor import SharedExecutor, SharedExecutor as Executor
from excel_dbapi.parser import (
    _expression_to_sql_for_order_by,
    _validate_join_column_reference,
    _validate_join_where_node,
    parse_sql,
)


def _create_workbook(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "t"
    sheet.append(["id", "name", "txt", "d", "v"])
    sheet.append([1, "alpha", "10", date(2024, 1, 15), 10])
    sheet.append([2, "beta", "x", date(2024, 1, 16), 20])
    sheet.append([3, None, None, None, None])

    u = workbook.create_sheet("u")
    u.append(["id", "v2"])
    u.append([1, 100])
    u.append([2, 200])

    empty = workbook.create_sheet("empty")
    empty.append(["id", "value"])

    workbook.save(path)


def test_scalar_edges_and_date_part_null_paths(tmp_path: Path) -> None:
    file_path = tmp_path / "cov_scalar_edges.xlsx"
    _create_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT "
            "COALESCE(NULL, NULL), "
            "SUBSTR('hello', -2), "
            "SUBSTR('hello', 0, 2), "
            "SUBSTR('hello', 2, 0), "
            "SUBSTR('hello', 2, NULL), "
            "YEAR(NULL), MONTH(NULL), DAY(NULL), "
            "YEAR(d), MONTH('2024-01-31'), DAY('2024-01-15T11:22:33Z') "
            "FROM t WHERE id = 1"
        )
        assert cur.fetchall() == [
            (None, "lo", "he", "", None, None, None, None, 2024, 1, 15)
        ]


def test_scalar_argument_validation_paths(tmp_path: Path) -> None:
    file_path = tmp_path / "cov_scalar_validate.xlsx"
    _create_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cur = conn.cursor()
        with pytest.raises(ProgrammingError, match="Invalid arguments for SUBSTR"):
            cur.execute("SELECT SUBSTR('abc', '', 1) FROM t")

        with pytest.raises(ProgrammingError, match="Unsupported function"):
            cur.execute("SELECT MADE_UP_FN(1) FROM t")

        with pytest.raises(ProgrammingError, match="expects at least"):
            conn._executor._call_function("LENGTH", [])

        with pytest.raises(ProgrammingError, match="expects at most"):
            conn._executor._call_function("NULLIF", [1, 1, 1])

        with pytest.raises(ProgrammingError, match="Invalid arguments for SUBSTR"):
            conn._executor._call_function("SUBSTR", ["abc", True, 1])


def test_like_escape_error_paths(tmp_path: Path) -> None:
    file_path = tmp_path / "cov_like_escape.xlsx"
    _create_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cur = conn.cursor()
        with pytest.raises(
            ProgrammingError, match="ESCAPE requires a single character"
        ):
            cur.execute("SELECT id FROM t WHERE name LIKE 'a%' ESCAPE 'xx'")

        with pytest.raises(ProgrammingError, match="trailing ESCAPE character"):
            cur.execute("SELECT id FROM t WHERE name LIKE 'abc!' ESCAPE '!'")


def test_cast_error_and_conversion_paths(tmp_path: Path) -> None:
    file_path = tmp_path / "cov_cast_paths.xlsx"
    _create_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cur = conn.cursor()

        with pytest.raises(ProgrammingError, match="Cannot cast empty string to DATE"):
            cur.execute("SELECT CAST('' AS DATE) FROM t")

        cur.execute("SELECT CAST('2024-01-15Z' AS DATE) FROM t LIMIT 1")
        assert cur.fetchone() == (date(2024, 1, 15),)

        with pytest.raises(ProgrammingError, match="Unsupported CAST target type"):
            cur.execute("SELECT CAST(1 AS BLOB) FROM t")


def test_window_rank_without_order_and_running_sum(tmp_path: Path) -> None:
    file_path = tmp_path / "cov_window_paths.xlsx"
    _create_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cur = conn.cursor()
        cur.execute("SELECT id, RANK() OVER () FROM t ORDER BY id")
        assert cur.fetchall() == [(1, 1), (2, 1), (3, 1)]

        cur.execute("SELECT id, SUM(v) OVER (ORDER BY id) FROM t ORDER BY id")
        assert cur.fetchall() == [(1, 10.0), (2, 30.0), (3, 30.0)]


def test_window_programming_error_paths(tmp_path: Path) -> None:
    file_path = tmp_path / "cov_window_errors.xlsx"
    _create_workbook(file_path)
    backend = OpenpyxlBackend(str(file_path))
    executor = SharedExecutor(backend)
    rows = [
        {"t.id": 1, "id": 1, "v": 10},
        {"t.id": 2, "id": 2, "v": 20},
    ]

    with pytest.raises(ProgrammingError, match="requires an argument"):
        executor._apply_window_functions(
            rows,
            [{"type": "window_function", "func": "SUM", "args": [], "order_by": []}],
            None,
        )

    with pytest.raises(ProgrammingError, match="Unsupported window function"):
        executor._apply_window_functions(
            rows,
            [
                {
                    "type": "window_function",
                    "func": "MEDIAN",
                    "args": ["v"],
                    "order_by": [],
                }
            ],
            None,
        )


def test_expression_and_where_sql_helpers() -> None:
    window_expr = {
        "type": "window_function",
        "func": "SUM",
        "args": ["t.v"],
        "distinct": True,
        "filter": {"column": "t.id", "operator": ">", "value": 0},
        "order_by": [
            "bad-order-item",
            {"__expression__": {"type": "column", "source": "t", "name": "id"}},
            {"column": "__expr__:t.v + 1", "direction": "DESC"},
        ],
    }
    sql = SharedExecutor._expression_to_sql(window_expr)
    assert "DISTINCT" in sql
    assert "FILTER" in sql
    assert "OVER" in sql

    case_sql = SharedExecutor._expression_to_sql(
        {
            "type": "case",
            "mode": "searched",
            "whens": [
                {
                    "condition": {"type": "exists"},
                    "result": {"type": "literal", "value": 1},
                }
            ],
            "else": {"type": "literal", "value": 0},
        }
    )
    assert case_sql.startswith("CASE")

    assert (
        SharedExecutor._where_operand_to_sql({"type": "exists"}, is_column=False)
        == "EXISTS (SUBQUERY)"
    )
    assert SharedExecutor._where_to_sql({"type": "exists"}) == "EXISTS (SUBQUERY)"
    assert SharedExecutor._contains_arithmetic_expression({"type": "mystery"}) is False


def test_collect_expression_refs_paths() -> None:
    expr = {
        "type": "window_function",
        "args": ["t.v", "*", {"type": "column", "source": "u", "name": "v2"}],
        "partition_by": [{"type": "column", "source": "t", "name": "id"}],
        "order_by": [
            "noise",
            {"__expression__": {"type": "column", "source": "t", "name": "v"}},
            {"column": "t.id"},
        ],
        "filter": {
            "column": {"type": "column", "source": "u", "name": "id"},
            "operator": ">",
            "value": 0,
        },
    }
    refs = SharedExecutor._collect_expression_column_refs(expr)
    assert "t.v" in refs
    assert "u.v2" in refs
    assert "t.id" in refs
    assert "u.id" in refs

    aggregate_refs = SharedExecutor._collect_expression_column_refs(
        {
            "type": "aggregate",
            "arg": "t.v",
            "filter": {"type": "exists"},
        }
    )
    assert aggregate_refs == {"t.v"}


def test_collect_where_refs_paths() -> None:
    where = {
        "conditions": [
            {"type": "exists"},
            {
                "column": {"type": "column", "source": "t", "name": "id"},
                "operator": "IN",
                "value": [
                    {"type": "column", "source": "u", "name": "id"},
                    1,
                ],
            },
        ]
    }
    refs = SharedExecutor._collect_where_column_refs(where)
    assert "t.id" in refs
    assert "u.id" in refs


def test_aggregate_select_validation_paths(tmp_path: Path) -> None:
    file_path = tmp_path / "cov_agg_validation.xlsx"
    _create_workbook(file_path)
    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cur = conn.cursor()

        with pytest.raises(ProgrammingError, match="Non-aggregate columns"):
            cur.execute("SELECT id, COUNT(*) FROM t")

        with pytest.raises(ProgrammingError, match=r"Unknown column\(s\): nope"):
            cur.execute("SELECT SUM(v) FROM t GROUP BY nope")

        with pytest.raises(ProgrammingError, match="Unknown column: t.nope"):
            cur.execute("SELECT SUM(t.nope) FROM t")

        with pytest.raises(ProgrammingError, match="must be a GROUP BY column"):
            cur.execute("SELECT id, COUNT(*) FROM t GROUP BY id HAVING name = 'alpha'")


def test_aggregate_distinct_dedupe_path(tmp_path: Path) -> None:
    file_path = tmp_path / "cov_agg_distinct.xlsx"
    _create_workbook(file_path)
    backend = OpenpyxlBackend(str(file_path))
    executor = SharedExecutor(backend)
    data = backend.read_sheet("t")
    headers = list(data.headers)
    rows = [
        executor._build_scoped_row(executor._row_from_values(headers, list(r)))
        for r in data.rows
    ]
    result = executor._execute_aggregate_select(
        "SELECT",
        {"distinct": True, "order_by": None, "limit": None, "offset": None},
        headers,
        rows,
        [{"type": "aggregate", "func": "COUNT", "arg": "*"}],
        None,
        None,
    )
    assert result.rows == [(3,)]


def test_join_validation_paths_with_custom_ast(tmp_path: Path) -> None:
    file_path = tmp_path / "cov_join_validation.xlsx"
    _create_workbook(file_path)
    backend = OpenpyxlBackend(str(file_path))
    executor = SharedExecutor(backend)
    left_data = backend.read_sheet("t")

    parsed = {
        "action": "SELECT",
        "table": "t",
        "from": {"table": "t", "ref": "t"},
        "joins": [
            {
                "type": "INNER",
                "source": {"table": "u", "ref": "u"},
                "on": {"column": "t.id", "operator": "=", "value": "u.id"},
            }
        ],
        "columns": [
            None,
            {"type": "literal", "value": 1},
            {
                "type": "unary_op",
                "operand": {"type": "column", "source": "t", "name": "v"},
            },
            {
                "type": "binary_op",
                "left": {"type": "column", "source": "t", "name": "v"},
                "right": {"type": "column", "source": "u", "name": "v2"},
            },
            {
                "type": "function",
                "args": [{"type": "column", "source": "t", "name": "id"}],
            },
            {"type": "cast", "value": {"type": "column", "source": "u", "name": "v2"}},
            {
                "type": "case",
                "mode": "simple",
                "value": {"type": "column", "source": "t", "name": "id"},
                "whens": [
                    "skip",
                    {
                        "match": {"type": "column", "source": "u", "name": "id"},
                        "result": {"type": "literal", "value": 1},
                    },
                ],
                "else": {"type": "subquery"},
            },
            {
                "type": "window_function",
                "args": ["*", "t.id", {"type": "column", "source": "u", "name": "v2"}],
                "partition_by": [{"type": "column", "source": "t", "name": "id"}],
                "order_by": [
                    "junk",
                    {"__expression__": {"type": "column", "source": "t", "name": "id"}},
                    {"column": "__expr__:t.id + 1"},
                    {"column": "t.id"},
                ],
                "filter": {"column": "u.id", "operator": ">", "value": 0},
            },
        ],
        "where": None,
        "group_by": None,
        "having": None,
        "order_by": [{"column": "t.id", "direction": "ASC"}],
    }

    with pytest.raises(Exception):
        executor._execute_join_select("SELECT", parsed, "t", left_data)


def test_misc_internal_paths() -> None:
    backend = OpenpyxlBackend("tests/data/sample.xlsx")
    executor = SharedExecutor(backend)

    assert (
        SharedExecutor._normalize_single_source_aggregate_arg("t.value", ["a.b"])
        == "t.value"
    )
    assert (
        SharedExecutor._normalize_single_source_aggregate_arg("t.value", ["value"])
        == "value"
    )
    assert (
        SharedExecutor._normalize_single_source_aggregate_arg("t.value", ["x"])
        == "t.value"
    )

    assert SharedExecutor._output_name({"type": "case"}) == "case_expr"
    assert SharedExecutor._source_key(
        {"type": "binary_op", "left": 1, "right": 2, "op": "+"}
    ).startswith("__expr__:")

    assert executor._sort_key(object())[0] == 0
    assert SharedExecutor._coerce_temporal_value(date(2024, 1, 1)) is not None
    assert SharedExecutor._parse_datetime_string("") is None
    assert SharedExecutor._parse_datetime_string("2024-01-01Z") is not None
    assert SharedExecutor._parse_datetime_string("2024-01-01") is not None

    row = executor._build_scoped_row({"id": 1}, headers=["id", "v"], source_refs={"t"})
    assert row["t.id"] == 1
    assert row["t.v"] is None

    with pytest.raises(DatabaseError, match="Unknown source reference"):
        executor._resolve_join_column({"t": {"id": 1}}, {"source": "x", "name": "id"})

    assert (
        executor._matches_join_on_condition({"t": {"id": 1}}, {"u": {"id": 1}}, None)
        is True
    )


def test_missing_sheet_errors_include_available_names(tmp_path: Path) -> None:
    file_path = tmp_path / "cov_missing_sheet.xlsx"
    _create_workbook(file_path)

    with ExcelConnection(str(file_path), engine="openpyxl") as conn:
        cur = conn.cursor()
        with pytest.raises(ProgrammingError, match="Available sheets"):
            cur.execute("SELECT * FROM nope")
        with pytest.raises(ProgrammingError, match="Available sheets"):
            cur.execute("UPDATE nope SET id = 1")
        with pytest.raises(ProgrammingError, match="Available sheets"):
            cur.execute("DELETE FROM nope")
        with pytest.raises(ProgrammingError, match="Available sheets"):
            cur.execute("INSERT INTO nope VALUES (1)")


def test_insert_and_alter_validation_paths(tmp_path: Path) -> None:
    file_path = tmp_path / "cov_insert_alter.xlsx"
    _create_workbook(file_path)
    backend = OpenpyxlBackend(str(file_path))
    executor = SharedExecutor(backend)

    parsed_insert = {
        "action": "INSERT",
        "table": "t",
        "values": 123,
    }
    with pytest.raises(DatabaseError, match="Invalid INSERT values format"):
        executor.execute(parsed_insert)

    parsed_insert_select = {
        "action": "INSERT",
        "table": "t",
        "columns": ["id"],
        "values": {
            "type": "subquery",
            "query": {
                "action": "SELECT",
                "table": "t",
                "from": {"table": "t", "ref": "t"},
                "columns": ["id", "v"],
                "where": None,
                "order_by": None,
                "limit": None,
                "offset": None,
                "distinct": False,
            },
        },
    }
    with pytest.raises(DatabaseError, match="column count mismatch"):
        executor.execute(parsed_insert_select)

    parsed_on_conflict_bad_action = {
        "action": "INSERT",
        "table": "t",
        "columns": ["id", "name", "txt", "d", "v"],
        "values": [[1, "x", "1", None, 10]],
        "on_conflict": {
            "target_columns": ["id"],
            "action": "merge",
            "set": [],
        },
    }
    with pytest.raises(DatabaseError, match="Invalid ON CONFLICT action"):
        executor.execute(parsed_on_conflict_bad_action)

    parsed_alter_bad_op = {
        "action": "ALTER",
        "table": "t",
        "operation": "NOPE",
    }
    with pytest.raises(DatabaseError, match="Unsupported ALTER operation"):
        executor.execute(parsed_alter_bad_op)


def test_direct_helper_branches_and_subquery_errors(tmp_path: Path) -> None:
    file_path = tmp_path / "cov_helper_branches.xlsx"
    _create_workbook(file_path)
    backend = OpenpyxlBackend(str(file_path))
    executor = SharedExecutor(backend)

    with pytest.raises(ProgrammingError, match="expected numeric value"):
        executor._call_function("SUBSTR", ["abc", " ", 1])

    assert executor._call_function("SUBSTR", [None, 1]) is None
    assert executor._call_function("YEAR", [date(2024, 1, 1)]) == 2024
    with pytest.raises(ProgrammingError, match="Invalid arguments for YEAR"):
        executor._call_function("YEAR", [""])
    with pytest.raises(ProgrammingError, match="Invalid arguments for DAY"):
        executor._call_function("DAY", [object()])

    with pytest.raises(DatabaseError, match="Invalid CTE definition"):
        executor.execute(
            {"ctes": [{"name": 1, "query": {}}], "action": "SELECT", "table": "t"}
        )

    with pytest.raises(ProgrammingError, match="Unsupported subquery node type"):
        executor._eval_subquery({"type": "weird"})
    with pytest.raises(ProgrammingError, match="missing parsed query"):
        executor._eval_subquery({"type": "subquery", "mode": "scalar", "query": None})
    with pytest.raises(ProgrammingError, match="Unsupported subquery mode"):
        executor._eval_subquery(
            {
                "type": "subquery",
                "mode": "bad",
                "query": {
                    "action": "SELECT",
                    "table": "t",
                    "from": {"table": "t", "ref": "t"},
                    "columns": ["id"],
                    "where": None,
                    "order_by": None,
                    "limit": None,
                    "offset": None,
                    "distinct": False,
                },
            }
        )


def test_condition_and_expression_normalization_branches(tmp_path: Path) -> None:
    file_path = tmp_path / "cov_condition_branches.xlsx"
    _create_workbook(file_path)
    backend = OpenpyxlBackend(str(file_path))
    executor = SharedExecutor(backend)

    row = {"x": 2, "vals": [1, 2, 3], "other": (4, 5), "z": "alpha"}
    assert (
        executor._evaluate_condition(
            row, {"column": "x", "operator": "IN", "value": None}
        )
        is False
    )
    assert (
        executor._evaluate_condition(
            row, {"column": "x", "operator": "IN", "value": [1, 2, 3]}
        )
        is True
    )
    assert (
        executor._evaluate_condition(
            row, {"column": "x", "operator": "NOT IN", "value": None}
        )
        is True
    )
    assert (
        executor._evaluate_condition(
            row, {"column": "x", "operator": "NOT IN", "value": (2, 9)}
        )
        is False
    )

    with pytest.raises(DatabaseError, match="single character"):
        executor._evaluate_condition(
            row,
            {"column": "z", "operator": "LIKE", "value": "a%", "escape": "!!"},
        )

    assert executor._eval_cast(date(2024, 1, 1), "TEXT") == "2024-01-01"
    assert executor._eval_cast("2024-01-02", "DATE") == date(2024, 1, 2)
    with pytest.raises(ProgrammingError, match="Cannot cast value"):
        executor._eval_cast(123, "DATE")

    assert executor._eval_expression(123, {}, lambda c: c) == 123
    assert (
        executor._eval_expression(
            {"type": "alias", "expression": {"type": "literal", "value": 7}},
            {},
            lambda c: c,
        )
        == 7
    )
    with pytest.raises(ProgrammingError, match="missing source"):
        executor._eval_expression({"type": "column", "name": "id"}, {}, lambda c: c)
    with pytest.raises(ProgrammingError, match="requires numeric operands"):
        executor._eval_expression(
            {
                "type": "unary_op",
                "op": "-",
                "operand": {"type": "literal", "value": "x"},
            },
            {},
            lambda c: c,
        )


def test_join_empty_headers_and_resolve_join_column_success(tmp_path: Path) -> None:
    file_path = tmp_path / "cov_join_empty_headers.xlsx"
    workbook = Workbook()
    left = workbook.active
    assert left is not None
    left.title = "left"
    workbook.create_sheet("right")
    workbook.save(file_path)

    backend = OpenpyxlBackend(str(file_path))
    executor = SharedExecutor(backend)
    left_data = TableData(headers=[], rows=[])
    parsed = {
        "from": {"table": "left", "ref": "left"},
        "joins": [],
        "columns": ["*"],
        "where": None,
        "order_by": None,
    }
    result = executor._execute_join_select("SELECT", parsed, "left", left_data)
    assert result.rows == []

    assert (
        executor._resolve_join_column({"l": {"id": 1}}, {"source": "l", "name": "id"})
        == 1
    )



def test_update_missing_table_shows_available(tmp_path: object) -> None:

    file_path = pathlib.Path(str(tmp_path)) / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id"])
    ws.append([1])
    wb.save(file_path)

    with ExcelConnection(str(file_path)) as conn:
        cursor = conn.cursor()
        with pytest.raises(Exception, match="Available sheets"):
            cursor.execute("UPDATE NonExistent SET id = 1")

def test_select_missing_table_shows_available(tmp_path: object) -> None:

    file_path = pathlib.Path(str(tmp_path)) / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id"])
    ws.append([1])
    wb.save(file_path)

    with ExcelConnection(str(file_path)) as conn:
        cursor = conn.cursor()
        with pytest.raises(Exception, match="Available sheets"):
            cursor.execute("SELECT * FROM NonExistent")

def test_delete_missing_table_shows_available(tmp_path: object) -> None:

    file_path = pathlib.Path(str(tmp_path)) / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id"])
    ws.append([1])
    wb.save(file_path)

    with ExcelConnection(str(file_path)) as conn:
        cursor = conn.cursor()
        with pytest.raises(Exception, match="Available sheets"):
            cursor.execute("DELETE FROM NonExistent")

def test_insert_invalid_subquery_format(tmp_path: object) -> None:
    """INSERT with dict values that isn't a proper subquery."""

    file_path = pathlib.Path(str(tmp_path)) / "test.xlsx"
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id", "name"])
    wb.save(file_path)

    backend = OpenpyxlBackend(str(file_path))
    executor = Executor(backend)
    # Directly call execute with a crafted parsed dict
    with pytest.raises(DatabaseError, match="Invalid INSERT subquery format"):
        executor.execute(
            {
                "action": "INSERT",
                "table": "Sheet1",
                "columns": None,
                "values": {"type": "invalid"},
            }
        )
    backend.close()



def _make_join_workbook(path: Path) -> OpenpyxlBackend:
    workbook = Workbook()
    left = workbook.active
    assert left is not None
    left.title = "t1"
    left.append(["id", "name", "value"])
    left.append([1, "A", 10])
    left.append([2, "B", 20])

    right = workbook.create_sheet("t2")
    right.append(["id", "label", "score"])
    right.append([1, "x", 5])
    right.append([2, "y", 7])

    workbook.save(path)
    return OpenpyxlBackend(str(path))

def test_execute_compound_structure_guards(tmp_path: Path) -> None:
    backend = _make_join_workbook(tmp_path / "compound_guards.xlsx")
    executor = SharedExecutor(backend)
    with pytest.raises(DatabaseError, match="at least one SELECT"):
        executor._execute_compound({"queries": []})
    with pytest.raises(DatabaseError, match="valid operator"):
        executor._execute_compound({"queries": [{"action": "SELECT"}]})
    with pytest.raises(DatabaseError, match="Invalid COMPOUND query structure"):
        executor._execute_compound(
            {
                "queries": [
                    parse_sql("SELECT id FROM t1"),
                    parse_sql("SELECT id FROM t2"),
                ],
                "operators": [],
            }
        )
    backend.close()

def test_execute_compound_invalid_operator_and_order_resolution(tmp_path: Path) -> None:
    backend = _make_join_workbook(tmp_path / "compound_ops.xlsx")
    executor = SharedExecutor(backend)

    with pytest.raises(DatabaseError, match="Unsupported compound operator"):
        executor.execute(
            {
                "action": "COMPOUND",
                "operators": ["MERGE"],
                "queries": [
                    parse_sql("SELECT id FROM t1"),
                    parse_sql("SELECT id FROM t2"),
                ],
            }
        )

    with pytest.raises(DatabaseError, match="ORDER BY column 'missing' not found"):
        executor.execute(
            parse_sql("SELECT id FROM t1 UNION SELECT id FROM t2 ORDER BY missing")
        )

    rows = executor.execute(
        parse_sql("SELECT id FROM t1 UNION ALL SELECT id FROM t2 OFFSET 1")
    ).rows
    assert rows == [(2,), (1,), (2,)]
    backend.close()

def test_executor_expression_serialization_case_and_where_shapes() -> None:
    simple_case = {
        "type": "case",
        "mode": "simple",
        "value": {"type": "column", "source": "t1", "name": "id"},
        "whens": [
            {
                "match": {"type": "literal", "value": 1},
                "result": {"type": "literal", "value": "one"},
            }
        ],
        "else": {"type": "literal", "value": "other"},
    }
    searched_case = {
        "type": "case",
        "mode": "searched",
        "whens": [
            {
                "condition": {
                    "type": "not",
                    "operand": {
                        "type": "compound",
                        "conditions": [
                            {"column": "t1.id", "operator": "IN", "value": [1, 2]}
                        ],
                        "conjunctions": ["AND"],
                    },
                },
                "result": {"type": "literal", "value": "x"},
            }
        ],
        "else": {"type": "literal", "value": "y"},
    }

    assert SharedExecutor._output_name(simple_case) == "case_expr"
    assert SharedExecutor._source_key(
        {"type": "alias", "alias": "k", "expression": simple_case}
    ).startswith("__expr__:")
    assert (
        SharedExecutor._expression_to_sql(
            {
                "type": "alias",
                "expression": {"type": "aggregate", "func": "COUNT", "arg": "*"},
            }
        )
        == "COUNT(*)"
    )

    assert "NOT" in SharedExecutor._where_to_sql(
        {"type": "not", "operand": {"column": "id", "operator": "=", "value": 1}}
    )
    assert SharedExecutor._where_to_sql({"type": "not", "operand": "bad"}) == "NOT"
    assert SharedExecutor._where_to_sql(
        {
            "type": "compound",
            "conditions": [{"column": "id", "operator": "=", "value": 1}],
            "conjunctions": ["AND"],
        }
    ).startswith("(")
    assert "(SUBQUERY)" in SharedExecutor._where_operand_to_sql(
        {"type": "subquery"}, is_column=False
    )
    assert "BETWEEN" in SharedExecutor._where_to_sql(
        {"column": "id", "operator": "BETWEEN", "value": [1]}
    )
    assert "CASE" in SharedExecutor._expression_to_sql(searched_case)

def test_executor_collect_expression_and_where_refs_paths() -> None:
    expr = {
        "type": "alias",
        "alias": "z",
        "expression": {
            "type": "binary_op",
            "op": "+",
            "left": {"type": "column", "source": "t1", "name": "id"},
            "right": {"type": "literal", "value": 1},
        },
    }
    where = {
        "type": "not",
        "operand": {
            "conditions": [
                {
                    "column": {"type": "column", "source": "t1", "name": "id"},
                    "operator": "IN",
                    "value": [
                        {"type": "column", "source": "t2", "name": "id"},
                        {"type": "literal", "value": 2},
                    ],
                }
            ]
        },
    }

    assert SharedExecutor._contains_arithmetic_expression(
        {
            "type": "alias",
            "alias": "a",
            "expression": {"type": "alias", "alias": "b", "expression": expr},
        }
    )
    assert SharedExecutor._collect_expression_column_refs("t1.id") == {"t1.id"}
    assert SharedExecutor._collect_expression_column_refs(3.14) == set()
    refs = SharedExecutor._collect_where_column_refs(where)
    assert "t1.id" in refs
    assert "t2.id" in refs

def test_executor_apply_order_by_and_sort_key_edges(tmp_path: Path) -> None:
    backend = _make_join_workbook(tmp_path / "sort_edges.xlsx")
    executor = SharedExecutor(backend)

    assert executor._apply_order_by([], None, value_getter=lambda row, col: None) == []

    single = [{"id": 1}]
    ordered = executor._apply_order_by(
        single,
        [{"column": "id", "direction": "ASC"}],
        value_getter=lambda row, col: row[col],
        available_columns={"id"},
    )
    assert ordered == single

    assert executor._sort_key(None) == (1, (0, ""))
    assert executor._sort_key("12") == (0, (0, 12.0))
    assert executor._sort_key(True) == (0, (1, 1))
    backend.close()

def test_executor_validate_join_where_node_case_branches() -> None:
    seen: list[tuple[str, str, str]] = []

    def _validate(source: str, name: str, context: str) -> None:
        seen.append((source, name, context))

    node = {
        "type": "not",
        "operand": {
            "conditions": [
                {
                    "column": {
                        "type": "case",
                        "mode": "searched",
                        "whens": [
                            "skip",
                            {
                                "condition": {
                                    "column": {
                                        "type": "column",
                                        "source": "t1",
                                        "name": "id",
                                    },
                                    "operator": "=",
                                    "value": {
                                        "type": "column",
                                        "source": "t2",
                                        "name": "id",
                                    },
                                },
                                "result": {"type": "literal", "value": 1},
                            },
                        ],
                    },
                    "operator": "IN",
                    "value": [
                        {
                            "type": "alias",
                            "alias": "a",
                            "expression": {
                                "type": "column",
                                "source": "t2",
                                "name": "score",
                            },
                        },
                        None,
                    ],
                }
            ]
        },
    }
    SharedExecutor._validate_join_where_node(node, _validate)
    assert ("t1", "id", "WHERE") in seen
    assert ("t2", "id", "WHERE") in seen
    assert ("t2", "score", "WHERE") in seen

def test_executor_join_validation_and_case_expression_errors(tmp_path: Path) -> None:
    backend = _make_join_workbook(tmp_path / "join_validate.xlsx")
    executor = SharedExecutor(backend)

    with pytest.raises(DatabaseError, match="Aggregate arguments in JOIN queries must be qualified"):
        executor.execute(parse_sql("SELECT SUM(id) FROM t1 a JOIN t2 b ON a.id = b.id"))

    with pytest.raises(DatabaseError, match="requires qualified column names in JOIN queries"):
        executor.execute(
            parse_sql(
                "SELECT CASE id WHEN 1 THEN 'x' ELSE 'y' END FROM t1 a JOIN t2 b ON a.id = b.id"
            )
        )

    with pytest.raises(DatabaseError, match="Invalid source reference in WHERE"):
        executor.execute(
            parse_sql("SELECT a.id FROM t1 a JOIN t2 b ON a.id = b.id WHERE c.id = 1")
        )

    backend.close()

def test_executor_eval_expression_and_condition_error_paths(tmp_path: Path) -> None:
    backend = _make_join_workbook(tmp_path / "eval_expr.xlsx")
    executor = SharedExecutor(backend)
    row = {"t1.id": 1, "t1.value": "text", "x": None}

    with pytest.raises(ProgrammingError, match="Unsupported unary operator"):
        executor._eval_expression(
            {"type": "unary_op", "op": "~", "operand": {"type": "literal", "value": 1}},
            row,
            lambda c: row.get(c),
        )
    with pytest.raises(ProgrammingError, match="requires numeric operands"):
        executor._eval_expression(
            {
                "type": "binary_op",
                "op": "+",
                "left": {"type": "literal", "value": "a"},
                "right": {"type": "literal", "value": "b"},
            },
            row,
            lambda c: row.get(c),
        )
    with pytest.raises(ProgrammingError, match="Unsupported arithmetic operator"):
        executor._eval_expression(
            {
                "type": "binary_op",
                "op": "%",
                "left": {"type": "literal", "value": 3},
                "right": {"type": "literal", "value": 2},
            },
            row,
            lambda c: row.get(c),
        )
    with pytest.raises(
        ProgrammingError, match="Aggregate expressions are not supported"
    ):
        executor._eval_expression(
            {"type": "aggregate", "func": "COUNT", "arg": "*"},
            row,
            lambda c: row.get(c),
        )
    with pytest.raises(ProgrammingError, match="Unsupported expression type"):
        executor._eval_expression({"type": "mystery"}, row, lambda c: row.get(c))

    assert (
        executor._evaluate_condition(
            row, {"column": "x", "operator": "NOT IN", "value": (1, 2)}
        )
        is None
    )  # SQL UNKNOWN
    assert (
        executor._evaluate_condition(
            row, {"column": "x", "operator": "NOT BETWEEN", "value": (1, 2)}
        )
        is None
    )  # SQL UNKNOWN
    assert (
        executor._evaluate_condition(
            {"x": 5}, {"column": "x", "operator": "NOT BETWEEN", "value": (None, 2)}
        )
        is None
    )  # SQL UNKNOWN
    with pytest.raises(DatabaseError, match="Unsupported LIKE pattern type"):
        executor._evaluate_condition(
            {"x": "abc"}, {"column": "x", "operator": "NOT LIKE", "value": 5}
        )

    backend.close()

def test_executor_aggregate_edge_paths(tmp_path: Path) -> None:
    backend = _make_join_workbook(tmp_path / "agg_edges.xlsx")
    executor = SharedExecutor(backend)

    with pytest.raises(DatabaseError, match=r"COUNT\(DISTINCT \*\) is not supported"):
        executor._compute_aggregate("COUNT", "*", [{"*": 1}], distinct=True)
    with pytest.raises(DatabaseError, match="Unsupported aggregate function"):
        executor._compute_aggregate("MEDIAN", "id", [{"id": 1}], distinct=False)
    assert executor._aggregate_spec_from_label("COUNT()") is None
    assert executor._aggregate_spec_from_label("SUM(DISTINCT t1.id)") is None

    with pytest.raises(DatabaseError, match="must appear in GROUP BY"):
        executor.execute(parse_sql("SELECT id, COUNT(*) FROM t1 GROUP BY name"))
    with pytest.raises(DatabaseError, match="must be a GROUP BY column"):
        executor.execute(
            parse_sql("SELECT name, COUNT(*) FROM t1 GROUP BY name HAVING id > 1")
        )

    backend.close()

def test_executor_compound_manual_description_mismatch(tmp_path: Path) -> None:
    backend = _make_join_workbook(tmp_path / "compound_desc_mismatch.xlsx")

    class _FakeExecutor(SharedExecutor):
        def __init__(self, backend_obj: OpenpyxlBackend) -> None:
            super().__init__(backend_obj)
            self._idx = 0

        def execute(self, parsed: dict[str, Any], **kwargs: Any) -> ExecutionResult:  # type: ignore[override]
            self._idx += 1
            if self._idx == 1:
                return ExecutionResult(
                    "SELECT", [(1,)], [("id", None, None, None, None, None, None)], 1
                )
            return ExecutionResult(
                "SELECT",
                [(1, "x")],
                [
                    ("id", None, None, None, None, None, None),
                    ("name", None, None, None, None, None, None),
                ],
                1,
            )

    fake = _FakeExecutor(backend)
    with pytest.raises(DatabaseError, match="matching column counts"):
        fake._execute_compound(
            {
                "queries": [{"action": "SELECT"}, {"action": "SELECT"}],
                "operators": ["UNION"],
            }
        )
    backend.close()

def test_executor_join_helper_edges_without_upsert(tmp_path: Path) -> None:
    backend = _make_join_workbook(tmp_path / "join_helpers.xlsx")
    executor = SharedExecutor(backend)

    seen: list[tuple[str, str, str]] = []

    def _validate(source: str, name: str, context: str) -> None:
        seen.append((source, name, context))

    SharedExecutor._validate_join_where_refs(
        {
            "conditions": [
                {
                    "column": {
                        "type": "unary_op",
                        "op": "-",
                        "operand": {"type": "column", "source": "a", "name": "id"},
                    },
                    "operator": "=",
                    "value": {
                        "type": "binary_op",
                        "op": "+",
                        "left": {"type": "column", "source": "b", "name": "id"},
                        "right": {"type": "literal", "value": 1},
                    },
                }
            ]
        },
        _validate,
    )
    assert ("a", "id", "WHERE") in seen
    assert ("b", "id", "WHERE") in seen

    with pytest.raises(DatabaseError, match="Unknown source reference"):
        executor._resolve_join_column({}, {"source": "missing", "name": "id"})
    flattened = executor._flatten_join_row({"a": {"id": 1}, "bad": 42})
    assert flattened == {"a.id": 1}
    backend.close()

def test_executor_sql_rendering_helper_branches() -> None:
    searched_case = {
        "type": "case",
        "mode": "searched",
        "whens": [
            {"condition": "bad", "result": {"type": "literal", "value": 1}},
            {
                "condition": {"column": "a.id", "operator": "=", "value": 1},
                "result": {"type": "literal", "value": 2},
            },
        ],
    }
    simple_case = {
        "type": "case",
        "mode": "simple",
        "value": {"type": "column", "source": "a", "name": "id"},
        "whens": [
            {
                "match": {"type": "literal", "value": 1},
                "result": {"type": "literal", "value": "x"},
            }
        ],
        "else": {"type": "literal", "value": "y"},
    }
    assert "CASE" in SharedExecutor._expression_to_sql(searched_case)
    assert "WHEN" in SharedExecutor._expression_to_sql(simple_case)
    assert SharedExecutor._expression_to_sql(123) == "123"
    assert (
        SharedExecutor._where_operand_to_sql(
            {"type": "literal", "value": 3}, is_column=False
        )
        == "3"
    )
    assert SharedExecutor._where_to_sql({"conditions": []}) == ""
    assert "AND" in SharedExecutor._where_to_sql(
        {
            "conditions": [
                {"column": "a.id", "operator": "=", "value": 1},
                {"column": "a.id", "operator": "=", "value": 2},
            ],
            "conjunctions": ["AND"],
        }
    )
    assert "(SUBQUERY)" in SharedExecutor._where_to_sql(
        {"column": "a.id", "operator": "IN", "value": {"type": "subquery"}}
    )
    assert "(3)" in SharedExecutor._where_to_sql(
        {"column": "a.id", "operator": "NOT IN", "value": 3}
    )
    assert "AND" in SharedExecutor._where_to_sql(
        {"column": "a.id", "operator": "BETWEEN", "value": [1, 5]}
    )
    assert SharedExecutor._contains_arithmetic_expression(
        {
            "type": "alias",
            "expression": {
                "type": "alias",
                "expression": {"type": "literal", "value": 1},
            },
        }
    )
    refs = SharedExecutor._collect_expression_column_refs(
        {"type": "alias", "expression": {"type": "column", "source": "a", "name": "id"}}
    )
    assert refs == {"a.id"}

def test_parser_join_validation_and_expression_to_sql_branches() -> None:
    allowed = {"a", "b"}
    _validate_join_column_reference(
        {
            "type": "case",
            "mode": "simple",
            "value": {"type": "column", "source": "a", "name": "id"},
            "whens": [
                "skip",
                {
                    "match": {"type": "column", "source": "a", "name": "id"},
                    "result": {"type": "column", "source": "b", "name": "id"},
                },
            ],
            "else": {"type": "column", "source": "a", "name": "id"},
        },
        allowed,
        "SELECT",
    )
    _validate_join_column_reference(
        {
            "type": "case",
            "mode": "searched",
            "whens": [
                {
                    "condition": {
                        "column": "a.id",
                        "operator": "IN",
                        "value": ({"type": "column", "source": "b", "name": "id"},),
                    },
                    "result": {"type": "column", "source": "a", "name": "id"},
                }
            ],
        },
        allowed,
        "SELECT",
    )
    _validate_join_where_node(
        {
            "column": "a.id",
            "operator": "IN",
            "value": ({"type": "column", "source": "b", "name": "id"},),
        },
        allowed,
    )

    assert (
        _expression_to_sql_for_order_by(
            {"type": "aggregate", "func": "COUNT", "arg": "*"}
        )
        == "COUNT(*)"
    )
    assert _expression_to_sql_for_order_by({"type": "literal", "value": "x"}) == "'x'"
    assert (
        _expression_to_sql_for_order_by(
            {"type": "unary_op", "operand": {"type": "literal", "value": 1}}
        )
        == "-1"
    )
    assert (
        _expression_to_sql_for_order_by(
            {
                "type": "binary_op",
                "op": "+",
                "left": {"type": "literal", "value": 1},
                "right": {"type": "literal", "value": 2},
            }
        )
        == "(1 + 2)"
    )



class MemoryBackend(WorkbookBackend):
    def __init__(self, sheets: dict[str, TableData], readonly: bool = False) -> None:
        super().__init__("memory.xlsx")
        self._sheets = sheets
        self._readonly = readonly

    @property
    def readonly(self) -> bool:
        return self._readonly

    @property
    def supports_transactions(self) -> bool:
        return True


    def load(self) -> None:
        return None

    def save(self) -> None:
        return None

    def snapshot(self) -> dict[str, TableData]:
        return self._sheets

    def restore(self, snapshot: Any) -> None:
        self._sheets = snapshot

    def list_sheets(self) -> list[str]:
        return list(self._sheets.keys())

    def read_sheet(self, sheet_name: str) -> TableData:
        if sheet_name not in self._sheets:
            raise ValueError(f"Sheet '{sheet_name}' not found in Excel")
        return self._sheets[sheet_name]

    def write_sheet(self, sheet_name: str, data: TableData) -> None:
        self._sheets[sheet_name] = data

    def append_row(self, sheet_name: str, row: list[Any]) -> int:
        self._sheets[sheet_name].rows.append(row)
        return len(self._sheets[sheet_name].rows) + 1

    def create_sheet(self, name: str, headers: list[str]) -> None:
        self._sheets[name] = TableData(headers=headers, rows=[])

    def drop_sheet(self, name: str) -> None:
        del self._sheets[name]

def _xlsx(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    ws.append(["id", "name"])
    ws.append([1, "Alice"])
    wb.save(path)

def test_executor_error_paths_and_utility_paths() -> None:
    backend = MemoryBackend(
        {
            "Users": TableData(headers=["id", "name", "extra"], rows=[[1, "A"]]),
            "Empty": TableData(headers=[], rows=[]),
        }
    )
    executor = SharedExecutor(backend)

    with pytest.raises(DatabaseError, match="Sheet 'Missing' not found"):
        executor.execute(
            {
                "action": "UPDATE",
                "table": "Missing",
                "set": [{"column": "name", "value": "B"}],
                "where": None,
            }
        )

    with pytest.raises(DatabaseError, match="No columns defined in sheet 'Empty'"):
        executor.execute(
            {
                "action": "UPDATE",
                "table": "Empty",
                "set": [{"column": "name", "value": "B"}],
                "where": None,
            }
        )

    with pytest.raises(DatabaseError, match="Unknown column"):
        executor.execute(
            {
                "action": "UPDATE",
                "table": "Users",
                "set": [{"column": "missing", "value": "B"}],
                "where": None,
            }
        )

    updated = executor.execute(
        {
            "action": "UPDATE",
            "table": "Users",
            "set": [{"column": "extra", "value": "Z"}],
            "where": {
                "conditions": [{"column": "id", "operator": "=", "value": 1}],
                "conjunctions": [],
            },
        }
    )
    assert updated.rowcount == 1
    assert backend.read_sheet("Users").rows[0] == [1, "A", "Z"]

    with pytest.raises(DatabaseError, match="Sheet 'Missing' not found"):
        executor.execute({"action": "DELETE", "table": "Missing", "where": None})

    delete_empty = executor.execute(
        {"action": "DELETE", "table": "Empty", "where": None}
    )
    assert delete_empty.rowcount == 0

    with pytest.raises(DatabaseError, match="Sheet 'Missing' not found"):
        executor.execute(
            {
                "action": "INSERT",
                "table": "Missing",
                "columns": None,
                "values": [[1]],
            }
        )

    with pytest.raises(DatabaseError, match="without headers"):
        executor.execute(
            {
                "action": "INSERT",
                "table": "Empty",
                "columns": None,
                "values": [[1]],
            }
        )

    with pytest.raises(DatabaseError, match="header count"):
        executor.execute(
            {
                "action": "INSERT",
                "table": "Users",
                "columns": None,
                "values": [[1]],
            }
        )

    with pytest.raises(DatabaseError, match="Unknown column"):
        executor.execute(
            {
                "action": "INSERT",
                "table": "Users",
                "columns": ["missing"],
                "values": [[1]],
            }
        )

    assert (
        executor._matches_where(
            {"id": 1}, {"column": "id", "operator": "=", "value": 1}
        )
        is True
    )
    assert (
        executor._evaluate_condition(
            {"id": None}, {"column": "id", "operator": "BETWEEN", "value": (1, 2)}
        )
        is None  # SQL UNKNOWN
    )
    assert (
        executor._evaluate_condition(
            {"id": 5}, {"column": "id", "operator": "BETWEEN", "value": (None, 9)}
        )
        is None  # SQL UNKNOWN
    )
    assert (
        executor._evaluate_condition(
            {"id": 1}, {"column": "id", "operator": "<", "value": 2}
        )
        is True
    )
    with pytest.raises(DatabaseError, match="Unsupported operator"):
        executor._evaluate_condition(
            {"id": 1}, {"column": "id", "operator": "~~~", "value": 2}
        )
    assert executor._sort_key(None) == (1, (0, ""))
    assert executor._to_number(True) is None
    assert executor._to_number({"x": 1}) is None

    with pytest.raises(DatabaseError, match="Unsupported action"):
        executor.execute({"action": "BOOM", "table": "Users"})
