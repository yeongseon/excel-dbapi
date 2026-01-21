from typing import Any, Dict, Optional
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook

from .base import BaseEngine
from .executor import execute_query
from .result import ExecutionResult
from .parser import parse_sql


class OpenpyxlEngine(BaseEngine):
    """
    OpenpyxlEngine is responsible for loading, executing, and saving Excel data
    using the openpyxl library. It implements the BaseEngine interface.
    """

    def __init__(self, file_path: str):
        """
        Initialize the OpenpyxlEngine with the given file path.

        Args:
            file_path (str): Path to the Excel (.xlsx) file.
        """
        self.file_path = file_path
        self.workbook: Workbook | None = None
        self.data = self.load()

    def load(self) -> Dict[str, Any]:
        """
        Load the Excel workbook into memory.

        Returns:
            Dict[str, Any]: A dictionary mapping sheet names to openpyxl Worksheet objects.
        """
        self.workbook = load_workbook(self.file_path, data_only=True)
        return {sheet: self.workbook[sheet] for sheet in self.workbook.sheetnames}

    def save(self) -> None:
        """
        Save any in-memory changes back to the Excel file.

        Notes:
            - This method reloads the workbook, updates each worksheet's cells
              with the current in-memory data, and saves it back to the file.
            - Intended for future use when INSERT, UPDATE, DELETE operations are supported.
        """
        if self.workbook is None:
            raise ValueError("Workbook is not loaded")
        self.workbook.save(self.file_path)

    def snapshot(self) -> BytesIO:
        if self.workbook is None:
            raise ValueError("Workbook is not loaded")
        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer

    def restore(self, snapshot: BytesIO) -> None:
        snapshot.seek(0)
        self.workbook = load_workbook(snapshot, data_only=True)
        self.data = {sheet: self.workbook[sheet] for sheet in self.workbook.sheetnames}

    def execute(self, query: str) -> ExecutionResult:
        """
        Execute a SQL-like query on the loaded Excel data.

        Args:
            query (str): A SQL-like query string (e.g., "SELECT * FROM Sheet1 WHERE id = '1'").

        Returns:
            ExecutionResult: Query results with rows, description, and rowcount.
        """
        parsed = parse_sql(query)
        return execute_query(parsed, self.data, self.workbook)

    def execute_with_params(self, query: str, params: Optional[tuple] = None) -> ExecutionResult:
        parsed = parse_sql(query, params)
        return execute_query(parsed, self.data, self.workbook)
